"""
Endpoints do módulo AJUS — catálogo de códigos de andamento + fila.

Auth: JWT obrigatório, permissão `prazos_iniciais` (mesmo escopo do
intake — quem trata prazos é quem decide quais andamentos vão pra AJUS).
Operações de catálogo (CRUD de cod_andamento) também respeitam a
permissão; operações destrutivas (delete) podem ser restritas a admin
no futuro.
"""

from __future__ import annotations

import io
import logging
from datetime import date, time
from pathlib import Path as FilePath
from typing import Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Path as FastapiPath,
    Query,
    UploadFile,
)
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core import auth as auth_security
from app.core.dependencies import get_db
from app.models.ajus import (
    AJUS_ACCOUNT_EXECUTANDO,
    AJUS_ACCOUNT_ONLINE,
    AJUS_CLASSIF_PENDENTE,
    AJUS_CLASSIF_PROCESSANDO,
    AJUS_QUEUE_PENDENTE,
    AJUS_QUEUE_STATUSES,
    AjusAndamentoQueue,
    AjusClassificacaoQueue,
    AjusClassificationBlocklist,
    AjusCodAndamento,
    AjusSessionAccount,
)
from app.services.ajus.crypto import is_configured as ajus_crypto_configured
from app.services.ajus.session_service import (
    AjusSessionService,
    has_storage_state,
)
from app.models.legal_one import LegalOneUser
from app.services.ajus.classificacao_service import (
    XLSX_HEADERS,
    AjusClassificacaoService,
    XlsxRow,
)
from app.services.ajus.legal_one_export import (
    convert_legal_one_export_to_xlsx_rows,
    is_legal_one_export,
)
from app.services.ajus.queue_service import (
    MAX_ITENS_POR_REQUEST,
    AjusQueueService,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ajus", tags=["AJUS"])


# ═══════════════════════════════════════════════════════════════════════
# Schemas Pydantic
# ═══════════════════════════════════════════════════════════════════════


class CodAndamentoIn(BaseModel):
    codigo: str = Field(..., max_length=64)
    label: str = Field(..., max_length=200)
    descricao: Optional[str] = None
    situacao: str = Field(default="A", pattern="^[AC]$")
    dias_agendamento_offset_uteis: int = Field(default=3, ge=-365, le=365)
    dias_fatal_offset_uteis: int = Field(default=15, ge=-365, le=365)
    informacao_template: str = Field(default="Andamento — processo {cnj}.")
    is_default: bool = False
    # Pin019: marcar este código como o usado pelo fluxo de devolução
    # automática (POST /prazos-iniciais/intake/devolucao). Apenas 1 ativo
    # pode ter is_devolucao=true por vez.
    is_devolucao: bool = False
    is_active: bool = True


class CodAndamentoOut(BaseModel):
    id: int
    codigo: str
    label: str
    descricao: Optional[str]
    situacao: str
    dias_agendamento_offset_uteis: int
    dias_fatal_offset_uteis: int
    informacao_template: str
    is_default: bool
    is_devolucao: bool
    is_active: bool

    class Config:
        from_attributes = True


class AndamentoQueueOut(BaseModel):
    id: int
    intake_id: int
    cnj_number: str
    cod_andamento_id: int
    cod_andamento_codigo: Optional[str] = None
    cod_andamento_label: Optional[str] = None
    situacao: str
    data_evento: date
    data_agendamento: date
    data_fatal: date
    hora_agendamento: Optional[time]
    informacao: str
    has_pdf: bool
    status: str
    cod_informacao_judicial: Optional[str]
    error_message: Optional[str]
    created_at: str
    dispatched_at: Optional[str]
    # Marca o item como bloqueado por classificacao pendente (CNJ esta no
    # ajus_classification_blocklist). UI usa pra desabilitar acoes de
    # disparo / mostrar badge "Class. pendente".
    is_blocked_classification: bool = False


class AndamentoQueueListResponse(BaseModel):
    total: int
    items: list[AndamentoQueueOut]


class DispatchBatchResponse(BaseModel):
    candidates: int
    success_count: int
    error_count: int
    success_ids: list[int]
    errored: list[dict]


class BackfillCompletedRequest(BaseModel):
    """
    Filtros opcionais pro backfill da fila AJUS a partir dos intakes
    de prazos iniciais ja' classificados (qualquer status pos-classificacao,
    incluindo erros e concluidos).
    """
    statuses: Optional[list[str]] = Field(
        default=None,
        description=(
            "Lista de status do intake. None = usa o set padrao "
            "(qualquer status pos-classificacao)."
        ),
    )
    from_date: Optional[date] = Field(
        default=None,
        description="Filtra intakes criados a partir desta data (inclusivo).",
    )
    to_date: Optional[date] = Field(
        default=None,
        description="Filtra intakes criados ate esta data (inclusivo).",
    )
    dry_run: bool = Field(
        default=False,
        description="Se True, so' conta candidatos sem mexer na fila.",
    )
    limit: Optional[int] = Field(
        default=None,
        ge=1,
        le=10000,
        description="Limite de intakes processados nessa chamada.",
    )


class BackfillNoPdfItem(BaseModel):
    intake_id: int
    cnj_number: str


class BackfillCompletedResponse(BaseModel):
    candidates: int
    enqueued: int
    skipped_already: int
    skipped_other: int
    intake_ids_enqueued: list[int]
    enqueued_without_pdf: list[BackfillNoPdfItem]
    dry_run: bool
    error: Optional[str] = None


# ═══════════════════════════════════════════════════════════════════════
# Helpers de serialização
# ═══════════════════════════════════════════════════════════════════════


def _queue_to_out(
    item: AjusAndamentoQueue,
    blocked_cnj_digits: Optional[set[str]] = None,
) -> AndamentoQueueOut:
    cod = item.cod_andamento
    digits = "".join(c for c in (item.cnj_number or "") if c.isdigit())
    is_blocked = bool(blocked_cnj_digits) and digits in (blocked_cnj_digits or set())
    return AndamentoQueueOut(
        id=item.id,
        intake_id=item.intake_id,
        cnj_number=item.cnj_number,
        cod_andamento_id=item.cod_andamento_id,
        cod_andamento_codigo=cod.codigo if cod else None,
        cod_andamento_label=cod.label if cod else None,
        situacao=item.situacao,
        data_evento=item.data_evento,
        data_agendamento=item.data_agendamento,
        data_fatal=item.data_fatal,
        hora_agendamento=item.hora_agendamento,
        informacao=item.informacao,
        has_pdf=bool(item.pdf_path),
        status=item.status,
        cod_informacao_judicial=item.cod_informacao_judicial,
        error_message=item.error_message,
        created_at=item.created_at.isoformat() if item.created_at else "",
        dispatched_at=(
            item.dispatched_at.isoformat() if item.dispatched_at else None
        ),
        is_blocked_classification=is_blocked,
    )


# ═══════════════════════════════════════════════════════════════════════
# Catálogo de códigos de andamento (CRUD)
# ═══════════════════════════════════════════════════════════════════════


@router.get("/cod-andamento", response_model=list[CodAndamentoOut])
def list_cod_andamento(
    only_active: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    q = db.query(AjusCodAndamento)
    if only_active:
        q = q.filter(AjusCodAndamento.is_active.is_(True))
    return [
        CodAndamentoOut.model_validate(c)
        for c in q.order_by(AjusCodAndamento.codigo.asc()).all()
    ]


@router.post(
    "/cod-andamento", response_model=CodAndamentoOut, status_code=201,
)
def create_cod_andamento(
    payload: CodAndamentoIn,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    # Se está marcado como default, garante que nenhum outro fique default
    if payload.is_default:
        db.query(AjusCodAndamento).filter(
            AjusCodAndamento.is_default.is_(True),
        ).update({"is_default": False})
    # Idem pra is_devolucao — apenas 1 ativo (pin019).
    if payload.is_devolucao:
        db.query(AjusCodAndamento).filter(
            AjusCodAndamento.is_devolucao.is_(True),
            AjusCodAndamento.is_active.is_(True),
        ).update({"is_devolucao": False})
    obj = AjusCodAndamento(**payload.model_dump())
    db.add(obj)
    try:
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail=f"Falha ao criar código (provavelmente duplicado): {exc}",
        ) from exc
    db.refresh(obj)
    return CodAndamentoOut.model_validate(obj)


@router.put("/cod-andamento/{cod_id}", response_model=CodAndamentoOut)
def update_cod_andamento(
    payload: CodAndamentoIn,
    cod_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = db.get(AjusCodAndamento, cod_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Código não encontrado.")
    if payload.is_default and not obj.is_default:
        # Garante que só um seja default
        db.query(AjusCodAndamento).filter(
            AjusCodAndamento.is_default.is_(True),
            AjusCodAndamento.id != cod_id,
        ).update({"is_default": False})
    # Idem pra is_devolucao — só um ativo por vez (pin019).
    if payload.is_devolucao and not obj.is_devolucao:
        db.query(AjusCodAndamento).filter(
            AjusCodAndamento.is_devolucao.is_(True),
            AjusCodAndamento.is_active.is_(True),
            AjusCodAndamento.id != cod_id,
        ).update({"is_devolucao": False})
    for k, v in payload.model_dump().items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return CodAndamentoOut.model_validate(obj)


@router.delete("/cod-andamento/{cod_id}", status_code=204)
def delete_cod_andamento(
    cod_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = db.get(AjusCodAndamento, cod_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Código não encontrado.")
    # Bloqueia delete se há itens em fila usando — mantém integridade
    in_use = (
        db.query(AjusAndamentoQueue)
        .filter(AjusAndamentoQueue.cod_andamento_id == cod_id)
        .first()
    )
    if in_use is not None:
        raise HTTPException(
            status_code=409,
            detail=(
                "Código está em uso por itens da fila — não é possível deletar. "
                "Desative com `is_active=false` se quiser parar de usá-lo."
            ),
        )
    db.delete(obj)
    db.commit()
    return None


# ═══════════════════════════════════════════════════════════════════════
# Fila de andamentos
# ═══════════════════════════════════════════════════════════════════════


@router.get("/andamentos", response_model=AndamentoQueueListResponse)
def list_andamentos(
    status: Optional[str] = Query(default=None, description="CSV de status."),
    cnj_number: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    q = db.query(AjusAndamentoQueue)
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        invalid = set(statuses) - AJUS_QUEUE_STATUSES
        if invalid:
            raise HTTPException(
                status_code=422,
                detail=f"Status inválido(s): {sorted(invalid)}",
            )
        if len(statuses) == 1:
            q = q.filter(AjusAndamentoQueue.status == statuses[0])
        else:
            q = q.filter(AjusAndamentoQueue.status.in_(statuses))
    if cnj_number:
        normalized = "".join(c for c in cnj_number if c.isdigit())
        if normalized:
            q = q.filter(AjusAndamentoQueue.cnj_number.like(f"%{normalized}%"))

    total = q.count()
    items = (
        q.order_by(AjusAndamentoQueue.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    # Pre-computa CNJs bloqueados por classificacao pendente (1 query soh
    # com IN-list dos digitos da pagina atual). Frontend usa pra renderizar
    # badge e desabilitar botoes de disparo nos items afetados.
    from app.models.ajus import AjusClassificationBlocklist
    digit_set: set[str] = set()
    for it in items:
        d = "".join(c for c in (it.cnj_number or "") if c.isdigit())
        if d:
            digit_set.add(d)
    blocked_digits: set[str] = set()
    if digit_set:
        rows = (
            db.query(AjusClassificationBlocklist.cnj_number)
            .filter(AjusClassificationBlocklist.cnj_number.in_(digit_set))
            .all()
        )
        blocked_digits = {r[0] for r in rows}
    return AndamentoQueueListResponse(
        total=total,
        items=[_queue_to_out(i, blocked_digits) for i in items],
    )


@router.post(
    "/andamentos/dispatch-pending",
    response_model=DispatchBatchResponse,
    summary=(
        "Dispara em lote os próximos N itens em status 'pendente'. "
        f"Limite máximo por chamada: {MAX_ITENS_POR_REQUEST} (limite AJUS)."
    ),
)
def dispatch_pending(
    batch_limit: int = Query(default=MAX_ITENS_POR_REQUEST, ge=1, le=MAX_ITENS_POR_REQUEST),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    result = service.dispatch_pending_batch(batch_limit=batch_limit)
    return DispatchBatchResponse(**result)


@router.post(
    "/andamentos/backfill-from-intakes",
    response_model=BackfillCompletedResponse,
    summary=(
        "Enfileira no AJUS os intakes de prazos iniciais ja' classificados "
        "que ainda nao tem item na fila. Inclui status de erro pos-classificacao "
        "e finalizados. Intakes sem PDF entram na fila marcados pra anexo manual. "
        "Idempotente -- rodar 2x nao duplica. Use `dry_run=true` pra preview."
    ),
)
def backfill_from_intakes(
    payload: BackfillCompletedRequest = BackfillCompletedRequest(),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    result = service.backfill_completed_intakes(
        statuses=payload.statuses,
        from_date=payload.from_date,
        to_date=payload.to_date,
        dry_run=payload.dry_run,
        limit=payload.limit,
    )
    return BackfillCompletedResponse(**result)



class DispatchSelectedRequest(BaseModel):
    item_ids: list[int] = Field(
        ...,
        min_length=1,
        max_length=MAX_ITENS_POR_REQUEST,
        description=(
            f"Ids dos itens a disparar (max {MAX_ITENS_POR_REQUEST} por "
            "request, limite AJUS). Todos devem estar em pendente ou erro."
        ),
    )


@router.post(
    "/andamentos/dispatch-selected",
    response_model=DispatchBatchResponse,
    summary=(
        "Dispatcha em UMA request um conjunto de itens escolhidos pelo "
        f"operador. Limite: {MAX_ITENS_POR_REQUEST} itens por chamada."
    ),
)
def dispatch_selected_andamentos(
    payload: DispatchSelectedRequest,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    try:
        result = service.dispatch_selected(payload.item_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return DispatchBatchResponse(**result)


@router.get(
    "/andamentos/{item_id}/pdf",
    summary=(
        "Devolve o PDF da habilitacao anexado ao item. 404 se item nao "
        "existe; 410 se item nao tem PDF anexado (ainda)."
    ),
)
def download_andamento_pdf(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    try:
        pdf_bytes, filename = service.get_pdf_bytes(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except FileNotFoundError as exc:
        # 410 Gone -- semanticamente correto: existe o item, mas o anexo
        # nao esta disponivel. Frontend pode tratar como "anexar via upload".
        raise HTTPException(status_code=410, detail=str(exc)) from exc
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@router.post(
    "/andamentos/{item_id}/pdf",
    response_model=AndamentoQueueOut,
    summary=(
        "Anexa um PDF a um item existente que nao tem anexo. Limpa o "
        "prefixo de 'sem anexo' do informacao e, se em ERRO, volta pra "
        "PENDENTE pra re-disparar."
    ),
)
async def upload_andamento_pdf(
    item_id: int = FastapiPath(..., ge=1),
    file: UploadFile = File(..., description="PDF da habilitacao."),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        pdf_bytes = await file.read()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"Falha ao ler arquivo: {exc}",
        ) from exc
    if len(pdf_bytes) > _BULK_MAX_FILE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=(
                f"Arquivo excede 10MB (limite AJUS): {len(pdf_bytes)} bytes."
            ),
        )
    service = AjusQueueService(db)
    try:
        item = service.attach_pdf(item_id, pdf_bytes)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except OSError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Falha ao gravar PDF no storage: {exc}",
        ) from exc
    return _queue_to_out(item)


@router.post("/andamentos/{item_id}/cancel", response_model=AndamentoQueueOut)
def cancel_andamento(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    try:
        item = service.cancel(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _queue_to_out(item)


@router.post("/andamentos/{item_id}/retry", response_model=AndamentoQueueOut)
def retry_andamento(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusQueueService(db)
    try:
        item = service.retry(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _queue_to_out(item)


class DispatchOneResponse(BaseModel):
    """Resposta do dispatch pontual (1 item).

    `success` traduz o status_final pra um booleano amigavel —
    True quando AJUS confirmou insercao (`inserido=true`), False
    nos demais casos. `msg` carrega a mensagem do erro AJUS quando
    success=False.
    """
    item_id: int
    status_final: str
    success: bool
    msg: Optional[str] = None
    cod_informacao_judicial: Optional[str] = None


@router.post(
    "/andamentos/{item_id}/dispatch",
    response_model=DispatchOneResponse,
    summary=(
        "Dispatcha 1 item pontual da fila pro AJUS — debug-friendly. "
        "Aceita item em status pendente ou erro; bloqueia sucesso/cancelado/enviando."
    ),
)
def dispatch_one_andamento(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Util pra:
      - Testar credenciais AJUS com 1 caso conhecido antes de soltar lote.
      - Re-disparar pontual depois de operador corrigir manualmente
        algum dado do item (ex.: data_evento errada).
      - Debug operacional ("processar este aqui agora, não esperar fila").

    HTTP 502 quando a chamada AJUS em si falha (token invalido, 5xx do
    AJUS, timeout). HTTP 409 quando status do item nao permite dispatch.
    """
    from app.services.ajus.ajus_client import AjusApiError, AjusConfigError

    service = AjusQueueService(db)
    try:
        result = service.dispatch_one(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except (AjusConfigError, AjusApiError) as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return DispatchOneResponse(**result)


# ═══════════════════════════════════════════════════════════════════════
# Upload em lote (manual) — arquivos com CNJ no titulo OU lista de CNJs
# ═══════════════════════════════════════════════════════════════════════
# Permite enfileirar N andamentos AJUS sem passar por intake. Dois modos:
#   1) bulk-upload (multipart): N PDFs, CNJ extraido do nome do arquivo,
#      o PDF vira anexo do andamento.
#   2) bulk-cnj (JSON): lista de CNJs sem arquivo (andamento de capa).
# Variaveis comuns (cod_andamento_id, situacao, datas, hora, template
# de informacao) se aplicam ao lote inteiro. Itens criados com
# intake_id=NULL e status=pendente; aparecem na fila normal e o
# operador clica "Enviar próximos 20" como sempre.


class BulkAndamentoVarsBase(BaseModel):
    cod_andamento_id: int = Field(..., ge=1)
    situacao: Optional[str] = Field(default=None, pattern="^[AC]$")
    data_evento: Optional[date] = None
    data_agendamento: Optional[date] = None
    data_fatal: Optional[date] = None
    hora_agendamento: Optional[time] = None
    informacao_template_override: Optional[str] = None


class BulkCnjIn(BulkAndamentoVarsBase):
    cnj_list: list[str] = Field(
        ...,
        min_length=1,
        max_length=500,
        description=(
            "Lista de CNJs (com ou sem mascara). Cada item gera 1 "
            "andamento sem anexo de PDF."
        ),
    )


class BulkAndamentoSkipped(BaseModel):
    cnj: str
    filename: Optional[str] = None
    reason: str


class BulkAndamentoResponse(BaseModel):
    created: int
    updated: int = 0
    skipped: list[BulkAndamentoSkipped]
    item_ids: list[int]
    created_ids: list[int] = []
    updated_ids: list[int] = []


def _resolve_cod_andamento_or_404(
    db: Session, cod_andamento_id: int,
) -> AjusCodAndamento:
    cod = db.get(AjusCodAndamento, cod_andamento_id)
    if cod is None:
        raise HTTPException(
            status_code=404,
            detail=f"Codigo de andamento {cod_andamento_id} nao encontrado.",
        )
    if not cod.is_active:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Codigo de andamento {cod.codigo} esta inativo. "
                "Reative-o ou escolha outro."
            ),
        )
    return cod


# Limite local: AJUS aceita 10MB por arquivo, mantemos o mesmo aqui.
_BULK_MAX_FILE_BYTES = 10 * 1024 * 1024


@router.post(
    "/andamentos/bulk-upload",
    response_model=BulkAndamentoResponse,
    summary=(
        "Upload em lote de PDFs com CNJ no nome do arquivo. Cria 1 item "
        "na fila por arquivo, com PDF anexado e variaveis comuns ao lote."
    ),
)
async def bulk_upload_andamentos(
    files: list[UploadFile] = File(..., description="PDFs com CNJ no nome."),
    cod_andamento_id: int = Form(..., ge=1),
    situacao: Optional[str] = Form(default=None, pattern="^[AC]$"),
    data_evento: Optional[date] = Form(default=None),
    data_agendamento: Optional[date] = Form(default=None),
    data_fatal: Optional[date] = Form(default=None),
    hora_agendamento: Optional[time] = Form(default=None),
    informacao_template_override: Optional[str] = Form(default=None),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Multipart: variaveis em form fields, arquivos em `files`. CNJ
    extraido do nome de cada arquivo (regex CNJ; aceita com ou sem
    mascara). Arquivo com CNJ nao encontrado vira `skipped`.
    Tamanho > 10MB (limite AJUS) tambem skipa.

    HTTP 404 se cod_andamento_id nao existe; 409 se inativo.
    """
    from app.services.ajus.queue_service import (
        AjusQueueService,
        extract_cnj_from_filename,
    )

    if not files:
        raise HTTPException(status_code=400, detail="Nenhum arquivo enviado.")

    cod = _resolve_cod_andamento_or_404(db, cod_andamento_id)

    pre_skipped: list[dict] = []
    valid_entries: list[dict] = []
    for upload in files:
        filename = upload.filename or ""
        cnj = extract_cnj_from_filename(filename)
        try:
            content = await upload.read()
        except Exception as exc:  # noqa: BLE001
            pre_skipped.append({
                "cnj": cnj or "",
                "filename": filename,
                "reason": f"Falha ao ler arquivo: {exc}",
            })
            continue
        if len(content) > _BULK_MAX_FILE_BYTES:
            pre_skipped.append({
                "cnj": cnj or "",
                "filename": filename,
                "reason": (
                    f"Arquivo excede 10MB (limite AJUS): {len(content)} bytes."
                ),
            })
            continue
        valid_entries.append({
            "cnj": cnj or "",
            "filename": filename,
            "pdf_bytes": content,
        })

    service = AjusQueueService(db)
    result = service.bulk_enqueue(
        cod_andamento=cod,
        entries=valid_entries,
        situacao=situacao,
        data_evento=data_evento,
        data_agendamento=data_agendamento,
        data_fatal=data_fatal,
        hora_agendamento=hora_agendamento,
        informacao_template_override=informacao_template_override,
    )
    return BulkAndamentoResponse(
        created=result["created"],
        updated=result.get("updated", 0),
        skipped=[
            BulkAndamentoSkipped(**s) for s in pre_skipped + result["skipped"]
        ],
        item_ids=result["item_ids"],
        created_ids=result.get("created_ids", []),
        updated_ids=result.get("updated_ids", []),
    )


@router.post(
    "/andamentos/bulk-cnj",
    response_model=BulkAndamentoResponse,
    summary=(
        "Cria N itens da fila a partir de uma lista de CNJs (sem anexo)."
    ),
)
def bulk_cnj_andamentos(
    payload: BulkCnjIn,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus.queue_service import AjusQueueService

    cod = _resolve_cod_andamento_or_404(db, payload.cod_andamento_id)

    entries: list[dict] = [
        {"cnj": raw.strip(), "filename": None, "pdf_bytes": None}
        for raw in payload.cnj_list if raw and raw.strip()
    ]
    if not entries:
        raise HTTPException(
            status_code=400,
            detail="Lista de CNJs vazia (apos limpar linhas em branco).",
        )

    service = AjusQueueService(db)
    result = service.bulk_enqueue(
        cod_andamento=cod,
        entries=entries,
        situacao=payload.situacao,
        data_evento=payload.data_evento,
        data_agendamento=payload.data_agendamento,
        data_fatal=payload.data_fatal,
        hora_agendamento=payload.hora_agendamento,
        informacao_template_override=payload.informacao_template_override,
    )
    return BulkAndamentoResponse(
        created=result["created"],
        updated=result.get("updated", 0),
        skipped=[BulkAndamentoSkipped(**s) for s in result["skipped"]],
        item_ids=result["item_ids"],
        created_ids=result.get("created_ids", []),
        updated_ids=result.get("updated_ids", []),
    )


# ═══════════════════════════════════════════════════════════════════════
# Classificação (Chunk 1 — fila + defaults; Playwright runner no Chunk 2)
# ═══════════════════════════════════════════════════════════════════════


class ClassifDefaultsIn(BaseModel):
    default_matter: Optional[str] = Field(default=None, max_length=255)
    default_risk_loss_probability: Optional[str] = Field(
        default=None, max_length=255,
    )


class ClassifDefaultsOut(BaseModel):
    default_matter: Optional[str]
    default_risk_loss_probability: Optional[str]
    updated_at: Optional[str] = None
    is_paused: bool = False
    paused_at: Optional[str] = None
    paused_by: Optional[str] = None


class ClassifQueueOut(BaseModel):
    id: int
    cnj_number: str
    intake_id: Optional[int]
    origem: str
    uf: Optional[str]
    comarca: Optional[str]
    matter: Optional[str]
    justice_fee: Optional[str]
    risk_loss_probability: Optional[str]
    status: str
    error_message: Optional[str]
    last_log: Optional[str]
    executed_at: Optional[str]
    created_at: Optional[str]
    updated_at: Optional[str]
    # Conta que claimou esse item (preserva mesmo apos sucesso/erro pra
    # frontend conseguir buscar screenshots de debug por item).
    dispatched_by_account_id: Optional[int] = None


class ClassifQueueListResponse(BaseModel):
    total: int
    items: list[ClassifQueueOut]


class ClassifQueueUpdateIn(BaseModel):
    uf: Optional[str] = Field(default=None, max_length=8)
    comarca: Optional[str] = Field(default=None, max_length=255)
    matter: Optional[str] = Field(default=None, max_length=255)
    justice_fee: Optional[str] = Field(default=None, max_length=255)
    risk_loss_probability: Optional[str] = Field(default=None, max_length=255)


class ClassifUploadResponse(BaseModel):
    created: int
    updated: int
    skipped: list[dict]


def _classif_to_out(item: AjusClassificacaoQueue) -> ClassifQueueOut:
    return ClassifQueueOut(
        id=item.id,
        cnj_number=item.cnj_number,
        intake_id=item.intake_id,
        origem=item.origem,
        uf=item.uf,
        comarca=item.comarca,
        matter=item.matter,
        justice_fee=item.justice_fee,
        risk_loss_probability=item.risk_loss_probability,
        status=item.status,
        error_message=item.error_message,
        last_log=item.last_log,
        executed_at=item.executed_at.isoformat() if item.executed_at else None,
        created_at=item.created_at.isoformat() if item.created_at else None,
        updated_at=item.updated_at.isoformat() if item.updated_at else None,
        dispatched_by_account_id=item.dispatched_by_account_id,
    )


# ── Defaults singleton ──────────────────────────────────────────────


@router.get("/classificacao/defaults", response_model=ClassifDefaultsOut)
def get_classif_defaults(
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = AjusClassificacaoService(db).get_defaults()
    return _defaults_to_out(obj)


@router.put("/classificacao/defaults", response_model=ClassifDefaultsOut)
def update_classif_defaults(
    payload: ClassifDefaultsIn,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = AjusClassificacaoService(db).update_defaults(
        default_matter=payload.default_matter,
        default_risk_loss_probability=payload.default_risk_loss_probability,
    )
    return _defaults_to_out(obj)


def _defaults_to_out(obj) -> ClassifDefaultsOut:
    return ClassifDefaultsOut(
        default_matter=obj.default_matter,
        default_risk_loss_probability=obj.default_risk_loss_probability,
        updated_at=obj.updated_at.isoformat() if obj.updated_at else None,
        is_paused=bool(obj.is_paused),
        paused_at=obj.paused_at.isoformat() if obj.paused_at else None,
        paused_by=obj.paused_by,
    )


# ── Pause / Resume / Cancelar pendentes (controle global) ──────────


class ClassifPauseIn(BaseModel):
    paused: bool = True


class ClassifCancelOut(BaseModel):
    cancelled: int
    ids: list[int]


@router.post(
    "/classificacao/pause",
    response_model=ClassifDefaultsOut,
    summary=(
        "Pausa o dispatcher AJUS. Itens em curso terminam normalmente; "
        "novos batches NAO sao claimados ate /resume."
    ),
)
def pause_classif(
    payload: ClassifPauseIn = ClassifPauseIn(),
    db: Session = Depends(get_db),
    user: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = AjusClassificacaoService(db).set_paused(
        paused=bool(payload.paused),
        by_user=getattr(user, "login", None) or getattr(user, "email", None),
    )
    return _defaults_to_out(obj)


@router.post(
    "/classificacao/resume",
    response_model=ClassifDefaultsOut,
    summary="Retoma o dispatcher AJUS apos uma pausa.",
)
def resume_classif(
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    obj = AjusClassificacaoService(db).set_paused(paused=False)
    return _defaults_to_out(obj)


@router.post(
    "/classificacao/cancel-pendentes",
    response_model=ClassifCancelOut,
    summary=(
        "Cancela TODOS os itens em status=pendente que ainda nao foram "
        "claimados por uma conta. NAO interrompe o que ja esta em curso."
    ),
)
def cancel_classif_pendentes(
    db: Session = Depends(get_db),
    user: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    res = AjusClassificacaoService(db).cancel_pendentes(
        by_user=getattr(user, "login", None) or getattr(user, "email", None),
    )
    return ClassifCancelOut(**res)


# ── Lista + detalhe ─────────────────────────────────────────────────


@router.get("/classificacao", response_model=ClassifQueueListResponse)
def list_classif(
    status: Optional[str] = Query(default=None, description="CSV de status."),
    origem: Optional[str] = Query(
        default=None, description="intake_auto | planilha",
    ),
    cnj_search: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    statuses: Optional[list[str]] = None
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
    try:
        total, items = AjusClassificacaoService(db).list(
            statuses=statuses,
            origem=origem,
            cnj_search=cnj_search,
            limit=limit,
            offset=offset,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return ClassifQueueListResponse(
        total=total, items=[_classif_to_out(i) for i in items],
    )


# ── Planilha modelo (download) ─ DECLARADO ANTES DE /{item_id} pra
#    evitar que o validador de int em item_id capture "template.xlsx".


@router.get(
    "/classificacao/template.xlsx",
    summary="Baixa a planilha modelo com cabeçalhos e exemplo.",
)
def download_classif_template(
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Gera on-the-fly uma planilha modelo com:
      - Linha 1: cabeçalhos (CNJ, UF, Comarca, Matéria, Justiça/Honorário,
        Risco/Probabilidade Perda).
      - Linha 2: exemplo (com valores plausíveis).
      - Aba 'Instruções' explicando preenchimento.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=(
                "openpyxl não instalado no container — adicionar ao "
                "requirements.txt."
            ),
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Classificações"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")  # cinza escuro
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    widths = (28, 8, 28, 32, 28, 28)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value="0001234-56.2026.8.05.0001")
    ws.cell(row=2, column=2, value="BA")
    ws.cell(row=2, column=3, value="SALVADOR")
    ws.cell(row=2, column=4, value="Cumprimento de Sentença")
    ws.cell(row=2, column=5, value="Justiça Estadual")
    ws.cell(row=2, column=6, value="Remoto")

    ws_help = wb.create_sheet("Instruções")
    ws_help.column_dimensions["A"].width = 25
    ws_help.column_dimensions["B"].width = 90
    ws_help.cell(row=1, column=1, value="Campo").font = Font(bold=True)
    ws_help.cell(row=1, column=2, value="Como preencher").font = Font(bold=True)

    instr = [
        ("CNJ", "Número do processo (com ou sem máscara)."),
        ("UF", "Sigla da UF (ex.: BA). Se vazio, será derivado do CNJ."),
        ("Comarca", "Nome da comarca (ex.: SALVADOR)."),
        ("Matéria",
         "Texto exato como aparece na capa do AJUS "
         "(ex.: Cumprimento de Sentença)."),
        ("Justiça/Honorário",
         "Texto exato como aparece na capa do AJUS "
         "(ex.: Justiça Estadual / Juizado Especial Cível)."),
        ("Risco/Probabilidade Perda",
         "Texto exato como aparece na capa do AJUS "
         "(ex.: Remoto / Possível / Provável)."),
    ]
    for idx, (campo, desc) in enumerate(instr, start=2):
        ws_help.cell(row=idx, column=1, value=campo)
        c = ws_help.cell(row=idx, column=2, value=desc)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="ajus-classificacao-modelo.xlsx"'
            ),
        },
    )


# ── Upload da planilha ─ TAMBÉM declarado antes de /{item_id}.


@router.post(
    "/classificacao/upload-xlsx", response_model=ClassifUploadResponse,
)
async def upload_classif_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Upload da planilha de classificação. Espera os cabeçalhos do
    template (CNJ, UF, Comarca, Matéria, Justiça/Honorário,
    Risco/Probabilidade Perda) na linha 1.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail="openpyxl não instalado no container.",
        ) from exc

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400,
            detail=f"Falha ao abrir XLSX: {exc}",
        ) from exc

    ws = wb.active
    if ws is None or ws.title.lower().startswith("instru"):
        for sheet in wb.worksheets:
            if not sheet.title.lower().startswith("instru"):
                ws = sheet
                break
    if ws is None:
        raise HTTPException(status_code=400, detail="Nenhuma aba de dados.")

    rows: list[XlsxRow]
    if is_legal_one_export(ws):
        # "Listagem de Acoes Judiciais" do Legal One: convertemos na
        # hora aplicando os defaults do playbook do MDR:
        #   - Materia = "Consumidor"
        #   - Risco = "Remoto"
        #   - Justica/Honorario derivado de "Tipo de Acao"
        #     ("Juizado Especial Civel" se conter "Juizado", senao
        #     "Justica Comum")
        #   - UF vazia -> inferida do segmento TR do CNJ (Estadual)
        #   - Comarca vazia / "Nao Informada" -> capital da UF
        rows = convert_legal_one_export_to_xlsx_rows(ws)
    else:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_norm = [
            (str(h).strip().lower() if h is not None else "")
            for h in header_row
        ]
        expected = [h.lower() for h in XLSX_HEADERS]
        if len(header_norm) < len(expected) or header_norm[: len(expected)] != expected:
            raise HTTPException(
                status_code=422,
                detail=(
                    "Cabeçalhos não batem com o template. Esperado: "
                    + " | ".join(XLSX_HEADERS)
                ),
            )

        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if raw is None:
                continue
            if all((v is None or str(v).strip() == "") for v in raw):
                continue
            padded = list(raw) + [None] * (len(expected) - len(raw))
            rows.append(
                XlsxRow(
                    cnj_number=str(padded[0] or "").strip(),
                    uf=(str(padded[1]).strip() if padded[1] is not None else None),
                    comarca=(
                        str(padded[2]).strip() if padded[2] is not None else None
                    ),
                    matter=(
                        str(padded[3]).strip() if padded[3] is not None else None
                    ),
                    justice_fee=(
                        str(padded[4]).strip() if padded[4] is not None else None
                    ),
                    risk_loss_probability=(
                        str(padded[5]).strip() if padded[5] is not None else None
                    ),
                ),
            )

    if not rows:
        raise HTTPException(status_code=400, detail="Sem linhas de dados.")

    result = AjusClassificacaoService(db).enqueue_from_xlsx_rows(rows)
    return ClassifUploadResponse(**result)


# ═══════════════════════════════════════════════════════════════════════
# Sessões AJUS — multi-conta (Chunk 2a)
# ═══════════════════════════════════════════════════════════════════════


class SessionAccountIn(BaseModel):
    label: str = Field(..., min_length=1, max_length=64)
    login: str = Field(..., min_length=1, max_length=128)
    password: str = Field(..., min_length=1)


class SessionAccountUpdateIn(BaseModel):
    label: Optional[str] = Field(default=None, max_length=64)
    login: Optional[str] = Field(default=None, max_length=128)
    password: Optional[str] = None
    is_active: Optional[bool] = None


class SessionAccountOut(BaseModel):
    id: int
    label: str
    login: str
    status: str
    has_storage_state: bool
    has_pending_ip_code: bool
    last_error_message: Optional[str]
    last_error_at: Optional[str]
    last_used_at: Optional[str]
    is_active: bool
    created_at: Optional[str]
    updated_at: Optional[str]


class SessionConfigOut(BaseModel):
    crypto_configured: bool
    portal_base_url: str


class IpCodeIn(BaseModel):
    code: str = Field(..., min_length=1, max_length=32)


def _session_to_out(obj: AjusSessionAccount) -> SessionAccountOut:
    return SessionAccountOut(
        id=obj.id,
        label=obj.label,
        login=obj.login,
        status=obj.status,
        has_storage_state=has_storage_state(obj),
        has_pending_ip_code=bool(obj.pending_ip_code),
        last_error_message=obj.last_error_message,
        last_error_at=(
            obj.last_error_at.isoformat() if obj.last_error_at else None
        ),
        last_used_at=(
            obj.last_used_at.isoformat() if obj.last_used_at else None
        ),
        is_active=obj.is_active,
        created_at=obj.created_at.isoformat() if obj.created_at else None,
        updated_at=obj.updated_at.isoformat() if obj.updated_at else None,
    )


@router.get(
    "/classificacao/sessions/config", response_model=SessionConfigOut,
)
def get_session_config(
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus import portal_constants as _portal
    return SessionConfigOut(
        crypto_configured=ajus_crypto_configured(),
        portal_base_url=_portal.PORTAL_BASE_URL,
    )


@router.get(
    "/classificacao/sessions", response_model=list[SessionAccountOut],
)
def list_sessions(
    only_active: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    accounts = AjusSessionService(db).list_accounts(only_active=only_active)
    return [_session_to_out(a) for a in accounts]


@router.post(
    "/classificacao/sessions",
    response_model=SessionAccountOut,
    status_code=201,
)
def create_session(
    payload: SessionAccountIn,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    if not ajus_crypto_configured():
        raise HTTPException(
            status_code=503,
            detail=(
                "AJUS_FERNET_KEY não configurada. Adicione a variável "
                "no painel do Coolify antes de cadastrar contas."
            ),
        )
    try:
        obj = AjusSessionService(db).create_account(
            label=payload.label,
            login=payload.login,
            password=payload.password,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=409,
            detail=f"Falha ao criar conta (label duplicado?): {exc}",
        ) from exc
    return _session_to_out(obj)


@router.put(
    "/classificacao/sessions/{account_id}",
    response_model=SessionAccountOut,
)
def update_session(
    payload: SessionAccountUpdateIn,
    account_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        obj = AjusSessionService(db).update_account(
            account_id,
            label=payload.label,
            login=payload.login,
            password=payload.password,
            is_active=payload.is_active,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return _session_to_out(obj)


@router.delete(
    "/classificacao/sessions/{account_id}", status_code=204,
)
def delete_session(
    account_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        AjusSessionService(db).delete_account(account_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return None


@router.post(
    "/classificacao/sessions/{account_id}/login",
    response_model=SessionAccountOut,
    summary="Solicita que o runner faça login nessa conta.",
)
def request_login(
    account_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        obj = AjusSessionService(db).request_login(account_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _session_to_out(obj)


@router.post(
    "/classificacao/sessions/{account_id}/ip-code",
    response_model=SessionAccountOut,
)
def submit_ip_code(
    payload: IpCodeIn,
    account_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        obj = AjusSessionService(db).submit_ip_code(account_id, payload.code)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _session_to_out(obj)


@router.post(
    "/classificacao/sessions/{account_id}/logout",
    response_model=SessionAccountOut,
)
def request_logout(
    account_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        obj = AjusSessionService(db).request_logout(account_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _session_to_out(obj)


# ── Dispatch manual da classificação ───────────────────────────────


class ClassifDispatchOut(BaseModel):
    candidates: int
    success_count: int
    error_count: int
    success_ids: list[int]
    errored: list[dict]
    accounts_used: list[int]
    # Campos novos pro modo soft-trigger (endpoint /dispatch). Quando
    # o endpoint apenas sinaliza (sem rodar Playwright), candidates
    # vira count de pendentes e accepted=True se ha trabalho pra fazer.
    accepted: bool = False
    accounts_online: int = 0
    message: str = ""


@router.post(
    "/classificacao/dispatch", response_model=ClassifDispatchOut,
    summary=(
        "Sinaliza pro ajus-runner pegar a fila imediatamente. NAO roda "
        "Playwright nesse container (API). O ajus-runner faz fast-poll "
        "de 2s e processa em ate ~2s apos a chamada. Retorna sumario do "
        "que foi sinalizado."
    ),
)
def dispatch_classif(
    db: Session = Depends(get_db),
    user: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    # IMPORTANTE: este endpoint vive no container `api` que NAO tem
    # Playwright instalado. Chamar AjusClassifDispatcher direto aqui
    # estourava `ModuleNotFoundError: No module named 'playwright'`
    # quando o runner tentava abrir o browser. A arquitetura correta
    # eh: API so SINALIZA; o container `ajus-runner` (em loop) processa.

    # Auto-despausa: o estado padrao do sistema eh "pausado" (modo
    # manual). Quando o operador clica "Disparar pendentes", a
    # intencao eh COMECAR a rodar — entao o endpoint /dispatch
    # despausa automaticamente antes de sinalizar. Se o operador
    # quiser parar de novo, usa o botao "Pausar" explicitamente.
    classif_service_for_pause = AjusClassificacaoService(db)
    was_paused = classif_service_for_pause.is_paused()
    if was_paused:
        classif_service_for_pause.set_paused(
            paused=False,
            by_user=getattr(user, "login", None) or getattr(user, "email", None),
        )

    # Auto-heal de fantasmas: itens com status=pendente mas com
    # dispatched_by_account_id preenchido podem ser residuos de crash
    # do dispatcher (ex.: ModuleNotFoundError do Playwright em deploy
    # antigo). MAS — atencao — se a conta dona desses itens estiver
    # AGORA em EXECUTANDO, eles NAO sao fantasmas: o runner os claimou
    # e ainda nao moveu pra status=processando (janela de ms entre
    # _claim_pending_items e mark_processing). Zerar nesse caso causa
    # corrida — o item pode ser re-claimado por outra iteracao do
    # dispatch_all enquanto o runner antigo ainda processa.
    #
    # Logica corrigida: liberar somente itens cuja conta dona NAO esta
    # em EXECUTANDO. Itens dispatched a contas online/erro/offline sao
    # fantasmas legitimos e podem ser liberados.
    executando_account_ids = [
        a_id for (a_id,) in (
            db.query(AjusSessionAccount.id)
            .filter(AjusSessionAccount.status == AJUS_ACCOUNT_EXECUTANDO)
            .all()
        )
    ]
    ghost_q = (
        db.query(AjusClassificacaoQueue)
        .filter(AjusClassificacaoQueue.status == AJUS_CLASSIF_PENDENTE)
        .filter(AjusClassificacaoQueue.dispatched_by_account_id.isnot(None))
    )
    if executando_account_ids:
        ghost_q = ghost_q.filter(
            ~AjusClassificacaoQueue.dispatched_by_account_id.in_(
                executando_account_ids,
            )
        )
    ghost_count = ghost_q.update(
        {AjusClassificacaoQueue.dispatched_by_account_id: None},
        synchronize_session=False,
    )
    if ghost_count:
        db.commit()
        logger.info(
            "AJUS dispatch auto-heal: %d fantasma(s) liberado(s) "
            "(dispatched_by_account_id zerado em itens pendentes; "
            "%d conta(s) em EXECUTANDO foram preservadas)",
            ghost_count, len(executando_account_ids),
        )

    pendentes = (
        db.query(AjusClassificacaoQueue)
        .filter(AjusClassificacaoQueue.status == AJUS_CLASSIF_PENDENTE)
        .filter(AjusClassificacaoQueue.dispatched_by_account_id.is_(None))
        .count()
    )
    processando = (
        db.query(AjusClassificacaoQueue)
        .filter(AjusClassificacaoQueue.status == AJUS_CLASSIF_PROCESSANDO)
        .count()
    )
    # Itens claimed pelo runner mas ainda em status=pendente (janela
    # entre _claim_pending_items e mark_processing). Conta como "em
    # curso" pra UI nao gritar "fila vazia" quando ha trabalho rodando.
    em_curso_pendente_claimed = (
        db.query(AjusClassificacaoQueue)
        .filter(AjusClassificacaoQueue.status == AJUS_CLASSIF_PENDENTE)
        .filter(AjusClassificacaoQueue.dispatched_by_account_id.isnot(None))
        .count()
    )
    online = (
        db.query(AjusSessionAccount)
        .filter(AjusSessionAccount.is_active.is_(True))
        .filter(AjusSessionAccount.status == AJUS_ACCOUNT_ONLINE)
        .count()
    )
    executando = len(executando_account_ids)

    em_curso = processando + em_curso_pendente_claimed
    # Apos o auto-despausa no inicio do endpoint, is_paused agora eh
    # False. Mantemos a referencia `was_paused` pra incluir na mensagem
    # quando relevante (deixa o operador saber que estava parado e
    # acabou de iniciar).
    is_paused = AjusClassificacaoService(db).is_paused()

    if is_paused:
        # Defesa: nao deveria acontecer apos auto-despausa, mas se algo
        # racou e ainda esta pausado, preserva a mensagem antiga.
        if em_curso > 0:
            msg = (
                f"Dispatcher PAUSADO. {em_curso} item(ns) em curso vao "
                f"terminar; novos batches nao serao claimados ate retomar. "
                f"({pendentes} pendente(s) na fila aguardando)."
            )
        else:
            msg = (
                f"Dispatcher PAUSADO. {pendentes} item(ns) na fila aguardando "
                "retomada."
            )
    elif pendentes == 0 and em_curso == 0:
        msg = "Nenhum item pendente na fila."
    elif pendentes == 0 and em_curso > 0:
        # Runner ja esta cuidando — nao tem o que sinalizar.
        msg = (
            f"Runner ja esta processando ({em_curso} item(ns) em curso "
            f"em {executando} conta(s)). Aguarde — nada novo na fila."
        )
    elif online == 0 and executando == 0:
        msg = (
            f"{pendentes} item(ns) pendente(s) mas nenhuma conta online. "
            "Acesse o card 'Sessoes AJUS' e faca login em pelo menos uma."
        )
    elif online == 0 and executando > 0:
        # Tudo executando, sem conta livre. Operador clicou querendo
        # acelerar — o que ja esta executando vai acabar e o worker
        # periodico pega o resto.
        msg = (
            f"{pendentes} item(ns) pendente(s) na fila, {executando} conta(s) "
            f"ja em execucao ({em_curso} item(ns) em curso). O proximo lote "
            "sera pego automaticamente quando uma conta liberar."
        )
    else:
        prefix = "Iniciado" if was_paused else "Sinalizado"
        msg = (
            f"{prefix}: {pendentes} item(ns) na fila, {online} conta(s) "
            f"online"
            + (
                f" ({executando} ja em execucao com {em_curso} item(ns))"
                if executando else ""
            )
            + ". O ajus-runner pega em ate ~2s. "
              "Pra parar, clique 'Pausar'."
        )

    return ClassifDispatchOut(
        candidates=pendentes,
        success_count=0,
        error_count=0,
        success_ids=[],
        errored=[],
        accounts_used=[],
        accepted=(pendentes > 0 and (online > 0 or executando > 0) and not is_paused),
        accounts_online=online,
        message=msg,
    )


# ── Debug screenshots dos logins (volume `ajus_session`, read-only) ─


@router.get(
    "/classificacao/sessions/{account_id}/debug-screenshots",
    summary="Lista screenshots de debug de uma conta (gerados em falhas de login).",
)
def list_debug_screenshots(
    account_id: int = FastapiPath(..., ge=1),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Retorna lista de PNGs salvos pelo runner no volume da conta.
    O runner cria esses arquivos quando o login falha — ajudam o
    operador a ver o que o headless Chromium estava vendo.
    """
    from app.core.config import settings as _s
    base = FilePath(_s.ajus_session_path) / str(account_id)
    if not base.exists() or not base.is_dir():
        return {"files": []}
    files = sorted(
        (
            {
                "name": f.name,
                "size": f.stat().st_size,
                "mtime": f.stat().st_mtime,
            }
            for f in base.iterdir()
            if f.is_file() and f.name.startswith("debug-") and f.name.endswith(".png")
        ),
        key=lambda d: d["mtime"],
        reverse=True,
    )
    return {"files": files}


@router.get(
    "/classificacao/sessions/{account_id}/debug-screenshots/{filename}",
    summary="Serve o PNG de debug por nome de arquivo.",
)
def get_debug_screenshot(
    account_id: int = FastapiPath(..., ge=1),
    filename: str = FastapiPath(..., min_length=1, max_length=128),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Serve o PNG de debug. Valida nome do arquivo pra evitar path
    traversal — soh aceita o padrao `debug-...png`.
    """
    from app.core.config import settings as _s
    if (
        ".." in filename
        or "/" in filename
        or "\\" in filename
        or not filename.startswith("debug-")
        or not filename.endswith(".png")
    ):
        raise HTTPException(status_code=400, detail="Nome de arquivo invalido.")
    fp = FilePath(_s.ajus_session_path) / str(account_id) / filename
    if not fp.exists() or not fp.is_file():
        raise HTTPException(status_code=404, detail="Screenshot nao encontrado.")
    return FileResponse(
        path=str(fp),
        media_type="image/png",
        filename=filename,
    )


# ── Detalhe + mutações por item (DEPOIS dos paths estáticos) ────────


@router.get("/classificacao/{item_id}", response_model=ClassifQueueOut)
def get_classif(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    try:
        item = AjusClassificacaoService(db).get(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return _classif_to_out(item)


@router.put("/classificacao/{item_id}", response_model=ClassifQueueOut)
@router.put("/classificacao/{item_id}", response_model=ClassifQueueOut)
def update_classif(
    payload: ClassifQueueUpdateIn,
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusClassificacaoService(db)
    try:
        item = service.update(
            item_id,
            uf=payload.uf,
            comarca=payload.comarca,
            matter=payload.matter,
            justice_fee=payload.justice_fee,
            risk_loss_probability=payload.risk_loss_probability,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _classif_to_out(item)


@router.post("/classificacao/{item_id}/cancel", response_model=ClassifQueueOut)
def cancel_classif(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusClassificacaoService(db)
    try:
        item = service.cancel(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _classif_to_out(item)


@router.post("/classificacao/{item_id}/retry", response_model=ClassifQueueOut)
def retry_classif(
    item_id: int = FastapiPath(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    service = AjusClassificacaoService(db)
    try:
        item = service.retry(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    return _classif_to_out(item)


# -- Retry em massa --------------------------------------------------


class ClassifRetryBulkIn(BaseModel):
    """
    Se `item_ids` vier, restringe ao conjunto (apenas status=erro sao reenfileirados).
    Se for None, retry em TODOS os itens em status 'erro'.
    """
    item_ids: Optional[list[int]] = None


class ClassifRetryBulkOut(BaseModel):
    retried: int
    ids: list[int]


@router.post(
    "/classificacao/retry-errors",
    response_model=ClassifRetryBulkOut,
    summary=(
        "Retry em massa de itens em status 'erro'. Sem body retoma "
        "todos; com `item_ids` restringe ao conjunto."
    ),
)
def retry_classif_bulk(
    payload: ClassifRetryBulkIn = ClassifRetryBulkIn(),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    result = AjusClassificacaoService(db).retry_errors_bulk(
        item_ids=payload.item_ids,
    )
    return ClassifRetryBulkOut(**result)


# ═══════════════════════════════════════════════════════════════════════
# Blocklist de classificacao pendente — upload XLSX, listagem, stats
# ═══════════════════════════════════════════════════════════════════════


class BlocklistItemOut(BaseModel):
    id: int
    cnj_number: str
    cod_ajus: Optional[str] = None
    materia: Optional[str] = None
    first_seen_at: str
    last_seen_at: str


class BlocklistListResponse(BaseModel):
    total: int
    items: list[BlocklistItemOut]


class BlocklistUploadResponse(BaseModel):
    """Resumo do upload — quantos foram add/update/remove + total final."""
    added: int
    updated: int
    removed: int
    total_after: int


class BlocklistStatsResponse(BaseModel):
    total_no_blocklist: int
    items_fila_bloqueados: int
    ultimo_upload_at: Optional[str] = None


def _blocklist_to_out(row: AjusClassificationBlocklist) -> BlocklistItemOut:
    return BlocklistItemOut(
        id=row.id,
        cnj_number=row.cnj_number,
        cod_ajus=row.cod_ajus,
        materia=row.materia,
        first_seen_at=row.first_seen_at.isoformat() if row.first_seen_at else "",
        last_seen_at=row.last_seen_at.isoformat() if row.last_seen_at else "",
    )


@router.post(
    "/classification-blocklist/upload",
    response_model=BlocklistUploadResponse,
    summary=(
        "Sobe planilha XLSX com CNJs de classificacao pendente e SUBSTITUI "
        "o blocklist atual. CNJs que sumirem do upload voltam a poder ser "
        "disparados; CNJs novos passam a ser pulados no dispatch."
    ),
)
async def upload_classification_blocklist(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus.classification_blocklist_service import (
        AjusClassificationBlocklistService,
        BlocklistParseError,
        parse_xlsx,
    )

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Arquivo deve ser .xlsx (ou .xlsm).",
        )
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")
    try:
        parsed = parse_xlsx(content)
    except BlocklistParseError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    service = AjusClassificationBlocklistService(db)
    result = service.replace_blocklist(parsed)
    logger.info(
        "Blocklist upload: file=%s parsed=%d %s",
        file.filename, len(parsed), result,
    )
    return BlocklistUploadResponse(**result)


@router.get(
    "/classification-blocklist",
    response_model=BlocklistListResponse,
    summary="Lista o blocklist atual com paginacao + filtro opcional por CNJ.",
)
def list_classification_blocklist(
    cnj_number: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus.classification_blocklist_service import (
        AjusClassificationBlocklistService,
    )

    service = AjusClassificationBlocklistService(db)
    total, rows = service.list_all(
        limit=limit, offset=offset, cnj_filter=cnj_number,
    )
    return BlocklistListResponse(
        total=total,
        items=[_blocklist_to_out(r) for r in rows],
    )


@router.get(
    "/classification-blocklist/stats",
    response_model=BlocklistStatsResponse,
    summary=(
        "Stats agregados do blocklist: total no blocklist, items da fila "
        "atualmente bloqueados, e timestamp do ultimo upload."
    ),
)
def stats_classification_blocklist(
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus.classification_blocklist_service import (
        AjusClassificationBlocklistService,
    )

    service = AjusClassificationBlocklistService(db)
    s = service.stats()
    return BlocklistStatsResponse(
        total_no_blocklist=s["total_no_blocklist"],
        items_fila_bloqueados=s["items_fila_bloqueados"],
        ultimo_upload_at=(
            s["ultimo_upload_at"].isoformat()
            if s["ultimo_upload_at"] else None
        ),
    )


@router.delete(
    "/classification-blocklist",
    response_model=BlocklistUploadResponse,
    summary="Apaga TODO o blocklist (escape hatch — nao usar em rotina).",
)
def clear_classification_blocklist(
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.services.ajus.classification_blocklist_service import (
        AjusClassificationBlocklistService,
    )

    service = AjusClassificationBlocklistService(db)
    n = service.clear()
    return BlocklistUploadResponse(
        added=0, updated=0, removed=n, total_after=0,
    )

