"""Endpoints da varredura de andamentos.

Feature incidental (sem deploy em main): operador clica "Nova varredura"
selecionando offices responsaveis com polo passivo, e o sistema raspa a
pagina DetailsAndamentos do L1 atras de eventos relevantes nos ultimos
N dias (default 30).
"""

from __future__ import annotations

from typing import Optional

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core.dependencies import get_db
from app.models.legal_one import LegalOneOffice
from app.models.varredura import (
    ALL_TIPOS_EVENTO,
    VarreduraAchado,
    VarreduraProcessado,
    VarreduraRun,
)
from app.services.varredura.regex_eventos import list_pattern_descriptions
from app.services.varredura.varredura_service import VarreduraService

router = APIRouter()


# ── Schemas ───────────────────────────────────────────────────────────


class VarreduraOfficeOption(BaseModel):
    external_id: int
    name: str
    path: Optional[str] = None
    polo_scope: str = "ambos"


class VarreduraRunCreatePayload(BaseModel):
    responsible_office_ids: list[int] = Field(..., min_length=1)
    window_days: int = Field(30, ge=1, le=365)
    max_processos: Optional[int] = Field(
        30,
        ge=1,
        le=5000,
        description=(
            "Limite global de processos a varrer (soma de todos os offices). "
            "Default 30 = ~1 pagina API L1. Aumente conforme o operador queira."
        ),
    )


class VarreduraRunFromListPayload(BaseModel):
    identifiers: list[str] = Field(
        ...,
        min_length=1,
        description=(
            "Lista de CNJs (com pontuacao ou apenas digitos) ou lawsuit_ids "
            "(numeros inteiros). Pode misturar os formatos."
        ),
    )
    window_days: int = Field(30, ge=1, le=365)


class VarreduraRunFromListResponse(BaseModel):
    run: "VarreduraRunOut"
    unresolved: list[str] = Field(
        default_factory=list,
        description="CNJs/identificadores que nao foram encontrados no L1.",
    )


class VarreduraRunOut(BaseModel):
    id: int
    status: str
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    responsible_office_ids: list[int] = Field(default_factory=list)
    window_days: int
    total_processos: int
    total_processados: int
    total_achados: int
    total_falhas: int
    triggered_by: Optional[str] = None
    error_message: Optional[str] = None


class VarreduraRunListOut(BaseModel):
    total: int
    items: list[VarreduraRunOut]


class VarreduraProcessadoOut(BaseModel):
    id: int
    run_id: int
    lawsuit_id: int
    cnj_number: Optional[str] = None
    office_id: Optional[int] = None
    queue_status: str
    attempt_count: int
    last_attempt_at: Optional[str] = None
    completed_at: Optional[str] = None
    total_andamentos_lidos: int
    total_achados: int
    last_error: Optional[str] = None
    last_reason: Optional[str] = None


class VarreduraProcessadoListOut(BaseModel):
    total: int
    items: list[VarreduraProcessadoOut]


class VarreduraAchadoOut(BaseModel):
    id: int
    run_id: int
    processado_id: int
    lawsuit_id: int
    cnj_number: Optional[str] = None
    andamento_data: Optional[str] = None
    andamento_hora: Optional[str] = None
    andamento_tipo: Optional[str] = None
    andamento_texto: str
    andamento_movimentado_por: Optional[str] = None
    tipo_evento: str
    regex_matched: Optional[str] = None
    tratado: bool
    tratado_em: Optional[str] = None
    tratado_por: Optional[str] = None
    observacao: Optional[str] = None
    created_at: Optional[str] = None


class VarreduraAchadoListOut(BaseModel):
    total: int
    items: list[VarreduraAchadoOut]


class VarreduraAchadoUpdatePayload(BaseModel):
    tratado: bool
    observacao: Optional[str] = None


class VarreduraRecoverResponse(BaseModel):
    recovered_count: int
    threshold_minutes: int


class VarreduraPatternsOut(BaseModel):
    patterns: list[dict]


# ── Endpoints ─────────────────────────────────────────────────────────


@router.get(
    "/varredura/offices-disponiveis",
    response_model=list[VarreduraOfficeOption],
    summary="Lista offices responsaveis disponiveis pra varredura",
    tags=["Varredura"],
)
def list_offices_disponiveis(db: Session = Depends(get_db)):
    """Retorna offices ativos pra UI selecionar. Pre-marca os 'passivo'
    no frontend (UI decide qual default selecionar)."""
    rows = (
        db.query(LegalOneOffice)
        .filter(LegalOneOffice.is_active == True)  # noqa: E712
        .order_by(LegalOneOffice.path)
        .all()
    )
    return [
        VarreduraOfficeOption(
            external_id=o.external_id,
            name=o.name,
            path=o.path,
            polo_scope=o.polo_scope or "ambos",
        )
        for o in rows
    ]


@router.get(
    "/varredura/patterns",
    response_model=VarreduraPatternsOut,
    summary="Lista regex/eventos detectados pela varredura",
    tags=["Varredura"],
)
def get_patterns():
    return {"patterns": list_pattern_descriptions()}


@router.post(
    "/varredura/runs",
    response_model=VarreduraRunOut,
    status_code=201,
    summary="Cria nova varredura e dispara processamento em background",
    tags=["Varredura"],
)
def create_run(
    payload: VarreduraRunCreatePayload,
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    try:
        run = svc.create_run(
            responsible_office_ids=payload.responsible_office_ids,
            window_days=payload.window_days,
            max_processos=payload.max_processos,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    return svc._run_to_dict(run)


@router.post(
    "/varredura/runs/from-list",
    response_model=VarreduraRunFromListResponse,
    status_code=201,
    summary="Cria varredura a partir de lista de CNJs/lawsuit_ids",
    tags=["Varredura"],
)
def create_run_from_list(
    payload: VarreduraRunFromListPayload,
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    try:
        run, unresolved = svc.create_run_from_list(
            identifiers=payload.identifiers,
            window_days=payload.window_days,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    return {"run": svc._run_to_dict(run), "unresolved": unresolved}


@router.get(
    "/varredura/runs",
    response_model=VarreduraRunListOut,
    summary="Lista runs paginadas",
    tags=["Varredura"],
)
def list_runs(
    status: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    return svc.list_runs(status=status, limit=limit, offset=offset)


@router.get(
    "/varredura/runs/{run_id}",
    response_model=VarreduraRunOut,
    summary="Detalhe de uma run",
    tags=["Varredura"],
)
def get_run(run_id: int, db: Session = Depends(get_db)):
    svc = VarreduraService(db)
    data = svc.get_run(run_id)
    if data is None:
        raise HTTPException(404, "Varredura nao encontrada.")
    return data


@router.get(
    "/varredura/runs/{run_id}/processados",
    response_model=VarreduraProcessadoListOut,
    summary="Items processados (fila) de uma run",
    tags=["Varredura"],
)
def list_processados(
    run_id: int,
    queue_status: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    return svc.list_processados(
        run_id=run_id,
        queue_status=queue_status,
        limit=limit,
        offset=offset,
    )


@router.post(
    "/varredura/runs/{run_id}/cancel",
    response_model=VarreduraRunOut,
    summary="Cancela uma run em execucao",
    tags=["Varredura"],
)
def cancel_run(run_id: int, db: Session = Depends(get_db)):
    svc = VarreduraService(db)
    data = svc.cancel_run(run_id)
    if data is None:
        raise HTTPException(404, "Varredura nao encontrada.")
    return data


@router.post(
    "/varredura/runs/{run_id}/recover-zombies",
    response_model=VarreduraRecoverResponse,
    summary="Recupera items presos em PROCESSANDO",
    tags=["Varredura"],
)
def recover_zombies(
    run_id: int,
    threshold_minutes: int = Query(default=10, ge=1, le=120),
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    # run_id e' ignorado no service (recover roda global), mas mantemos no
    # path pra clareza da UI. Se quiser run-specific, filtra no service.
    _ = run_id
    return svc.recover_zombies(threshold_minutes=threshold_minutes)


@router.get(
    "/varredura/achados",
    response_model=VarreduraAchadoListOut,
    summary="Lista achados com filtros",
    tags=["Varredura"],
)
def list_achados(
    run_id: Optional[int] = Query(default=None),
    tipo_evento: Optional[str] = Query(default=None),
    tratado: Optional[bool] = Query(default=None),
    cnj_search: Optional[str] = Query(default=None),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
):
    if tipo_evento and tipo_evento not in ALL_TIPOS_EVENTO:
        raise HTTPException(
            400,
            f"tipo_evento invalido. Aceitos: {sorted(ALL_TIPOS_EVENTO)}",
        )
    svc = VarreduraService(db)
    return svc.list_achados(
        run_id=run_id,
        tipo_evento=tipo_evento,
        tratado=tratado,
        cnj_search=cnj_search,
        limit=limit,
        offset=offset,
    )


@router.patch(
    "/varredura/achados/{achado_id}",
    response_model=VarreduraAchadoOut,
    summary="Atualiza um achado (marca tratado e/ou observacao)",
    tags=["Varredura"],
)
def update_achado(
    achado_id: int,
    payload: VarreduraAchadoUpdatePayload,
    db: Session = Depends(get_db),
):
    svc = VarreduraService(db)
    data = svc.update_achado(
        achado_id,
        tratado=payload.tratado,
        observacao=payload.observacao,
    )
    if data is None:
        raise HTTPException(404, "Achado nao encontrado.")
    return data


# ── Export XLSX ───────────────────────────────────────────────────────


_EVENTO_LABEL = {
    "audiencia_designada": "Audiência designada",
    "audiencia_cancelada": "Audiência cancelada",
    "sentenca": "Sentença",
    "revelia": "Revelia",
    "transito_julgado": "Trânsito em julgado",
    "arquivamento": "Arquivamento",
}


def _build_xlsx_payload(
    db: Session,
    *,
    run: VarreduraRun,
    achados: list[VarreduraAchado],
    processados: list[VarreduraProcessado],
) -> bytes:
    """Monta um XLSX com 3 abas: Resumo, Achados, Processos."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Aba 1: Resumo da run ──────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    bold = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid",
    )

    ws_resumo["A1"] = "Varredura de Andamentos — Resumo"
    ws_resumo["A1"].font = Font(bold=True, size=14)
    ws_resumo.merge_cells("A1:B1")

    summary_rows = [
        ("Run ID", run.id),
        ("Status", run.status),
        ("Iniciada em", run.started_at.isoformat() if run.started_at else "—"),
        (
            "Concluída em",
            run.completed_at.isoformat() if run.completed_at else "—",
        ),
        (
            "Offices selecionados",
            ", ".join(str(x) for x in (run.responsible_office_ids or [])),
        ),
        ("Janela (dias)", run.window_days),
        ("Total processos", run.total_processos),
        ("Total processados", run.total_processados),
        ("Total achados", run.total_achados),
        ("Total falhas", run.total_falhas),
        ("Disparada por", run.triggered_by or "—"),
    ]
    for i, (k, v) in enumerate(summary_rows, start=3):
        ws_resumo.cell(row=i, column=1, value=k).font = bold
        ws_resumo.cell(row=i, column=2, value=v)

    # Contagem por tipo de evento
    base_row = len(summary_rows) + 5
    ws_resumo.cell(row=base_row, column=1, value="Tipo de evento").font = bold
    ws_resumo.cell(row=base_row, column=2, value="Qtd").font = bold
    by_type: dict[str, int] = {}
    for a in achados:
        by_type[a.tipo_evento] = by_type.get(a.tipo_evento, 0) + 1
    for offset, tipo in enumerate(sorted(by_type.keys()), start=1):
        ws_resumo.cell(
            row=base_row + offset,
            column=1,
            value=_EVENTO_LABEL.get(tipo, tipo),
        )
        ws_resumo.cell(row=base_row + offset, column=2, value=by_type[tipo])

    ws_resumo.column_dimensions["A"].width = 28
    ws_resumo.column_dimensions["B"].width = 50

    # ── Aba 2: Achados ────────────────────────────────────────────────
    ws_ach = wb.create_sheet("Achados")
    headers_ach = [
        "ID",
        "CNJ",
        "Lawsuit ID",
        "Data andamento",
        "Hora",
        "Tipo do evento",
        "Trecho (regex)",
        "Movimentado por",
        "Texto completo do andamento",
        "Tratado",
        "Tratado em",
        "Observação",
    ]
    for col, h in enumerate(headers_ach, start=1):
        cell = ws_ach.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for r_idx, a in enumerate(achados, start=2):
        ws_ach.cell(row=r_idx, column=1, value=a.id)
        ws_ach.cell(row=r_idx, column=2, value=a.cnj_number or "")
        ws_ach.cell(row=r_idx, column=3, value=a.lawsuit_id)
        ws_ach.cell(
            row=r_idx,
            column=4,
            value=(
                a.andamento_data.strftime("%d/%m/%Y")
                if a.andamento_data
                else ""
            ),
        )
        ws_ach.cell(row=r_idx, column=5, value=a.andamento_hora or "")
        ws_ach.cell(
            row=r_idx,
            column=6,
            value=_EVENTO_LABEL.get(a.tipo_evento, a.tipo_evento),
        )
        ws_ach.cell(row=r_idx, column=7, value=a.regex_matched or "")
        ws_ach.cell(
            row=r_idx, column=8, value=a.andamento_movimentado_por or "",
        )
        texto_cell = ws_ach.cell(
            row=r_idx, column=9, value=a.andamento_texto or "",
        )
        texto_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_ach.cell(row=r_idx, column=10, value="Sim" if a.tratado else "Não")
        ws_ach.cell(
            row=r_idx,
            column=11,
            value=a.tratado_em.isoformat() if a.tratado_em else "",
        )
        ws_ach.cell(row=r_idx, column=12, value=a.observacao or "")

    # Larguras + freeze
    widths_ach = [8, 28, 12, 14, 8, 22, 32, 28, 80, 10, 22, 32]
    for i, w in enumerate(widths_ach, start=1):
        ws_ach.column_dimensions[get_column_letter(i)].width = w
    ws_ach.freeze_panes = "A2"
    ws_ach.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers_ach))}{max(2, len(achados) + 1)}"
    )

    # ── Aba 3: Processos (todos os items processados da fila) ─────────
    ws_proc = wb.create_sheet("Processos")
    headers_proc = [
        "Processo ID",
        "CNJ",
        "Lawsuit ID",
        "Office ID",
        "Status",
        "Tentativas",
        "Última tentativa",
        "Concluído em",
        "Andamentos lidos",
        "Achados",
        "Motivo",
        "Erro",
    ]
    for col, h in enumerate(headers_proc, start=1):
        cell = ws_proc.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
    for r_idx, p in enumerate(processados, start=2):
        ws_proc.cell(row=r_idx, column=1, value=p.id)
        ws_proc.cell(row=r_idx, column=2, value=p.cnj_number or "")
        ws_proc.cell(row=r_idx, column=3, value=p.lawsuit_id)
        ws_proc.cell(row=r_idx, column=4, value=p.office_id or "")
        ws_proc.cell(row=r_idx, column=5, value=p.queue_status)
        ws_proc.cell(row=r_idx, column=6, value=p.attempt_count)
        ws_proc.cell(
            row=r_idx,
            column=7,
            value=p.last_attempt_at.isoformat() if p.last_attempt_at else "",
        )
        ws_proc.cell(
            row=r_idx,
            column=8,
            value=p.completed_at.isoformat() if p.completed_at else "",
        )
        ws_proc.cell(row=r_idx, column=9, value=p.total_andamentos_lidos)
        ws_proc.cell(row=r_idx, column=10, value=p.total_achados)
        ws_proc.cell(row=r_idx, column=11, value=p.last_reason or "")
        ws_proc.cell(row=r_idx, column=12, value=p.last_error or "")

    widths_proc = [10, 28, 12, 10, 14, 10, 22, 22, 14, 10, 22, 50]
    for i, w in enumerate(widths_proc, start=1):
        ws_proc.column_dimensions[get_column_letter(i)].width = w
    ws_proc.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get(
    "/varredura/runs/{run_id}/achados.xlsx",
    summary="Exporta planilha XLSX com achados da run",
    tags=["Varredura"],
)
def export_run_xlsx(run_id: int, db: Session = Depends(get_db)):
    run = (
        db.query(VarreduraRun)
        .filter(VarreduraRun.id == run_id)
        .first()
    )
    if run is None:
        raise HTTPException(404, "Varredura nao encontrada.")
    achados = (
        db.query(VarreduraAchado)
        .filter(VarreduraAchado.run_id == run_id)
        .order_by(
            VarreduraAchado.andamento_data.desc().nullslast(),
            VarreduraAchado.id.desc(),
        )
        .all()
    )
    processados = (
        db.query(VarreduraProcessado)
        .filter(VarreduraProcessado.run_id == run_id)
        .order_by(VarreduraProcessado.id.asc())
        .all()
    )
    content = _build_xlsx_payload(
        db, run=run, achados=achados, processados=processados,
    )
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"varredura-{run_id}-{ts}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
