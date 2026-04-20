import csv
import json
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional

from fastapi import APIRouter, BackgroundTasks, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload

from app.api.v1.schemas import (
    BatchInteractiveCreationRequest,
    BatchTaskCreationRequest,
    TaskTriggerPayload,
    ValidatePublicationTasksRequest,
)
from app.core import auth as auth_security
from app.core.config import settings
from app.core.dependencies import (
    get_api_client,
    get_batch_task_creation_service,
    get_db,
    get_orchestration_service,
    get_task_rule_service,
)
from app.core.uploads import validate_spreadsheet_file_metadata
from app.models.batch_execution import (
    BATCH_STATUS_CANCELLED,
    BATCH_STATUS_COMPLETED,
    BATCH_STATUS_COMPLETED_WITH_ERRORS,
    BATCH_STATUS_PAUSED,
    BATCH_STATUS_PENDING,
    BATCH_STATUS_PROCESSING,
    BatchExecution,
    BatchExecutionItem,
)
from app.models.legal_one import LegalOneOffice, LegalOneTaskType, LegalOneUser
from app.services.batch_strategies.spreadsheet_strategy import SpreadsheetStrategy
from app.services.batch_task_creation_service import BatchTaskCreationService
from app.services.legal_one_client import LegalOneApiClient
from app.services.orchestration_service import (
    MissingResponsibleUserError,
    OrchestrationService,
    ProcessNotFoundError,
)
from app.services.task_creation_service import (
    InvalidDataError,
    LawsuitNotFoundError,
    TaskCreationError,
    TaskCreationRequest,
    TaskCreationService,
    TaskLinkingError,
)
from app.services.task_rule_service import TaskRuleService

router = APIRouter()


class RetryRequest(BaseModel):
    item_ids: Optional[List[int]] = None


class LegalOnePositionFixControlRequest(BaseModel):
    action: Literal["pause", "resume"]


class SpreadsheetRow(BaseModel):
    row_id: int
    data: Dict[str, Any]


class SpreadsheetAnalysisResponse(BaseModel):
    filename: str
    headers: List[str]
    rows: List[SpreadsheetRow]


class SpreadsheetPreviewRow(BaseModel):
    row_id: int
    process_number: Optional[str] = None
    is_valid: bool
    errors: List[str]
    warnings: List[str]
    fingerprint: Optional[str] = None
    data: Dict[str, Any]


class SpreadsheetPreviewSummary(BaseModel):
    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_rows_in_file: int
    duplicate_rows_in_history: int


class SpreadsheetPreviewResponse(BaseModel):
    filename: str
    headers: List[str]
    summary: SpreadsheetPreviewSummary
    rows: List[SpreadsheetPreviewRow]


class SubTypeSchema(BaseModel):
    id: int
    external_id: int
    name: str

    class Config:
        from_attributes = True


class HierarchicalTaskTypeSchema(BaseModel):
    id: int
    name: str
    sub_types: List[SubTypeSchema]

    class Config:
        from_attributes = True


class UserForTaskForm(BaseModel):
    id: int
    external_id: int
    name: str
    squads: List[Dict[str, Any]]

    class Config:
        from_attributes = True


class OfficeForTaskForm(BaseModel):
    id: int
    external_id: int
    name: str

    class Config:
        from_attributes = True


class TaskCreationDataResponse(BaseModel):
    task_types: List[HierarchicalTaskTypeSchema]
    offices: List[OfficeForTaskForm]
    users: List[UserForTaskForm]
    task_statuses: List[Dict[str, Any]]


def _validate_spreadsheet_upload(file: UploadFile, file_content: bytes) -> None:
    validate_spreadsheet_file_metadata(
        file.filename,
        file.content_type,
        len(file_content),
        max_size_bytes=settings.spreadsheet_max_size_bytes,
        allowed_content_types=settings.allowed_spreadsheet_content_types,
    )


def _build_template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Agendamentos"
    headers = [
        "ESCRITORIO",
        "CNJ",
        "PUBLISH_DATE",
        "SUBTIPO",
        "EXECUTANTE",
        "PRAZO",
        "DATA_TAREFA",
        "HORARIO",
        "OBSERVACAO",
        "DESCRICAO",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "Juridico / Filial SP / Contencioso",
            "0000000-00.2026.8.00.0000",
            "18/03/2026",
            "Audiencia",
            "Maria Silva",
            "25/03/2026",
            "25/03/2026",
            "14:00",
            "Observacoes opcionais",
            "Complemento opcional",
        ]
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _build_error_report_csv(items: list[BatchExecutionItem]) -> str:
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow(["item_id", "cnj", "status", "task_id", "erro"])
    for item in items:
        writer.writerow(
            [
                item.id,
                item.process_number,
                item.status,
                item.created_task_id or "",
                item.error_message or "",
            ]
        )
    return stream.getvalue()


def _resolve_legal_one_position_fix_status_file() -> Path:
    if settings.legal_one_position_fix_status_file:
        return Path(settings.legal_one_position_fix_status_file)
    return Path(__file__).resolve().parents[4] / "output" / "playwright" / "legalone" / "position-fix-live.json"


def _resolve_legal_one_position_fix_control_file() -> Path:
    status_file = _resolve_legal_one_position_fix_status_file()
    if status_file.exists():
        try:
            payload = json.loads(status_file.read_text(encoding="utf-8"))
            normalized = _normalize_legal_one_position_fix_payload(payload)
            control_file = normalized.get("controlFile")
            if control_file:
                return Path(control_file)
        except (OSError, ValueError, json.JSONDecodeError):
            pass

    return status_file.with_name("position-fix.control")


def _read_legal_one_position_fix_control_signal() -> str:
    control_file = _resolve_legal_one_position_fix_control_file()
    if not control_file.exists():
        return "run"

    try:
        signal = control_file.read_text(encoding="utf-8").strip().lower()
    except OSError:
        return "run"

    if signal in {"pause", "stop"}:
        return signal
    return "run"


def set_legal_one_position_fix_control(action: Literal["pause", "resume"]) -> Dict[str, Any]:
    control_file = _resolve_legal_one_position_fix_control_file()
    desired_signal = "pause" if action == "pause" else "run"

    control_file.parent.mkdir(parents=True, exist_ok=True)
    control_file.write_text(desired_signal, encoding="utf-8")

    return {
        "message": "Sinal aplicado com sucesso.",
        "action": action,
        "signal": desired_signal,
        "control_file": str(control_file),
    }


def _normalize_legal_one_position_fix_payload(payload: Any) -> Dict[str, Any]:
    if isinstance(payload, list):
        items = payload
        updated_count = len([item for item in items if item.get("status") == "updated"])
        failed_count = len([item for item in items if item.get("status") in {"error", "verify_failed"}])
        retry_pending_count = len([item for item in items if item.get("status") == "scheduled_retry" or item.get("retryPending")])
        return {
            "generatedAt": None,
            "state": None,
            "batchSize": None,
            "currentBatch": None,
            "totalBatches": None,
            "sleepUntil": None,
            "controlFile": None,
            "totalItems": len(items),
            "processedItems": len(items),
            "updatedCount": updated_count,
            "failedCount": failed_count,
            "retryPendingCount": retry_pending_count,
            "remainingItems": max(0, len(items) - updated_count - failed_count),
            "activeQueueType": None,
            "retryPass": None,
            "maxAttempts": None,
            "workers": [],
            "items": items,
        }

    if isinstance(payload, dict):
        items = payload.get("items", [])
        total_items = payload.get("totalItems", len(items))
        processed_items = payload.get("processedItems", len(items))
        updated_count = payload.get(
            "updatedCount",
            len([item for item in items if item.get("status") == "updated"]),
        )
        failed_count = payload.get(
            "failedCount",
            len([item for item in items if item.get("status") in {"error", "verify_failed"}]),
        )
        retry_pending_count = payload.get(
            "retryPendingCount",
            len([item for item in items if item.get("status") == "scheduled_retry" or item.get("retryPending")]),
        )
        return {
            "generatedAt": payload.get("generatedAt"),
            "state": payload.get("state"),
            "batchSize": payload.get("batchSize"),
            "currentBatch": payload.get("currentBatch"),
            "totalBatches": payload.get("totalBatches"),
            "sleepUntil": payload.get("sleepUntil"),
            "controlFile": payload.get("controlFile"),
            "totalItems": total_items,
            "processedItems": processed_items,
            "updatedCount": updated_count,
            "failedCount": failed_count,
            "retryPendingCount": retry_pending_count,
            "remainingItems": payload.get("remainingItems", max(0, total_items - updated_count - failed_count)),
            "activeQueueType": payload.get("activeQueueType"),
            "retryPass": payload.get("retryPass"),
            "maxAttempts": payload.get("maxAttempts"),
            "workers": payload.get("workers", []),
            "items": items,
        }

    raise ValueError("Formato de progresso invalido.")


def _parse_iso_datetime(value: Any) -> Optional[datetime]:
    if not value or not isinstance(value, str):
        return None

    normalized = value.strip()
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"

    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed


def _build_legal_one_position_fix_timing_metrics(
    items: List[Dict[str, Any]],
    processed_items: int,
    remaining_items: int,
    generated_at: Optional[str],
) -> Dict[str, Any]:
    durations_in_seconds: List[float] = []
    started_timestamps: List[datetime] = []

    for item in items:
        started_at = _parse_iso_datetime(item.get("startedAt"))
        finished_at = _parse_iso_datetime(item.get("finishedAt"))

        if started_at is not None:
            started_timestamps.append(started_at)

        if started_at is None or finished_at is None:
            continue

        duration = (finished_at - started_at).total_seconds()
        if duration >= 0:
            durations_in_seconds.append(duration)

    average_update_seconds = (
        sum(durations_in_seconds) / len(durations_in_seconds)
        if durations_in_seconds
        else None
    )

    generated_at_dt = _parse_iso_datetime(generated_at) or datetime.now(timezone.utc)
    effective_average_seconds = None
    estimated_remaining_seconds = None
    estimated_completion_at = None

    if processed_items > 0 and started_timestamps:
        first_started_at = min(started_timestamps)
        elapsed_seconds = max(0.0, (generated_at_dt - first_started_at).total_seconds())
        effective_average_seconds = elapsed_seconds / processed_items if elapsed_seconds > 0 else 0.0
        estimated_remaining_seconds = effective_average_seconds * remaining_items
        estimated_completion_at = (generated_at_dt + timedelta(seconds=estimated_remaining_seconds)).isoformat()

    return {
        "averageUpdateSeconds": average_update_seconds,
        "effectiveAverageSeconds": effective_average_seconds,
        "estimatedRemainingSeconds": estimated_remaining_seconds,
        "estimatedCompletionAt": estimated_completion_at,
    }


@router.get(
    "/task-creation-data",
    response_model=TaskCreationDataResponse,
    summary="Obter dados para o formulario de criacao de tarefas",
)
def get_data_for_task_form(db: Session = Depends(get_db)):
    task_types_query = (
        db.query(LegalOneTaskType)
        .options(joinedload(LegalOneTaskType.subtypes))
        .order_by(LegalOneTaskType.name)
        .all()
    )
    formatted_task_types = [
        HierarchicalTaskTypeSchema(
            id=parent.id,
            name=parent.name,
            sub_types=[
                SubTypeSchema(id=sub.id, external_id=sub.external_id, name=sub.name)
                for sub in sorted(parent.subtypes, key=lambda item: item.name)
            ],
        )
        for parent in task_types_query
    ]

    offices = (
        db.query(LegalOneOffice)
        .filter(LegalOneOffice.is_active == True)
        .order_by(LegalOneOffice.name)
        .all()
    )
    users = db.query(LegalOneUser)
    users_for_form = [
        UserForTaskForm(
            id=user.id,
            external_id=user.external_id,
            name=user.name,
            squads=[{"id": member.squad.id, "name": member.squad.name} for member in user.squad_members],
        )
        for user in users
    ]
    task_statuses = [
        {"id": 0, "name": "Pendente"},
        {"id": 1, "name": "Cumprido"},
        {"id": 2, "name": "Nao cumprido"},
        {"id": 3, "name": "Cancelado"},
    ]

    return TaskCreationDataResponse(
        task_types=formatted_task_types,
        offices=[OfficeForTaskForm(id=office.id, name=office.name, external_id=office.external_id) for office in offices],
        users=users_for_form,
        task_statuses=task_statuses,
    )


@router.get("/spreadsheet-template", summary="Baixar modelo oficial da planilha")
def download_spreadsheet_template():
    return StreamingResponse(
        BytesIO(_build_template_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_agendamento_planilha.xlsx"'},
    )


@router.post(
    "/preview-spreadsheet",
    response_model=SpreadsheetPreviewResponse,
    summary="Pre-validar planilha antes do agendamento",
)
async def preview_spreadsheet(
    file: UploadFile = File(...),
    service: BatchTaskCreationService = Depends(get_batch_task_creation_service),
):
    file_content = await file.read()
    try:
        _validate_spreadsheet_upload(file, file_content)
        strategy = SpreadsheetStrategy(service.db, service.client)
        preview = await strategy.build_preview(file_content)
        return SpreadsheetPreviewResponse(
            filename=file.filename or "planilha.xlsx",
            headers=preview["headers"],
            summary=SpreadsheetPreviewSummary(**preview["summary"]),
            rows=[SpreadsheetPreviewRow(**row) for row in preview["rows"]],
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/trigger/task", tags=["Tasks"])
def trigger_task_creation(
    payload: TaskTriggerPayload,
    orchestrator: OrchestrationService = Depends(get_orchestration_service),
):
    try:
        result = orchestrator.handle_task_trigger(payload)
        return JSONResponse(status_code=201, content=result)
    except ProcessNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except MissingResponsibleUserError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Ocorreu um erro interno inesperado: {exc}") from exc


@router.get("/search-lawsuit", summary="Buscar processo por CNJ no Legal One", tags=["Tasks"])
def search_lawsuit(
    cnj: str = Query(..., description="Numero CNJ a ser pesquisado."),
    legal_one_client: LegalOneApiClient = Depends(get_api_client),
):
    lawsuit = legal_one_client.search_lawsuit_by_cnj(cnj)
    if not lawsuit:
        raise HTTPException(status_code=404, detail=f"Nenhum processo encontrado para o CNJ: {cnj}")
    return lawsuit


@router.post("/create-full-process", summary="Criar tarefa (processo completo)", tags=["Tasks"])
def create_full_task(
    request: TaskCreationRequest,
    db: Session = Depends(get_db),
):
    task_service = TaskCreationService(db)
    try:
        result = task_service.create_full_task_process(request)
        return JSONResponse(status_code=201, content=result)
    except LawsuitNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except InvalidDataError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except (TaskCreationError, TaskLinkingError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Ocorreu um erro interno inesperado: {exc}") from exc


@router.post(
    "/batch-create",
    status_code=202,
    summary="Criar tarefas em lote a partir de uma fonte externa",
)
def create_batch_tasks(
    request: BatchTaskCreationRequest,
    background_tasks: BackgroundTasks,
    service: BatchTaskCreationService = Depends(get_batch_task_creation_service),
):
    background_tasks.add_task(service.process_batch_request, request)
    return {
        "status": "recebido",
        "message": "A solicitacao de criacao de tarefas em lote foi recebida e esta sendo processada em segundo plano.",
    }


@router.post(
    "/batch-create-from-spreadsheet",
    status_code=202,
    summary="Criar tarefas em lote a partir de uma planilha",
)
async def create_batch_tasks_from_spreadsheet(
    file: UploadFile = File(...),
    service: BatchTaskCreationService = Depends(get_batch_task_creation_service),
    current_user: LegalOneUser = Depends(auth_security.get_current_user),
):
    file_content = await file.read()

    try:
        _validate_spreadsheet_upload(file, file_content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    try:
        execution_log = service.create_spreadsheet_execution(
            file_content=file_content,
            source_filename=file.filename,
            requested_by_email=current_user.email,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "status": execution_log.status,
        "message": "Lote recebido e colocado na fila de processamento.",
        "batch_id": execution_log.id,
    }


@router.post(
    "/analyze-spreadsheet",
    response_model=SpreadsheetAnalysisResponse,
    summary="Analisar planilha para criacao de tarefas interativas",
)
async def analyze_spreadsheet(file: UploadFile = File(...)):
    workbook = None
    try:
        content = await file.read()
        _validate_spreadsheet_upload(file, content)

        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        raw_headers = [cell.value for cell in sheet[1]]
        headers = [str(header).strip() if header is not None else "" for header in raw_headers]

        rows_data = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") and str(value).strip() for value in row):
                continue
            cleaned_row = ["" if value is None else value for value in row]
            rows_data.append(SpreadsheetRow(row_id=index, data=dict(zip(headers, cleaned_row))))

        return SpreadsheetAnalysisResponse(
            filename=file.filename or "planilha.xlsx",
            headers=headers,
            rows=rows_data,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Nao foi possivel processar a planilha: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


@router.post(
    "/batch-create-interactive",
    status_code=202,
    summary="Criar tarefas em lote a partir da interface interativa",
)
def create_batch_tasks_interactive(
    request: BatchInteractiveCreationRequest,
    service: BatchTaskCreationService = Depends(get_batch_task_creation_service),
    current_user: LegalOneUser = Depends(auth_security.get_current_user),
):
    execution_log = service.create_interactive_execution(
        request,
        requested_by_email=current_user.email,
    )
    return {
        "status": execution_log.status,
        "message": "A solicitacao foi recebida e entrou na fila de agendamento.",
        "batch_id": execution_log.id,
    }


@router.post(
    "/validate-publication-tasks",
    summary="Validar regras de negocio para um conjunto de tarefas de uma publicacao",
)
def validate_publication_tasks(
    request: ValidatePublicationTasksRequest,
    rule_service: TaskRuleService = Depends(get_task_rule_service),
):
    try:
        tasks_as_dicts = [task.model_dump(by_alias=True) for task in request.tasks]
        rule_service.validate_co_requisites(tasks_as_dicts)
        return {"status": "success", "message": "As regras de negocio foram atendidas."}
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Ocorreu um erro inesperado durante a validacao: {exc}") from exc


@router.get("/batch/status/{batch_id}", summary="Verificar progresso de um lote (resumo)")
def get_batch_status(batch_id: int, db: Session = Depends(get_db)):
    execution = db.query(BatchExecution).filter(BatchExecution.id == batch_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Lote nao encontrado")

    processed_count = (
        db.query(BatchExecutionItem)
        .filter(
            BatchExecutionItem.execution_id == batch_id,
            BatchExecutionItem.status.in_(["SUCESSO", "FALHA"]),
        )
        .count()
    )

    total = execution.total_items or 0
    percentage = int((processed_count / total) * 100) if total > 0 else 0
    remaining_items = max(0, total - processed_count)

    return {
        "id": execution.id,
        "status": execution.status,
        "total_items": total,
        "processed_items": processed_count,
        "remaining_items": remaining_items,
        "success_count": execution.success_count,
        "failure_count": execution.failure_count,
        "percentage": percentage,
        "source_filename": execution.source_filename,
        "requested_by_email": execution.requested_by_email,
        "can_pause": execution.status in {BATCH_STATUS_PENDING, BATCH_STATUS_PROCESSING},
        "can_resume": execution.status == BATCH_STATUS_PAUSED,
        "can_cancel": execution.status in {BATCH_STATUS_PENDING, BATCH_STATUS_PROCESSING, BATCH_STATUS_PAUSED},
    }


@router.post("/executions/{execution_id}/pause", summary="Pausar processamento do lote")
def pause_execution(execution_id: int, db: Session = Depends(get_db)):
    execution = db.query(BatchExecution).filter(BatchExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execucao nao encontrada")
    if execution.status not in {BATCH_STATUS_PENDING, BATCH_STATUS_PROCESSING}:
        raise HTTPException(status_code=409, detail="Somente lotes pendentes ou em processamento podem ser pausados.")

    execution.status = BATCH_STATUS_PAUSED
    execution.worker_id = None
    execution.heartbeat_at = None
    execution.lease_expires_at = None
    db.commit()
    return {"message": "Lote pausado com sucesso.", "status": execution.status}


@router.post("/executions/{execution_id}/resume", summary="Retomar processamento do lote")
def resume_execution(execution_id: int, db: Session = Depends(get_db)):
    execution = db.query(BatchExecution).filter(BatchExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execucao nao encontrada")
    if execution.status != BATCH_STATUS_PAUSED:
        raise HTTPException(status_code=409, detail="Somente lotes pausados podem ser retomados.")

    execution.status = BATCH_STATUS_PENDING
    execution.end_time = None
    execution.worker_id = None
    execution.heartbeat_at = None
    execution.lease_expires_at = None
    db.commit()
    return {"message": "Lote retornou para a fila e sera retomado pelo worker.", "status": execution.status}


@router.post("/executions/{execution_id}/cancel", summary="Cancelar processamento do lote")
def cancel_execution(execution_id: int, db: Session = Depends(get_db)):
    execution = db.query(BatchExecution).filter(BatchExecution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execucao nao encontrada")
    if execution.status in {BATCH_STATUS_COMPLETED, BATCH_STATUS_COMPLETED_WITH_ERRORS, BATCH_STATUS_CANCELLED}:
        raise HTTPException(status_code=409, detail="Este lote ja foi finalizado.")

    execution.status = BATCH_STATUS_CANCELLED
    execution.worker_id = None
    execution.heartbeat_at = None
    execution.lease_expires_at = None
    if execution.end_time is None:
        execution.end_time = datetime.now(timezone.utc)
    db.commit()
    return {"message": "Lote cancelado com sucesso.", "status": execution.status}


@router.get("/executions/{execution_id}/error-report", summary="Baixar relatorio CSV de erros do lote")
def download_error_report(
    execution_id: int,
    db: Session = Depends(get_db),
):
    execution = (
        db.query(BatchExecution)
        .options(joinedload(BatchExecution.items))
        .filter(BatchExecution.id == execution_id)
        .first()
    )
    if not execution:
        raise HTTPException(status_code=404, detail="Execucao nao encontrada")

    failed_items = [item for item in execution.items if item.status == "FALHA"]
    if not failed_items:
        raise HTTPException(status_code=404, detail="Esta execucao nao possui itens com falha.")

    csv_content = _build_error_report_csv(failed_items)
    return StreamingResponse(
        iter([csv_content.encode("utf-8")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="lote_{execution_id}_erros.csv"'},
    )


@router.get("/executions/{execution_id}", summary="Obter detalhes completos da execucao")
def get_execution_details(
    execution_id: int,
    db: Session = Depends(get_db),
):
    execution = (
        db.query(BatchExecution)
        .options(joinedload(BatchExecution.items))
        .filter(BatchExecution.id == execution_id)
        .first()
    )
    if not execution:
        raise HTTPException(status_code=404, detail="Execucao nao encontrada")
    return execution


@router.get("/legal-one-position-fix/status", summary="Acompanhar correcao de posicao do cliente principal")
def get_legal_one_position_fix_status():
    status_file = _resolve_legal_one_position_fix_status_file()
    if not status_file.exists():
        return {
            "available": False,
            "file_path": str(status_file),
            "generated_at": None,
            "state": None,
            "batch_size": None,
            "current_batch": None,
            "total_batches": None,
            "sleep_until": None,
            "control_file": None,
            "control_signal": _read_legal_one_position_fix_control_signal(),
            "total_items": 0,
            "processed_items": 0,
            "updated_count": 0,
            "failed_count": 0,
            "retry_pending_count": 0,
            "remaining_items": 0,
            "progress_percentage": 0,
            "average_update_seconds": None,
            "effective_average_seconds": None,
            "estimated_remaining_seconds": None,
            "estimated_completion_at": None,
            "active_queue_type": None,
            "retry_pass": None,
            "max_attempts": None,
            "workers": [],
            "items": [],
        }

    try:
        payload = json.loads(status_file.read_text(encoding="utf-8"))
        normalized = _normalize_legal_one_position_fix_payload(payload)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=500, detail=f"Nao foi possivel ler o andamento da correcao: {exc}") from exc

    total_items = normalized["totalItems"] or 0
    processed_items = normalized["processedItems"] or 0
    percentage = int((processed_items / total_items) * 100) if total_items > 0 else 0
    remaining_items = normalized["remainingItems"]
    timing_metrics = _build_legal_one_position_fix_timing_metrics(
        normalized["items"],
        processed_items=processed_items,
        remaining_items=remaining_items,
        generated_at=normalized["generatedAt"],
    )
    recent_items = list(reversed(normalized["items"][-25:]))

    return {
        "available": True,
        "file_path": str(status_file),
        "generated_at": normalized["generatedAt"],
        "state": normalized["state"],
        "batch_size": normalized["batchSize"],
        "current_batch": normalized["currentBatch"],
        "total_batches": normalized["totalBatches"],
        "sleep_until": normalized["sleepUntil"],
        "control_file": normalized["controlFile"],
        "control_signal": _read_legal_one_position_fix_control_signal(),
        "total_items": total_items,
        "processed_items": processed_items,
        "updated_count": normalized["updatedCount"],
        "failed_count": normalized["failedCount"],
        "retry_pending_count": normalized["retryPendingCount"],
        "remaining_items": remaining_items,
        "progress_percentage": percentage,
        "average_update_seconds": timing_metrics["averageUpdateSeconds"],
        "effective_average_seconds": timing_metrics["effectiveAverageSeconds"],
        "estimated_remaining_seconds": timing_metrics["estimatedRemainingSeconds"],
        "estimated_completion_at": timing_metrics["estimatedCompletionAt"],
        "active_queue_type": normalized["activeQueueType"],
        "retry_pass": normalized["retryPass"],
        "max_attempts": normalized["maxAttempts"],
        "workers": normalized["workers"],
        "items": recent_items,
    }


@router.post("/executions/{execution_id}/retry", summary="Reprocessar itens falhos")
async def retry_execution(
    execution_id: int,
    retry_data: Optional[RetryRequest] = Body(None),
    service: BatchTaskCreationService = Depends(get_batch_task_creation_service),
):
    target_ids = retry_data.item_ids if retry_data else None
    await service.retry_failed_items(
        original_execution_id=execution_id,
        target_item_ids=target_ids,
    )
    return {"message": "Itens elegiveis reenfileirados para reprocessamento."}
