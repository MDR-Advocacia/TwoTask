"""
Endpoints do fluxo "Agendar Prazos Iniciais".

Divisão em dois sub-routers:

* `intake_router` — ingresso da automação externa (autenticação por
  API key). Um único endpoint `POST /intake` que aceita multipart
  com o payload JSON + PDF da habilitação.
* `router` — operações internas (JWT + permissão `prazos_iniciais`):
  listagem, detalhe, preview do PDF e controle do ciclo de vida.

O `intake_router` é registrado sem `protected_dependencies` em main.py;
o `router` interno entra no mesmo padrão dos demais (JWT obrigatório).
"""

from __future__ import annotations

import json
import logging
from datetime import date, datetime
from typing import Any, List, Optional

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    Header,
    HTTPException,
    Path,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy.orm import Session, selectinload

from app.core import auth as auth_security
from app.core.config import settings
from app.core.dependencies import get_db
from app.models.legal_one import LegalOneOffice, LegalOneTaskSubType, LegalOneUser
from app.models.prazo_inicial import (
    INTAKE_SOURCE_EXTERNAL_API,
    INTAKE_SOURCE_USER_UPLOAD,
    INTAKE_STATUS_AWAITING_TEMPLATE_CONFIG,
    INTAKE_STATUS_CANCELLED,
    INTAKE_STATUS_CLASSIFICATION_ERROR,
    INTAKE_STATUS_CLASSIFIED,
    INTAKE_STATUS_IN_REVIEW,
    INTAKE_STATUS_LAWSUIT_NOT_FOUND,
    INTAKE_STATUS_RECEIVED,
    PrazoInicialIntake,
    PrazoInicialSugestao,
)
from app.models.prazo_inicial_task_template import PrazoInicialTaskTemplate
from app.services.classifier.prazos_iniciais_classifier import (
    PrazosIniciaisBatchClassifier,
)
from app.models.prazo_inicial import (
    INTAKE_STATUS_READY_TO_CLASSIFY,
    PIN_BATCH_STATUS_APPLIED,
    PIN_BATCH_STATUS_FAILED,
    PIN_BATCH_STATUS_IN_PROGRESS,
    PIN_BATCH_STATUS_READY,
    PIN_BATCH_STATUS_SUBMITTED,
    PrazoInicialBatch,
)
from app.services.classifier.prazos_iniciais_schema import (
    NATUREZAS_VALIDAS,
    PRODUTOS_VALIDOS,
    TIPO_PRAZO_AUDIENCIA,
    TIPO_PRAZO_JULGAMENTO,
    TIPOS_PRAZO_VALIDOS,
)
from app.services.prazos_iniciais.intake_service import IntakeService
from app.services.prazos_iniciais.storage import (
    PdfValidationError,
    resolve_pdf_path,
)

logger = logging.getLogger(__name__)


# ─── Helpers privados de parsing CSV ─────────────────────────────────
# Usados pelos filtros multi-valor (status, treated_by_user_id, etc.).
# Espelham os helpers de `publication_search_service` mas mantemos uma
# cópia local pra evitar dependência cruzada API → service.

def _parse_csv_ints(raw) -> list[int]:
    """
    Aceita None, str ("5,8"), int, ou lista[int|str]. Retorna lista de
    inteiros descartando entradas inválidas (não bloqueia a request).
    """
    if raw is None:
        return []
    if isinstance(raw, int):
        return [raw]
    if isinstance(raw, list):
        out: list[int] = []
        for item in raw:
            try:
                out.append(int(item))
            except (TypeError, ValueError):
                continue
        return out
    out: list[int] = []
    for chunk in str(raw).split(","):
        chunk = chunk.strip()
        if chunk.isdigit() or (chunk.startswith("-") and chunk[1:].isdigit()):
            out.append(int(chunk))
    return out


def _parse_csv_strs(raw) -> list[str]:
    """Aceita None, str ("a,b") ou lista. Retorna lista filtrada de strings."""
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    return [chunk.strip() for chunk in str(raw).split(",") if chunk.strip()]


# ═══════════════════════════════════════════════════════════════════════
# Schemas Pydantic
# ═══════════════════════════════════════════════════════════════════════


class ParteProcessual(BaseModel):
    """Representa uma parte do processo (autor ou réu)."""

    nome: str
    documento: Optional[str] = None  # CPF/CNPJ


class CapaProcesso(BaseModel):
    """Dados de capa extraídos pela automação externa."""

    tribunal: Optional[str] = None
    vara: Optional[str] = None
    classe: Optional[str] = None
    assunto: Optional[str] = None
    valor_causa: Optional[float] = None
    data_distribuicao: Optional[date] = None
    polo_ativo: list[ParteProcessual] = Field(default_factory=list)
    polo_passivo: list[ParteProcessual] = Field(default_factory=list)
    segredo_justica: bool = False

    # Permite que a automação envie campos adicionais sem quebrar o contrato.
    class Config:
        extra = "allow"


class PrazoInicialIntakePayload(BaseModel):
    """Body (JSON) do endpoint de ingestão. Acompanha o PDF no multipart."""

    external_id: str = Field(min_length=1, max_length=255)
    cnj_number: str = Field(min_length=1, max_length=32)
    capa: CapaProcesso
    # íntegra em blocos com data — estrutura ainda em definição, aceita qualquer JSON.
    integra_json: dict = Field(default_factory=dict)
    metadata: Optional[dict] = None


class IntakeResponse(BaseModel):
    intake_id: int
    external_id: str
    status: str
    pdf_stored_path: Optional[str] = None
    already_existed: bool = False


class SugestaoOut(BaseModel):
    id: int
    tipo_prazo: str
    subtipo: Optional[str]
    data_base: Optional[date]
    prazo_dias: Optional[int]
    prazo_tipo: Optional[str]
    data_final_calculada: Optional[date]
    # Campos novos Bloco B — prazo fatal.
    prazo_fatal_data: Optional[date] = None
    prazo_fatal_fundamentacao: Optional[str] = None
    prazo_base_decisao: Optional[str] = None
    audiencia_data: Optional[date]
    audiencia_hora: Optional[str]
    audiencia_link: Optional[str]
    confianca: Optional[str]
    justificativa: Optional[str]
    responsavel_sugerido_id: Optional[int]
    task_type_id: Optional[int]
    task_subtype_id: Optional[int]
    payload_proposto: Optional[dict]
    review_status: str
    reviewed_by_email: Optional[str]
    reviewed_at: Optional[datetime]
    created_task_id: Optional[int]
    created_at: datetime

    class Config:
        from_attributes = True


class IntakeSummary(BaseModel):
    id: int
    external_id: str
    cnj_number: str
    lawsuit_id: Optional[int]
    office_id: Optional[int]
    status: str
    # Classificação preliminar (Fase 3c). None enquanto o intake não foi
    # classificado ou veio de um fluxo pré-3c.
    natureza_processo: Optional[str] = None
    produto: Optional[str] = None
    # Campos novos Bloco C — info de agravo.
    agravo_processo_origem_cnj: Optional[str] = None
    agravo_decisao_agravada_resumo: Optional[str] = None
    # Campos novos Bloco E — agregados globais.
    valor_total_pedido: Optional[float] = None
    valor_total_estimado: Optional[float] = None
    aprovisionamento_sugerido: Optional[float] = None
    probabilidade_exito_global: Optional[str] = None
    analise_estrategica: Optional[str] = None
    error_message: Optional[str]
    pdf_filename_original: Optional[str]
    pdf_bytes: Optional[int]
    ged_document_id: Optional[int]
    ged_uploaded_at: Optional[datetime]
    received_at: datetime
    updated_at: datetime
    sugestoes_count: int
    # Tipos de prazo distintos das sugestoes do intake — usado pela UI
    # de listagem pra exibir a "classificacao" (ex.: ["CONTESTAR",
    # "AUDIENCIA"]). Mantem ordem de criacao das sugestoes pra
    # estabilidade visual. Lista vazia se intake nao foi classificado.
    tipos_prazo: list[str] = []
    # Data fatal mais proxima (min) entre as sugestoes do intake. Usada
    # pela UI de listagem pra exibir urgencia (cor por proximidade).
    # NULL se nenhuma sugestao tem prazo_fatal_data preenchido.
    prazo_fatal_mais_proximo: Optional[date] = None
    # Tratado por (pin011) — quem confirmou agendamentos OU finalizou
    # sem providência. NULL enquanto intake estiver em fluxo.
    treated_by_user_id: Optional[int] = None
    treated_by_email: Optional[str] = None
    treated_by_name: Optional[str] = None
    treated_at: Optional[datetime] = None
    # Disparo desacoplado (pin012, Onda 3 #5).
    dispatch_pending: bool = False
    dispatched_at: Optional[datetime] = None
    dispatch_error_message: Optional[str] = None
    # Origem do intake (pin016).
    source: str = INTAKE_SOURCE_EXTERNAL_API
    source_provider_name: Optional[str] = None
    submitted_by_user_id: Optional[int] = None
    submitted_by_email: Optional[str] = None
    submitted_by_name: Optional[str] = None
    submitted_at: Optional[datetime] = None
    pdf_extraction_failed: bool = False
    extractor_used: Optional[str] = None
    extraction_confidence: Optional[str] = None
    has_habilitacao_pdf: bool = False
    habilitacao_pdf_filename_original: Optional[str] = None
    habilitacao_pdf_bytes: Optional[int] = None
    # Patrocínio (pin018) — só presente quando o intake bateu com
    # vinculada Master. Operador vê na listagem como badge sumário.
    patrocinio_decisao: Optional[str] = None
    patrocinio_suspeita_devolucao: bool = False
    patrocinio_review_status: Optional[str] = None

    class Config:
        from_attributes = True


class PatrocinioOut(BaseModel):
    """Bloco completo de patrocínio do intake — vai no IntakeDetail."""
    id: int
    intake_id: int
    decisao: str
    outro_escritorio_nome: Optional[str] = None
    outro_advogado_nome: Optional[str] = None
    outro_advogado_oab: Optional[str] = None
    outro_advogado_data_habilitacao: Optional[date] = None
    suspeita_devolucao: bool = False
    motivo_suspeita: Optional[str] = None
    natureza_acao: Optional[str] = None
    polo_passivo_confirmado: bool = True
    polo_passivo_observacao: Optional[str] = None
    confianca: Optional[str] = None
    fundamentacao: Optional[str] = None
    review_status: str
    reviewed_by_user_id: Optional[int] = None
    reviewed_by_email: Optional[str] = None
    reviewed_by_name: Optional[str] = None
    reviewed_at: Optional[datetime] = None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class IntakeDetail(IntakeSummary):
    capa_json: dict
    metadata_json: Optional[dict]
    sugestoes: list[SugestaoOut]
    # Pedidos extraídos da PI (Bloco D2). Lista vazia se intake não
    # classificado ou classificado em versão antiga (pré-D2).
    pedidos: list[dict] = []
    # Patrocínio (pin018) — bloco opcional, presente apenas quando o
    # intake bateu com vinculada Master.
    patrocinio: Optional[PatrocinioOut] = None


class IntakeListResponse(BaseModel):
    total: int
    items: list[IntakeSummary]


# ═══════════════════════════════════════════════════════════════════════
# Intake router (API externa — auth por API key)
# ═══════════════════════════════════════════════════════════════════════

intake_router = APIRouter(prefix="/prazos-iniciais", tags=["Prazos Iniciais"])


def _validate_api_key(
    x_intake_api_key: Optional[str] = Header(default=None, alias="X-Intake-Api-Key"),
) -> str:
    """
    Dependency que autentica a automação externa por header `X-Intake-Api-Key`.

    Aceita múltiplas chaves configuradas em `PRAZOS_INICIAIS_API_KEY`
    (separadas por vírgula) para permitir rotação sem downtime.
    """
    valid_keys = settings.prazos_iniciais_api_keys
    if not valid_keys:
        logger.error(
            "PRAZOS_INICIAIS_API_KEY não configurada — intake rejeitado."
        )
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Endpoint de intake não configurado.",
        )
    if not x_intake_api_key or x_intake_api_key not in valid_keys:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="API key inválida ou ausente.",
        )
    return x_intake_api_key


@intake_router.post(
    "/intake",
    response_model=IntakeResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Recebe um processo novo da automação externa (multipart).",
)
async def ingest_intake(
    background_tasks: BackgroundTasks,
    payload: str = Form(
        ...,
        description="JSON serializado conforme PrazoInicialIntakePayload.",
    ),
    habilitacao: UploadFile = File(
        ...,
        description="PDF da habilitação nos autos (application/pdf).",
    ),
    _: str = Depends(_validate_api_key),
    db: Session = Depends(get_db),
):
    """
    Recebe um processo novo para a fila de classificação.

    Request: multipart/form-data com:
      - `payload`: JSON serializado com os dados do processo
      - `habilitacao`: PDF da habilitação (≤ PRAZOS_INICIAIS_MAX_PDF_MB)

    Resposta: 202 Accepted em criação nova, 200 OK em reenvio idempotente
    (mesmo `external_id`). Em ambos os casos, o `intake_id` retornado é
    estável para o ciclo de vida do processo.
    """
    # 1. Parse do JSON (o FastAPI já validou que o campo existe, mas é string).
    # Recovery de encoding: python-multipart decoda form fields em Latin-1
    # quando o cliente não envia charset explícito, corrompendo UTF-8
    # (Í vira "Ã\x8D" etc.). Re-encodamos pra Latin-1 → decodamos como
    # UTF-8 — operação idempotente pra strings já corretas (cai no
    # UnicodeDecodeError do fallback).
    try:
        recovered = payload.encode("latin-1").decode("utf-8")
        payload = recovered
    except (UnicodeEncodeError, UnicodeDecodeError):
        # Já era UTF-8 válido, ou contém caracteres fora de Latin-1 —
        # segue com o payload original sem tocar.
        pass

    try:
        payload_dict = json.loads(payload)
        intake_payload = PrazoInicialIntakePayload.model_validate(payload_dict)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Campo 'payload' não é JSON válido: {exc}",
        )
    except ValidationError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=exc.errors(),
        )

    # 2. Valida MIME declarado (a validação real — magic bytes — é no storage).
    content_type = (habilitacao.content_type or "").lower()
    if content_type and content_type not in ("application/pdf", "application/octet-stream"):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"Content-Type não suportado: {content_type}. Esperado application/pdf.",
        )

    # 3. Lê os bytes do PDF (com limite defensivo).
    max_bytes = settings.prazos_iniciais_max_pdf_bytes
    pdf_bytes = await habilitacao.read()
    if len(pdf_bytes) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"PDF excede {settings.prazos_iniciais_max_pdf_mb} MB "
                f"(recebido: {len(pdf_bytes)} bytes)."
            ),
        )

    service = IntakeService(db=db)

    # 4. Se já existe, retorna sem reprocessar (mesmo sem ler o PDF adiante).
    existing = service.get_by_external_id(intake_payload.external_id)
    if existing is not None:
        return IntakeResponse(
            intake_id=existing.id,
            external_id=existing.external_id,
            status=existing.status,
            pdf_stored_path=existing.pdf_path,
            already_existed=True,
        )

    # 5. Cria o intake (inclui gravação + validação dos bytes do PDF).
    try:
        result = service.create_intake(
            external_id=intake_payload.external_id,
            cnj_number=intake_payload.cnj_number,
            capa_json=intake_payload.capa.model_dump(mode="json"),
            integra_json=intake_payload.integra_json,
            metadata_json=intake_payload.metadata,
            pdf_bytes=pdf_bytes,
            pdf_filename_original=habilitacao.filename,
            source=INTAKE_SOURCE_EXTERNAL_API,
        )
    except PdfValidationError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )
    except ValueError as exc:
        # normalize_cnj e afins
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )

    # 6. Dispara resolução do lawsuit em background (não bloqueia a resposta).
    background_tasks.add_task(
        service.resolve_lawsuit_for_intake, result.intake.id
    )

    return IntakeResponse(
        intake_id=result.intake.id,
        external_id=result.intake.external_id,
        status=result.intake.status,
        pdf_stored_path=result.intake.pdf_path,
        already_existed=False,
    )


# ═══════════════════════════════════════════════════════════════════════
# Router interno (JWT + permissão `prazos_iniciais`)
# ═══════════════════════════════════════════════════════════════════════

router = APIRouter(prefix="/prazos-iniciais", tags=["Prazos Iniciais"])


def _intake_to_summary(intake: PrazoInicialIntake) -> IntakeSummary:
    # Agrega tipos_prazo distintos das sugestoes preservando a ordem de
    # criacao (estabilidade visual na listagem). Lista vazia se intake
    # nao foi classificado ou nao tem sugestoes ainda.
    seen_tipos: set[str] = set()
    tipos_prazo: list[str] = []
    prazos_fatais: list[date] = []
    for s in intake.sugestoes or []:
        tp = s.tipo_prazo
        if tp and tp not in seen_tipos:
            seen_tipos.add(tp)
            tipos_prazo.append(tp)
        if s.prazo_fatal_data is not None:
            prazos_fatais.append(s.prazo_fatal_data)
    # Min = data mais proxima/iminente (prazo fatal mais critico).
    prazo_fatal_mais_proximo = min(prazos_fatais) if prazos_fatais else None

    return IntakeSummary(
        id=intake.id,
        external_id=intake.external_id,
        cnj_number=intake.cnj_number,
        lawsuit_id=intake.lawsuit_id,
        office_id=intake.office_id,
        status=intake.status,
        natureza_processo=intake.natureza_processo,
        produto=intake.produto,
        agravo_processo_origem_cnj=intake.agravo_processo_origem_cnj,
        agravo_decisao_agravada_resumo=intake.agravo_decisao_agravada_resumo,
        valor_total_pedido=float(intake.valor_total_pedido) if intake.valor_total_pedido is not None else None,
        valor_total_estimado=float(intake.valor_total_estimado) if intake.valor_total_estimado is not None else None,
        aprovisionamento_sugerido=float(intake.aprovisionamento_sugerido) if intake.aprovisionamento_sugerido is not None else None,
        probabilidade_exito_global=intake.probabilidade_exito_global,
        analise_estrategica=intake.analise_estrategica,
        error_message=intake.error_message,
        pdf_filename_original=intake.pdf_filename_original,
        pdf_bytes=intake.pdf_bytes,
        ged_document_id=intake.ged_document_id,
        ged_uploaded_at=intake.ged_uploaded_at,
        received_at=intake.received_at,
        updated_at=intake.updated_at,
        sugestoes_count=len(intake.sugestoes or []),
        tipos_prazo=tipos_prazo,
        prazo_fatal_mais_proximo=prazo_fatal_mais_proximo,
        treated_by_user_id=intake.treated_by_user_id,
        treated_by_email=intake.treated_by_email,
        treated_by_name=intake.treated_by_name,
        treated_at=intake.treated_at,
        dispatch_pending=bool(getattr(intake, "dispatch_pending", False)),
        dispatched_at=getattr(intake, "dispatched_at", None),
        dispatch_error_message=getattr(intake, "dispatch_error_message", None),
        source=getattr(intake, "source", INTAKE_SOURCE_EXTERNAL_API) or INTAKE_SOURCE_EXTERNAL_API,
        source_provider_name=getattr(intake, "source_provider_name", None),
        submitted_by_user_id=getattr(intake, "submitted_by_user_id", None),
        submitted_by_email=getattr(intake, "submitted_by_email", None),
        submitted_by_name=getattr(intake, "submitted_by_name", None),
        submitted_at=getattr(intake, "submitted_at", None),
        pdf_extraction_failed=bool(getattr(intake, "pdf_extraction_failed", False)),
        extractor_used=getattr(intake, "extractor_used", None),
        extraction_confidence=getattr(intake, "extraction_confidence", None),
        has_habilitacao_pdf=bool(getattr(intake, "habilitacao_pdf_path", None)),
        habilitacao_pdf_filename_original=getattr(
            intake, "habilitacao_pdf_filename_original", None
        ),
        habilitacao_pdf_bytes=getattr(intake, "habilitacao_pdf_bytes", None),
        patrocinio_decisao=(
            intake.patrocinio.decisao if getattr(intake, "patrocinio", None) else None
        ),
        patrocinio_suspeita_devolucao=bool(
            intake.patrocinio.suspeita_devolucao
            if getattr(intake, "patrocinio", None) else False
        ),
        patrocinio_review_status=(
            intake.patrocinio.review_status if getattr(intake, "patrocinio", None) else None
        ),
    )


def _sugestao_to_out(sugestao: PrazoInicialSugestao) -> SugestaoOut:
    return SugestaoOut(
        id=sugestao.id,
        tipo_prazo=sugestao.tipo_prazo,
        subtipo=sugestao.subtipo,
        data_base=sugestao.data_base,
        prazo_dias=sugestao.prazo_dias,
        prazo_tipo=sugestao.prazo_tipo,
        data_final_calculada=sugestao.data_final_calculada,
        prazo_fatal_data=sugestao.prazo_fatal_data,
        prazo_fatal_fundamentacao=sugestao.prazo_fatal_fundamentacao,
        prazo_base_decisao=sugestao.prazo_base_decisao,
        audiencia_data=sugestao.audiencia_data,
        audiencia_hora=(
            sugestao.audiencia_hora.strftime("%H:%M")
            if sugestao.audiencia_hora
            else None
        ),
        audiencia_link=sugestao.audiencia_link,
        confianca=sugestao.confianca,
        justificativa=sugestao.justificativa,
        responsavel_sugerido_id=sugestao.responsavel_sugerido_id,
        task_type_id=sugestao.task_type_id,
        task_subtype_id=sugestao.task_subtype_id,
        payload_proposto=sugestao.payload_proposto,
        review_status=sugestao.review_status,
        reviewed_by_email=sugestao.reviewed_by_email,
        reviewed_at=sugestao.reviewed_at,
        created_task_id=sugestao.created_task_id,
        created_at=sugestao.created_at,
    )


class EnumsResponse(BaseModel):
    """
    Dicionário de valores válidos pros selects da UI (CRUD de templates,
    filtros de intakes, etc). Público dentro do router protegido; a UI
    carrega uma vez e cacheia.
    """

    tipos_prazo: list[str]
    naturezas: list[str]
    produtos: list[str]
    subtipos_audiencia: list[str]
    subtipos_julgamento: list[str]
    priorities: list[str]
    due_date_references: list[str]


@router.get(
    "/enums",
    response_model=EnumsResponse,
    summary="Enumerações válidas do fluxo (para popular selects da UI).",
)
def get_enums(
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Retorna as enumerações usadas pelos formulários de CRUD de templates
    e filtros de intakes. Evita espalhar as listas hardcoded no frontend.

    Convenção: listas ordenadas alfabeticamente; strings vazias NÃO são
    incluídas (a UI usa `null` pra "qualquer"/"global").
    """
    return EnumsResponse(
        tipos_prazo=sorted(TIPOS_PRAZO_VALIDOS),
        naturezas=sorted(NATUREZAS_VALIDAS),
        produtos=sorted(PRODUTOS_VALIDOS),
        subtipos_audiencia=sorted(SUBTIPOS_AUDIENCIA),
        subtipos_julgamento=sorted(SUBTIPOS_JULGAMENTO),
        priorities=sorted(PRIORIDADES_VALIDAS),
        due_date_references=sorted(DUE_DATE_REFERENCES_VALIDAS),
    )


def _parse_csv_strs(value: Optional[str]) -> list[str]:
    """Divide string CSV em lista de strings limpas. 'a,b, c ' → ['a','b','c']."""
    if not value:
        return []
    return [v.strip() for v in value.split(",") if v.strip()]


def _parse_csv_ints(value: Optional[str]) -> list[int]:
    """Divide string CSV em lista de ints. '61, 62' → [61, 62]. Tolerante a erros."""
    out: list[int] = []
    for v in _parse_csv_strs(value):
        try:
            out.append(int(v))
        except (ValueError, TypeError):
            continue
    return out


@router.get(
    "/intakes",
    response_model=IntakeListResponse,
    summary="Lista intakes com filtros.",
)
def list_intakes(
    status_filter: Optional[str] = Query(
        default=None,
        alias="status",
        description="Aceita CSV: 'CLASSIFICADO,AGENDADO'.",
    ),
    office_id: Optional[str] = Query(
        default=None,
        description="CSV de office_ids: '61,62'.",
    ),
    cnj_number: Optional[str] = Query(default=None),
    natureza_processo: Optional[str] = Query(
        default=None,
        description="CSV: 'COMUM,JUIZADO,AGRAVO_INSTRUMENTO,OUTRO'.",
    ),
    produto: Optional[str] = Query(
        default=None,
        description="CSV: 'SUPERENDIVIDAMENTO,CREDCESTA,...'.",
    ),
    probabilidade_exito_global: Optional[str] = Query(
        default=None,
        description="CSV: 'remota,possivel,provavel'.",
    ),
    date_from: Optional[str] = Query(
        default=None,
        description="Data início (YYYY-MM-DD). Filtra por received_at >=.",
    ),
    date_to: Optional[str] = Query(
        default=None,
        description="Data fim (YYYY-MM-DD). Filtra por received_at < date_to+1dia.",
    ),
    has_error: Optional[bool] = Query(
        default=None,
        description="true = só com error_message; false = só sem; omitido = ambos.",
    ),
    batch_id: Optional[int] = Query(
        default=None,
        description="Filtra intakes de um batch de classificação específico.",
    ),
    treated_by_user_id: Optional[str] = Query(
        default=None,
        description="CSV de user_ids: '5,8'. Filtra por quem confirmou/finalizou.",
    ),
    dispatch_pending: Optional[bool] = Query(
        default=None,
        description="true = só pendentes de disparo (Tratamento Web); false = já disparados; omitido = ambos.",
    ),
    source: Optional[str] = Query(
        default=None,
        description="CSV de origens: 'EXTERNAL_API,USER_UPLOAD'.",
    ),
    submitted_by_user_id: Optional[str] = Query(
        default=None,
        description="CSV de user_ids: '5,8'. Filtra por quem submeteu (USER_UPLOAD). Atalho 'Minha fila'.",
    ),
    pdf_extraction_failed: Optional[bool] = Query(
        default=None,
        description="true = só uploads com extração falha (classificação manual). Omitido = ambos.",
    ),
    patrocinio_decisao: Optional[str] = Query(
        default=None,
        description="CSV de decisões: 'MDR_ADVOCACIA,OUTRO_ESCRITORIO,CONDUCAO_INTERNA'.",
    ),
    patrocinio_suspeita_devolucao: Optional[bool] = Query(
        default=None,
        description="true = só intakes marcados pra devolução; false = só não-suspeitos.",
    ),
    patrocinio_review_status: Optional[str] = Query(
        default=None,
        description="CSV: 'pendente,aprovado,editado,rejeitado'.",
    ),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    # selectinload das sugestoes evita N+1 quando o summary agrega
    # tipos_prazo. selectin (em vez de joined) gera 1 SELECT extra com
    # IN no intake_ids — mais leve que multiplicar linhas via JOIN
    # quando ha muitas sugestoes por intake.
    query = db.query(PrazoInicialIntake).options(
        selectinload(PrazoInicialIntake.sugestoes),
        selectinload(PrazoInicialIntake.patrocinio),
    )

    # status — CSV, IN quando múltiplo
    status_list = _parse_csv_strs(status_filter)
    if status_list:
        if len(status_list) == 1:
            query = query.filter(PrazoInicialIntake.status == status_list[0])
        else:
            query = query.filter(PrazoInicialIntake.status.in_(status_list))

    # office_id — CSV de ints
    office_ids = _parse_csv_ints(office_id)
    if office_ids:
        if len(office_ids) == 1:
            query = query.filter(PrazoInicialIntake.office_id == office_ids[0])
        else:
            query = query.filter(PrazoInicialIntake.office_id.in_(office_ids))

    if cnj_number:
        # aceita busca por pedaço do CNJ (sem máscara)
        normalized = "".join(c for c in cnj_number if c.isdigit())
        if normalized:
            query = query.filter(
                PrazoInicialIntake.cnj_number.like(f"%{normalized}%")
            )

    # natureza_processo — CSV
    natureza_list = _parse_csv_strs(natureza_processo)
    if natureza_list:
        if len(natureza_list) == 1:
            query = query.filter(
                PrazoInicialIntake.natureza_processo == natureza_list[0]
            )
        else:
            query = query.filter(
                PrazoInicialIntake.natureza_processo.in_(natureza_list)
            )

    # produto — CSV
    produto_list = _parse_csv_strs(produto)
    if produto_list:
        if len(produto_list) == 1:
            query = query.filter(PrazoInicialIntake.produto == produto_list[0])
        else:
            query = query.filter(PrazoInicialIntake.produto.in_(produto_list))

    # probabilidade_exito_global — CSV
    prob_list = [p.lower() for p in _parse_csv_strs(probabilidade_exito_global)]
    if prob_list:
        if len(prob_list) == 1:
            query = query.filter(
                PrazoInicialIntake.probabilidade_exito_global == prob_list[0]
            )
        else:
            query = query.filter(
                PrazoInicialIntake.probabilidade_exito_global.in_(prob_list)
            )

    # Data range por received_at — comparação lexicográfica segura pq é timestamptz
    if date_from:
        query = query.filter(PrazoInicialIntake.received_at >= date_from)
    if date_to:
        # Inclusivo no dia final: compara com "<= date_to 23:59:59".
        query = query.filter(PrazoInicialIntake.received_at <= f"{date_to}T23:59:59.999999+00:00")

    # Presença de erro (independente do status — útil pra triagem)
    if has_error is True:
        query = query.filter(
            PrazoInicialIntake.error_message.isnot(None),
            PrazoInicialIntake.error_message != "",
        )
    elif has_error is False:
        query = query.filter(
            (PrazoInicialIntake.error_message.is_(None))
            | (PrazoInicialIntake.error_message == "")
        )

    # Batch de classificação
    if batch_id is not None:
        query = query.filter(PrazoInicialIntake.classification_batch_id == batch_id)

    # Tratado por (CSV de user_ids)
    treated_by_ids = _parse_csv_ints(treated_by_user_id)
    if treated_by_ids:
        if len(treated_by_ids) == 1:
            query = query.filter(
                PrazoInicialIntake.treated_by_user_id == treated_by_ids[0]
            )
        else:
            query = query.filter(
                PrazoInicialIntake.treated_by_user_id.in_(treated_by_ids)
            )

    # Onda 3 #5 — pendentes de disparo (Tratamento Web)
    if dispatch_pending is True:
        query = query.filter(PrazoInicialIntake.dispatch_pending.is_(True))
    elif dispatch_pending is False:
        query = query.filter(PrazoInicialIntake.dispatch_pending.is_(False))

    # Origem do intake (pin016)
    source_list = _parse_csv_strs(source)
    if source_list:
        if len(source_list) == 1:
            query = query.filter(PrazoInicialIntake.source == source_list[0])
        else:
            query = query.filter(PrazoInicialIntake.source.in_(source_list))

    # Quem submeteu (USER_UPLOAD) — atalho "Minha fila"
    submitted_by_ids = _parse_csv_ints(submitted_by_user_id)
    if submitted_by_ids:
        if len(submitted_by_ids) == 1:
            query = query.filter(
                PrazoInicialIntake.submitted_by_user_id == submitted_by_ids[0]
            )
        else:
            query = query.filter(
                PrazoInicialIntake.submitted_by_user_id.in_(submitted_by_ids)
            )

    # Extração mecânica falhou (operador classifica manualmente)
    if pdf_extraction_failed is True:
        query = query.filter(PrazoInicialIntake.pdf_extraction_failed.is_(True))
    elif pdf_extraction_failed is False:
        query = query.filter(PrazoInicialIntake.pdf_extraction_failed.is_(False))

    # Patrocínio (pin018) — filtros via join lazy na tabela 1:1.
    patrocinio_filtros_aplicados = (
        bool(_parse_csv_strs(patrocinio_decisao))
        or patrocinio_suspeita_devolucao is not None
        or bool(_parse_csv_strs(patrocinio_review_status))
    )
    if patrocinio_filtros_aplicados:
        from app.models.prazo_inicial_patrocinio import PrazoInicialPatrocinio
        query = query.join(
            PrazoInicialPatrocinio,
            PrazoInicialPatrocinio.intake_id == PrazoInicialIntake.id,
        )
        decisao_list = _parse_csv_strs(patrocinio_decisao)
        if decisao_list:
            if len(decisao_list) == 1:
                query = query.filter(PrazoInicialPatrocinio.decisao == decisao_list[0])
            else:
                query = query.filter(PrazoInicialPatrocinio.decisao.in_(decisao_list))
        if patrocinio_suspeita_devolucao is True:
            query = query.filter(
                PrazoInicialPatrocinio.suspeita_devolucao.is_(True)
            )
        elif patrocinio_suspeita_devolucao is False:
            query = query.filter(
                PrazoInicialPatrocinio.suspeita_devolucao.is_(False)
            )
        review_list = _parse_csv_strs(patrocinio_review_status)
        if review_list:
            if len(review_list) == 1:
                query = query.filter(
                    PrazoInicialPatrocinio.review_status == review_list[0]
                )
            else:
                query = query.filter(
                    PrazoInicialPatrocinio.review_status.in_(review_list)
                )

    total = query.count()
    items = (
        query.order_by(PrazoInicialIntake.received_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return IntakeListResponse(
        total=total,
        items=[_intake_to_summary(item) for item in items],
    )


@router.get(
    "/intakes/{intake_id}",
    response_model=IntakeDetail,
    summary="Detalhe de um intake, com sugestões.",
)
def get_intake(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")

    # Auto-recompute defensivo (Onda 2): se por qualquer motivo os
    # agregados do intake ficarem dessincronizados dos pedidos (apply
    # antigo, exception silenciosa, PATCH direto no banco), o GET
    # reconcilia transparentemente. Custo: 1 passada sobre pedidos já
    # eager-loaded. Sem I/O extra além do UPDATE quando algo mudou.
    try:
        from app.services.classifier.prazos_iniciais_classifier import (
            PrazosIniciaisBatchClassifier,
        )
        clf = PrazosIniciaisBatchClassifier.__new__(PrazosIniciaisBatchClassifier)
        clf.db = db
        # Snapshot do estado antes pra só commitar se mudou algo.
        before = (
            intake.valor_total_pedido,
            intake.valor_total_estimado,
            intake.aprovisionamento_sugerido,
            intake.probabilidade_exito_global,
        )
        clf._compute_intake_globals(intake)
        after = (
            intake.valor_total_pedido,
            intake.valor_total_estimado,
            intake.aprovisionamento_sugerido,
            intake.probabilidade_exito_global,
        )
        if before != after:
            db.commit()
            db.refresh(intake)
    except Exception:  # noqa: BLE001
        logger.exception(
            "Auto-recompute de agregados falhou (intake %s) — seguindo com valores antigos.",
            intake_id,
        )
        db.rollback()

    summary = _intake_to_summary(intake)
    sugestoes_sorted = sorted(
        intake.sugestoes or [],
        key=lambda s: (s.review_status != "pendente", s.id),
    )
    pedidos_serialized = [
        {
            "id": p.id,
            "intake_id": p.intake_id,
            "tipo_pedido": p.tipo_pedido,
            "natureza": p.natureza,
            "valor_indicado": float(p.valor_indicado) if p.valor_indicado is not None else None,
            "valor_estimado": float(p.valor_estimado) if p.valor_estimado is not None else None,
            "fundamentacao_valor": p.fundamentacao_valor,
            "probabilidade_perda": p.probabilidade_perda,
            "aprovisionamento": float(p.aprovisionamento) if p.aprovisionamento is not None else None,
            "fundamentacao_risco": p.fundamentacao_risco,
        }
        for p in (intake.pedidos or [])
    ]
    patrocinio_out: Optional[PatrocinioOut] = None
    if getattr(intake, "patrocinio", None) is not None:
        patrocinio_out = PatrocinioOut.model_validate(intake.patrocinio)

    return IntakeDetail(
        **summary.model_dump(),
        capa_json=intake.capa_json,
        metadata_json=intake.metadata_json,
        sugestoes=[_sugestao_to_out(s) for s in sugestoes_sorted],
        pedidos=pedidos_serialized,
        patrocinio=patrocinio_out,
    )


@router.get(
    "/intakes/{intake_id}/pdf",
    summary="Stream do PDF da habilitação (preview na UI).",
)
def get_intake_pdf(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")
    if not intake.pdf_path:
        raise HTTPException(
            status_code=410,
            detail="PDF não está mais disponível localmente (retenção expirada).",
        )

    try:
        absolute = resolve_pdf_path(intake.pdf_path)
    except ValueError:
        logger.error(
            "pdf_path inválido no intake %d: %r", intake_id, intake.pdf_path
        )
        raise HTTPException(status_code=500, detail="Caminho do PDF inválido.")

    if not absolute.exists():
        raise HTTPException(status_code=404, detail="Arquivo PDF não encontrado.")

    filename = intake.pdf_filename_original or f"habilitacao-{intake_id}.pdf"
    return FileResponse(
        path=absolute,
        media_type="application/pdf",
        filename=filename,
    )


@router.get(
    "/intakes/{intake_id}/habilitacao-pdf",
    summary="Stream do PDF da habilitação MDR (separado do processo).",
)
def get_intake_habilitacao_pdf(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Serve o PDF de habilitação MDR (procuração + carta de preposição)
    quando enviado junto com USER_UPLOAD. Distinto de
    GET /intakes/{id}/pdf — esse último é o "PDF principal" do intake
    (habilitação no fluxo EXTERNAL_API legado, ou processo na íntegra
    quando a extração USER_UPLOAD falhou e o arquivo foi preservado).
    """
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")
    habilitacao_path = getattr(intake, "habilitacao_pdf_path", None)
    if not habilitacao_path:
        raise HTTPException(
            status_code=404,
            detail="Este intake não tem PDF de habilitação anexado.",
        )

    try:
        absolute = resolve_pdf_path(habilitacao_path)
    except ValueError:
        logger.error(
            "habilitacao_pdf_path inválido no intake %d: %r",
            intake_id, habilitacao_path,
        )
        raise HTTPException(status_code=500, detail="Caminho do PDF inválido.")

    if not absolute.exists():
        raise HTTPException(status_code=404, detail="Arquivo PDF não encontrado.")

    filename = (
        getattr(intake, "habilitacao_pdf_filename_original", None)
        or f"habilitacao-{intake_id}.pdf"
    )
    return FileResponse(
        path=absolute,
        media_type="application/pdf",
        filename=filename,
    )


# ─── USER_UPLOAD: operador sobe processo na íntegra (pin016) ───────

class UserUploadResponse(BaseModel):
    intake_id: int
    external_id: str
    status: str
    extractor_used: Optional[str] = None
    extraction_confidence: Optional[str] = None
    pdf_extraction_failed: bool = False
    has_habilitacao_pdf: bool = False
    already_existed: bool = False
    # Mensagem traduzida pra UI exibir como "info" ou "warning". Vazia
    # quando a extração foi 100% mecânica e sem alertas.
    user_message: Optional[str] = None


@router.post(
    "/intake/upload",
    response_model=UserUploadResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary=(
        "Operador sobe PDF do processo na íntegra (extração mecânica)."
    ),
)
async def upload_intake_pdf(
    background_tasks: BackgroundTasks,
    processo_pdf: UploadFile = File(
        ...,
        description="PDF do processo na íntegra (≤ PRAZOS_INICIAIS_MAX_UPLOAD_PDF_MB).",
    ),
    habilitacao_pdf: Optional[UploadFile] = File(
        default=None,
        description="(Opcional) PDF de habilitação MDR — preservado pro GED+AJUS.",
    ),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    """
    Cria um intake a partir de um PDF subido pelo operador.

    Fluxo:
    1. Lê o PDF do processo, valida tamanho.
    2. Roda o motor de extração (pdfplumber + extractor PJe TJBA ou
       fallback texto cru).
    3. Se extração OK: cria intake com `source=USER_UPLOAD`,
       `submitted_by_*` do JWT; PDF do processo NÃO é gravado em disco
       (só os bytes/SHA pra auditoria) — economiza armazenamento.
    4. Se extração falha (PDF escaneado/sem texto): cria intake com
       `pdf_extraction_failed=True`, MANTÉM o PDF salvo, capa/integra
       vazias. Operador classifica manualmente no HITL.
    5. Se `habilitacao_pdf` foi enviado: SEMPRE preservado em
       `habilitacao_pdf_path` (vai pro GED L1 + AJUS).
    6. Idempotência: SHA256 do PDF do processo é a base do
       `external_id` (`upload-{sha8}`). Re-subir o mesmo PDF retorna
       o intake existente.
    """
    from app.services.prazos_iniciais.pdf_extractor import extract as pdf_extract
    from app.services.prazos_iniciais.storage import (
        validate_pdf_bytes,
    )
    import hashlib as _hashlib

    # 1. Lê PDF do processo (limite específico do upload manual).
    max_bytes = settings.prazos_iniciais_max_upload_pdf_bytes
    processo_bytes = await processo_pdf.read()
    if not processo_bytes:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="O arquivo do processo está vazio.",
        )
    if len(processo_bytes) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"O PDF do processo excede {settings.prazos_iniciais_max_upload_pdf_mb} MB "
                f"(recebido: {len(processo_bytes) / 1024 / 1024:.1f} MB)."
            ),
        )

    # Magic bytes — falha cedo se não for PDF.
    try:
        validate_pdf_bytes(processo_bytes)
    except PdfValidationError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Arquivo do processo inválido: {exc}",
        )

    # 2. Habilitação (opcional).
    habilitacao_bytes: Optional[bytes] = None
    habilitacao_filename: Optional[str] = None
    if habilitacao_pdf is not None:
        habilitacao_bytes = await habilitacao_pdf.read()
        if habilitacao_bytes:
            if len(habilitacao_bytes) > settings.prazos_iniciais_max_pdf_bytes:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail=(
                        f"PDF de habilitação excede {settings.prazos_iniciais_max_pdf_mb} MB."
                    ),
                )
            try:
                validate_pdf_bytes(habilitacao_bytes)
            except PdfValidationError as exc:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"Arquivo de habilitação inválido: {exc}",
                )
            habilitacao_filename = habilitacao_pdf.filename
        else:
            habilitacao_bytes = None  # arquivo vazio = ignora silenciosamente

    # 3. Idempotência por SHA256 do PDF do processo.
    sha = _hashlib.sha256(processo_bytes).hexdigest()
    external_id = f"upload-{sha[:16]}"

    service = IntakeService(db=db)
    existing = service.get_by_external_id(external_id)
    if existing is not None:
        return UserUploadResponse(
            intake_id=existing.id,
            external_id=existing.external_id,
            status=existing.status,
            extractor_used=existing.extractor_used,
            extraction_confidence=existing.extraction_confidence,
            pdf_extraction_failed=bool(existing.pdf_extraction_failed),
            has_habilitacao_pdf=bool(existing.habilitacao_pdf_path),
            already_existed=True,
            user_message="Este PDF já tinha sido cadastrado antes.",
        )

    # 4. Roda o motor de extração.
    extraction = pdf_extract(processo_bytes)

    # 5. Resolve CNJ — preferir o do extractor; sem ele, criar intake
    #    sem CNJ válido é problemático porque cnj_number é NOT NULL.
    cnj_number = extraction.cnj_number
    if not cnj_number:
        # Sem CNJ não dá pra criar intake (campo NOT NULL e o L1 não
        # vai resolver nada). Devolvemos erro traduzido pro operador.
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "Não foi possível identificar o número do processo (CNJ) "
                "neste PDF. Verifique se o arquivo está completo e que a "
                "primeira página contém a capa do processo."
            ),
        )

    # 6. Cria intake.
    pdf_extraction_failed = not extraction.success

    try:
        result = service.create_intake(
            external_id=external_id,
            cnj_number=cnj_number,
            capa_json=extraction.capa_json or {},
            integra_json=extraction.integra_json or {},
            metadata_json={
                "user_upload": True,
                "extractor_used": extraction.extractor_used,
                "extraction_confidence": extraction.confidence,
                "pdf_sha256": sha,
                "pdf_size_bytes": len(processo_bytes),
                "submitted_via": "upload_endpoint_v1",
            },
            pdf_bytes=processo_bytes,
            pdf_filename_original=processo_pdf.filename,
            source=INTAKE_SOURCE_USER_UPLOAD,
            submitted_by_user_id=current_user.id,
            submitted_by_email=current_user.email,
            submitted_by_name=current_user.name,
            pdf_extraction_failed=pdf_extraction_failed,
            extractor_used=extraction.extractor_used,
            extraction_confidence=extraction.confidence,
            habilitacao_pdf_bytes=habilitacao_bytes,
            habilitacao_pdf_filename_original=habilitacao_filename,
            # Quando a extração foi bem-sucedida, descarta o PDF do
            # processo (capa+integra já estão no JSON, não precisamos
            # do arquivo bruto). Falha de extração mantém o PDF pra
            # operador conseguir baixar e classificar manualmente.
            skip_pdf_storage=not pdf_extraction_failed,
        )
    except PdfValidationError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )

    # 7. Resolve lawsuit em background (não bloqueia a resposta).
    background_tasks.add_task(
        service.resolve_lawsuit_for_intake, result.intake.id
    )

    # 8. Mensagem pro operador (sem JSON cru — apenas info acionável).
    if pdf_extraction_failed:
        user_message = (
            "Processo cadastrado, mas o texto não pôde ser extraído "
            "automaticamente (PDF escaneado ou sem texto). "
            "Você precisa classificar manualmente na tela de tratamento."
        )
    elif extraction.confidence == "low":
        user_message = (
            "Processo cadastrado. A extração capturou pouca informação — "
            "o motor de classificação vai tentar completar."
        )
    elif extraction.confidence == "partial":
        user_message = (
            "Processo cadastrado. Alguns campos da capa não foram capturados "
            "automaticamente — o motor de classificação vai completar."
        )
    else:
        user_message = "Processo cadastrado. Classificação em andamento."

    return UserUploadResponse(
        intake_id=result.intake.id,
        external_id=result.intake.external_id,
        status=result.intake.status,
        extractor_used=result.intake.extractor_used,
        extraction_confidence=result.intake.extraction_confidence,
        pdf_extraction_failed=bool(result.intake.pdf_extraction_failed),
        has_habilitacao_pdf=bool(result.intake.habilitacao_pdf_path),
        already_existed=False,
        user_message=user_message,
    )


class PatrocinioPatchRequest(BaseModel):
    """
    Body do PATCH /intakes/{id}/patrocinio. Operador pode aprovar (sem
    alteração), editar (mudar campos) ou rejeitar (`review_action='rejeitado'`).

    Quando `review_action='aprovado'`, o backend ignora qualquer campo
    de dados e só carimba o status. Edição obriga ao menos um campo
    de dado preenchido.
    """
    review_action: str = Field(
        ...,
        description="aprovado | editado | rejeitado",
    )
    decisao: Optional[str] = None
    outro_escritorio_nome: Optional[str] = None
    outro_advogado_nome: Optional[str] = None
    outro_advogado_oab: Optional[str] = None
    outro_advogado_data_habilitacao: Optional[date] = None
    suspeita_devolucao: Optional[bool] = None
    motivo_suspeita: Optional[str] = None
    natureza_acao: Optional[str] = None
    polo_passivo_confirmado: Optional[bool] = None
    polo_passivo_observacao: Optional[str] = None


@router.patch(
    "/intakes/{intake_id}/patrocinio",
    response_model=PatrocinioOut,
    summary="HITL — operador aprova/edita/rejeita decisão de patrocínio.",
)
def patch_intake_patrocinio(
    payload: PatrocinioPatchRequest,
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    from app.models.prazo_inicial_patrocinio import (
        PATROCINIO_DECISOES_VALIDAS,
        PATROCINIO_NATUREZAS_VALIDAS,
        PATROCINIO_REVIEW_APPROVED,
        PATROCINIO_REVIEW_EDITED,
        PATROCINIO_REVIEW_REJECTED,
        PATROCINIO_REVIEW_STATUSES_VALIDOS,
        PrazoInicialPatrocinio,
    )

    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")

    patrocinio = (
        db.query(PrazoInicialPatrocinio)
        .filter(PrazoInicialPatrocinio.intake_id == intake_id)
        .first()
    )
    if patrocinio is None:
        raise HTTPException(
            status_code=404,
            detail="Este intake não tem registro de patrocínio.",
        )

    review_action = payload.review_action.strip().lower()
    if review_action not in PATROCINIO_REVIEW_STATUSES_VALIDOS or review_action == "pendente":
        raise HTTPException(
            status_code=422,
            detail="review_action deve ser 'aprovado', 'editado' ou 'rejeitado'.",
        )

    if review_action == PATROCINIO_REVIEW_EDITED:
        # Validações dos campos editados (só os que vieram).
        if payload.decisao is not None:
            if payload.decisao not in PATROCINIO_DECISOES_VALIDAS:
                raise HTTPException(
                    status_code=422,
                    detail=f"decisao inválida: {payload.decisao}",
                )
            patrocinio.decisao = payload.decisao
        if payload.natureza_acao is not None:
            if payload.natureza_acao not in PATROCINIO_NATUREZAS_VALIDAS:
                raise HTTPException(
                    status_code=422,
                    detail=f"natureza_acao inválida: {payload.natureza_acao}",
                )
            patrocinio.natureza_acao = payload.natureza_acao
        if payload.outro_escritorio_nome is not None:
            patrocinio.outro_escritorio_nome = payload.outro_escritorio_nome or None
        if payload.outro_advogado_nome is not None:
            patrocinio.outro_advogado_nome = payload.outro_advogado_nome or None
        if payload.outro_advogado_oab is not None:
            patrocinio.outro_advogado_oab = payload.outro_advogado_oab or None
        if payload.outro_advogado_data_habilitacao is not None:
            patrocinio.outro_advogado_data_habilitacao = (
                payload.outro_advogado_data_habilitacao
            )
        if payload.suspeita_devolucao is not None:
            patrocinio.suspeita_devolucao = bool(payload.suspeita_devolucao)
        if payload.motivo_suspeita is not None:
            patrocinio.motivo_suspeita = payload.motivo_suspeita or None
        if payload.polo_passivo_confirmado is not None:
            patrocinio.polo_passivo_confirmado = bool(payload.polo_passivo_confirmado)
        if payload.polo_passivo_observacao is not None:
            patrocinio.polo_passivo_observacao = payload.polo_passivo_observacao or None

    patrocinio.review_status = review_action
    patrocinio.reviewed_by_user_id = current_user.id
    patrocinio.reviewed_by_email = current_user.email
    patrocinio.reviewed_by_name = current_user.name
    patrocinio.reviewed_at = datetime.utcnow()

    db.commit()
    db.refresh(patrocinio)
    return PatrocinioOut.model_validate(patrocinio)


@router.post(
    "/intakes/{intake_id}/reprocessar-cnj",
    response_model=IntakeSummary,
    summary="Força nova tentativa de resolução do lawsuit no L1.",
)
def reprocess_intake_cnj(
    background_tasks: BackgroundTasks,
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")
    if intake.status not in (
        INTAKE_STATUS_LAWSUIT_NOT_FOUND,
        INTAKE_STATUS_RECEIVED,
    ):
        raise HTTPException(
            status_code=409,
            detail=(
                f"Reprocessamento permitido apenas em RECEBIDO / "
                f"PROCESSO_NAO_ENCONTRADO. Status atual: {intake.status}."
            ),
        )
    # Volta pra RECEBIDO pra liberar a resolução.
    intake.status = INTAKE_STATUS_RECEIVED
    intake.error_message = None
    db.commit()
    db.refresh(intake)

    service = IntakeService(db=db)
    background_tasks.add_task(service.resolve_lawsuit_for_intake, intake.id)
    return _intake_to_summary(intake)


@router.post(
    "/intakes/{intake_id}/cancelar",
    response_model=IntakeSummary,
    summary="Cancela manualmente um intake.",
)
def cancel_intake(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")
    intake.status = INTAKE_STATUS_CANCELLED
    intake.error_message = f"Cancelado por {current_user.email}."
    db.commit()
    db.refresh(intake)
    return _intake_to_summary(intake)


# ── Status elegiveis pra reclassificacao ─────────────────────────────
# Reclassificar = jogar o intake de volta pra PRONTO_PARA_CLASSIFICAR
# e apagar todas as sugestoes/pedidos atuais. Util pra:
#   1. Casos antigos com SEM_DETERMINACAO legado que voce quer
#      reclassificar pelo prompt novo (split SEM_PRAZO/INDETERMINADO).
#   2. INDETERMINADO em que voce ajustou a integra externamente.
#   3. AGUARDANDO_CONFIG_TEMPLATE em que cadastrou template depois.
#   4. ERRO_CLASSIFICACAO pra retentar.
# Bloqueia em estados terminais (AGENDADO, CONCLUIDO, CANCELADO,
# CONCLUIDO_SEM_PROVIDENCIA, GED_ENVIADO) - dai precisa cancelar
# antes ou usar o botao "Reprocessar CNJ" se for resolucao.
_RECLASSIFY_ALLOWED_STATUSES = frozenset({
    INTAKE_STATUS_CLASSIFIED,
    INTAKE_STATUS_AWAITING_TEMPLATE_CONFIG,
    INTAKE_STATUS_IN_REVIEW,
    INTAKE_STATUS_CLASSIFICATION_ERROR,
})


@router.post(
    "/intakes/{intake_id}/reclassify",
    response_model=IntakeSummary,
    summary="Volta intake pra PRONTO_PARA_CLASSIFICAR e apaga sugestoes/pedidos atuais.",
)
def reclassify_intake(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    """
    Re-encaminha o intake para nova classificacao. Apaga sugestoes e
    pedidos persistidos da rodada anterior, limpa campos derivados pelo
    classifier (natureza, produto, agravo, agregados), e volta o status
    pra PRONTO_PARA_CLASSIFICAR pra entrar no proximo batch.

    NAO mexe no PDF da habilitacao, no lawsuit_id, no office_id nem na
    fila do Tratamento Web - se ja foi enfileirado pra cancel da legada,
    continua la (operador decide se cancela ou nao).
    """
    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake nao encontrado.")
    if intake.status not in _RECLASSIFY_ALLOWED_STATUSES:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Reclassificacao permitida apenas em CLASSIFICADO / "
                f"AGUARDANDO_CONFIG_TEMPLATE / EM_REVISAO / ERRO_CLASSIFICACAO. "
                f"Status atual: {intake.status}."
            ),
        )

    # Apaga sugestoes (cascade=all,delete-orphan cuida via relationship,
    # mas explicito e mais seguro contra reordenacao do SA).
    sugestoes_apagadas = 0
    for old_sug in list(intake.sugestoes):
        db.delete(old_sug)
        sugestoes_apagadas += 1

    # Apaga pedidos da rodada anterior (cascade tambem cuida, idem).
    pedidos_apagados = 0
    for old_ped in list(intake.pedidos):
        db.delete(old_ped)
        pedidos_apagados += 1

    # Limpa campos derivados da classificacao anterior. NAO mexe em
    # lawsuit_id/office_id (resolucao do L1 nao precisa rodar de novo).
    intake.status = INTAKE_STATUS_READY_TO_CLASSIFY
    intake.classification_batch_id = None
    intake.error_message = None
    intake.natureza_processo = None
    intake.produto = None
    intake.agravo_processo_origem_cnj = None
    intake.agravo_decisao_agravada_resumo = None
    intake.valor_total_pedido = None
    intake.valor_total_estimado = None
    intake.aprovisionamento_sugerido = None
    intake.probabilidade_exito_global = None
    intake.analise_estrategica = None

    db.commit()
    db.refresh(intake)
    logger.info(
        "Intake %d reclassificado por %s: %d sugestao(oes) e %d pedido(s) apagados.",
        intake.id, current_user.email, sugestoes_apagadas, pedidos_apagados,
    )
    return _intake_to_summary(intake)


# ── Reaplicar templates em lote ──────────────────────────────────────
# Re-roda match_templates nas sugestoes existentes dos intakes filtrados
# (sem chamar IA de novo). Usado quando o operador cadastra/edita
# template novo e quer aplicar no backlog em AGUARDANDO_CONFIG_TEMPLATE.
# Suporta dry_run pra preview do impacto antes de confirmar.

class ReapplyTemplatesRequest(BaseModel):
    """Body do POST /intakes/reapply-templates."""

    # Status dos intakes elegiveis. Default: so AGUARDANDO_CONFIG_TEMPLATE
    # (caso de uso primario). Operador pode marcar CLASSIFICADO/EM_REVISAO
    # pra reaplicar tambem em intakes ja com template antigo.
    status_in: list[str] = Field(
        default_factory=lambda: ["AGUARDANDO_CONFIG_TEMPLATE"],
        min_length=1,
    )
    office_ids: Optional[list[int]] = None
    tipos_prazo: Optional[list[str]] = None
    # Se True, calcula metricas mas faz rollback no fim. Usado pelo
    # botao "Visualizar impacto" na UI.
    dry_run: bool = False


class ReapplyTemplatesResponse(BaseModel):
    """Metricas devolvidas pelo reapply (dry_run ou real)."""

    intakes_processed: int
    intakes_promoted: int
    sugestoes_updated: int
    sugestoes_skipped_already_in_l1: int
    sugestoes_skipped_edited: int
    sugestoes_no_match: int
    intake_ids_processed: list[int]
    intake_ids_promoted: list[int]
    dry_run: bool


@router.post(
    "/intakes/reapply-templates",
    response_model=ReapplyTemplatesResponse,
    summary="Re-roda match_templates em sugestoes existentes (sem chamar IA).",
)
def reapply_templates_endpoint(
    body: ReapplyTemplatesRequest,
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    """
    Reaplica templates em lote nas sugestoes ja materializadas dos
    intakes filtrados. NAO chama a IA de novo — so re-roda
    `match_templates` com a configuracao atual de templates e atualiza
    `task_subtype_id`/`responsavel_sugerido_id`/`payload_proposto`.

    Salvaguardas (sempre aplicadas):
    - Sugestoes com `created_task_id` NOT NULL: puladas (task ja
      existe no L1).
    - Sugestoes com `review_status='editado'`: puladas (operador
      ajustou na mao).

    Promocao de status: intakes em AGUARDANDO_CONFIG_TEMPLATE cuja
    TODAS as sugestoes passaram a ter template (ou skip_task_creation)
    sao promovidos pra CLASSIFICADO. Operador confirma na tela como
    sempre (esse endpoint NAO cria tarefa no L1).
    """
    from app.services.prazos_iniciais.template_reapply_service import (
        reapply_templates_bulk,
    )

    try:
        metrics = reapply_templates_bulk(
            db,
            status_in=body.status_in,
            office_ids=body.office_ids,
            tipos_prazo=body.tipos_prazo,
            dry_run=body.dry_run,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    logger.info(
        "Reapply templates por %s: dry_run=%s metrics=%s",
        current_user.email, body.dry_run, metrics.to_dict(),
    )
    return ReapplyTemplatesResponse(dry_run=body.dry_run, **metrics.to_dict())


@router.delete(
    "/intakes/{intake_id}",
    status_code=204,
    summary="HARD DELETE de um intake (admin only). Apaga registro + PDF + cascata.",
)
def delete_intake_admin(
    intake_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(auth_security.get_current_user),
):
    """
    Apaga fisicamente o intake (e tudo em cascata via ondelete=CASCADE:
    sugestoes, pedidos, legacy_task_queue). Tambem deleta o PDF fisico.

    Usado durante testes pra reinjetar o mesmo processo do zero. Vai
    virar arquivamento depois — esse endpoint sera removido / trocado
    por soft delete.

    SOMENTE admins podem chamar. `require_permission` nao basta porque
    qualquer usuario com `can_use_prazos_iniciais` passaria por ele.
    """
    if getattr(current_user, "role", "user") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas administradores podem deletar intakes.",
        )

    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")

    # Best-effort: deleta PDF fisico antes de apagar a row no banco.
    pdf_path = getattr(intake, "pdf_path", None)
    if pdf_path:
        try:
            from app.services.prazos_iniciais import storage as pdf_storage
            pdf_storage.delete_pdf(pdf_path)
        except Exception:
            # Nao impede o delete do intake. Worker de cleanup pega depois.
            pass

    db.delete(intake)
    db.commit()
    return None


# ═══════════════════════════════════════════════════════════════════════
# Trigger e batches de classificação
# ═══════════════════════════════════════════════════════════════════════


class ClassifyPendingResponse(BaseModel):
    submitted: bool
    batch_id: Optional[int] = None
    anthropic_batch_id: Optional[str] = None
    intakes_count: int = 0
    message: str


class BatchSummary(BaseModel):
    id: int
    anthropic_batch_id: Optional[str]
    status: str
    anthropic_status: Optional[str]
    total_records: int
    succeeded_count: int
    errored_count: int
    expired_count: int
    canceled_count: int
    model_used: Optional[str]
    requested_by_email: Optional[str]
    intake_ids: Optional[list[int]]
    created_at: Optional[str]
    submitted_at: Optional[str]
    ended_at: Optional[str]
    applied_at: Optional[str]


class BatchListResponse(BaseModel):
    total: int
    items: list[BatchSummary]


class ApplyBatchResponse(BaseModel):
    succeeded: int
    failed: int
    skipped: int
    total_results: int
    total_sugestoes: int


@router.post(
    "/classificar-pendentes",
    response_model=ClassifyPendingResponse,
    summary="Submete um batch com todos os intakes em PRONTO_PARA_CLASSIFICAR.",
)
async def submit_pending_classification(
    limit: Optional[int] = Query(
        default=None, ge=1, le=500,
        description="Limite opcional de intakes nesta submissão.",
    ),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(
        auth_security.require_permission("prazos_iniciais")
    ),
):
    """
    Coleta intakes em PRONTO_PARA_CLASSIFICAR e dispara um batch da
    Anthropic. Usa Claude Sonnet (configurável em settings).

    Usado tanto pelo botão da UI quanto pelo worker periódico (via call
    direto ao service, sem passar por HTTP).
    """
    classifier = PrazosIniciaisBatchClassifier(db=db)
    cap = limit or settings.prazos_iniciais_batch_max_size
    intakes = classifier.collect_pending_intakes(limit=cap)
    if not intakes:
        return ClassifyPendingResponse(
            submitted=False,
            intakes_count=0,
            message="Nenhum intake em PRONTO_PARA_CLASSIFICAR.",
        )

    try:
        batch = await classifier.submit_batch(
            intakes=intakes,
            requested_by_email=current_user.email,
        )
    except Exception as exc:
        logger.exception("Falha ao submeter batch de prazos iniciais.")
        raise HTTPException(
            status_code=502,
            detail=f"Falha ao enviar batch para Anthropic: {exc}",
        )

    return ClassifyPendingResponse(
        submitted=True,
        batch_id=batch.id,
        anthropic_batch_id=batch.anthropic_batch_id,
        intakes_count=len(intakes),
        message=f"Batch criado com {len(intakes)} intake(s).",
    )


@router.get(
    "/batches",
    response_model=BatchListResponse,
    summary="Lista batches de classificação.",
)
def list_classification_batches(
    limit: int = Query(default=50, ge=1, le=500),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    classifier = PrazosIniciaisBatchClassifier(db=db)
    batches = classifier.list_batches(limit=limit)
    return BatchListResponse(
        total=len(batches),
        items=[BatchSummary(**classifier.batch_to_dict(b)) for b in batches],
    )


@router.post(
    "/batches/{batch_id}/refresh",
    response_model=BatchSummary,
    summary="Consulta o status de um batch e atualiza contadores.",
)
async def refresh_classification_batch(
    batch_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    classifier = PrazosIniciaisBatchClassifier(db=db)
    batch = classifier.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado.")
    if not batch.anthropic_batch_id:
        raise HTTPException(
            status_code=409,
            detail="Batch não tem anthropic_batch_id (provavelmente falhou no submit).",
        )
    try:
        batch = await classifier.refresh_batch_status(batch)
    except Exception as exc:
        logger.exception("Falha ao consultar batch %s", batch_id)
        raise HTTPException(status_code=502, detail=f"Falha ao consultar batch: {exc}")
    return BatchSummary(**classifier.batch_to_dict(batch))


@router.post(
    "/batches/{batch_id}/apply",
    response_model=ApplyBatchResponse,
    summary="Baixa os resultados do batch e materializa as sugestões.",
)
async def apply_classification_batch(
    batch_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    classifier = PrazosIniciaisBatchClassifier(db=db)
    batch = classifier.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch não encontrado.")
    if batch.status == PIN_BATCH_STATUS_APPLIED:
        raise HTTPException(
            status_code=409,
            detail="Batch já foi aplicado; reaplicar exigiria limpar sugestões manualmente.",
        )
    try:
        summary = await classifier.apply_batch_results(batch)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except Exception as exc:
        logger.exception("Falha ao aplicar batch %s", batch_id)
        raise HTTPException(status_code=502, detail=f"Falha ao aplicar batch: {exc}")
    return ApplyBatchResponse(**summary)


# ═══════════════════════════════════════════════════════════════════════
# CRUD de Templates (prazo_inicial_task_templates)
# ═══════════════════════════════════════════════════════════════════════
#
# Os templates governam o casamento (tipo_prazo, subtipo,
# natureza_aplicavel, office) → tarefa L1. Só admin/operador com permissão
# `prazos_iniciais` pode tocar. A UI de Fase 4a+ consome estes endpoints.
#
# IMPORTANTE: múltiplos templates por combinação são permitidos (cada um
# gera uma sugestão separada, igual `task_templates` das publicações). Não
# há UniqueConstraint na chave de casamento — ver pin005.
#
# Regras de validação replicadas na camada HTTP:
#   - `tipo_prazo` ∈ TIPOS_PRAZO_VALIDOS
#   - `subtipo`:
#       * AUDIENCIA  → {"conciliacao","instrucao","una","outra"} ou None
#       * JULGAMENTO → {"merito","extincao_sem_merito","outro"} ou None
#       * demais tipos → obrigatoriamente None (o casamento do classifier
#         só usa subtipo para AUDIENCIA/JULGAMENTO)
#   - `natureza_aplicavel` ∈ NATUREZAS_VALIDAS ou None
#   - `priority` ∈ {"Low","Normal","High"}
#   - `due_date_reference` ∈ {"data_base","data_final_calculada",
#                             "today","audiencia_data"}
#   - `due_business_days` ∈ [-365, 30] (negativo = D-N; ver CheckConstraint
#     no pin005)
#   - FKs (office_external_id, task_subtype_external_id,
#     responsible_user_external_id) validadas com SELECT antes de
#     persistir, pra devolver 422 com mensagem clara (em vez de deixar
#     estourar IntegrityError no commit).


SUBTIPOS_AUDIENCIA = frozenset({"conciliacao", "instrucao", "una", "outra"})
SUBTIPOS_JULGAMENTO = frozenset({"merito", "extincao_sem_merito", "outro"})
PRIORIDADES_VALIDAS = frozenset({"Low", "Normal", "High"})
DUE_DATE_REFERENCES_VALIDAS = frozenset({
    "data_base",
    "data_final_calculada",
    "today",
    "audiencia_data",
})


class TemplateBase(BaseModel):
    """Campos comuns a create/update — toda validação cruzada vive aqui."""

    name: str = Field(..., min_length=1, max_length=255)
    tipo_prazo: str = Field(..., max_length=64)
    subtipo: Optional[str] = Field(default=None, max_length=128)
    # NULL = template casa em qualquer natureza. NOT NULL = só casa em
    # intakes daquela natureza específica. Diferente de office, não há
    # regra de override nessa dimensão (ver template_matching_service).
    natureza_aplicavel: Optional[str] = Field(default=None, max_length=64)
    office_external_id: Optional[int] = Field(default=None, ge=1)
    # Template "no-op" (pin014): quando True, casa normal mas NAO cria
    # tarefa no L1 — intake vira CONCLUIDO_SEM_PROVIDENCIA na confirmação.
    # task_subtype_external_id e responsible_user_external_id ficam NULL
    # (validado em `_validate_skip_task_creation_fields`).
    skip_task_creation: bool = False
    # Optional pra suportar template no-op. Em template normal, exigido
    # via `_validate_skip_task_creation_fields`.
    task_subtype_external_id: Optional[int] = Field(default=None, ge=1)
    responsible_user_external_id: Optional[int] = Field(default=None, ge=1)
    priority: str = Field(default="Normal")
    # Offset em dias úteis a partir da `due_date_reference`. Negativo = antes
    # (D-N, caso típico em prazos iniciais); 0 = no dia; positivo = depois.
    # Range -365..+30 replicado na CheckConstraint do banco (pin005).
    due_business_days: int = Field(default=-3, ge=-365, le=30)
    due_date_reference: str = Field(default="data_base")
    description_template: Optional[str] = None
    notes_template: Optional[str] = None
    is_active: bool = True
    # 'principal' (default) ou 'assistente'. Quando 'assistente', o
    # backend redireciona a tarefa pro assistente da squad do
    # `responsible_user_external_id` no momento de criar no L1.
    target_role: str = Field(default="principal", pattern="^(principal|assistente)$")
    # Quando setado, aponta pra uma squad de suporte (kind='support').
    # Combinado com `target_role`: 'principal'=lider, 'assistente'=assistente.
    target_squad_id: Optional[int] = Field(default=None, ge=1)


def _validate_tipo_subtipo(tipo_prazo: str, subtipo: Optional[str]) -> None:
    """
    Validação cruzada de tipo_prazo × subtipo. Levanta HTTPException 422.

    - AUDIENCIA e JULGAMENTO aceitam um dos subtipos categorizados ou NULL.
    - Demais tipos (CONTESTAR, LIMINAR, MANIFESTACAO_AVULSA,
      SEM_DETERMINACAO, CONTRARRAZOES) exigem subtipo=NULL. O matching
      desses tipos no classifier só usa o `tipo_prazo` (ver
      `_derive_subtipo_for_matching`).
    """
    if tipo_prazo not in TIPOS_PRAZO_VALIDOS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"tipo_prazo inválido: {tipo_prazo!r}. "
                f"Válidos: {sorted(TIPOS_PRAZO_VALIDOS)}."
            ),
        )
    if tipo_prazo == TIPO_PRAZO_AUDIENCIA:
        if subtipo is not None and subtipo not in SUBTIPOS_AUDIENCIA:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"subtipo inválido para AUDIENCIA: {subtipo!r}. "
                    f"Válidos: {sorted(SUBTIPOS_AUDIENCIA)} ou null."
                ),
            )
    elif tipo_prazo == TIPO_PRAZO_JULGAMENTO:
        if subtipo is not None and subtipo not in SUBTIPOS_JULGAMENTO:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"subtipo inválido para JULGAMENTO: {subtipo!r}. "
                    f"Válidos: {sorted(SUBTIPOS_JULGAMENTO)} ou null."
                ),
            )
    else:
        if subtipo is not None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"subtipo só é permitido para AUDIENCIA ou JULGAMENTO "
                    f"(tipo_prazo atual: {tipo_prazo}). Envie subtipo=null."
                ),
            )


def _validate_natureza_aplicavel(natureza: Optional[str]) -> None:
    """
    Valida o campo `natureza_aplicavel`. NULL é aceito (template genérico
    — casa em qualquer natureza). Se vier valor, precisa estar em
    NATUREZAS_VALIDAS (COMUM / JUIZADO / AGRAVO_INSTRUMENTO / OUTRO).
    """
    if natureza is None:
        return
    if natureza not in NATUREZAS_VALIDAS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"natureza_aplicavel inválida: {natureza!r}. "
                f"Válidas: {sorted(NATUREZAS_VALIDAS)} ou null."
            ),
        )


def _validate_priority_and_due_ref(priority: str, due_ref: str) -> None:
    if priority not in PRIORIDADES_VALIDAS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"priority inválida: {priority!r}. "
                f"Válidas: {sorted(PRIORIDADES_VALIDAS)}."
            ),
        )
    if due_ref not in DUE_DATE_REFERENCES_VALIDAS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"due_date_reference inválida: {due_ref!r}. "
                f"Válidas: {sorted(DUE_DATE_REFERENCES_VALIDAS)}."
            ),
        )


def _validate_foreign_keys(
    db: Session,
    *,
    office_external_id: Optional[int],
    task_subtype_external_id: Optional[int],
    responsible_user_external_id: Optional[int],
) -> None:
    """
    Confirma que os external_ids existem nas tabelas do L1. 422 em
    caso de ausência. IDs None são pulados — usado por templates no-op
    (skip_task_creation=True), onde task_subtype/responsible são NULL.
    """
    if office_external_id is not None:
        exists = (
            db.query(LegalOneOffice.id)
            .filter(LegalOneOffice.external_id == office_external_id)
            .first()
        )
        if not exists:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"office_external_id não encontrado: {office_external_id}.",
            )

    if task_subtype_external_id is not None:
        exists = (
            db.query(LegalOneTaskSubType.id)
            .filter(LegalOneTaskSubType.external_id == task_subtype_external_id)
            .first()
        )
        if not exists:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"task_subtype_external_id não encontrado: "
                    f"{task_subtype_external_id}."
                ),
            )

    if responsible_user_external_id is not None:
        exists = (
            db.query(LegalOneUser.id)
            .filter(LegalOneUser.external_id == responsible_user_external_id)
            .first()
        )
        if not exists:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    f"responsible_user_external_id não encontrado: "
                    f"{responsible_user_external_id}."
                ),
            )


def _validate_skip_task_creation_fields(
    skip: bool,
    task_subtype_external_id: Optional[int],
    responsible_user_external_id: Optional[int],
) -> None:
    """
    Garante consistência entre skip_task_creation e os campos de tarefa:
    - skip=True: task_subtype/responsible devem ser NULL (operador
      enviou IDs por engano? rejeita pra evitar dados orfãos no banco).
    - skip=False: ambos devem estar preenchidos (template "normal"
      precisa criar tarefa no L1).
    Espelha a CheckConstraint `ck_pin_task_templates_skip_or_task_fields`
    do banco — falha cedo no 422 em vez de IntegrityError no commit.
    """
    if skip:
        if (
            task_subtype_external_id is not None
            or responsible_user_external_id is not None
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "Template no-op (skip_task_creation=true) não aceita "
                    "task_subtype_external_id nem responsible_user_external_id. "
                    "Envie ambos como null."
                ),
            )
    else:
        if task_subtype_external_id is None or responsible_user_external_id is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "Template normal (skip_task_creation=false) exige "
                    "task_subtype_external_id e responsible_user_external_id "
                    "preenchidos."
                ),
            )


class TemplateCreate(TemplateBase):
    """Body do POST /prazos-iniciais/templates."""


class TemplateUpdate(BaseModel):
    """
    Body do PATCH /prazos-iniciais/templates/{id}.

    Todos os campos são opcionais — só o que vier é atualizado. Validações
    de tipo/subtipo são re-executadas considerando os valores resultantes
    (os que vieram + os atuais do registro) pra garantir consistência.
    """

    name: Optional[str] = Field(default=None, min_length=1, max_length=255)
    tipo_prazo: Optional[str] = Field(default=None, max_length=64)
    subtipo: Optional[str] = Field(default=None, max_length=128)
    natureza_aplicavel: Optional[str] = Field(default=None, max_length=64)
    office_external_id: Optional[int] = Field(default=None, ge=1)
    skip_task_creation: Optional[bool] = None
    task_subtype_external_id: Optional[int] = Field(default=None, ge=1)
    responsible_user_external_id: Optional[int] = Field(default=None, ge=1)
    priority: Optional[str] = None
    # Mesmo range do create; ver TemplateBase.due_business_days.
    due_business_days: Optional[int] = Field(default=None, ge=-365, le=30)
    due_date_reference: Optional[str] = None
    description_template: Optional[str] = None
    notes_template: Optional[str] = None
    is_active: Optional[bool] = None
    target_role: Optional[str] = Field(default=None, pattern="^(principal|assistente)$")
    target_squad_id: Optional[int] = Field(default=None, ge=1)
    # Sentinelas pra permitir explicitamente "setar para NULL" em PATCH.
    # Na prática os campos acima já são Optional; se o cliente quer zerar,
    # manda null. Pra diferenciar "não mexer" de "zerar" usamos
    # `model_fields_set` (Pydantic v2) no handler.

    class Config:
        extra = "forbid"


class TemplateResponse(BaseModel):
    id: int
    name: str
    tipo_prazo: str
    subtipo: Optional[str]
    natureza_aplicavel: Optional[str]
    office_external_id: Optional[int]
    office_name: Optional[str] = None
    skip_task_creation: bool = False
    task_subtype_external_id: Optional[int] = None
    task_subtype_name: Optional[str] = None
    responsible_user_external_id: Optional[int] = None
    responsible_user_name: Optional[str] = None
    priority: str
    due_business_days: int
    due_date_reference: str
    description_template: Optional[str]
    notes_template: Optional[str]
    is_active: bool
    target_role: str = "principal"
    target_squad_id: Optional[int] = None
    target_squad_name: Optional[str] = None
    created_at: Optional[datetime]
    updated_at: Optional[datetime]


class TemplateListResponse(BaseModel):
    total: int
    items: list[TemplateResponse]


def _template_to_response(
    t: PrazoInicialTaskTemplate,
    *,
    office_name: Optional[str] = None,
    task_subtype_name: Optional[str] = None,
    responsible_user_name: Optional[str] = None,
    target_squad_name: Optional[str] = None,
) -> TemplateResponse:
    return TemplateResponse(
        id=t.id,
        name=t.name,
        tipo_prazo=t.tipo_prazo,
        subtipo=t.subtipo,
        natureza_aplicavel=t.natureza_aplicavel,
        office_external_id=t.office_external_id,
        office_name=office_name,
        skip_task_creation=bool(getattr(t, "skip_task_creation", False)),
        task_subtype_external_id=t.task_subtype_external_id,
        task_subtype_name=task_subtype_name,
        responsible_user_external_id=t.responsible_user_external_id,
        responsible_user_name=responsible_user_name,
        priority=t.priority,
        due_business_days=t.due_business_days,
        due_date_reference=t.due_date_reference,
        description_template=t.description_template,
        notes_template=t.notes_template,
        is_active=t.is_active,
        target_role=getattr(t, "target_role", None) or "principal",
        target_squad_id=getattr(t, "target_squad_id", None),
        target_squad_name=target_squad_name,
        created_at=t.created_at,
        updated_at=t.updated_at,
    )


def _enrich_template_response(
    db: Session, t: PrazoInicialTaskTemplate
) -> TemplateResponse:
    """
    Resolve os nomes legíveis de office / task_subtype / responsável pra
    exibir na UI sem forçar join. Um SELECT por campo (3 queries no pior
    caso) — trivial pra um CRUD de baixa frequência.
    """
    office_name = None
    if t.office_external_id is not None:
        office = (
            db.query(LegalOneOffice.name)
            .filter(LegalOneOffice.external_id == t.office_external_id)
            .first()
        )
        office_name = office[0] if office else None
    # Templates no-op podem ter task_subtype/responsible NULL — pula
    # lookups quando ausentes (sem nome pra exibir).
    sub_name = None
    if t.task_subtype_external_id is not None:
        sub = (
            db.query(LegalOneTaskSubType.name)
            .filter(LegalOneTaskSubType.external_id == t.task_subtype_external_id)
            .first()
        )
        sub_name = sub[0] if sub else None
    user_name = None
    if t.responsible_user_external_id is not None:
        user = (
            db.query(LegalOneUser.name)
            .filter(LegalOneUser.external_id == t.responsible_user_external_id)
            .first()
        )
        user_name = user[0] if user else None
    target_squad_name = None
    if getattr(t, "target_squad_id", None):
        from app.models.rules import Squad
        sq = (
            db.query(Squad.name)
            .filter(Squad.id == t.target_squad_id)
            .first()
        )
        target_squad_name = sq[0] if sq else None
    return _template_to_response(
        t,
        office_name=office_name,
        task_subtype_name=sub_name,
        responsible_user_name=user_name,
        target_squad_name=target_squad_name,
    )


@router.get(
    "/templates",
    response_model=TemplateListResponse,
    summary="Lista templates de tarefa (prazos iniciais), com filtros.",
)
def list_templates(
    tipo_prazo: Optional[str] = Query(default=None),
    subtipo: Optional[str] = Query(default=None),
    natureza_aplicavel: Optional[str] = Query(
        default=None,
        description=(
            "Filtra por natureza. Use string vazia ('') pra pegar apenas "
            "templates genéricos (natureza NULL)."
        ),
    ),
    office_external_id: Optional[int] = Query(default=None),
    is_active: Optional[bool] = Query(default=None),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Lista templates. Filtros opcionais combinam com AND. Ordenação estável
    por (tipo_prazo, subtipo NULLs first, natureza NULLs first,
    office NULLs last, id) pra facilitar leitura na UI de admin.
    """
    q = db.query(PrazoInicialTaskTemplate)
    if tipo_prazo:
        q = q.filter(PrazoInicialTaskTemplate.tipo_prazo == tipo_prazo)
    if subtipo is not None:
        # Cliente mandou string vazia → filtra por subtipo NULL.
        if subtipo == "":
            q = q.filter(PrazoInicialTaskTemplate.subtipo.is_(None))
        else:
            q = q.filter(PrazoInicialTaskTemplate.subtipo == subtipo)
    if natureza_aplicavel is not None:
        # Cliente mandou string vazia → filtra por natureza NULL (genérico).
        if natureza_aplicavel == "":
            q = q.filter(PrazoInicialTaskTemplate.natureza_aplicavel.is_(None))
        else:
            q = q.filter(
                PrazoInicialTaskTemplate.natureza_aplicavel == natureza_aplicavel
            )
    if office_external_id is not None:
        if office_external_id == 0:
            # Convenção: office_external_id=0 → global (NULL).
            q = q.filter(PrazoInicialTaskTemplate.office_external_id.is_(None))
        else:
            q = q.filter(
                PrazoInicialTaskTemplate.office_external_id == office_external_id
            )
    if is_active is not None:
        q = q.filter(PrazoInicialTaskTemplate.is_active.is_(is_active))

    total = q.count()
    items = (
        q.order_by(
            PrazoInicialTaskTemplate.tipo_prazo.asc(),
            PrazoInicialTaskTemplate.subtipo.asc().nullsfirst(),
            PrazoInicialTaskTemplate.natureza_aplicavel.asc().nullsfirst(),
            PrazoInicialTaskTemplate.office_external_id.asc().nullslast(),
            PrazoInicialTaskTemplate.id.asc(),
        )
        .offset(offset)
        .limit(limit)
        .all()
    )
    return TemplateListResponse(
        total=total,
        items=[_enrich_template_response(db, t) for t in items],
    )


@router.post(
    "/templates",
    response_model=TemplateResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Cria um template de tarefa (prazos iniciais).",
)
def create_template(
    body: TemplateCreate,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    _validate_tipo_subtipo(body.tipo_prazo, body.subtipo)
    _validate_natureza_aplicavel(body.natureza_aplicavel)
    _validate_priority_and_due_ref(body.priority, body.due_date_reference)
    _validate_skip_task_creation_fields(
        body.skip_task_creation,
        body.task_subtype_external_id,
        body.responsible_user_external_id,
    )
    _validate_foreign_keys(
        db,
        office_external_id=body.office_external_id,
        task_subtype_external_id=body.task_subtype_external_id,
        responsible_user_external_id=body.responsible_user_external_id,
    )

    # Múltiplos templates na MESMA combinação (tipo, subtipo, natureza,
    # office) são permitidos — cada um gera uma sugestão separada. Não há
    # checagem de unicidade aqui (ver pin005 e docstring do model).

    template = PrazoInicialTaskTemplate(
        name=body.name,
        tipo_prazo=body.tipo_prazo,
        subtipo=body.subtipo,
        natureza_aplicavel=body.natureza_aplicavel,
        office_external_id=body.office_external_id,
        skip_task_creation=body.skip_task_creation,
        task_subtype_external_id=body.task_subtype_external_id,
        responsible_user_external_id=body.responsible_user_external_id,
        priority=body.priority,
        due_business_days=body.due_business_days,
        due_date_reference=body.due_date_reference,
        description_template=body.description_template,
        notes_template=body.notes_template,
        is_active=body.is_active,
        target_role=body.target_role,
        target_squad_id=body.target_squad_id,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return _enrich_template_response(db, template)


@router.get(
    "/templates/{template_id}",
    response_model=TemplateResponse,
    summary="Detalhe de um template.",
)
def get_template(
    template_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    template = db.get(PrazoInicialTaskTemplate, template_id)
    if template is None:
        raise HTTPException(status_code=404, detail="Template não encontrado.")
    return _enrich_template_response(db, template)


@router.patch(
    "/templates/{template_id}",
    response_model=TemplateResponse,
    summary="Atualiza parcialmente um template (campos omitidos ficam como estão).",
)
def update_template(
    body: TemplateUpdate,
    template_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    template = db.get(PrazoInicialTaskTemplate, template_id)
    if template is None:
        raise HTTPException(status_code=404, detail="Template não encontrado.")

    fields_set = body.model_fields_set  # campos realmente enviados

    # ── Resolve valores efetivos (merge do que veio com o atual) ─────
    eff_tipo = body.tipo_prazo if "tipo_prazo" in fields_set else template.tipo_prazo
    eff_subtipo = body.subtipo if "subtipo" in fields_set else template.subtipo
    eff_natureza = (
        body.natureza_aplicavel
        if "natureza_aplicavel" in fields_set
        else template.natureza_aplicavel
    )
    eff_office = (
        body.office_external_id
        if "office_external_id" in fields_set
        else template.office_external_id
    )
    eff_task_sub = (
        body.task_subtype_external_id
        if "task_subtype_external_id" in fields_set
        else template.task_subtype_external_id
    )
    eff_resp = (
        body.responsible_user_external_id
        if "responsible_user_external_id" in fields_set
        else template.responsible_user_external_id
    )
    eff_priority = body.priority if "priority" in fields_set else template.priority
    eff_due_ref = (
        body.due_date_reference
        if "due_date_reference" in fields_set
        else template.due_date_reference
    )
    eff_skip = (
        body.skip_task_creation
        if "skip_task_creation" in fields_set
        else bool(getattr(template, "skip_task_creation", False))
    )

    _validate_tipo_subtipo(eff_tipo, eff_subtipo)
    _validate_natureza_aplicavel(eff_natureza)
    _validate_priority_and_due_ref(eff_priority, eff_due_ref)
    _validate_skip_task_creation_fields(eff_skip, eff_task_sub, eff_resp)
    # FKs: só valida o que mudou (reduz queries). task_sub/responsible
    # podem ser None em template no-op — _validate_foreign_keys pula
    # IDs None.
    _validate_foreign_keys(
        db,
        office_external_id=(
            eff_office
            if "office_external_id" in fields_set
            else None  # None aqui = não re-valida
        ),
        task_subtype_external_id=(
            eff_task_sub
            if "task_subtype_external_id" in fields_set
            else None
        ),
        responsible_user_external_id=(
            eff_resp
            if "responsible_user_external_id" in fields_set
            else None
        ),
    )

    # Sem checagem de unicidade na chave de casamento — múltiplos templates
    # na mesma combinação são válidos (ver pin005).

    # Aplica as mudanças.
    for fname in fields_set:
        setattr(template, fname, getattr(body, fname))

    db.commit()
    db.refresh(template)
    return _enrich_template_response(db, template)


@router.delete(
    "/templates/{template_id}",
    response_model=TemplateResponse,
    summary="Soft-delete de um template (is_active=False).",
)
def delete_template(
    template_id: int = Path(..., ge=1),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    """
    Soft-delete: marca `is_active=False` em vez de apagar. Templates
    inativos não casam mais sugestões novas, mas permanecem disponíveis
    para auditoria das sugestões antigas que apontam para eles via
    `payload_proposto.template_id`.
    """
    template = db.get(PrazoInicialTaskTemplate, template_id)
    if template is None:
        raise HTTPException(status_code=404, detail="Template não encontrado.")
    template.is_active = False
    db.commit()
    db.refresh(template)
    return _enrich_template_response(db, template)


# ─────────────────────────────────────────────────────────────────────
# Tipos de pedido (dicionário admin, usado pela análise estratégica
# de aprovisionamento / contingência pela IA).
# ─────────────────────────────────────────────────────────────────────

from app.models.prazo_inicial_tipo_pedido import PrazoInicialTipoPedido  # noqa: E402


class TipoPedidoResponse(BaseModel):
    id: int
    codigo: str
    nome: str
    naturezas: Optional[str] = None
    display_order: int
    is_active: bool


class TipoPedidoPatch(BaseModel):
    """Hoje só toggle is_active.  Mantido como objeto pra evolução futura
    (admin editar `nome` ou `naturezas` sem migration, se o negócio pedir)."""
    is_active: Optional[bool] = None
    nome: Optional[str] = None
    naturezas: Optional[str] = None
    display_order: Optional[int] = None


@router.get(
    "/tipos-pedido",
    response_model=List[TipoPedidoResponse],
    summary="Lista os tipos de pedido (taxonomia) usados pela análise estratégica.",
)
def list_tipos_pedido(
    only_active: bool = Query(default=False, description="Se True, retorna só is_active=True"),
    natureza: Optional[str] = Query(
        default=None,
        description=(
            "Se informado, filtra tipos cuja coluna `naturezas` contém essa string. "
            "Ex.: natureza='Consumidor' traz todos que aplicam em Consumidor."
        ),
    ),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    q = db.query(PrazoInicialTipoPedido)
    if only_active:
        q = q.filter(PrazoInicialTipoPedido.is_active.is_(True))
    if natureza:
        # Filtro simples por LIKE — naturezas é CSV ";". Maiúsculo pra
        # robustez a variações tipo "Cível" vs "cível".
        like = f"%{natureza}%"
        q = q.filter(PrazoInicialTipoPedido.naturezas.ilike(like))
    q = q.order_by(
        PrazoInicialTipoPedido.display_order.asc(),
        PrazoInicialTipoPedido.nome.asc(),
    )
    rows = q.all()
    return [
        TipoPedidoResponse(
            id=r.id,
            codigo=r.codigo,
            nome=r.nome,
            naturezas=r.naturezas,
            display_order=r.display_order,
            is_active=r.is_active,
        )
        for r in rows
    ]


@router.patch(
    "/tipos-pedido/{tipo_id}",
    response_model=TipoPedidoResponse,
    summary="Edita metadados de um tipo de pedido (toggle is_active, nome, naturezas).",
)
def patch_tipo_pedido(
    tipo_id: int,
    payload: TipoPedidoPatch,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    tipo = db.get(PrazoInicialTipoPedido, tipo_id)
    if tipo is None:
        raise HTTPException(status_code=404, detail="Tipo de pedido não encontrado.")
    updates = payload.dict(exclude_unset=True)
    for field, value in updates.items():
        setattr(tipo, field, value)
    db.commit()
    db.refresh(tipo)
    return TipoPedidoResponse(
        id=tipo.id,
        codigo=tipo.codigo,
        nome=tipo.nome,
        naturezas=tipo.naturezas,
        display_order=tipo.display_order,
        is_active=tipo.is_active,
    )



# ─────────────────────────────────────────────────────────────────────
# Pedidos extraídos da petição inicial (Bloco D2).
# ─────────────────────────────────────────────────────────────────────

from app.models.prazo_inicial_pedido import (  # noqa: E402
    PrazoInicialPedido,
    PROB_PERDA_VALIDAS,
)


class PedidoResponse(BaseModel):
    id: int
    intake_id: int
    tipo_pedido: str
    natureza: Optional[str] = None
    valor_indicado: Optional[float] = None
    valor_estimado: Optional[float] = None
    fundamentacao_valor: Optional[str] = None
    probabilidade_perda: Optional[str] = None
    aprovisionamento: Optional[float] = None
    fundamentacao_risco: Optional[str] = None


class PedidoPatch(BaseModel):
    """Operador pode editar qualquer campo no HITL — a persistência
    é auditada via updated_at. Mantemos também o código do tipo_pedido
    editável caso a IA tenha classificado mal."""
    tipo_pedido: Optional[str] = None
    natureza: Optional[str] = None
    valor_indicado: Optional[float] = None
    valor_estimado: Optional[float] = None
    fundamentacao_valor: Optional[str] = None
    probabilidade_perda: Optional[str] = None
    aprovisionamento: Optional[float] = None
    fundamentacao_risco: Optional[str] = None


def _pedido_to_response(p: PrazoInicialPedido) -> PedidoResponse:
    return PedidoResponse(
        id=p.id,
        intake_id=p.intake_id,
        tipo_pedido=p.tipo_pedido,
        natureza=p.natureza,
        valor_indicado=float(p.valor_indicado) if p.valor_indicado is not None else None,
        valor_estimado=float(p.valor_estimado) if p.valor_estimado is not None else None,
        fundamentacao_valor=p.fundamentacao_valor,
        probabilidade_perda=p.probabilidade_perda,
        aprovisionamento=float(p.aprovisionamento) if p.aprovisionamento is not None else None,
        fundamentacao_risco=p.fundamentacao_risco,
    )


@router.get(
    "/intakes/{intake_id}/pedidos",
    response_model=List[PedidoResponse],
    summary="Lista pedidos extraídos da petição inicial de um intake.",
)
def list_pedidos_do_intake(
    intake_id: int,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    rows = (
        db.query(PrazoInicialPedido)
        .filter(PrazoInicialPedido.intake_id == intake_id)
        .order_by(PrazoInicialPedido.id.asc())
        .all()
    )
    return [_pedido_to_response(p) for p in rows]


@router.patch(
    "/pedidos/{pedido_id}",
    response_model=PedidoResponse,
    summary="Edita um pedido (HITL — operador ajusta valor estimado, probabilidade, aprovisionamento).",
)
def patch_pedido(
    pedido_id: int,
    payload: PedidoPatch,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    pedido = db.get(PrazoInicialPedido, pedido_id)
    if pedido is None:
        raise HTTPException(status_code=404, detail="Pedido não encontrado.")

    updates = payload.dict(exclude_unset=True)

    # Valida probabilidade se foi informada
    if "probabilidade_perda" in updates and updates["probabilidade_perda"] is not None:
        if updates["probabilidade_perda"] not in PROB_PERDA_VALIDAS:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"probabilidade_perda inválida. Use um de: "
                    f"{sorted(PROB_PERDA_VALIDAS)}"
                ),
            )

    for field, value in updates.items():
        setattr(pedido, field, value)

    # Reaplica regra CPC 25 se operador mudou probabilidade mas esqueceu
    # de atualizar aprovisionamento: remota/possível = 0; provável =
    # valor_estimado. Operador pode sobrescrever depois com um PATCH novo
    # se quiser — aqui só garantimos coerência quando não foi informado.
    if (
        "probabilidade_perda" in updates
        and "aprovisionamento" not in updates
    ):
        if pedido.probabilidade_perda in ("remota", "possivel"):
            pedido.aprovisionamento = 0
        elif pedido.probabilidade_perda == "provavel" and pedido.valor_estimado is not None:
            pedido.aprovisionamento = pedido.valor_estimado

    # Recalcula agregados do intake ao editar qualquer pedido.  Mantem
    # valor_total_*, aprovisionamento_sugerido e probabilidade_exito_global
    # coerentes — evita dashboard/relatorio defasado apos correcao HITL.
    db.commit()
    db.refresh(pedido)

    try:
        from app.services.classifier.prazos_iniciais_classifier import (
            PrazosIniciaisBatchClassifier,
        )
        # Reusa o helper (metodo de instancia, mas independe do state
        # do classifier — só precisa de uma sessao db).
        clf = PrazosIniciaisBatchClassifier.__new__(PrazosIniciaisBatchClassifier)
        clf.db = db
        intake = db.get(type(pedido).__mapper__.class_, pedido.intake_id)  # hack pra nao importar
        # Melhor explicito:
        from app.models.prazo_inicial import PrazoInicialIntake
        intake = db.get(PrazoInicialIntake, pedido.intake_id)
        if intake is not None:
            clf._compute_intake_globals(intake)
            db.commit()
    except Exception:
        # Nao bloqueia o PATCH se o recalculo falhar — operador pode
        # acionar Reanalisar depois.
        logger.exception(
            "Falha ao recalcular agregados do intake %s apos PATCH do pedido %s",
            pedido.intake_id, pedido.id,
        )

    return _pedido_to_response(pedido)



# ─────────────────────────────────────────────────────────────────────
# Reanalisar intake (Bloco F): reseta status pra reprocessar,
# útil pra popular campos novos (pedidos, prazo_fatal_*, agravo_*,
# agregados globais) em intakes classificados antes desses blocos.
# ─────────────────────────────────────────────────────────────────────


@router.post(
    "/intakes/{intake_id}/reanalisar",
    summary=(
        "Reseta um intake pra reprocessar na próxima janela de classificação. "
        "Limpa sugestões, pedidos e agregados antigos (cascade)."
    ),
)
def reanalisar_intake(
    intake_id: int,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.models.prazo_inicial import (
        PrazoInicialIntake,
        INTAKE_STATUS_READY_TO_CLASSIFY,
        INTAKE_STATUS_IN_CLASSIFICATION,
        INTAKE_STATUS_RECEIVED,
    )

    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")

    # Estados não-reanalisáveis (ou seria atropelo num batch ativo).
    if intake.status in {INTAKE_STATUS_RECEIVED, INTAKE_STATUS_IN_CLASSIFICATION}:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Intake está em '{intake.status}' — aguarde o ciclo atual "
                "terminar antes de reanalisar."
            ),
        )

    # Limpa sugestões e pedidos (cascade delete-orphan cuidaria, mas
    # explícito é mais seguro e documentado).
    for s in list(intake.sugestoes):
        db.delete(s)
    for p in list(intake.pedidos):
        db.delete(p)

    # Limpa agregados derivados (Bloco E) e enriquecimentos (Blocos B+C).
    intake.valor_total_pedido = None
    intake.valor_total_estimado = None
    intake.aprovisionamento_sugerido = None
    intake.probabilidade_exito_global = None
    intake.analise_estrategica = None
    intake.agravo_processo_origem_cnj = None
    intake.agravo_decisao_agravada_resumo = None
    intake.error_message = None

    # Reseta status pra próximo batch pegar.
    intake.status = INTAKE_STATUS_READY_TO_CLASSIFY
    intake.classification_batch_id = None

    db.commit()
    db.refresh(intake)
    return {
        "intake_id": intake.id,
        "status": intake.status,
        "message": (
            "Intake resetado. Será incluído na próxima janela de classificação."
        ),
    }


# ─────────────────────────────────────────────────────────────────────
# Recompute globals (Onda 2): reconsolida agregados do intake a partir
# dos pedidos atuais SEM re-rodar Sonnet. Útil pra corrigir intakes
# que ficaram com agregados NULL por apply antigo ou exception
# silenciosa. Idempotente, barato (sem I/O externo), sem gasto de
# tokens.
# ─────────────────────────────────────────────────────────────────────


@router.post(
    "/intakes/{intake_id}/recompute-globals",
    summary=(
        "Recalcula valor_total_pedido/estimado/aprovisionamento/"
        "probabilidade_exito_global a partir dos pedidos já persistidos."
    ),
)
def recompute_intake_globals(
    intake_id: int,
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from app.models.prazo_inicial import PrazoInicialIntake
    from app.services.classifier.prazos_iniciais_classifier import (
        PrazosIniciaisBatchClassifier,
    )

    intake = db.get(PrazoInicialIntake, intake_id)
    if intake is None:
        raise HTTPException(status_code=404, detail="Intake não encontrado.")

    clf = PrazosIniciaisBatchClassifier.__new__(PrazosIniciaisBatchClassifier)
    clf.db = db
    try:
        clf._compute_intake_globals(intake)
        db.commit()
        db.refresh(intake)
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        logger.exception(
            "Recompute globals falhou pro intake %s: %s", intake_id, exc,
        )
        raise HTTPException(
            status_code=500,
            detail=f"Falha ao recalcular agregados: {exc}",
        )

    return {
        "intake_id": intake.id,
        "valor_total_pedido": (
            float(intake.valor_total_pedido)
            if intake.valor_total_pedido is not None
            else None
        ),
        "valor_total_estimado": (
            float(intake.valor_total_estimado)
            if intake.valor_total_estimado is not None
            else None
        ),
        "aprovisionamento_sugerido": (
            float(intake.aprovisionamento_sugerido)
            if intake.aprovisionamento_sugerido is not None
            else None
        ),
        "probabilidade_exito_global": intake.probabilidade_exito_global,
        "pedidos_count": len(intake.pedidos or []),
    }


# ─────────────────────────────────────────────────────────────────────
# Export XLSX (Bloco F): relatório contábil com agregados + pedidos.
# Protegido por permission('prazos_iniciais'). Formato: 3 abas
# (Resumo, Sugestões, Pedidos) pra facilitar pivot no Excel.
# ─────────────────────────────────────────────────────────────────────

from fastapi.responses import StreamingResponse  # noqa: E402
import io  # noqa: E402


@router.get(
    "/intakes/export.xlsx",
    summary=(
        "Exporta intakes (com agregados, sugestões e pedidos) em XLSX. "
        "3 abas: Resumo, Sugestões, Pedidos. Útil pra relatório contábil "
        "de aprovisionamento (CPC 25)."
    ),
)
def export_intakes_xlsx(
    status: Optional[str] = Query(default=None, description="Filtra por status (ex.: CONCLUIDO)"),
    office_id: Optional[int] = Query(default=None, description="Filtra por office_id"),
    date_from: Optional[str] = Query(default=None, description="ISO YYYY-MM-DD, filtra received_at >="),
    date_to: Optional[str] = Query(default=None, description="ISO YYYY-MM-DD, filtra received_at <"),
    db: Session = Depends(get_db),
    _: LegalOneUser = Depends(auth_security.require_permission("prazos_iniciais")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.models.prazo_inicial import PrazoInicialIntake
    from app.models.prazo_inicial_pedido import PrazoInicialPedido
    from app.models.legal_one import LegalOneOffice
    from app.services.publication_search_service import uf_from_cnj
    from datetime import datetime

    q = db.query(PrazoInicialIntake)
    if status:
        q = q.filter(PrazoInicialIntake.status == status)
    if office_id:
        q = q.filter(PrazoInicialIntake.office_id == office_id)
    if date_from:
        q = q.filter(PrazoInicialIntake.received_at >= date_from)
    if date_to:
        q = q.filter(PrazoInicialIntake.received_at < date_to + "T99")

    intakes = q.order_by(PrazoInicialIntake.received_at.desc()).all()

    # Pré-carrega path dos escritórios pra mostrar a hierarquia completa
    # (mesmo padrão das outras telas — o `name` é só a folha).
    office_external_ids = {
        i.office_id for i in intakes if i.office_id is not None
    }
    office_paths: dict[int, str] = {}
    if office_external_ids:
        offices = (
            db.query(LegalOneOffice)
            .filter(LegalOneOffice.external_id.in_(office_external_ids))
            .all()
        )
        office_paths = {o.external_id: (o.path or o.name) for o in offices}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    def _write_header(ws, cols):
        for idx, title in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _autofit(ws, cols):
        for idx in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 18

    # ── Aba 1: Resumo (1 linha por intake) ──
    ws = wb.active
    ws.title = "Resumo"
    resumo_cols = [
        "Intake ID", "CNJ", "UF", "Status", "Natureza", "Produto",
        "Escritório responsável", "Lawsuit ID",
        "Prob. Êxito Global",
        "Valor Total Pedido", "Valor Total Estimado",
        "Aprovisionamento Sugerido",
        "Análise Estratégica",
        "Agravo Origem CNJ",
        "Recebido em", "Atualizado em",
    ]
    _write_header(ws, resumo_cols)
    for i, intake in enumerate(intakes, start=2):
        office_label = (
            office_paths.get(intake.office_id, str(intake.office_id))
            if intake.office_id is not None
            else ""
        )
        ws.cell(row=i, column=1, value=intake.id)
        ws.cell(row=i, column=2, value=intake.cnj_number)
        ws.cell(row=i, column=3, value=uf_from_cnj(intake.cnj_number) or "")
        ws.cell(row=i, column=4, value=intake.status)
        ws.cell(row=i, column=5, value=intake.natureza_processo)
        ws.cell(row=i, column=6, value=intake.produto)
        ws.cell(row=i, column=7, value=office_label)
        ws.cell(row=i, column=8, value=intake.lawsuit_id)
        ws.cell(row=i, column=9, value=intake.probabilidade_exito_global)
        ws.cell(row=i, column=10, value=float(intake.valor_total_pedido) if intake.valor_total_pedido is not None else None)
        ws.cell(row=i, column=11, value=float(intake.valor_total_estimado) if intake.valor_total_estimado is not None else None)
        ws.cell(row=i, column=12, value=float(intake.aprovisionamento_sugerido) if intake.aprovisionamento_sugerido is not None else None)
        ws.cell(row=i, column=13, value=intake.analise_estrategica)
        ws.cell(row=i, column=14, value=intake.agravo_processo_origem_cnj)
        ws.cell(row=i, column=15, value=intake.received_at.isoformat() if intake.received_at else None)
        ws.cell(row=i, column=16, value=intake.updated_at.isoformat() if intake.updated_at else None)
    _autofit(ws, resumo_cols)
    # Larguras específicas: análise estratégica ocupa mais espaço.
    ws.column_dimensions[get_column_letter(7)].width = 36   # escritório (path)
    ws.column_dimensions[get_column_letter(13)].width = 60  # análise estratégica

    # ── Aba 2: Sugestões (1 linha por sugestão) ──
    ws2 = wb.create_sheet(title="Sugestões")
    sug_cols = [
        "Sugestão ID", "Intake ID", "CNJ", "UF", "Escritório responsável",
        "Tipo Prazo", "Subtipo",
        "Data Base", "Prazo Dias", "Prazo Tipo",
        "Data Final Calculada",
        "Prazo Fatal Data", "Prazo Fatal Fundamentação",
        "Prazo Base Decisão",
        "Audiência Data", "Audiência Hora", "Audiência Link",
        "Confiança", "Justificativa", "Review Status",
    ]
    _write_header(ws2, sug_cols)
    row = 2
    for intake in intakes:
        intake_uf = uf_from_cnj(intake.cnj_number) or ""
        intake_office = (
            office_paths.get(intake.office_id, str(intake.office_id))
            if intake.office_id is not None
            else ""
        )
        for s in intake.sugestoes:
            ws2.cell(row=row, column=1, value=s.id)
            ws2.cell(row=row, column=2, value=intake.id)
            ws2.cell(row=row, column=3, value=intake.cnj_number)
            ws2.cell(row=row, column=4, value=intake_uf)
            ws2.cell(row=row, column=5, value=intake_office)
            ws2.cell(row=row, column=6, value=s.tipo_prazo)
            ws2.cell(row=row, column=7, value=s.subtipo)
            ws2.cell(row=row, column=8, value=s.data_base.isoformat() if s.data_base else None)
            ws2.cell(row=row, column=9, value=s.prazo_dias)
            ws2.cell(row=row, column=10, value=s.prazo_tipo)
            ws2.cell(row=row, column=11, value=s.data_final_calculada.isoformat() if s.data_final_calculada else None)
            ws2.cell(row=row, column=12, value=s.prazo_fatal_data.isoformat() if s.prazo_fatal_data else None)
            ws2.cell(row=row, column=13, value=s.prazo_fatal_fundamentacao)
            ws2.cell(row=row, column=14, value=s.prazo_base_decisao)
            ws2.cell(row=row, column=15, value=s.audiencia_data.isoformat() if s.audiencia_data else None)
            ws2.cell(row=row, column=16, value=s.audiencia_hora.isoformat() if s.audiencia_hora else None)
            ws2.cell(row=row, column=17, value=s.audiencia_link)
            ws2.cell(row=row, column=18, value=s.confianca)
            ws2.cell(row=row, column=19, value=s.justificativa)
            ws2.cell(row=row, column=20, value=s.review_status)
            row += 1
    _autofit(ws2, sug_cols)
    ws2.column_dimensions[get_column_letter(5)].width = 36  # escritório

    # ── Aba 3: Pedidos (1 linha por pedido) ──
    ws3 = wb.create_sheet(title="Pedidos")
    ped_cols = [
        "Pedido ID", "Intake ID", "CNJ", "Tipo Pedido", "Natureza",
        "Valor Indicado", "Valor Estimado", "Fundamentação Valor",
        "Probabilidade Perda", "Aprovisionamento", "Fundamentação Risco",
    ]
    _write_header(ws3, ped_cols)
    row = 2
    for intake in intakes:
        for p in intake.pedidos:
            ws3.cell(row=row, column=1, value=p.id)
            ws3.cell(row=row, column=2, value=intake.id)
            ws3.cell(row=row, column=3, value=intake.cnj_number)
            ws3.cell(row=row, column=4, value=p.tipo_pedido)
            ws3.cell(row=row, column=5, value=p.natureza)
            ws3.cell(row=row, column=6, value=float(p.valor_indicado) if p.valor_indicado is not None else None)
            ws3.cell(row=row, column=7, value=float(p.valor_estimado) if p.valor_estimado is not None else None)
            ws3.cell(row=row, column=8, value=p.fundamentacao_valor)
            ws3.cell(row=row, column=9, value=p.probabilidade_perda)
            ws3.cell(row=row, column=10, value=float(p.aprovisionamento) if p.aprovisionamento is not None else None)
            ws3.cell(row=row, column=11, value=p.fundamentacao_risco)
            row += 1
    _autofit(ws3, ped_cols)
    ws3.column_dimensions[get_column_letter(8)].width = 60
    ws3.column_dimensions[get_column_letter(11)].width = 60

    # Serializa pra memória e devolve como download.
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"prazos_iniciais_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
