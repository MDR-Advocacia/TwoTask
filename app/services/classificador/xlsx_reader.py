"""Leitor do XLSX do template do Classificador.

Template MINIMO da Fase 2 — operador sobe planilha onde cada linha vira
um processo no lote. Header e' detectado na linha 1 ou 2 (tolera
exports L1 que tem titulo na linha 1).

Colunas:
- OBRIGATORIA: `cnj` (com aliases: 'cnj', 'cnj number', 'numero do processo',
  'numero processo', 'n processo', 'numero', 'processo')
- OPCIONAIS: `cliente_externo_id`, `produto`, `observacao`

O grosso dos dados (capa, partes, valor causa, situacao) vem do REFRESH L1
que e' disparado depois do upload — nao precisa do operador preencher tudo,
basta o CNJ. O classifier do Classificador (Fase 3) preenche pedidos, PCOND
e prob_exito automaticamente.

Tolerante: ignora linhas com CNJ vazio (footer/rodape de planilha),
linhas com CNJ malformado entram com warning mas seguem (vao falhar no
refresh L1 individualmente, sem travar o lote inteiro).
"""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Iterator

import openpyxl


# Footer/rodape comuns em planilhas (TOTAIS, TOTAL, totais:) sao ignorados
# silenciosamente. Sao linhas de sumario, nao processos reais.
_FOOTER_CNJ_RE = re.compile(r"^\s*totais?\s*:?\s*$", re.IGNORECASE)

# CNJ valido tem 20 digitos formatados (NNNNNNN-DD.AAAA.J.TR.OOOO) ou
# 20 digitos crus. Validacao mais leniente aqui — so checa se sobra >=15
# digitos depois de remover nao-digitos.
_CNJ_DIGITS_MIN = 15
_CNJ_DIGITS_MAX = 25


def _normalize_header(s) -> str:
    """lower + sem acento + sem pontuacao + espacos colapsados."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    only_ascii = "".join(c for c in nfkd if not unicodedata.combining(c))
    cleaned = only_ascii.lower().replace(".", "").replace("/", " ")
    cleaned = cleaned.replace("-", " ").replace("_", " ")
    return " ".join(cleaned.split())


# Mapeamento canonico -> aliases aceitos (todos normalizados via _normalize_header).
COLUMN_ALIASES: dict[str, list[str]] = {
    "cnj_number": [
        "cnj",
        "cnj number",
        "numero do processo",
        "numero processo",
        "numeros processo",
        "no processo",
        "n processo",
        "no do processo",
        "numero",
        "processo",
        "numero unico",
    ],
    "cliente_externo_id": [
        "cliente externo id",
        "id cliente",
        "cliente id",
        "id externo",
        "external id",
        "id do cliente",
    ],
    "produto": [
        "produto",
        "produto bancario",
        "contrato",
        "modalidade",
    ],
    "observacao": [
        "observacao",
        "observacoes",
        "obs",
        "nota",
        "comentario",
        "comentarios",
    ],
}

REQUIRED_COLUMNS = frozenset({"cnj_number"})


class XlsxHeaderError(Exception):
    """Header invalido — falta coluna obrigatoria ou planilha vazia."""


def _normalize_cnj(value) -> str | None:
    """Extrai digitos do CNJ. Retorna None se vazio ou claramente invalido.

    Aceita 15-25 digitos pra tolerar variacoes (com/sem digito verificador
    completo, espacos, formatos). Validacao estrita ocorre no refresh L1
    (a L1 retorna 'nao encontrado' pra CNJ malformado).
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if _FOOTER_CNJ_RE.match(s):
        return None
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return None
    if len(digits) < _CNJ_DIGITS_MIN or len(digits) > _CNJ_DIGITS_MAX:
        return s  # devolve raw mesmo invalido — vamos registrar como warning
    return digits


def _detect_header(ws) -> tuple[int, dict[str, int]]:
    """Detecta linha de header (1 ou 2) e mapeia coluna -> indice (1-based).

    Retorna (header_row, mapping) onde mapping[canonical_name] = column_idx.
    Raises XlsxHeaderError se cnj nao encontrado nas 2 primeiras linhas.
    """
    # Acumula aliases normalizados -> canonical_name
    alias_to_canonical: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_to_canonical[_normalize_header(alias)] = canonical

    best_row = 0
    best_mapping: dict[str, int] = {}
    for try_row in (1, 2):
        row_vals = next(
            ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True), None
        )
        if not row_vals:
            continue
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row_vals, start=1):
            normalized = _normalize_header(cell)
            if not normalized:
                continue
            canonical = alias_to_canonical.get(normalized)
            if canonical and canonical not in mapping:
                mapping[canonical] = col_idx
        # Heuristica: header valido tem pelo menos o cnj_number
        if "cnj_number" in mapping:
            if len(mapping) > len(best_mapping):
                best_row = try_row
                best_mapping = mapping

    if not best_mapping or "cnj_number" not in best_mapping:
        raise XlsxHeaderError(
            "Planilha invalida — nao encontrei coluna 'CNJ' (ou aliases: "
            "'numero do processo', 'cnj number', 'processo'). Header esperado "
            "na linha 1 ou 2."
        )
    return best_row, best_mapping


def read_classificador_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    """Le bytes de xlsx e retorna (warnings, rows).

    Cada row do retorno e' um dict com chaves canonicas + valor raw da
    celula (None se vazia). Coluna ausente no header = chave ausente no dict.

    Linhas com CNJ vazio sao ignoradas (sem warning). Linhas com CNJ
    malformado entram com warning mas sao incluidas (pra falhar no
    refresh L1 individualmente sem bloquear o lote).
    """
    if not content:
        raise XlsxHeaderError("Arquivo vazio.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise XlsxHeaderError(f"Nao foi possivel ler o arquivo como XLSX: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise XlsxHeaderError("Planilha sem aba ativa.")

    header_row, mapping = _detect_header(ws)
    data_start_row = header_row + 1

    warnings: list[str] = []
    rows: list[dict] = []
    seen_cnjs: set[str] = set()

    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue  # linha completamente vazia

        # Extrai por canonical name
        out: dict = {}
        for canonical, col_idx in mapping.items():
            if col_idx <= len(row_vals):
                out[canonical] = row_vals[col_idx - 1]
            else:
                out[canonical] = None

        cnj_raw = out.get("cnj_number")
        cnj_norm = _normalize_cnj(cnj_raw)
        if cnj_norm is None:
            continue  # linha sem CNJ — provavel rodape, ignora silenciosamente

        # Detecta CNJ malformado (devolveu raw em vez de digitos)
        if not cnj_norm.isdigit() or len(cnj_norm) < _CNJ_DIGITS_MIN:
            warnings.append(
                f"Linha {row_idx}: CNJ '{cnj_raw}' parece malformado — sera "
                "incluido mas pode falhar no refresh L1."
            )

        # Dedup dentro do mesmo arquivo
        if cnj_norm in seen_cnjs:
            warnings.append(
                f"Linha {row_idx}: CNJ {cnj_norm} duplicado na planilha — "
                "ignorando duplicata."
            )
            continue
        seen_cnjs.add(cnj_norm)

        out["cnj_number"] = cnj_norm
        # Stringify campos textuais opcionais
        for key in ("cliente_externo_id", "produto", "observacao"):
            if key in out and out[key] is not None:
                out[key] = str(out[key]).strip() or None

        rows.append(out)

    wb.close()
    return warnings, rows
