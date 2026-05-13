"""Gerador de relatorio XLSX multi-aba do Classificador.

9 abas:
  1. Capa Executiva  — KPIs grandes + meta info
  2. Sumario Categoria
  3. Sumario Subcategoria
  4. Sumario Patrocinio  (MDR vs OUTRO vs CONDUCAO)
  5. Sumario Produto
  6. Sumario UF / Tribunal
  7. Top 20 por Valor
  8. Detalhamento Processos  (1 row/processo, autofilter, freeze pane)
  9. Pedidos Detalhados  (1 row/pedido)
  10. Sentencas + Transito  (counts)
  11. Analise Estrategica  (texto consolidado)

Entrada: dict retornado por `report_data.build_report_data`.
Saida: bytes do xlsx (pra salvar via report_storage.save_report).

Sem dependencia nova — usa openpyxl que ja esta instalado.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ─── Estilos compartilhados ──────────────────────────────────────────

COR_HEADER_BG = "1A365D"       # azul-escuro MDR-ish
COR_HEADER_FG = "FFFFFF"
COR_KPI_BG = "EBF4FF"          # azul claro pra KPIs
COR_ROW_ALT = "F7FAFC"         # cinza levissimo

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COR_HEADER_FG)
FONT_KPI_LABEL = Font(name="Calibri", size=10, bold=False, color="4A5568")
FONT_KPI_VALUE = Font(name="Calibri", size=18, bold=True, color="1A365D")
FONT_TITLE = Font(name="Calibri", size=20, bold=True, color="1A365D")
FONT_SUBTITLE = Font(name="Calibri", size=12, bold=False, color="4A5568")
FONT_BODY = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)

FILL_HEADER = PatternFill("solid", fgColor=COR_HEADER_BG)
FILL_KPI = PatternFill("solid", fgColor=COR_KPI_BG)

BORDER_THIN = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FMT_BRL = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_PERCENT = "0.0%"
FMT_DATE = "DD/MM/YYYY"


def _set_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 26


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_row_border(ws: Worksheet, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).border = BORDER_THIN


# ─── Aba 1: Capa Executiva ──────────────────────────────────────────


def _build_capa(ws: Worksheet, data: dict) -> None:
    lote = data["lote"]
    kpis = data["kpis"]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28

    ws["A1"] = "DIAGNOSTICO DE CARTEIRA"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:D1")

    cliente = lote.get("cliente_nome") or "(sem cliente)"
    ws["A2"] = f"Cliente: {cliente}"
    ws["A2"].font = FONT_SUBTITLE
    ws.merge_cells("A2:D2")

    ws["A3"] = (
        f"Lote #{lote['id']} — {lote['nome']} · "
        f"Snapshot: {_fmt_dt(lote.get('snapshot_at'))} · "
        f"Gerado: {_fmt_dt(data.get('generated_at'))}"
    )
    ws["A3"].font = FONT_BODY
    ws["A3"].alignment = ALIGN_LEFT
    ws.merge_cells("A3:D3")
    ws.row_dimensions[1].height = 32

    # KPIs em grid 2x4
    kpi_cards = [
        ("Total de processos", kpis.get("total_processos"), None),
        ("Classificados", kpis.get("total_classificados"), None),
        ("Valor total estimado", kpis.get("valor_total_estimado"), FMT_BRL),
        ("Aprovisionamento (PCOND)", kpis.get("pcond_total"), FMT_BRL),
        ("Valor total da causa", kpis.get("valor_total_causa"), FMT_BRL),
        ("Probabilidade media de exito", kpis.get("prob_exito_medio"), FMT_PERCENT),
        ("Com erro", kpis.get("total_com_erro"), None),
        ("Confianca media", None, None),  # placeholder
    ]

    row = 5
    for i, (label, value, fmt) in enumerate(kpi_cards):
        col_a = 1 if i % 2 == 0 else 3
        col_b = col_a + 1
        if i > 0 and i % 2 == 0:
            row += 3
        label_cell = ws.cell(row=row, column=col_a, value=label)
        label_cell.font = FONT_KPI_LABEL
        label_cell.fill = FILL_KPI
        label_cell.alignment = ALIGN_LEFT
        label_cell.border = BORDER_THIN

        val_cell = ws.cell(row=row + 1, column=col_a, value=value if value is not None else "—")
        val_cell.font = FONT_KPI_VALUE
        val_cell.fill = FILL_KPI
        val_cell.alignment = ALIGN_LEFT
        val_cell.border = BORDER_THIN
        if fmt and value is not None:
            val_cell.number_format = fmt

        # mescla 2 colunas pra cada card (label + value ocupam 2-col)
        ws.merge_cells(start_row=row, start_column=col_a,
                       end_row=row, end_column=col_b)
        ws.merge_cells(start_row=row + 1, start_column=col_a,
                       end_row=row + 1, end_column=col_b)
        ws.row_dimensions[row + 1].height = 30

    # Sentencas + transito (mini-resumo)
    base_row = row + 4
    ws.cell(row=base_row, column=1, value="Sentencas no lote").font = FONT_BODY_BOLD
    sent_resumo = data.get("sentencas_resumo") or {}
    for j, (tipo, count) in enumerate(sent_resumo.items(), start=1):
        ws.cell(row=base_row + j, column=1, value=tipo).font = FONT_BODY
        ws.cell(row=base_row + j, column=2, value=count).font = FONT_BODY
        ws.cell(row=base_row + j, column=2).alignment = ALIGN_RIGHT

    transit = data.get("transito_julgado_resumo") or {}
    ws.cell(row=base_row, column=3, value="Transito em julgado").font = FONT_BODY_BOLD
    ws.cell(row=base_row + 1, column=3, value="Transitados").font = FONT_BODY
    ws.cell(row=base_row + 1, column=4, value=transit.get("transitados", 0)).font = FONT_BODY
    ws.cell(row=base_row + 1, column=4).alignment = ALIGN_RIGHT
    ws.cell(row=base_row + 2, column=3, value="Nao transitados").font = FONT_BODY
    ws.cell(row=base_row + 2, column=4, value=transit.get("nao_transitados", 0)).font = FONT_BODY
    ws.cell(row=base_row + 2, column=4).alignment = ALIGN_RIGHT


def _fmt_dt(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso


# ─── Abas de sumario (estrutura comum) ───────────────────────────────


def _build_sumario_simple(
    ws: Worksheet,
    title: str,
    label_header: str,
    rows: list[dict],
) -> None:
    """Sumario padrao 5-cols: label, qtd, valor_estimado, pcond, prob_exito."""
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    _set_header_row(ws, 3, [label_header, "Qtd", "% do total", "Valor estimado", "PCOND", "Prob. exito media"])

    total_qtd = sum(r.get("qtd", 0) for r in rows) or 1
    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r.get("label"))
        ws.cell(row=i, column=2, value=r.get("qtd"))
        ws.cell(row=i, column=3, value=(r.get("qtd", 0) / total_qtd) if total_qtd else 0).number_format = FMT_PERCENT
        ws.cell(row=i, column=4, value=r.get("valor_estimado") or 0).number_format = FMT_BRL
        ws.cell(row=i, column=5, value=r.get("pcond") or 0).number_format = FMT_BRL
        pe = r.get("prob_exito_medio")
        c = ws.cell(row=i, column=6, value=pe if pe is not None else "—")
        if pe is not None:
            c.number_format = FMT_PERCENT
        # estilo zebra
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_ROW_ALT)
        _apply_row_border(ws, i, 6)

    _autosize(ws, {1: 42, 2: 10, 3: 12, 4: 20, 5: 20, 6: 18})
    ws.freeze_panes = "A4"


# ─── Aba 7: Top 20 por valor ────────────────────────────────────────


def _build_top_n(ws: Worksheet, top_n: list[dict]) -> None:
    ws["A1"] = "TOP 20 PROCESSOS POR VALOR ESTIMADO"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:G1")

    _set_header_row(ws, 3, ["#", "CNJ", "Tribunal", "Categoria",
                             "Valor estimado", "PCOND", "Prob. exito"])
    for i, p in enumerate(top_n, start=4):
        ws.cell(row=i, column=1, value=p.get("id"))
        ws.cell(row=i, column=2, value=p.get("cnj_number"))
        ws.cell(row=i, column=3, value=p.get("tribunal"))
        ws.cell(row=i, column=4, value=p.get("categoria"))
        c = ws.cell(row=i, column=5, value=p.get("valor_estimado") or 0)
        c.number_format = FMT_BRL
        c2 = ws.cell(row=i, column=6, value=p.get("pcond_sugerido") or 0)
        c2.number_format = FMT_BRL
        pe = p.get("prob_exito")
        c3 = ws.cell(row=i, column=7, value=pe if pe is not None else "—")
        if pe is not None:
            c3.number_format = FMT_PERCENT
        if i % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_ROW_ALT)
        _apply_row_border(ws, i, 7)

    _autosize(ws, {1: 6, 2: 28, 3: 12, 4: 38, 5: 20, 6: 20, 7: 14})
    ws.freeze_panes = "A4"


# ─── Aba 8: Detalhamento (1 row/processo) ────────────────────────────


_DETALHE_COLS = [
    ("id", "#", 6),
    ("cnj_number", "CNJ", 24),
    ("tribunal", "Tribunal", 10),
    ("uf", "UF", 6),
    ("vara", "Vara", 28),
    ("classe", "Classe", 22),
    ("valor_causa", "Valor da causa", 18),
    ("polo", "Polo MDR", 8),
    ("natureza_processo", "Natureza", 14),
    ("produto", "Produto", 18),
    ("categoria", "Categoria", 32),
    ("subcategoria", "Subcategoria", 30),
    ("valor_estimado", "Valor estimado", 18),
    ("pcond_sugerido", "PCOND", 18),
    ("prob_exito", "Prob. exito", 12),
    ("confianca", "Confianca", 12),
    ("patrocinio_decisao", "Patrocinio", 16),
    ("patrocinio_outro_advogado", "Outro advogado", 26),
    ("patrocinio_outro_oab", "OAB", 14),
    ("patrocinio_outro_escritorio", "Outro escritorio", 26),
    ("patrocinio_suspeita_devolucao", "Susp. devolucao", 12),
    ("contestacao_existe", "Tem contestacao?", 12),
    ("contestacao_por_mdr", "Por MDR?", 10),
    ("contestacao_por_nome", "Contestada por", 26),
    ("contestacao_generica", "Generica?", 10),
    ("sentenca_existe", "Tem sentenca?", 12),
    ("sentenca_tipo", "Tipo sentenca", 22),
    ("sentenca_data", "Data sentenca", 14),
    ("sentenca_valor", "Valor condenacao", 18),
    ("transito_julgado", "Transitou?", 10),
    ("primeira_hab_master_nome", "1a hab Master — nome", 26),
    ("primeira_hab_master_oab", "1a hab Master — OAB", 14),
    ("primeira_hab_master_data", "1a hab Master — data", 14),
    ("analise_estrategica", "Analise estrategica", 50),
    ("status", "Status processo", 18),
    ("extractor_used", "Extractor", 14),
    ("extraction_confidence", "Conf. extracao", 12),
]


def _build_detalhamento(ws: Worksheet, processos: list[dict]) -> None:
    ws["A1"] = "DETALHAMENTO POR PROCESSO"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_DETALHE_COLS))

    headers = [h[1] for h in _DETALHE_COLS]
    _set_header_row(ws, 3, headers)

    for i, p in enumerate(processos, start=4):
        for j, (key, _label, _width) in enumerate(_DETALHE_COLS, start=1):
            val = p.get(key)
            # Booleans -> "Sim"/"Nao"
            if isinstance(val, bool):
                val = "Sim" if val else "Nao"
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            if key in ("valor_causa", "valor_estimado", "pcond_sugerido", "sentenca_valor"):
                if val not in (None, ""):
                    cell.number_format = FMT_BRL
            elif key in ("prob_exito", "confianca"):
                if val not in (None, ""):
                    cell.number_format = FMT_PERCENT
            elif key in ("sentenca_data", "primeira_hab_master_data"):
                if isinstance(val, str):
                    cell.alignment = ALIGN_LEFT
        if i % 2 == 0:
            for col in range(1, len(_DETALHE_COLS) + 1):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_ROW_ALT)

    widths = {idx + 1: col[2] for idx, col in enumerate(_DETALHE_COLS)}
    _autosize(ws, widths)
    ws.freeze_panes = "C4"  # congela ate CNJ
    # AutoFilter
    last_col = get_column_letter(len(_DETALHE_COLS))
    last_row = max(3, 3 + len(processos))
    ws.auto_filter.ref = f"A3:{last_col}{last_row}"


# ─── Aba 9: Pedidos detalhados ───────────────────────────────────────


def _build_pedidos(ws: Worksheet, pedidos: list[dict]) -> None:
    ws["A1"] = "PEDIDOS DETALHADOS"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:H1")

    _set_header_row(ws, 3, [
        "Processo #", "CNJ", "Tipo pedido", "Natureza",
        "Valor indicado", "Valor estimado", "Prob. perda", "Aprovisionamento",
    ])
    for i, p in enumerate(pedidos, start=4):
        ws.cell(row=i, column=1, value=p.get("processo_id"))
        ws.cell(row=i, column=2, value=p.get("cnj_number"))
        ws.cell(row=i, column=3, value=p.get("tipo_pedido"))
        ws.cell(row=i, column=4, value=p.get("natureza"))
        c1 = ws.cell(row=i, column=5, value=p.get("valor_indicado") or 0)
        c1.number_format = FMT_BRL
        c2 = ws.cell(row=i, column=6, value=p.get("valor_estimado") or 0)
        c2.number_format = FMT_BRL
        ws.cell(row=i, column=7, value=p.get("probabilidade_perda"))
        c3 = ws.cell(row=i, column=8, value=p.get("aprovisionamento") or 0)
        c3.number_format = FMT_BRL
        if i % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_ROW_ALT)
        _apply_row_border(ws, i, 8)

    _autosize(ws, {1: 10, 2: 24, 3: 28, 4: 14, 5: 18, 6: 18, 7: 12, 8: 18})
    ws.freeze_panes = "C4"
    last_row = max(3, 3 + len(pedidos))
    ws.auto_filter.ref = f"A3:H{last_row}"


# ─── Aba 10: Pedidos por tipo (resumo) ───────────────────────────────


def _build_pedidos_resumo(ws: Worksheet, rows: list[dict]) -> None:
    ws["A1"] = "PEDIDOS POR TIPO"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:E1")

    _set_header_row(ws, 3, ["Tipo de pedido", "Qtd", "Valor indicado", "Valor estimado", "PCOND"])
    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r.get("tipo_pedido"))
        ws.cell(row=i, column=2, value=r.get("qtd"))
        c1 = ws.cell(row=i, column=3, value=r.get("valor_indicado") or 0)
        c1.number_format = FMT_BRL
        c2 = ws.cell(row=i, column=4, value=r.get("valor_estimado") or 0)
        c2.number_format = FMT_BRL
        c3 = ws.cell(row=i, column=5, value=r.get("pcond") or 0)
        c3.number_format = FMT_BRL
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=COR_ROW_ALT)
        _apply_row_border(ws, i, 5)

    _autosize(ws, {1: 36, 2: 10, 3: 18, 4: 18, 5: 18})
    ws.freeze_panes = "A4"


# ─── Aba 11: Analise estrategica ─────────────────────────────────────


def _build_analise(ws: Worksheet, data: dict) -> None:
    ws["A1"] = "ANALISE ESTRATEGICA"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:B1")

    carteira = data.get("analise_estrategica_carteira")
    if carteira:
        ws["A3"] = "Sintese da carteira"
        ws["A3"].font = FONT_BODY_BOLD
        ws["A4"] = carteira
        ws["A4"].alignment = ALIGN_LEFT
        ws.merge_cells("A4:B4")
        ws.row_dimensions[4].height = 80

    # Por processo (so quando tem analise individual)
    ws["A6"] = "Analises por processo"
    ws["A6"].font = FONT_BODY_BOLD
    _set_header_row(ws, 8, ["CNJ", "Analise estrategica"])

    row = 9
    for p in data.get("processos", []):
        if not p.get("analise_estrategica"):
            continue
        ws.cell(row=row, column=1, value=p.get("cnj_number"))
        c = ws.cell(row=row, column=2, value=p.get("analise_estrategica"))
        c.alignment = ALIGN_LEFT
        _apply_row_border(ws, row, 2)
        row += 1

    _autosize(ws, {1: 26, 2: 90})


# ─── Top-level ───────────────────────────────────────────────────────


def generate_xlsx_report(data: dict) -> bytes:
    """Gera o xlsx multi-aba a partir do payload do `build_report_data`.

    Returns:
        bytes do arquivo xlsx (pra salvar via report_storage.save_report).
    """
    wb = Workbook()

    # Remove a aba "Sheet" default
    default_ws = wb.active
    wb.remove(default_ws)

    # 1. Capa
    ws = wb.create_sheet("Capa Executiva")
    _build_capa(ws, data)

    # 2. Categoria
    ws = wb.create_sheet("Por Categoria")
    _build_sumario_simple(ws, "SUMARIO POR CATEGORIA", "Categoria",
                           data.get("por_categoria", []))

    # 3. Subcategoria
    ws = wb.create_sheet("Por Subcategoria")
    _build_sumario_simple(ws, "SUMARIO POR SUBCATEGORIA", "Categoria / Sub",
                           data.get("por_subcategoria", []))

    # 4. Patrocinio
    ws = wb.create_sheet("Por Patrocinio")
    _build_sumario_simple(ws, "SUMARIO POR PATROCINIO", "Decisao",
                           data.get("por_patrocinio", []))

    # 5. Produto
    ws = wb.create_sheet("Por Produto")
    _build_sumario_simple(ws, "SUMARIO POR PRODUTO", "Produto",
                           data.get("por_produto", []))

    # 6. UF
    ws = wb.create_sheet("Por UF")
    _build_sumario_simple(ws, "SUMARIO POR UF", "UF / Tribunal",
                           data.get("por_uf", []))

    # 7. Tribunal
    ws = wb.create_sheet("Por Tribunal")
    _build_sumario_simple(ws, "SUMARIO POR TRIBUNAL", "Tribunal",
                           data.get("por_tribunal", []))

    # 8. Pedidos por tipo (resumo)
    ws = wb.create_sheet("Pedidos por tipo")
    _build_pedidos_resumo(ws, data.get("pedidos_por_tipo", []))

    # 9. Top 20
    ws = wb.create_sheet("Top 20")
    _build_top_n(ws, data.get("top_n_valor", []))

    # 10. Detalhamento processos
    ws = wb.create_sheet("Detalhamento")
    _build_detalhamento(ws, data.get("processos", []))

    # 11. Pedidos
    ws = wb.create_sheet("Pedidos")
    _build_pedidos(ws, data.get("pedidos", []))

    # 12. Analise estrategica
    ws = wb.create_sheet("Analise")
    _build_analise(ws, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
