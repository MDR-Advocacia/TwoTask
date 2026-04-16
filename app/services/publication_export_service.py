"""
Exportação XLSX das publicações listadas em "Processos com Publicações".

Mantém os mesmos filtros usados em ``list_records_grouped`` e gera um
arquivo organizado, com uma linha por publicação (layout plano) para
facilitar filtros dentro do Excel.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.legal_one import LegalOneOffice
from app.models.publication_search import PublicationRecord


# Ordem e cabeçalho das colunas exportadas. Espelha os campos visíveis
# na tela "Processos com Publicações", mais contexto útil da classificação.
EXPORT_COLUMNS: list[tuple[str, str, int]] = [
    # (chave interna, título, largura aproximada em caracteres)
    ("processo", "Processo (CNJ)", 24),
    ("escritorio", "Escritório", 28),
    ("publication_date", "Data da Publicação", 20),
    ("creation_date", "Data no Ajus", 20),
    ("status", "Status", 14),
    ("category", "Classificação", 24),
    ("subcategory", "Subclassificação", 24),
    ("polo", "Polo", 10),
    ("confianca", "Confiança", 12),
    ("audiencia_data", "Audiência (Data)", 16),
    ("audiencia_hora", "Audiência (Hora)", 12),
    ("justificativa", "Justificativa da IA", 60),
    ("descricao", "Texto da Publicação", 80),
    ("proposta", "Proposta de Tarefa (Template)", 32),
    ("legal_one_update_id", "ID L1", 12),
    ("link", "Link no Legal One", 48),
]


def _apply_filters(query, *, search_id, status, linked_office_id, date_from, date_to, category):
    query = query.filter(PublicationRecord.is_duplicate == False)  # noqa: E712
    if search_id is not None:
        query = query.filter(PublicationRecord.search_id == search_id)
    if status:
        query = query.filter(PublicationRecord.status == status)
    if linked_office_id is not None:
        query = query.filter(PublicationRecord.linked_office_id == linked_office_id)
    if date_from:
        query = query.filter(PublicationRecord.creation_date >= date_from)
    if date_to:
        # Mantém o mesmo comportamento de list_records_grouped (inclui o dia inteiro).
        query = query.filter(PublicationRecord.creation_date < date_to + "T99")
    if category:
        query = query.filter(PublicationRecord.category == category)
    return query


def _primary_classification(record: PublicationRecord) -> dict[str, Any]:
    """Extrai a classificação primária (primeiro item do JSON ``classifications``)."""
    if isinstance(record.classifications, list) and record.classifications:
        first = record.classifications[0]
        if isinstance(first, dict):
            return first
    return {}


def _proposal_label(record: PublicationRecord) -> Optional[str]:
    """Retorna o nome do template da proposta de tarefa, se houver."""
    raw = getattr(record, "raw_relationships", None)
    if not isinstance(raw, dict):
        return None
    proposal = raw.get("_proposed_task")
    if isinstance(proposal, dict):
        template = proposal.get("template_name") or (
            proposal.get("payload", {}).get("template_name")
            if isinstance(proposal.get("payload"), dict) else None
        )
        if template:
            return str(template)
    proposals = raw.get("_proposed_tasks")
    if isinstance(proposals, list):
        names = []
        for rp in proposals:
            if isinstance(rp, dict):
                name = rp.get("template_name") or (
                    rp.get("payload", {}).get("template_name")
                    if isinstance(rp.get("payload"), dict) else None
                )
                if name:
                    names.append(str(name))
        if names:
            return " | ".join(dict.fromkeys(names))  # dedup mantendo ordem
    return None


def _truncate(value: Any, limit: int = 32_000) -> Any:
    """Excel tem limite de ~32767 caracteres por célula; corta com sufixo."""
    if isinstance(value, str) and len(value) > limit:
        return value[: limit - 20] + " …[truncado]"
    return value


def _row_for_record(
    record: PublicationRecord,
    office_names: dict[int, str],
) -> dict[str, Any]:
    cls = _primary_classification(record)
    return {
        "processo": record.linked_lawsuit_cnj or "",
        "escritorio": (
            office_names.get(record.linked_office_id, "")
            if record.linked_office_id is not None
            else ""
        ),
        "publication_date": record.publication_date or "",
        "creation_date": record.creation_date or "",
        "status": record.status or "",
        "category": record.category or "",
        "subcategory": record.subcategory or "",
        "polo": record.polo or "",
        "confianca": cls.get("confianca") if cls else None,
        "audiencia_data": record.audiencia_data or "",
        "audiencia_hora": record.audiencia_hora or "",
        "justificativa": _truncate(cls.get("justificativa") or ""),
        "descricao": _truncate(record.description or ""),
        "proposta": _proposal_label(record) or "",
        "legal_one_update_id": record.legal_one_update_id,
        "link": (
            f"https://firm.legalone.com.br/publications?publicationId={record.legal_one_update_id}&treatStatus=3"
            if record.legal_one_update_id
            else ""
        ),
    }


def export_records_grouped_xlsx(
    db: Session,
    *,
    search_id: Optional[int] = None,
    status: Optional[str] = None,
    linked_office_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    category: Optional[str] = None,
    uf: Optional[str] = None,
) -> tuple[bytes, str]:
    """Gera o XLSX e devolve ``(bytes, filename)``."""

    query = db.query(PublicationRecord)
    query = _apply_filters(
        query,
        search_id=search_id,
        status=status,
        linked_office_id=linked_office_id,
        date_from=date_from,
        date_to=date_to,
        category=category,
    )
    # Filtro UF via coluna materializada (SQL) — antes era em memória.
    if uf:
        query = query.filter(PublicationRecord.uf == uf.strip().upper())

    records = (
        query
        .order_by(
            PublicationRecord.linked_lawsuit_cnj.asc().nullslast(),
            PublicationRecord.publication_date.desc(),
        )
        .all()
    )

    # Pré-carrega nomes de escritórios para mapear id → nome sem N+1.
    office_ids = {r.linked_office_id for r in records if r.linked_office_id is not None}
    office_names: dict[int, str] = {}
    if office_ids:
        offices = (
            db.query(LegalOneOffice)
            .filter(LegalOneOffice.external_id.in_(office_ids))
            .all()
        )
        office_names = {o.external_id: o.name for o in offices}

    wb = Workbook()
    ws = wb.active
    ws.title = "Publicações"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    # Cabeçalho
    for idx, (_key, title, width) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Linhas
    for row_idx, record in enumerate(records, start=2):
        row = _row_for_record(record, office_names)
        for col_idx, (key, _title, _width) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key))
            cell.alignment = body_align

    # Autofiltro sobre todo o range usado
    if records:
        last_col = get_column_letter(len(EXPORT_COLUMNS))
        ws.auto_filter.ref = f"A1:{last_col}{len(records) + 1}"

    # Aba de metadados — deixa rastreável qual filtro gerou o arquivo.
    meta = wb.create_sheet("Filtros")
    meta_rows = [
        ("Gerado em (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Total de publicações exportadas", len(records)),
        ("search_id", search_id if search_id is not None else ""),
        ("status", status or ""),
        ("linked_office_id", linked_office_id if linked_office_id is not None else ""),
        ("date_from", date_from or ""),
        ("date_to", date_to or ""),
        ("category", category or ""),
        ("uf", uf or ""),
    ]
    for row_idx, (label, value) in enumerate(meta_rows, start=1):
        meta.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        meta.cell(row=row_idx, column=2, value=value)
    meta.column_dimensions["A"].width = 36
    meta.column_dimensions["B"].width = 48

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"publicacoes-{stamp}.xlsx"
    return buffer.read(), filename
