"""Geracao XLSX dos 6 templates de relatorio do Base Processual (Chunk 5).

Cada template recebe (db, params) e devolve (xlsx_bytes, total_rows, params_normalized).
A funcao publica `dispatch_template(name, db, params)` roteia por nome.

V1: SINCRONO — pra carteira de ~6k processos cada template gera em < 5s.
V2: APScheduler pra >50k processos + queueing.
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Callable, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import types as sa_types
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.models.base_processual import (
    BaseProcessualEvento,
    BaseProcessualProcesso,
    BaseProcessualSnapshot,
    EVENTO_ATUALIZADO,
    EVENTO_ATUALIZADO_MANUAL,
    EVENTO_ENTROU,
    EVENTO_SAIU,
    EXPORT_TPL_CARTEIRA_RESPONSAVEL,
    EXPORT_TPL_CARTEIRA_UF_COMARCA,
    EXPORT_TPL_MOVIMENTACAO_SEMANAL,
    EXPORT_TPL_SNAPSHOT_COMPLETO,
    EXPORT_TPL_SUMICOS_PERIODO,
    EXPORT_TPL_VARIACAO_VALORES,
    PRESENCA_ATIVO,
    PRESENCA_REMOVIDO,
)

logger = logging.getLogger(__name__)


# Styling helpers
HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
HEADER_FONT = Font(bold=True, color="111827")


def _new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # remove a sheet default — vamos criar tudo via _add_sheet
    if wb.active is not None:
        wb.remove(wb.active)
    return wb


def _add_sheet(
    wb: openpyxl.Workbook,
    name: str,
    headers: list[str],
    rows: list[list],
) -> int:
    """Cria sheet com header estilizado + rows. Retorna numero de rows escritas."""
    ws = wb.create_sheet(title=name[:31])
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in rows:
        ws.append(r)
    # Auto widths conservadores: limita 8..60.
    for i, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = 8
        for r in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=r, column=i).value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[col_letter].width = min(60, max_len + 2)
    return len(rows)


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _parse_date(value: Any) -> Optional[date]:
    """Parse 'YYYY-MM-DD' ou ISO datetime ou date direto."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _money(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _partes_nomes(partes) -> str:
    if not partes:
        return ""
    if isinstance(partes, list):
        return "; ".join(p.get("nome") or "" for p in partes if isinstance(p, dict))
    return str(partes)


def _iso_or_blank(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, (datetime, date)):
        return dt.isoformat()
    return str(dt)


# ============================================================================
# Templates
# ============================================================================


def gen_movimentacao_semanal(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Sumario por dia + listas Entraram/Sairam/Atualizados no periodo.

    Default: ultimos 7 dias (incluindo hoje).
    """
    from_d = _parse_date(params.get("from_date"))
    to_d = _parse_date(params.get("to_date"))
    if to_d is None:
        to_d = datetime.utcnow().date()
    if from_d is None:
        from_d = to_d - timedelta(days=6)
    start = datetime(from_d.year, from_d.month, from_d.day)
    end = datetime(to_d.year, to_d.month, to_d.day) + timedelta(days=1)
    norm_params = {"from_date": from_d.isoformat(), "to_date": to_d.isoformat()}

    wb = _new_workbook()

    # Sumario por dia
    day_col = sa_func.cast(BaseProcessualEvento.created_at, sa_types.Date)
    rows_sum = (
        db.query(
            day_col.label("d"),
            BaseProcessualEvento.tipo_evento,
            sa_func.count().label("total"),
        )
        .filter(BaseProcessualEvento.created_at >= start)
        .filter(BaseProcessualEvento.created_at < end)
        .group_by("d", BaseProcessualEvento.tipo_evento)
        .all()
    )
    pivot: dict[date, dict[str, int]] = {}
    for d_val, tipo, total in rows_sum:
        dd = d_val if isinstance(d_val, date) else d_val.date()
        pivot.setdefault(dd, {})[tipo] = int(total)
    sumario_rows = []
    cur = from_d
    while cur <= to_d:
        bucket = pivot.get(cur, {})
        sumario_rows.append([
            cur.isoformat(),
            bucket.get(EVENTO_ENTROU, 0),
            bucket.get(EVENTO_SAIU, 0),
            bucket.get(EVENTO_ATUALIZADO, 0) + bucket.get(EVENTO_ATUALIZADO_MANUAL, 0),
        ])
        cur += timedelta(days=1)
    _add_sheet(
        wb, "Sumário",
        ["Data", "Entraram", "Saíram", "Atualizados"],
        sumario_rows,
    )

    # Listas detalhadas
    def _query_eventos(tipos: list[str]):
        return (
            db.query(BaseProcessualEvento, BaseProcessualProcesso)
            .join(
                BaseProcessualProcesso,
                BaseProcessualEvento.processo_id == BaseProcessualProcesso.id,
            )
            .filter(BaseProcessualEvento.created_at >= start)
            .filter(BaseProcessualEvento.created_at < end)
            .filter(BaseProcessualEvento.tipo_evento.in_(tipos))
            .order_by(BaseProcessualEvento.created_at.desc())
            .all()
        )

    headers_ev = [
        "Quando", "Cód AJUS", "CNJ", "Empresa", "UF",
        "Comarca", "Responsável", "Valor causa", "Tipo",
    ]
    entrou_list = []
    for e, p in _query_eventos([EVENTO_ENTROU]):
        entrou_list.append([
            _iso_or_blank(e.created_at), e.cod_ajus,
            p.numero_processo_mascarado or "", p.empresa or "", p.uf or "",
            p.comarca or "", p.usuario_responsavel or "",
            _money(p.valor_causa), e.tipo_evento,
        ])
    _add_sheet(wb, "Entraram", headers_ev, entrou_list)

    saiu_list = []
    for e, p in _query_eventos([EVENTO_SAIU]):
        saiu_list.append([
            _iso_or_blank(e.created_at), e.cod_ajus,
            p.numero_processo_mascarado or "", p.empresa or "", p.uf or "",
            p.comarca or "", p.usuario_responsavel or "",
            _money(p.valor_causa), e.tipo_evento,
        ])
    _add_sheet(wb, "Saíram", headers_ev, saiu_list)

    headers_atu = headers_ev + ["Campos mudados"]
    atu_list = []
    for e, p in _query_eventos([EVENTO_ATUALIZADO, EVENTO_ATUALIZADO_MANUAL]):
        campos = ", ".join((e.changed_fields or {}).keys()) if e.changed_fields else ""
        atu_list.append([
            _iso_or_blank(e.created_at), e.cod_ajus,
            p.numero_processo_mascarado or "", p.empresa or "", p.uf or "",
            p.comarca or "", p.usuario_responsavel or "",
            _money(p.valor_causa), e.tipo_evento, campos,
        ])
    _add_sheet(wb, "Atualizados", headers_atu, atu_list)

    total = len(entrou_list) + len(saiu_list) + len(atu_list)
    return _wb_to_bytes(wb), total, norm_params


def gen_carteira_responsavel(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Agrupado por usuario_responsavel da carteira ATIVA: totais e valores."""
    empresa = (params.get("empresa") or "").strip() or None
    norm_params = {"empresa": empresa} if empresa else {}

    q = (
        db.query(
            BaseProcessualProcesso.usuario_responsavel,
            sa_func.count().label("total_ativos"),
            sa_func.coalesce(sa_func.sum(BaseProcessualProcesso.valor_causa), 0).label(
                "soma_valor_causa"
            ),
            sa_func.coalesce(
                sa_func.sum(BaseProcessualProcesso.valor_contingencia), 0
            ).label("soma_valor_contingencia"),
        )
        .filter(BaseProcessualProcesso.presenca_status == PRESENCA_ATIVO)
        .group_by(BaseProcessualProcesso.usuario_responsavel)
        .order_by(sa_func.count().desc())
    )
    if empresa:
        q = q.filter(BaseProcessualProcesso.empresa == empresa)

    rows_data = q.all()
    rows = [
        [
            r[0] or "(sem responsável)",
            int(r[1]),
            _money(r[2]),
            _money(r[3]),
        ]
        for r in rows_data
    ]

    wb = _new_workbook()
    _add_sheet(
        wb, "Carteira por responsável",
        ["Responsável", "Ativos", "Σ Valor causa", "Σ Valor contingência"],
        rows,
    )
    return _wb_to_bytes(wb), len(rows), norm_params


def gen_sumicos_periodo(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Processos com presenca=REMOVIDO + SAIU dentro do periodo (default = mes corrente)."""
    from_d = _parse_date(params.get("from_date"))
    to_d = _parse_date(params.get("to_date"))
    today = datetime.utcnow().date()
    if from_d is None:
        from_d = today.replace(day=1)
    if to_d is None:
        to_d = today
    start = datetime(from_d.year, from_d.month, from_d.day)
    end = datetime(to_d.year, to_d.month, to_d.day) + timedelta(days=1)
    norm_params = {"from_date": from_d.isoformat(), "to_date": to_d.isoformat()}

    rows_data = (
        db.query(BaseProcessualEvento, BaseProcessualProcesso)
        .join(
            BaseProcessualProcesso,
            BaseProcessualEvento.processo_id == BaseProcessualProcesso.id,
        )
        .filter(BaseProcessualEvento.tipo_evento == EVENTO_SAIU)
        .filter(BaseProcessualEvento.created_at >= start)
        .filter(BaseProcessualEvento.created_at < end)
        .filter(BaseProcessualProcesso.presenca_status == PRESENCA_REMOVIDO)
        .order_by(BaseProcessualEvento.created_at.desc())
        .all()
    )
    rows = []
    for e, p in rows_data:
        rows.append([
            _iso_or_blank(e.created_at),
            p.cod_ajus,
            p.numero_processo_mascarado or "",
            p.empresa or "",
            p.uf or "",
            p.comarca or "",
            p.usuario_responsavel or "",
            _money(p.valor_causa),
            _partes_nomes(p.autores_json),
        ])

    wb = _new_workbook()
    _add_sheet(
        wb, "Sumiços",
        ["Saiu em", "Cód AJUS", "CNJ", "Empresa", "UF", "Comarca",
         "Responsável", "Valor causa", "Autores"],
        rows,
    )
    return _wb_to_bytes(wb), len(rows), norm_params


def gen_variacao_valores(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Processos com mudanca de valor_causa >= threshold_pct comparando.

    Compara o snapshot ATUAL (current_snapshot_id) com o PRIMEIRO snapshot
    do processo (historico inteiro). Mostra variacao absoluta e percentual.
    Default threshold = 50%. Em v1 nao filtra por periodo — relatorio
    historico. v2 pode aceitar from_date pra restringir.
    """
    threshold_pct = float(params.get("threshold_pct") or 50.0)
    norm_params = {"threshold_pct": threshold_pct}

    # Pega todos os processos com current_snapshot_id NOT NULL (carteira ja' upada)
    processos = (
        db.query(BaseProcessualProcesso)
        .filter(BaseProcessualProcesso.current_snapshot_id.isnot(None))
        .all()
    )

    rows = []
    for p in processos:
        primeiro = (
            db.query(BaseProcessualSnapshot)
            .filter(BaseProcessualSnapshot.processo_id == p.id)
            .order_by(BaseProcessualSnapshot.captured_at.asc())
            .first()
        )
        if primeiro is None:
            continue
        try:
            antes_raw = (primeiro.payload_normalized or {}).get("valor_causa")
            antes = float(antes_raw) if antes_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            antes = 0.0
        depois = _money(p.valor_causa)
        if antes == 0 and depois == 0:
            continue
        delta = depois - antes
        # %: se antes=0 e depois>0, marca como "novo" (100% por convencao);
        # senao calcula percentual relativo a antes.
        if antes == 0:
            pct = 100.0 if depois > 0 else -100.0 if depois < 0 else 0.0
        else:
            pct = (delta / antes) * 100.0
        if abs(pct) < threshold_pct:
            continue
        rows.append([
            p.cod_ajus,
            p.numero_processo_mascarado or "",
            p.empresa or "",
            p.uf or "",
            p.usuario_responsavel or "",
            antes,
            depois,
            delta,
            round(pct, 2),
        ])
    rows.sort(key=lambda r: abs(r[8]), reverse=True)

    wb = _new_workbook()
    _add_sheet(
        wb, "Variação de valores",
        ["Cód AJUS", "CNJ", "Empresa", "UF", "Responsável",
         "Antes", "Depois", "Δ Absoluto", "Δ %"],
        rows,
    )
    return _wb_to_bytes(wb), len(rows), norm_params


def gen_carteira_uf_comarca(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Pivot por UF + comarca: total ATIVOS e soma valor causa.

    Sheet 1: UF-Comarca detalhado. Sheet 2: UF agregado.
    """
    empresa = (params.get("empresa") or "").strip() or None
    norm_params = {"empresa": empresa} if empresa else {}

    base = db.query(BaseProcessualProcesso).filter(
        BaseProcessualProcesso.presenca_status == PRESENCA_ATIVO
    )
    if empresa:
        base = base.filter(BaseProcessualProcesso.empresa == empresa)

    # UF + Comarca detalhado
    rows_detail = (
        base.with_entities(
            BaseProcessualProcesso.uf,
            BaseProcessualProcesso.comarca,
            sa_func.count().label("total"),
            sa_func.coalesce(sa_func.sum(BaseProcessualProcesso.valor_causa), 0).label(
                "soma"
            ),
        )
        .group_by(
            BaseProcessualProcesso.uf, BaseProcessualProcesso.comarca
        )
        .order_by(
            BaseProcessualProcesso.uf.asc().nullslast(),
            sa_func.count().desc(),
        )
        .all()
    )
    detail = [
        [r[0] or "—", r[1] or "—", int(r[2]), _money(r[3])]
        for r in rows_detail
    ]

    # UF agregado
    rows_uf = (
        base.with_entities(
            BaseProcessualProcesso.uf,
            sa_func.count().label("total"),
            sa_func.coalesce(sa_func.sum(BaseProcessualProcesso.valor_causa), 0).label(
                "soma"
            ),
        )
        .group_by(BaseProcessualProcesso.uf)
        .order_by(sa_func.count().desc())
        .all()
    )
    agreg = [
        [r[0] or "—", int(r[1]), _money(r[2])]
        for r in rows_uf
    ]

    wb = _new_workbook()
    _add_sheet(
        wb, "UF + Comarca",
        ["UF", "Comarca", "Ativos", "Σ Valor causa"],
        detail,
    )
    _add_sheet(wb, "UF", ["UF", "Ativos", "Σ Valor causa"], agreg)
    return _wb_to_bytes(wb), len(detail) + len(agreg), norm_params


def gen_snapshot_completo(db: Session, params: dict) -> tuple[bytes, int, dict]:
    """Estado atual de todos os processos da carteira (1 linha por processo).

    Default: presenca=ATIVO_NA_BASE. Para incluir REMOVIDOs, passe presenca_status=ambos.
    Colunas similares a planilha original do Reports do L1.
    """
    presenca = params.get("presenca_status") or PRESENCA_ATIVO
    norm_params = {"presenca_status": presenca}

    q = db.query(BaseProcessualProcesso).order_by(
        BaseProcessualProcesso.cod_ajus.asc()
    )
    if presenca != "ambos":
        q = q.filter(BaseProcessualProcesso.presenca_status == presenca)

    headers = [
        "Cód AJUS", "Nº Processo (CNJ)", "Nº Pasta", "Nº Interno",
        "Ação Principal", "Matéria", "Risco/Prob. Perda", "Tipo de Ação",
        "Polo", "Natureza", "Nº Vara", "Foro", "Comarca", "UF",
        "Empresa", "Grupo Responsável", "Usuário Responsável",
        "Escritório Responsável", "Situação", "Justiça/Honorário",
        "Valor Causa", "Valor Prev. Acordo", "Valor Acordo", "Valor Discutido",
        "Valor Êxito", "Valor Condenação", "Valor Contingência",
        "Últ. Andamento", "Data Últ. Andamento", "Dias Últ. Atualização",
        "Distribuído em", "Processo Virtual", "Nº Contrato",
        "Autores", "Réus", "Presença na base",
    ]
    rows = []
    for p in q.yield_per(500):
        rows.append([
            p.cod_ajus,
            p.numero_processo_mascarado or "",
            p.numero_pasta or "",
            p.numero_interno or "",
            p.acao_principal or "",
            p.materia or "",
            p.risco_prob_perda or "",
            p.tipo_acao or "",
            p.polo or "",
            p.natureza or "",
            p.numero_vara or "",
            p.foro or "",
            p.comarca or "",
            p.uf or "",
            p.empresa,
            p.grupo_responsavel or "",
            p.usuario_responsavel or "",
            p.escritorio_responsavel or "",
            p.situacao_processo,
            p.justica_honorario or "",
            _money(p.valor_causa),
            _money(p.valor_prev_acordo),
            _money(p.valor_acordo),
            _money(p.valor_discutido),
            _money(p.valor_exito),
            _money(p.valor_condenacao),
            _money(p.valor_contingencia),
            p.ult_andamento or "",
            _iso_or_blank(p.data_ult_andamento),
            p.dias_ult_atualizacao if p.dias_ult_atualizacao is not None else "",
            _iso_or_blank(p.distribuido_em),
            "Sim" if p.processo_virtual else ("Não" if p.processo_virtual is False else ""),
            p.numero_contrato or "",
            _partes_nomes(p.autores_json),
            _partes_nomes(p.reus_json),
            p.presenca_status,
        ])

    wb = _new_workbook()
    _add_sheet(wb, "Snapshot", headers, rows)
    return _wb_to_bytes(wb), len(rows), norm_params


# ============================================================================
# Roteador
# ============================================================================

_REGISTRY: dict[str, Callable[[Session, dict], tuple[bytes, int, dict]]] = {
    EXPORT_TPL_MOVIMENTACAO_SEMANAL: gen_movimentacao_semanal,
    EXPORT_TPL_CARTEIRA_RESPONSAVEL: gen_carteira_responsavel,
    EXPORT_TPL_SUMICOS_PERIODO: gen_sumicos_periodo,
    EXPORT_TPL_VARIACAO_VALORES: gen_variacao_valores,
    EXPORT_TPL_CARTEIRA_UF_COMARCA: gen_carteira_uf_comarca,
    EXPORT_TPL_SNAPSHOT_COMPLETO: gen_snapshot_completo,
}


def dispatch_template(
    template_name: str, db: Session, params: Optional[dict]
) -> tuple[bytes, int, dict]:
    """Dispara o template adequado. Levanta ValueError se nao existir."""
    fn = _REGISTRY.get(template_name)
    if fn is None:
        raise ValueError(
            f"Template desconhecido: {template_name!r}. "
            f"Validos: {sorted(_REGISTRY.keys())}"
        )
    return fn(db, params or {})


def list_templates() -> list[str]:
    return sorted(_REGISTRY.keys())
