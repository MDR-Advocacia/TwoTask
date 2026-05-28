"""Converte a Listagem de Acoes Judiciais (saida AJUS / RPA L1)
no XLSX de migracao do Legal One.

Port do script `gerar_planilha.py` (autor Jonilson Vilela) usado
historicamente fora do sistema. Mantem a logica:
- Banco Master sempre como Reu.
- Responsavel fixo "Jose Alberto Veloso de Carvalho".
- Escritorio fixo "MDR Advocacia / Area operacional / Banco Master / Reu".
- Marca "bmagravo" nas Observacoes quando a Acao = "Agravo de Instrumento"
  ou o numero do processo termina em ".0000"; caso contrario "bmcomum".

API publica: `gerar_planilha_l1(listagem_bytes) -> bytes`.
"""

from __future__ import annotations

import io
import re
import warnings as _warnings
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl import load_workbook

_warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

TEMPLATE_PATH = Path(__file__).parent / "templates" / "MODELO_LEGAL_ONE.xlsx"

RESPONSAVEIS_REU = ["Jose Alberto Veloso de Carvalho"]
ESCRITORIO_REU = "MDR Advocacia / Área operacional / Banco Master / Réu"
CLIENTE_PRINCIPAL = "Banco Master S.A. - Em Liquidação Extrajudicial"
CLIENTE_CNPJ = "33.923.798/0001-00"
ESCRITORIO_ORIGEM = "MDR Advocacia"

# Header da Listagem L1: mapeia nomes que aparecem na planilha para o
# nome canonico usado por _montar_linha.  Tolera variacoes do export.
COLUMN_ALIASES: dict[str, list[str]] = {
    "Processo": ["Números Processo", "Numeros Processo", "Número Processo"],
    "Ação": ["Tipo de Ação", "Ação"],
    "Polo": ["Polo"],
    "Autores": ["Autores - CNPJCPF", "Autores"],
    "Reus": ["Réus - CNPJCPF", "Reus - CNPJCPF", "Réus"],
    "UF": ["UF"],
    "Comarca": ["Comarca"],
    "Data ajuizamento": ["Distribuído em", "Distribuido em", "Data ajuizamento"],
    "Valor da Causa": ["Valor Causa", "Valor da Causa"],
}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().lower().split())


def _extrair_nome(texto: Any) -> str:
    if not isinstance(texto, str):
        return ""
    match = re.search(
        r"Nome:\s*(.*?)\s*(?:CNPJCPF:|\n|$)", texto, re.IGNORECASE
    )
    if match:
        return match.group(1).strip()
    return texto.strip()


def _limpar_data(valor: Any) -> Any:
    if valor is None:
        return None
    if str(valor).strip().lower() == "a cadastrar":
        return None
    return valor


def _eh_agravo(acao: Any, numero: Any) -> bool:
    acao_str = str(acao).strip().lower() if acao is not None else ""
    numero_str = str(numero).strip() if numero is not None else ""
    return acao_str == "agravo de instrumento" or numero_str.endswith(".0000")


def _converter_valor_causa(valor_original: Any) -> Optional[float]:
    if valor_original is None or valor_original == "":
        return None
    if isinstance(valor_original, (int, float)):
        return float(valor_original)
    valor_str = (
        str(valor_original)
        .replace("R$", "")
        .replace(".", "")
        .replace(",", ".")
        .strip()
    )
    try:
        return float(valor_str)
    except ValueError:
        return None


def _localizar_header_e_indices(
    ws,
) -> tuple[int, dict[str, int]]:
    """Localiza a linha de cabecalho contendo 'Polo' e devolve (linha, mapa).

    O mapa associa nome canonico (Processo, Acao, ...) ao indice 0-based
    da celula na linha. Levanta ValueError se nao achar a coluna 'Polo'.
    """
    aliases_norm: dict[str, list[str]] = {
        canon: [_norm(a) for a in aliases]
        for canon, aliases in COLUMN_ALIASES.items()
    }

    max_probe_rows = 30
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_probe_rows, values_only=True),
        start=1,
    ):
        if row is None:
            continue
        normalized = [_norm(c) for c in row]
        if "polo" not in normalized:
            continue

        idx_map: dict[str, int] = {}
        for canon, alist in aliases_norm.items():
            for cell_idx, cell in enumerate(normalized):
                if cell in alist:
                    idx_map[canon] = cell_idx
                    break
        return row_idx, idx_map

    raise ValueError(
        "Não foi possível localizar a linha do cabeçalho com a coluna 'Polo'. "
        "Confira se o arquivo é a Listagem de Ações Judiciais exportada do Legal One."
    )


def _montar_linha(
    raw: dict[str, Any],
    chave: int,
    responsavel: str,
    linha_modelo_len: int,
    data_extracao: str,
) -> list[Any]:
    adverso = raw.get("Autores", "")
    linha: list[Any] = [""] * linha_modelo_len

    linha[0] = chave                                          # Chave do Processo *
    linha[1] = None                                           # Chave Processo Pai
    linha[2] = "Processo"                                     # Tipo de Registro
    linha[3] = raw.get("Processo", "")                        # Número CNJ
    linha[4] = "Judicial"                                     # Tipo *
    linha[5] = "Ativo"                                        # Status *
    linha[6] = "Cível"                                        # Natureza *
    linha[7] = CLIENTE_PRINCIPAL                              # Cliente principal
    linha[8] = "Réu"                                          # Posição
    linha[9] = CLIENTE_CNPJ                                   # CPF/CNPJ cliente
    linha[10] = "PJ"                                          # Tipo cliente
    linha[11] = _extrair_nome(adverso)                        # Contrário principal
    linha[12] = None                                          # CPF/CNPJ contrário
    linha[13] = "PF" if linha[11] else None                   # Tipo contrário
    linha[14] = None                                          # Número Antigo
    linha[15] = _limpar_data(raw.get("Data ajuizamento", "")) # Distribuição
    linha[16] = raw.get("Ação", "")                           # Ação
    linha[17] = data_extracao                                 # Título
    linha[18] = ""                                            # Fase
    linha[19] = raw.get("UF", "")                             # UF
    linha[20] = raw.get("Comarca", "")                        # Cidade
    linha[21] = ""                                            # Órgão
    linha[22] = "Justiça Estadual"                            # Justiça (CNJ)
    linha[23] = None                                          # Instância (CNJ)
    linha[24] = _converter_valor_causa(raw.get("Valor da Causa", ""))
    linha[25] = None                                          # Data Encerramento
    linha[26] = responsavel                                   # Responsável
    linha[27] = ESCRITORIO_REU                                # Escritório responsável
    linha[28] = ESCRITORIO_ORIGEM                             # Escritório Origem
    linha[29] = "bmagravo" if _eh_agravo(
        raw.get("Ação", ""), raw.get("Processo", "")
    ) else "bmcomum"                                          # Observações

    return linha


def _ler_linhas_listagem(listagem_bytes: bytes) -> list[dict[str, Any]]:
    """Le o XLSX da Listagem L1 e devolve linhas como dicts canonicos."""
    wb = openpyxl.load_workbook(
        io.BytesIO(listagem_bytes), read_only=True, data_only=True
    )
    if not wb.sheetnames:
        raise ValueError("Planilha sem abas.")
    ws = wb[wb.sheetnames[0]]

    header_row, idx_map = _localizar_header_e_indices(ws)

    linhas: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue
        # Linha completamente vazia? pula.
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        raw: dict[str, Any] = {}
        for canon, idx in idx_map.items():
            raw[canon] = row[idx] if idx < len(row) else ""
        # Considera "linha valida" se tem numero de processo nao-vazio
        if not raw.get("Processo"):
            continue
        linhas.append(raw)

    wb.close()
    return linhas


def gerar_planilha_l1(
    listagem_bytes: bytes, *, agora: Optional[datetime] = None
) -> bytes:
    """Converte o XLSX da Listagem L1 em XLSX no formato do MODELO LEGAL ONE.

    Levanta ValueError se a entrada nao for reconhecida como Listagem
    (cabecalho "Polo" ausente). Outros erros (XLSX corrompido, template
    ausente) sobem como excecao original.
    """
    if not TEMPLATE_PATH.exists():
        raise RuntimeError(
            f"Template MODELO_LEGAL_ONE.xlsx não encontrado em {TEMPLATE_PATH}"
        )

    agora = agora or datetime.now()
    data_extracao = agora.strftime("%d/%m/%Y %H:%M:%S")

    linhas_input = _ler_linhas_listagem(listagem_bytes)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Processos"]

    # Linha-modelo (estilos) na linha 2 do template
    linha_modelo = list(ws.iter_rows(min_row=2, max_row=2))[0]
    n_cols = len(linha_modelo)

    # Limpa dados anteriores (linha 3 em diante)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    # Distribuicao circular entre responsaveis (mesmo que so haja 1)
    fila = RESPONSAVEIS_REU.copy()

    linhas_processos: list[list[Any]] = []
    for chave, raw in enumerate(linhas_input, start=1):
        if not fila:
            fila = RESPONSAVEIS_REU.copy()
        responsavel = fila.pop(0) if fila else ""
        if responsavel:
            fila.append(responsavel)
        linhas_processos.append(
            _montar_linha(
                raw=raw,
                chave=chave,
                responsavel=responsavel,
                linha_modelo_len=n_cols,
                data_extracao=data_extracao,
            )
        )

    linhas_processos.sort(key=lambda x: int(x[0]))

    # Escreve linhas mantendo estilo do modelo
    for indice_linha, linha_valores in enumerate(linhas_processos, start=3):
        for col_index, cell_model in enumerate(linha_modelo, start=1):
            new_cell = ws.cell(row=indice_linha, column=col_index)
            valor = (
                linha_valores[col_index - 1]
                if col_index - 1 < len(linha_valores)
                else ""
            )
            if col_index == 18 and valor is None:
                valor = ""
            new_cell.value = valor
            if cell_model.has_style:
                new_cell._style = copy(cell_model._style)

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def nome_saida(agora: Optional[datetime] = None) -> str:
    """Nome padrao do arquivo gerado (mesmo formato do script original)."""
    agora = agora or datetime.now()
    return f"PLANILHA_MIGRACAO_COMPLETA - {agora.strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
