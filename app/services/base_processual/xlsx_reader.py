"""Leitor do XLSX de 'Listagem de Acoes Judiciais' do Legal One.

Header vem na LINHA 2 do export (linha 1 e' titulo/vazia). Lookup de
colunas e' por NOME NORMALIZADO (lower + sem acento + espacos colapsados)
pra tolerar pequenas mudancas no Reports do L1 sem nos quebrar.

Se faltar uma coluna OBRIGATORIA, levanta XlsxHeaderError com mensagem
clara apontando o nome. Se faltar uma coluna OPCIONAL, registra warning
e segue (campo fica None na row dict).
"""

from __future__ import annotations

import io
import unicodedata
from typing import Iterator, Optional

import openpyxl


def _normalize_header(s) -> str:
    """lower + sem acento + sem pontuacao + espacos colapsados."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    only_ascii = "".join(c for c in nfkd if not unicodedata.combining(c))
    # remove pontuacao comum dos headers do L1: '.', '/'
    cleaned = only_ascii.lower().replace(".", "").replace("/", " ")
    return " ".join(cleaned.split())


# Mapeamento canonico -> aliases aceitos (todos serao normalizados).
# Adicione novos aliases conforme o Reports do L1 mudar.
COLUMN_ALIASES: dict[str, list[str]] = {
    "cod_ajus": ["cod ajus", "codigo ajus", "id ajus"],
    "acao_principal": ["acao principal", "acao"],
    "materia": ["materia"],
    "risco_prob_perda": [
        "risco prob perda",
        "risco prob  perda",
        "risco",
    ],
    "autores_raw": ["autores - cnpjcpf", "autores", "autores cnpjcpf"],
    "reus_raw": ["reus - cnpjcpf", "reus", "rus - cnpjcpf", "rus", "reus cnpjcpf"],
    # 'Nº' normaliza via NFKD pra 'No' (ordinal indicator 'º' decompoe pra 'o'),
    # entao incluimos tanto a forma com 'no' quanto a sem (caso alguem use 'N°' ou 'N').
    "numero_processo_mascarado": [
        "numeros processo",
        "numero processo",
        "numeros do processo",
        "no processo",
        "n processo",
    ],
    "numero_interno": ["no interno", "n interno", "numero interno"],
    "tipo_acao": ["tipo de acao", "tipo acao"],
    "polo": ["polo"],
    "natureza": ["natureza"],
    "numero_vara": ["no vara", "n vara", "numero vara"],
    "foro": ["foro"],
    "comarca": ["comarca"],
    "uf": ["uf"],
    "empresa": ["empresa", "cliente"],
    "numero_pasta": ["no pasta", "n pasta", "numero pasta"],
    "grupo_responsavel": ["grupo responsavel"],
    "usuario_responsavel": ["usuario responsavel"],
    "escritorio_responsavel": ["escritorio responsavel"],
    "situacao_processo": [
        "situacao processo",
        "situacao do processo",
        "situacao",
    ],
    "justica_honorario": ["justica honorario", "justica  honorario"],
    "valor_causa": ["valor causa"],
    "valor_prev_acordo": [
        "valor prev acordo",
        "valor previsao acordo",
        "valor previsto acordo",
    ],
    "valor_acordo": ["valor acordo"],
    "valor_discutido": ["valor discutido"],
    "valor_exito": ["valor exito"],
    "valor_condenacao": ["valor condenacao"],
    "valor_contingencia": ["valor contingencia"],
    "ult_andamento": ["ult andamento", "ultimo andamento"],
    "data_ult_andamento": [
        "data ult andamento",
        "data do ultimo andamento",
    ],
    "dias_ult_atualizacao": [
        "dias ult atualizacao",
        "dias da ultima atualizacao",
    ],
    "distribuido_em": ["distribuido em"],
    "processo_virtual": ["processo virtual"],
    "numero_contrato": ["no contrato", "n contrato", "numero contrato"],
    "usuario_cadastro_acao": ["usuario cadastro acao"],
    "data_cadastro_acao": [
        "data hora cadastro acao",
        "data cadastro acao",
    ],
}


# Obrigatorias pra entry continuar — se faltar, FALHOU com mensagem.
REQUIRED_CANONICAL = [
    "cod_ajus",
    "numero_processo_mascarado",
    "empresa",
    "situacao_processo",
]


class XlsxHeaderError(Exception):
    """Header invalido — faltando coluna obrigatoria ou layout diferente."""


def read_xlsx_rows(xlsx_bytes: bytes) -> tuple[list[str], Iterator[dict]]:
    """Le o XLSX e devolve (warnings, iterator de rows).

    warnings: lista de canonical names que nao casaram com nenhuma coluna do
        header (informativo, nao bloqueia).
    rows: iterator de dicts canonical_name -> raw_value. Colunas ausentes
        no header ficam None na row.

    Header detectado em linha 1 ou linha 2 (export do L1 costuma ter titulo
    na linha 1). Falha com XlsxHeaderError se nenhuma combinacao tiver as
    colunas obrigatorias.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if not wb.sheetnames:
        raise XlsxHeaderError("Planilha sem sheets.")
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)

    # tenta as 2 primeiras linhas como possivel header
    candidates: list[tuple] = []
    try:
        candidates.append(next(rows_iter))
    except StopIteration:
        raise XlsxHeaderError("Planilha vazia.")
    try:
        candidates.append(next(rows_iter))
    except StopIteration:
        pass

    header_map: Optional[dict[str, int]] = None
    header_idx = -1
    for idx, row in enumerate(candidates):
        mapping = _try_match_header(row)
        # Header valido = ao menos 20 colunas canonicas casaram
        matches = sum(1 for v in mapping.values() if v is not None)
        if matches >= 20:
            header_map = mapping
            header_idx = idx
            break

    if header_map is None:
        raise XlsxHeaderError(
            "Nao foi possivel identificar o header da planilha. "
            "Esperado linha 1 ou 2 com colunas como 'Cod AJUS', 'Numeros "
            "Processo', 'Empresa', 'Situacao Processo'."
        )

    missing_required = [c for c in REQUIRED_CANONICAL if header_map.get(c) is None]
    if missing_required:
        raise XlsxHeaderError(
            f"Colunas obrigatorias ausentes no header: {missing_required}. "
            "Conferir export do Legal One Reports."
        )

    warnings = [c for c in COLUMN_ALIASES if header_map.get(c) is None]

    # Se header foi a linha 1 (idx=0), e candidates[1] existe, ele e' a primeira row de dados.
    # Se header foi a linha 2 (idx=1), candidates[0] era titulo/vazio (descarta).
    first_extra = None
    if header_idx == 0 and len(candidates) > 1:
        first_extra = candidates[1]
    return warnings, _rows_iterator(header_map, first_extra, rows_iter)


def _try_match_header(row: tuple) -> dict[str, Optional[int]]:
    """Casa cada celula do row com um canonical name. Retorna canonical -> col_idx."""
    if not row:
        return {c: None for c in COLUMN_ALIASES}
    normalized_cells = [_normalize_header(c) for c in row]
    result: dict[str, Optional[int]] = {c: None for c in COLUMN_ALIASES}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_norm = _normalize_header(alias)
            for idx, cell in enumerate(normalized_cells):
                if cell == alias_norm:
                    result[canonical] = idx
                    break
            if result[canonical] is not None:
                break
    return result


def _rows_iterator(
    header_map: dict[str, Optional[int]],
    first_extra,
    rows_iter,
) -> Iterator[dict]:
    """Junta opcional primeira-row + iterator restante, transformando em dicts."""
    if first_extra is not None:
        d = _row_to_dict(first_extra, header_map)
        if any(v is not None for v in d.values()):
            yield d
    for row in rows_iter:
        if row is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        d = _row_to_dict(row, header_map)
        yield d


def _row_to_dict(row: tuple, header_map: dict[str, Optional[int]]) -> dict:
    out: dict = {}
    for canonical, idx in header_map.items():
        if idx is None or idx >= len(row):
            out[canonical] = None
        else:
            out[canonical] = row[idx]
    return out
