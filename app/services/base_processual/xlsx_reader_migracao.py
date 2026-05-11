"""Parser leve do PLANILHA_MIGRACAO_COMPLETA — template de migracao do L1.

Schema diferente da Listagem_de_Acoes_Judiciais:
- Multi-sheet: 'Processos', 'Envolvidos', 'Orientacoes', 'Tabelas'
- Header na linha 1 (nao linha 2)
- Coluna chave: 'Chave do Processo *' (col 1) — sequencial 1, 2, 3 (NAO e' cod_ajus)
- 'Numero CNJ' (col 4) e' a chave natural utilizavel

Pro backfill historico, usamos esse parser SO' pra CONTAR processos no lote.
Nao extrai detalhes individuais — esses entram na base quando aparecem na
Listagem_de_Acoes_Judiciais (com cod_ajus real do L1).
"""

from __future__ import annotations

import io
import logging

import openpyxl

logger = logging.getLogger(__name__)


class MigracaoSchemaError(Exception):
    """Schema nao bate com PLANILHA_MIGRACAO_COMPLETA."""


def count_processos_no_lote(xlsx_bytes: bytes) -> int:
    """Conta quantos processos (linhas com Chave do Processo nao-vazia) ha no lote.

    Procura sheet 'Processos' — se nao existir, levanta MigracaoSchemaError.
    Pula a linha de header (linha 1).
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes), read_only=True, data_only=True
    )
    if "Processos" not in wb.sheetnames:
        raise MigracaoSchemaError(
            f"Sheet 'Processos' ausente. Sheets encontradas: {wb.sheetnames}. "
            "Nao parece PLANILHA_MIGRACAO_COMPLETA."
        )
    ws = wb["Processos"]
    count = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            # Header — valida que tem a coluna esperada
            if not row or not row[0] or str(row[0]).strip().lower() not in {
                "chave do processo *", "chave do processo"
            }:
                raise MigracaoSchemaError(
                    f"Header da sheet 'Processos' nao reconhecido. "
                    f"Esperado 'Chave do Processo *', achei: {row[0]!r}."
                )
            continue
        # Conta apenas linhas com Chave do Processo (col 1) preenchida
        if row and row[0] is not None and str(row[0]).strip():
            count += 1
    return count


def detect_format(xlsx_bytes: bytes) -> str:
    """Heuristica leve pra escolher entre 'listagem' e 'migracao'.

    Retorna 'migracao' se ha sheet 'Processos' + header tipico do template L1.
    Retorna 'listagem' caso contrario (default — pipeline normal).
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
    except Exception:
        return "listagem"
    if "Processos" in wb.sheetnames:
        # checa header
        ws = wb["Processos"]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if row1 and row1[0] and "chave do processo" in str(row1[0]).strip().lower():
            return "migracao"
    return "listagem"
