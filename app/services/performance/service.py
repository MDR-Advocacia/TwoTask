"""Serviço de métricas do "Minha Equipe".

Lê de perf_l1_tarefa / perf_pessoa / perf_subtipo_categoria e devolve:
- equipe(days, cargo)      — lista de pessoas com produção/ritmo/prazo/backlog + KPIs.
- pessoa_detalhe(id, days) — mix de tarefas por subtipo + ritmo/ócio operacional.
- tipos(days)              — mapa de impacto: volume/cycle/natureza por subtipo.
- cargos()                 — cargos distintos (filtro).

Filosofia (ver docs/performance-equipes-plano.md): cadência/ócio só são confiáveis
no segmento operacional (tarefas back-to-back); trabalho profundo mede-se por
volume + cycle time + cumprimento de prazo. Por isso o detalhe sempre informa a
"fatia operacional" — quanto do ritmo/ócio é confiável.
"""

import datetime
import statistics
from collections import defaultdict

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.performance import PerfPessoa, PerfTarefa

try:
    from zoneinfo import ZoneInfo

    BRT = ZoneInfo("America/Sao_Paulo")
except Exception:  # pragma: no cover
    BRT = None

# Gaps acima disso são pausa/intervalo, não "tempo gasto por tarefa".
_CAP_SEG = 30 * 60


class PerformanceService:
    def __init__(self, db: Session):
        self.db = db

    def _period(self, days: int):
        now = datetime.datetime.now(tz=BRT) if BRT else datetime.datetime.now()
        start = (now - datetime.timedelta(days=days)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        return start, now

    # ── equipe ────────────────────────────────────────────────────────────
    def equipe(self, days: int = 30, cargo: str | None = None, team: str | None = None) -> dict:
        start, now = self._period(days)
        params = {"start": start, "end": now}
        cargo_clause = ""
        if cargo:
            cargo_clause = "AND p.cargo = :cargo"
            params["cargo"] = cargo
        team_clause = ""
        if team:
            team_clause = "AND p.equipe = :team"
            params["team"] = team

        rows = self.db.execute(
            text(
                f"""
                SELECT p.id, p.nome, p.cargo, p.squad, p.posicao,
                  COUNT(t.id) AS concluido,
                  COUNT(DISTINCT date(t.concluido_em AT TIME ZONE 'America/Sao_Paulo')) AS dias_ativos,
                  COUNT(*) FILTER (WHERE t.prazo_previsto IS NOT NULL AND t.concluido_em <= t.prazo_previsto) AS no_prazo,
                  COUNT(*) FILTER (WHERE t.prazo_previsto IS NOT NULL) AS com_prazo,
                  percentile_cont(0.5) WITHIN GROUP (ORDER BY EXTRACT(EPOCH FROM (t.concluido_em - t.cadastrado_em)))
                    FILTER (WHERE t.cadastrado_em IS NOT NULL) AS cycle_seg,
                  COUNT(*) FILTER (WHERE c.categoria = 'operacional') AS oper_n,
                  COUNT(*) FILTER (WHERE c.categoria = 'profundo') AS prof_n,
                  COUNT(*) FILTER (WHERE c.categoria = 'ruido') AS ruido_n
                FROM perf_pessoa p
                LEFT JOIN perf_l1_tarefa t
                  ON t.pessoa_id = p.id AND t.status = 'Cumprido'
                  AND t.concluido_em >= :start AND t.concluido_em <= :end
                LEFT JOIN perf_subtipo_categoria c ON c.subtipo = t.subtipo
                WHERE p.ativo {team_clause} {cargo_clause}
                GROUP BY p.id, p.nome, p.cargo, p.squad, p.posicao
                ORDER BY concluido DESC, p.nome
                """
            ),
            params,
        ).fetchall()

        backlog = {
            pid: n
            for pid, n in self.db.execute(
                text("SELECT pessoa_id, COUNT(*) FROM perf_l1_tarefa WHERE status='Pendente' GROUP BY pessoa_id")
            ).fetchall()
        }

        pessoas = []
        for r in rows:
            concluido = r.concluido or 0
            dias = r.dias_ativos or 0
            pessoas.append(
                {
                    "id": r.id,
                    "nome": r.nome,
                    "cargo": r.cargo,
                    "squad": r.squad,
                    "posicao": r.posicao,
                    "concluido": concluido,
                    "dias_ativos": dias,
                    "throughput_dia": round(concluido / dias, 1) if dias else 0.0,
                    "no_prazo_pct": round(100.0 * r.no_prazo / r.com_prazo) if r.com_prazo else None,
                    "cycle_dias": round(r.cycle_seg / 86400.0, 1) if r.cycle_seg is not None else None,
                    "backlog": int(backlog.get(r.id, 0)),
                    "operacional_n": r.oper_n or 0,
                    "profundo_n": r.prof_n or 0,
                    "ruido_n": r.ruido_n or 0,
                }
            )

        total_conc = sum(p["concluido"] for p in pessoas)
        tot_no = sum((r.no_prazo or 0) for r in rows)
        tot_com = sum((r.com_prazo or 0) for r in rows)
        kpis = {
            "concluido": total_conc,
            "backlog": sum(p["backlog"] for p in pessoas),
            "pessoas_ativas": sum(1 for p in pessoas if p["concluido"] > 0),
            "pessoas_total": len(pessoas),
            "no_prazo_pct": round(100.0 * tot_no / tot_com) if tot_com else None,
        }
        return {"periodo_dias": days, "kpis": kpis, "pessoas": pessoas}

    def cargos(self, team: str | None = None) -> list:
        sql = "SELECT DISTINCT cargo FROM perf_pessoa WHERE cargo IS NOT NULL"
        params: dict = {}
        if team:
            sql += " AND equipe = :team"
            params["team"] = team
        return [r[0] for r in self.db.execute(text(sql + " ORDER BY cargo"), params).fetchall()]

    # ── detalhe da pessoa ─────────────────────────────────────────────────
    def pessoa_detalhe(self, pessoa_id: int, days: int = 30) -> dict | None:
        start, now = self._period(days)
        p = self.db.execute(
            text("SELECT id, nome, cargo, squad, posicao FROM perf_pessoa WHERE id = :id"),
            {"id": pessoa_id},
        ).fetchone()
        if not p:
            return None

        mix = self.db.execute(
            text(
                """
                SELECT t.subtipo AS subtipo, COALESCE(c.categoria, 'profundo') AS categoria,
                  COUNT(*) AS vol,
                  percentile_cont(0.5) WITHIN GROUP (ORDER BY EXTRACT(EPOCH FROM (t.concluido_em - t.cadastrado_em)))
                    FILTER (WHERE t.cadastrado_em IS NOT NULL) AS cycle_seg,
                  COUNT(*) FILTER (WHERE t.prazo_previsto IS NOT NULL AND t.concluido_em <= t.prazo_previsto) AS no_prazo,
                  COUNT(*) FILTER (WHERE t.prazo_previsto IS NOT NULL) AS com_prazo
                FROM perf_l1_tarefa t
                LEFT JOIN perf_subtipo_categoria c ON c.subtipo = t.subtipo
                WHERE t.pessoa_id = :id AND t.status = 'Cumprido'
                  AND t.concluido_em >= :start AND t.concluido_em <= :end
                GROUP BY t.subtipo, c.categoria
                ORDER BY vol DESC
                """
            ),
            {"id": pessoa_id, "start": start, "end": now},
        ).fetchall()

        ritmo, tempo_sub = self._ritmo_jornada(pessoa_id, start, now)
        mix_out = [
            {
                "subtipo": m.subtipo or "(sem subtipo)",
                "categoria": m.categoria,
                "volume": m.vol,
                "cycle_dias": round(m.cycle_seg / 86400.0, 1) if m.cycle_seg is not None else None,
                "no_prazo_pct": round(100.0 * m.no_prazo / m.com_prazo) if m.com_prazo else None,
                "tempo_tarefa_seg": tempo_sub.get(m.subtipo),
            }
            for m in mix
        ]

        # ── PASSADO: resumo (KPIs do que já foi feito) ──
        resumo = self.db.execute(
            text(
                """
                SELECT COUNT(*) AS concluido,
                  COUNT(DISTINCT date(concluido_em AT TIME ZONE 'America/Sao_Paulo')) AS dias,
                  COUNT(*) FILTER (WHERE prazo_previsto IS NOT NULL AND concluido_em <= prazo_previsto) AS no_prazo,
                  COUNT(*) FILTER (WHERE prazo_previsto IS NOT NULL) AS com_prazo,
                  percentile_cont(0.5) WITHIN GROUP (ORDER BY EXTRACT(EPOCH FROM (concluido_em - cadastrado_em)))
                    FILTER (WHERE cadastrado_em IS NOT NULL) AS cycle_seg
                FROM perf_l1_tarefa
                WHERE pessoa_id = :id AND status = 'Cumprido'
                  AND concluido_em >= :start AND concluido_em <= :end
                """
            ),
            {"id": pessoa_id, "start": start, "end": now},
        ).fetchone()
        concluido = resumo.concluido or 0
        dias = resumo.dias or 0
        passado_kpis = {
            "concluido": concluido,
            "dias_ativos": dias,
            "throughput_dia": round(concluido / dias, 1) if dias else 0.0,
            "no_prazo_pct": round(100 * resumo.no_prazo / resumo.com_prazo) if resumo.com_prazo else None,
            "cycle_dias": round(resumo.cycle_seg / 86400.0, 1) if resumo.cycle_seg is not None else None,
        }

        # ── FUTURO: carga aberta (pendentes/atrasadas), por tipo e próximos prazos ──
        pend = self.db.execute(
            text(
                """
                SELECT COUNT(*) AS total,
                  COUNT(*) FILTER (WHERE prazo_previsto < now()) AS atrasado,
                  COUNT(*) FILTER (WHERE prazo_previsto IS NULL) AS sem_prazo
                FROM perf_l1_tarefa WHERE pessoa_id = :id AND status = 'Pendente'
                """
            ),
            {"id": pessoa_id},
        ).fetchone()

        pend_tipo = self.db.execute(
            text(
                """
                SELECT t.subtipo AS subtipo, COALESCE(c.categoria, 'profundo') AS categoria,
                  COUNT(*) AS total, COUNT(*) FILTER (WHERE t.prazo_previsto < now()) AS atrasado
                FROM perf_l1_tarefa t LEFT JOIN perf_subtipo_categoria c ON c.subtipo = t.subtipo
                WHERE t.pessoa_id = :id AND t.status = 'Pendente'
                GROUP BY t.subtipo, c.categoria
                ORDER BY total DESC
                """
            ),
            {"id": pessoa_id},
        ).fetchall()

        urgentes_rows = self.db.execute(
            text(
                """
                SELECT subtipo, prazo_previsto, cnj, pasta
                FROM perf_l1_tarefa
                WHERE pessoa_id = :id AND status = 'Pendente' AND prazo_previsto IS NOT NULL
                ORDER BY prazo_previsto ASC LIMIT 20
                """
            ),
            {"id": pessoa_id},
        ).fetchall()
        urgentes = []
        for u in urgentes_rows:
            pr = u.prazo_previsto
            dias_prazo = (pr - now).days if pr else None
            urgentes.append(
                {
                    "subtipo": u.subtipo or "(sem subtipo)",
                    "prazo": (pr.astimezone(BRT) if BRT else pr).strftime("%d/%m/%Y") if pr else None,
                    "dias": dias_prazo,
                    "atrasado": dias_prazo is not None and dias_prazo < 0,
                    "cnj": u.cnj,
                    "pasta": u.pasta,
                }
            )

        return {
            "pessoa": {"id": p.id, "nome": p.nome, "cargo": p.cargo, "squad": p.squad, "posicao": p.posicao},
            "periodo_dias": days,
            "passado": {
                "kpis": passado_kpis,
                "ritmo": ritmo,
                "mix": mix_out,
            },
            "futuro": {
                "pendente": pend.total or 0,
                "atrasado": pend.atrasado or 0,
                "sem_prazo": pend.sem_prazo or 0,
                "por_tipo": [
                    {"subtipo": r.subtipo or "(sem subtipo)", "categoria": r.categoria, "total": r.total, "atrasado": r.atrasado}
                    for r in pend_tipo
                ],
                "urgentes": urgentes,
            },
        }

    def _ritmo_jornada(self, pessoa_id: int, start, end):
        """Ritmo (tempo entre entregas), ócio, JORNADA (horário típico de chegada/
        saída) e o TEMPO DE DECISÃO por subtipo (quanto a pessoa leva, em média,
        pra concluir uma tarefa de cada tipo). Tudo sobre as conclusões do período;
        informa a fatia operacional (quanto maior, mais confiável a leitura).

        Devolve (ritmo: dict, tempo_por_subtipo: {subtipo: segundos}).
        """
        rows = self.db.execute(
            text(
                """
                SELECT (t.concluido_em AT TIME ZONE 'America/Sao_Paulo') AS c,
                       COALESCE(cat.categoria, 'profundo') AS categoria,
                       t.subtipo AS subtipo
                FROM perf_l1_tarefa t
                LEFT JOIN perf_subtipo_categoria cat ON cat.subtipo = t.subtipo
                WHERE t.pessoa_id = :id AND t.status = 'Cumprido'
                  AND t.concluido_em >= :start AND t.concluido_em <= :end
                ORDER BY t.concluido_em
                """
            ),
            {"id": pessoa_id, "start": start, "end": end},
        ).fetchall()

        items = [(r.c, r.categoria, r.subtipo) for r in rows if r.c]
        total = len(items)
        vazio = {
            "volume": total, "cadencia_seg": None, "ocio_pct": None, "dias": 0,
            "oper_share": None, "inicio_h": None, "fim_h": None,
        }
        if total < 2:
            return vazio, {}

        oper = sum(1 for _, cat, _ in items if cat == "operacional")
        by_day = defaultdict(list)
        for ts, _cat, sub in items:
            by_day[ts.date()].append((ts, sub))

        gaps, hands_on, window = [], 0.0, 0.0
        starts, ends = [], []
        tempo_sub = defaultdict(list)
        for _d, lst in by_day.items():
            lst.sort(key=lambda x: x[0])
            starts.append(lst[0][0].hour + lst[0][0].minute / 60.0)
            ends.append(lst[-1][0].hour + lst[-1][0].minute / 60.0)
            for i in range(1, len(lst)):
                g = (lst[i][0] - lst[i - 1][0]).total_seconds()
                if 0 < g <= _CAP_SEG:
                    gaps.append(g)
                    hands_on += g
                    if lst[i][1]:
                        tempo_sub[lst[i][1]].append(g)
            if len(lst) >= 2:
                window += (lst[-1][0] - lst[0][0]).total_seconds()

        ritmo = {
            "volume": total,
            "cadencia_seg": round(statistics.median(gaps)) if gaps else None,
            "ocio_pct": round(100.0 * (1 - hands_on / window)) if window > 0 else None,
            "dias": len(by_day),
            "oper_share": round(100.0 * oper / total),
            "inicio_h": round(statistics.median(starts), 2) if starts else None,
            "fim_h": round(statistics.median(ends), 2) if ends else None,
        }
        tempo_por_sub = {sub: round(statistics.median(v)) for sub, v in tempo_sub.items() if v}
        return ritmo, tempo_por_sub

    # ── mapa de impacto por tipo ──────────────────────────────────────────
    def tipos(self, days: int = 30, team: str | None = None) -> list:
        start, now = self._period(days)
        params = {"start": start, "end": now}
        team_join = ""
        if team:
            team_join = "JOIN perf_pessoa pp ON pp.id = t.pessoa_id AND pp.equipe = :team"
            params["team"] = team
        rows = self.db.execute(
            text(
                f"""
                SELECT t.subtipo AS subtipo, COALESCE(c.categoria, 'profundo') AS categoria,
                  COUNT(*) AS vol, COUNT(DISTINCT t.pessoa_id) AS pessoas,
                  percentile_cont(0.5) WITHIN GROUP (ORDER BY EXTRACT(EPOCH FROM (t.concluido_em - t.cadastrado_em)))
                    FILTER (WHERE t.cadastrado_em IS NOT NULL) AS cycle_seg,
                  c.densidade AS densidade
                FROM perf_l1_tarefa t
                {team_join}
                LEFT JOIN perf_subtipo_categoria c ON c.subtipo = t.subtipo
                WHERE t.status = 'Cumprido' AND t.concluido_em >= :start AND t.concluido_em <= :end
                  AND t.subtipo IS NOT NULL
                GROUP BY t.subtipo, c.categoria, c.densidade
                ORDER BY vol DESC
                """
            ),
            params,
        ).fetchall()
        return [
            {
                "subtipo": r.subtipo,
                "categoria": r.categoria,
                "volume": r.vol,
                "pessoas": r.pessoas,
                "cycle_dias": round(r.cycle_seg / 86400.0, 1) if r.cycle_seg is not None else None,
                "densidade": r.densidade,
            }
            for r in rows
        ]

    # ── painel do setor (dados pros gráficos) ─────────────────────────────
    def dashboard(self, days: int = 30, team: str | None = None) -> dict:
        """Agrega o setor pros gráficos: vazão, pool pendente/atrasado, jornada
        (início/fim do dia + hands-on), e top tipos. Tudo por pessoa."""
        start, now = self._period(days)
        tparam = {"team": team} if team else {}
        team_where = "AND equipe = :team" if team else ""
        team_t = "AND t.pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)" if team else ""
        team_nb = "AND pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)" if team else ""
        pessoas = {
            r.id: r
            for r in self.db.execute(
                text(f"SELECT id, nome, cargo, squad FROM perf_pessoa WHERE ativo {team_where}"),
                tparam,
            ).fetchall()
        }
        completions = self.db.execute(
            text(
                f"""
                SELECT t.pessoa_id AS pid,
                       (t.concluido_em AT TIME ZONE 'America/Sao_Paulo') AS c,
                       COALESCE(cat.categoria, 'profundo') AS categoria
                FROM perf_l1_tarefa t
                LEFT JOIN perf_subtipo_categoria cat ON cat.subtipo = t.subtipo
                WHERE t.status = 'Cumprido' AND t.pessoa_id IS NOT NULL
                  AND t.concluido_em >= :start AND t.concluido_em <= :end {team_t}
                ORDER BY t.pessoa_id, t.concluido_em
                """
            ),
            {"start": start, "end": now, **tparam},
        ).fetchall()

        per = defaultdict(lambda: {"times": [], "oper": 0, "total": 0})
        for r in completions:
            if not r.c:
                continue
            d = per[r.pid]
            d["times"].append(r.c)
            d["total"] += 1
            if r.categoria == "operacional":
                d["oper"] += 1

        backlog_map = {
            pid: (pend, atr)
            for pid, pend, atr in self.db.execute(
                text(
                    f"""
                    SELECT pessoa_id, COUNT(*) AS pend,
                           COUNT(*) FILTER (WHERE prazo_previsto < now()) AS atr
                    FROM perf_l1_tarefa
                    WHERE status = 'Pendente' AND pessoa_id IS NOT NULL {team_nb}
                    GROUP BY pessoa_id
                    """
                ),
                tparam,
            ).fetchall()
        }

        vazao, backlog, jornada = [], [], []
        for pid, p in pessoas.items():
            pend, atr = backlog_map.get(pid, (0, 0))
            backlog.append(
                {"id": pid, "nome": p.nome, "cargo": p.cargo, "backlog": int(pend), "atrasado": int(atr)}
            )
            d = per.get(pid)
            if not d or d["total"] == 0:
                continue
            by_day = defaultdict(list)
            for t in d["times"]:
                by_day[t.date()].append(t)
            starts, ends, hands = [], [], []
            window_total, hands_total = 0.0, 0.0
            for _, ts in by_day.items():
                ts.sort()
                starts.append(ts[0].hour + ts[0].minute / 60.0)
                ends.append(ts[-1].hour + ts[-1].minute / 60.0)
                ho = 0.0
                for i in range(1, len(ts)):
                    g = (ts[i] - ts[i - 1]).total_seconds()
                    if 0 < g <= _CAP_SEG:
                        ho += g
                hands.append(ho / 3600.0)
                if len(ts) >= 2:
                    window_total += (ts[-1] - ts[0]).total_seconds()
                    hands_total += ho
            dias = len(by_day)
            concluido = d["total"]
            vazao.append(
                {
                    "id": pid, "nome": p.nome, "cargo": p.cargo, "concluido": concluido,
                    "throughput_dia": round(concluido / dias, 1) if dias else 0.0,
                }
            )
            jornada.append(
                {
                    "id": pid, "nome": p.nome, "cargo": p.cargo,
                    "inicio_h": round(statistics.median(starts), 2),
                    "fim_h": round(statistics.median(ends), 2),
                    "hands_on_h": round(statistics.median(hands), 1),
                    "ocio_pct": round(100 * (1 - hands_total / window_total)) if window_total > 0 else None,
                    "dias": dias,
                    "oper_share": round(100 * d["oper"] / concluido) if concluido else 0,
                }
            )

        vazao.sort(key=lambda x: -x["concluido"])
        backlog.sort(key=lambda x: (-x["atrasado"], -x["backlog"]))
        jornada.sort(key=lambda x: x["inicio_h"])

        rows = self.db.execute(
            text(
                f"""
                SELECT t.subtipo AS subtipo, COALESCE(c.categoria, 'profundo') AS categoria,
                  COUNT(*) FILTER (
                    WHERE t.status = 'Cumprido' AND t.concluido_em >= :start AND t.concluido_em <= :end
                  ) AS vol,
                  COUNT(*) FILTER (WHERE t.status = 'Pendente') AS pendente,
                  COUNT(*) FILTER (WHERE t.status = 'Pendente' AND t.prazo_previsto < now()) AS atrasado
                FROM perf_l1_tarefa t
                LEFT JOIN perf_subtipo_categoria c ON c.subtipo = t.subtipo
                WHERE t.subtipo IS NOT NULL {team_t}
                GROUP BY t.subtipo, c.categoria
                """
            ),
            {"start": start, "end": now, **tparam},
        ).fetchall()
        by_sub = {
            r.subtipo: {
                "subtipo": r.subtipo,
                "categoria": r.categoria,
                "volume": r.vol,
                "pendente": r.pendente,
                "atrasado": r.atrasado,
            }
            for r in rows
        }
        # Board curado por time (perf_board_tarefa); sem curadoria → top-12 por volume.
        board_cfg: list = []
        if team:
            board_cfg = [
                row.subtipo
                for row in self.db.execute(
                    text("SELECT subtipo FROM perf_board_tarefa WHERE team = :team ORDER BY ordem, id"),
                    {"team": team},
                ).fetchall()
            ]
        if board_cfg:
            top_list = [
                by_sub.get(s, {"subtipo": s, "categoria": "profundo", "volume": 0, "pendente": 0, "atrasado": 0})
                for s in board_cfg
            ]
        else:
            top_list = sorted(by_sub.values(), key=lambda d: -d["volume"])[:12]

        return {
            "periodo_dias": days,
            "kpis": {
                "atrasado_total": sum(b["atrasado"] for b in backlog),
                "backlog_total": sum(b["backlog"] for b in backlog),
            },
            "vazao": vazao,
            "backlog": backlog,
            "jornada": jornada,
            "top_tipos": top_list,
            "board_curado": bool(board_cfg),
        }

    # ── detalhe de UM subtipo do board (capacity) ──
    def subtipo_setor_detalhe(self, subtipo: str, days: int = 30, team: str | None = None) -> dict:
        """Detalhe de um subtipo pro (i) do board 'Tarefas mais importantes':
        contagens (concluído no período / pendente / atrasado), % no prazo,
        quantas pessoas o executam, e os DOIS tempos que medem capacity:

          • tempo_conclusao_seg — mediana de (conclusão − cadastro): a LATÊNCIA da
            tarefa, quanto ela leva do cadastro até ser concluída;
          • tempo_trabalho_seg — mediana do gap entre conclusões consecutivas do
            tipo (por pessoa/dia, capado em 30min): o ESFORÇO efetivo por tarefa,
            o número que entra na conta de quantas cabem num dia de trabalho.
        """
        import statistics
        from collections import defaultdict

        start, now = self._period(days)
        team_t = "AND t.pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)" if team else ""
        tp: dict = {"sub": subtipo, "start": start, "end": now}
        if team:
            tp["team"] = team

        r = self.db.execute(
            text(
                f"""
                SELECT
                  COUNT(*) FILTER (WHERE t.status='Cumprido' AND t.concluido_em BETWEEN :start AND :end) AS concluido,
                  COUNT(*) FILTER (WHERE t.status='Pendente') AS pendente,
                  COUNT(*) FILTER (WHERE t.status='Pendente' AND t.prazo_previsto < now()) AS atrasado,
                  COUNT(*) FILTER (WHERE t.status='Cumprido' AND t.prazo_previsto IS NOT NULL
                                   AND t.concluido_em <= t.prazo_previsto
                                   AND t.concluido_em BETWEEN :start AND :end) AS no_prazo,
                  COUNT(*) FILTER (WHERE t.status='Cumprido' AND t.prazo_previsto IS NOT NULL
                                   AND t.concluido_em BETWEEN :start AND :end) AS com_prazo,
                  COUNT(DISTINCT t.pessoa_id) FILTER (WHERE t.status='Cumprido'
                                   AND t.concluido_em BETWEEN :start AND :end) AS pessoas,
                  percentile_cont(0.5) WITHIN GROUP (
                      ORDER BY EXTRACT(EPOCH FROM (t.concluido_em - t.cadastrado_em))
                  ) FILTER (WHERE t.status='Cumprido' AND t.cadastrado_em IS NOT NULL
                            AND t.concluido_em >= t.cadastrado_em
                            AND t.concluido_em BETWEEN :start AND :end) AS cycle_seg
                FROM perf_l1_tarefa t
                WHERE t.subtipo = :sub {team_t}
                """
            ),
            tp,
        ).fetchone()

        comps = self.db.execute(
            text(
                f"""
                SELECT t.pessoa_id AS pid, (t.concluido_em AT TIME ZONE 'America/Sao_Paulo') AS c
                FROM perf_l1_tarefa t
                WHERE t.subtipo = :sub AND t.status='Cumprido'
                  AND t.concluido_em BETWEEN :start AND :end {team_t}
                ORDER BY t.pessoa_id, t.concluido_em
                """
            ),
            tp,
        ).fetchall()
        by_pd: dict = defaultdict(list)
        for x in comps:
            if x.c:
                by_pd[(x.pid, x.c.date())].append(x.c)
        gaps: list = []
        for lst in by_pd.values():
            lst.sort()
            for i in range(1, len(lst)):
                g = (lst[i] - lst[i - 1]).total_seconds()
                if 0 < g <= _CAP_SEG:
                    gaps.append(g)

        cat = self.db.execute(
            text("SELECT categoria FROM perf_subtipo_categoria WHERE subtipo = :sub"),
            {"sub": subtipo},
        ).scalar()

        return {
            "subtipo": subtipo,
            "categoria": cat or "profundo",
            "concluido": r.concluido or 0,
            "pendente": r.pendente or 0,
            "atrasado": r.atrasado or 0,
            "pessoas": r.pessoas or 0,
            "no_prazo_pct": round(100.0 * r.no_prazo / r.com_prazo) if r.com_prazo else None,
            "tempo_conclusao_seg": round(r.cycle_seg) if r.cycle_seg is not None else None,
            "tempo_trabalho_seg": round(statistics.median(gaps)) if gaps else None,
            "amostra_trabalho": len(gaps),
            "periodo_dias": days,
        }

    # ── duplicadas: mesma pasta + mesmo subtipo (preview) ──
    def duplicadas_subtipo(self, subtipo: str, team: str | None = None) -> dict:
        """Pastas com MAIS DE UMA tarefa PENDENTE do mesmo subtipo — as duplicadas
        geradas pelo desvio de fluxo. Mantém a MAIS ANTIGA (original) e marca as
        criadas depois como canceláveis. Preview lido do snapshot; a cancelação
        real (fase B) reverifica cada tarefa ao vivo no L1 (pré-check de terminal).
        """
        from collections import defaultdict

        team_t = "AND t.pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)" if team else ""
        team_nb = "AND pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)" if team else ""
        tp: dict = {"sub": subtipo}
        if team:
            tp["team"] = team

        rows = self.db.execute(
            text(
                f"""
                SELECT t.pasta AS pasta, t.l1_task_id AS tid, t.cnj AS cnj,
                       t.cadastrado_em AS cad, t.prazo_previsto AS prazo, p.nome AS pessoa
                FROM perf_l1_tarefa t
                LEFT JOIN perf_pessoa p ON p.id = t.pessoa_id
                WHERE t.subtipo = :sub AND t.status = 'Pendente'
                  AND t.pasta IS NOT NULL AND t.l1_task_id IS NOT NULL {team_t}
                  AND t.pasta IN (
                    SELECT pasta FROM perf_l1_tarefa
                    WHERE subtipo = :sub AND status = 'Pendente' AND pasta IS NOT NULL
                      AND l1_task_id IS NOT NULL {team_nb}
                    GROUP BY pasta HAVING COUNT(*) > 1
                  )
                ORDER BY t.pasta, t.cadastrado_em ASC NULLS FIRST, t.l1_task_id ASC
                """
            ),
            tp,
        ).fetchall()

        grupos_map: dict = defaultdict(list)
        for r in rows:
            grupos_map[r.pasta].append(r)

        grupos: list = []
        total_cancelar = 0
        for pasta, lst in grupos_map.items():
            manter = lst[0]  # mais antiga = original
            cancelar = lst[1:]
            total_cancelar += len(cancelar)
            grupos.append(
                {
                    "pasta": pasta,
                    "cnj": manter.cnj,
                    "manter": {
                        "task_id": manter.tid,
                        "cadastrado_em": manter.cad,
                        "pessoa": manter.pessoa,
                        "prazo": manter.prazo,
                    },
                    "cancelar": [
                        {"task_id": x.tid, "cadastrado_em": x.cad, "pessoa": x.pessoa, "prazo": x.prazo}
                        for x in cancelar
                    ],
                }
            )
        grupos.sort(key=lambda g: -len(g["cancelar"]))
        return {
            "subtipo": subtipo,
            "total_grupos": len(grupos),
            "total_cancelar": total_cancelar,
            "grupos": grupos,
        }

    # ── curadoria do board 'Tarefas mais importantes' ──
    def board_listar(self, team: str) -> dict:
        rows = self.db.execute(
            text("SELECT subtipo FROM perf_board_tarefa WHERE team = :team ORDER BY ordem, id"),
            {"team": team},
        ).fetchall()
        return {"curado": bool(rows), "subtipos": [r.subtipo for r in rows]}

    def board_catalogo(self, team: str, busca: str = "") -> list:
        """Subtipos que o time tem (pra adicionar ao board), com volume, já sem os
        curados. Ordena por volume desc."""
        ja = {
            r.subtipo
            for r in self.db.execute(
                text("SELECT subtipo FROM perf_board_tarefa WHERE team = :team"), {"team": team}
            ).fetchall()
        }
        params: dict = {"team": team}
        like = ""
        if busca.strip():
            like = "AND t.subtipo ILIKE :busca"
            params["busca"] = f"%{busca.strip()}%"
        rows = self.db.execute(
            text(
                f"""
                SELECT t.subtipo AS subtipo, COUNT(*) AS vol
                FROM perf_l1_tarefa t
                WHERE t.subtipo IS NOT NULL {like}
                  AND t.pessoa_id IN (SELECT id FROM perf_pessoa WHERE equipe = :team)
                GROUP BY t.subtipo
                ORDER BY vol DESC
                LIMIT 100
                """
            ),
            params,
        ).fetchall()
        return [{"subtipo": r.subtipo, "volume": r.vol} for r in rows if r.subtipo not in ja]

    def board_adicionar(self, team: str, subtipo: str) -> dict:
        existe = self.db.execute(
            text("SELECT 1 FROM perf_board_tarefa WHERE team = :team AND subtipo = :sub"),
            {"team": team, "sub": subtipo},
        ).scalar()
        if not existe:
            nxt = self.db.execute(
                text("SELECT COALESCE(MAX(ordem), 0) + 1 FROM perf_board_tarefa WHERE team = :team"),
                {"team": team},
            ).scalar()
            self.db.execute(
                text("INSERT INTO perf_board_tarefa (team, subtipo, ordem) VALUES (:team, :sub, :ord)"),
                {"team": team, "sub": subtipo, "ord": nxt or 1},
            )
            self.db.commit()
        return self.board_listar(team)

    def board_remover(self, team: str, subtipo: str) -> dict:
        self.db.execute(
            text("DELETE FROM perf_board_tarefa WHERE team = :team AND subtipo = :sub"),
            {"team": team, "sub": subtipo},
        )
        self.db.commit()
        return self.board_listar(team)

    # ── export Excel de um recorte (tarefas que o operador precisa tratar) ──
    def export_xlsx(self, escopo: str = "atrasado", pessoa_id: int | None = None,
                    subtipo: str | None = None, days: int = 30, team: str | None = None) -> bytes:
        """Gera xlsx com as tarefas de um recorte. escopo: atrasado|pendente|concluido."""
        from io import BytesIO

        import openpyxl
        from openpyxl.utils import get_column_letter

        start, now = self._period(days)
        conds: list = []
        params: dict = {}
        if escopo == "concluido":
            conds.append("t.status = 'Cumprido' AND t.concluido_em >= :start AND t.concluido_em <= :end")
            params["start"] = start
            params["end"] = now
        elif escopo == "pendente":
            conds.append("t.status = 'Pendente'")
        else:  # atrasado
            conds.append("t.status = 'Pendente' AND t.prazo_previsto < now()")
        if pessoa_id:
            conds.append("t.pessoa_id = :pid")
            params["pid"] = pessoa_id
        if subtipo:
            conds.append("t.subtipo = :sub")
            params["sub"] = subtipo
        if team:
            conds.append("p.equipe = :team")
            params["team"] = team
        where = " AND ".join(conds)

        rows = self.db.execute(
            text(
                f"""
                SELECT p.nome AS responsavel, p.cargo AS cargo, t.l1_task_id AS l1_id,
                  t.pasta, t.cnj, t.uf, t.subtipo, t.status,
                  t.cadastrado_em, t.concluido_em, t.prazo_previsto
                FROM perf_l1_tarefa t
                LEFT JOIN perf_pessoa p ON p.id = t.pessoa_id
                WHERE {where}
                ORDER BY t.prazo_previsto ASC NULLS LAST, p.nome
                """
            ),
            params,
        ).fetchall()

        def _dt(dt, only_date=False):
            if not dt:
                return None
            d = dt.astimezone(BRT) if BRT else dt
            return d.strftime("%d/%m/%Y") if only_date else d.strftime("%d/%m/%Y %H:%M")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = {"atrasado": "Atrasadas", "pendente": "Pendentes", "concluido": "Concluidas"}.get(escopo, "Tarefas")[:31]
        headers = ["Responsável", "Cargo", "ID L1", "Pasta", "CNJ", "UF", "Subtipo", "Status",
                   "Cadastro", "Conclusão", "Prazo", "Situação"]
        ws.append(headers)
        for r in rows:
            sit = ""
            if r.prazo_previsto:
                dd = (r.prazo_previsto - now).days
                sit = f"atrasada há {abs(dd)}d" if dd < 0 else ("vence hoje" if dd == 0 else f"vence em {dd}d")
            ws.append([
                r.responsavel, r.cargo, r.l1_id, r.pasta, r.cnj, r.uf, r.subtipo, r.status,
                _dt(r.cadastrado_em), _dt(r.concluido_em), _dt(r.prazo_previsto, only_date=True), sit,
            ])
        for i, w in enumerate([26, 14, 9, 14, 24, 6, 36, 10, 17, 17, 12, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── manutenção do roster (editor de equipe) ───────────────────────────
    def roster(self, team: str) -> list:
        """Pessoas do time (visão editável), com contagem de tarefas pra referência."""
        rows = self.db.execute(
            text(
                """
                SELECT p.id, p.nome, p.cargo, p.equipe, p.is_supervisor, p.ativo, p.squad, p.posicao,
                  (SELECT COUNT(*) FROM perf_l1_tarefa t WHERE t.pessoa_id = p.id AND t.status = 'Cumprido') AS concluido,
                  (SELECT COUNT(*) FROM perf_l1_tarefa t WHERE t.pessoa_id = p.id AND t.status = 'Pendente') AS pendente
                FROM perf_pessoa p
                WHERE p.equipe = :team
                ORDER BY
                  p.is_supervisor DESC,
                  CASE
                    WHEN p.cargo ILIKE '%superv%' THEN 0
                    WHEN p.cargo ILIKE '%advog%' THEN 1
                    WHEN p.cargo ILIKE '%assist%' THEN 2
                    WHEN p.cargo ILIKE '%estag%' THEN 3
                    ELSE 4
                  END,
                  p.nome
                """
            ),
            {"team": team},
        ).fetchall()
        return [
            {
                "id": r.id, "nome": r.nome, "cargo": r.cargo, "equipe": r.equipe,
                "is_supervisor": r.is_supervisor, "ativo": r.ativo, "squad": r.squad,
                "posicao": r.posicao, "concluido": r.concluido, "pendente": r.pendente,
            }
            for r in rows
        ]

    def update_pessoa(self, pessoa_id: int, *, cargo=None, equipe=None,
                      is_supervisor=None, ativo=None):
        p = self.db.query(PerfPessoa).filter(PerfPessoa.id == pessoa_id).first()
        if not p:
            return None
        if cargo is not None:
            p.cargo = (cargo or "").strip() or None
        if equipe is not None:
            p.equipe = equipe
        if is_supervisor is not None:
            p.is_supervisor = bool(is_supervisor)
        if ativo is not None:
            p.ativo = bool(ativo)
        self.db.commit()
        return {
            "id": p.id, "nome": p.nome, "cargo": p.cargo, "equipe": p.equipe,
            "is_supervisor": p.is_supervisor, "ativo": p.ativo,
        }

    def candidatos(self, team: str, busca: str | None = None, limit: int = 80) -> list:
        """Pessoas do catálogo de usuários do L1 que NÃO estão neste time (pra adicionar)."""
        from app.models.legal_one import LegalOneUser
        from app.services.performance.seed import norm as _norm

        pessoa_eq = {p.nome_norm: p.equipe for p in self.db.query(PerfPessoa).all()}
        bnorm = _norm(busca) if busca else None
        out = []
        for (nome,) in self.db.query(LegalOneUser.name).filter(LegalOneUser.is_active.is_(True)).all():
            if not nome:
                continue
            nm = _norm(nome)
            if pessoa_eq.get(nm) == team:
                continue
            if bnorm and bnorm not in nm:
                continue
            out.append({"nome": nome, "equipe_atual": pessoa_eq.get(nm)})
            if len(out) >= limit:
                break
        out.sort(key=lambda x: x["nome"])
        return out

    def adicionar_pessoa(self, nome: str, team: str):
        from app.services.performance.seed import norm as _norm

        nm = _norm(nome)
        if not nm:
            return None
        p = self.db.query(PerfPessoa).filter(PerfPessoa.nome_norm == nm).first()
        if p is None:
            p = PerfPessoa(nome_norm=nm, nome=nome.strip())
            self.db.add(p)
        p.nome = p.nome or nome.strip()
        p.equipe = team
        p.ativo = True
        self.db.commit()
        return {"id": p.id, "nome": p.nome, "equipe": p.equipe}

    def excluir_pessoa(self, pessoa_id: int):
        """Exclui a pessoa do sistema (saiu do escritório). Desvincula as tarefas
        dela (ficam sem responsável) pra não violar a FK, e remove o registro."""
        p = self.db.query(PerfPessoa).filter(PerfPessoa.id == pessoa_id).first()
        if not p:
            return None
        self.db.query(PerfTarefa).filter(PerfTarefa.pessoa_id == pessoa_id).update(
            {PerfTarefa.pessoa_id: None}, synchronize_session=False
        )
        self.db.delete(p)
        self.db.commit()
        return {"id": pessoa_id, "excluido": True}
