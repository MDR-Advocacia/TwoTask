"""Seed do "Minha Equipe" a partir dos exports do L1 (histórico).

Lê dois arquivos no container:
  /tmp/squads.xlsx  — planilha "DIVISÃO DAS EQUIPES - SQUADS" (roster).
  /tmp/agenda.xlsx  — export "Agenda Analytics" (tarefas).

Popula perf_pessoa (roster), perf_l1_tarefa (só as tarefas das pessoas do
roster — resto é ruído) e perf_subtipo_categoria (natureza por subtipo).

É o "seed do histórico via export" do plano — a ingestão incremental via API
/Tasks entra numa fase seguinte. Idempotente: re-rodar limpa tarefas/categorias
e regrava; pessoas são upsertadas por nome normalizado.

Rodar:  docker exec onetask-api-1 python -m app.services.performance.seed
"""

import datetime
import unicodedata

import openpyxl
from sqlalchemy import text

from app.db.session import SessionLocal
from app.models.performance import (
    CAT_OPERACIONAL,
    CAT_PROFUNDO,
    CAT_RUIDO,
    PerfPessoa,
    PerfSubtipoCategoria,
    PerfTarefa,
)

try:
    from zoneinfo import ZoneInfo

    BRT = ZoneInfo("America/Sao_Paulo")
except Exception:  # pragma: no cover
    BRT = None

SQUADS_XLSX = "/tmp/squads.xlsx"
AGENDA_XLSX = "/tmp/agenda.xlsx"

# Índices de coluna do export "Agenda Analytics" (formato fixo).
ESC, ENV, TIPO, STATUS, CONCL, PRAZO, CUMPRIU, CAD, SUBTIPO = 1, 2, 4, 6, 7, 9, 14, 16, 17
ID, PASTA, CNJ, UF = 3, 10, 11, 12

_SHARED_POS = {"acordo", "encerramento"}

# Overrides manuais de cargo: a planilha traz o cargo FORMAL, mas a função REAL
# pode diferir. Ex.: Cinthia Samylle é "Assistente" formal, mas lidera squad como
# Advogado Responsável e atua como advogada — entra na análise como advogada.
# (No futuro isso vira edição no admin; por ora fica versionado aqui.)
_CARGO_OVERRIDE = {
    "cinthia samylle martins souza da silva": "Advogado(a)",
}

# Aba (sheet) da planilha de squads → setor (supervisão). Chave = nome da aba
# normalizado. Várias abas caem no MESMO setor (BB Réu = Defesa + Réu + Recursos).
# Abas fora deste mapa são ignoradas (ex.: BB CADASTRO vazia).
TAB_TO_SETOR = {
    "bb defesa": "bb-reu",
    "bb reu": "bb-reu",
    "recursos - reu": "bb-reu",
    "bb execucao&encerramento": "bb-execucao",
    "bb acordos": "bb-acordos",
    "bb estrategico": "bb-estrategico",
    "master reu": "master-reu",
    "ativos reu": "ativos-reu",
}


def norm(s) -> str:
    """Minúsculo, sem acento, espaços colapsados — chave de join com o L1."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.strip().lower().split())


def _aware(dt):
    if not isinstance(dt, datetime.datetime):
        return None
    if BRT is not None and dt.tzinfo is None:
        return dt.replace(tzinfo=BRT)
    return dt


def _find_col(header, *keys):
    """Índice da 1ª coluna cujo cabeçalho contém alguma das chaves (abas variam)."""
    for i, h in enumerate(header or ()):
        hn = norm(h)
        if hn and any(k in hn for k in keys):
            return i
    return None


def _canon_cargo(c):
    """Normaliza a grafia do cargo (as abas usam variações)."""
    cn = norm(c)
    if "advog" in cn:
        return "Advogado(a)"
    if "estag" in cn:
        return "Estagiário(a)"
    if "assist" in cn:
        return "Assistente"
    if "superv" in cn:
        return "Supervisor(a)"
    return (c or "").strip() or None


def seed_pessoas(db) -> dict:
    """Lê TODAS as abas da planilha de squads (cada aba → setor), detectando as
    colunas pelo cabeçalho (os layouts variam entre abas), e popula perf_pessoa
    com equipe (setor) + is_supervisor. Pessoa em abas do mesmo setor colapsa;
    em setores diferentes, fica no primeiro (ordem das abas)."""
    wb = openpyxl.load_workbook(SQUADS_XLSX, read_only=True, data_only=True)
    agg: dict = {}
    for sheet in wb.sheetnames:
        equipe = TAB_TO_SETOR.get(norm(sheet))
        if not equipe:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        ci_nome = _find_col(header, "nome")
        if ci_nome is None:
            continue
        ci_cargo = _find_col(header, "cargo")
        ci_squad = _find_col(header, "squad")
        ci_pos = _find_col(header, "posic")

        def _cell(r, i):
            return str(r[i]).strip() if i is not None and len(r) > i and r[i] else ""

        for r in rows[1:]:
            if not r or len(r) <= ci_nome or not r[ci_nome]:
                continue
            nome = str(r[ci_nome]).strip()
            nm = norm(nome)
            if not nm:
                continue
            cargo = _cell(r, ci_cargo)
            squad = _cell(r, ci_squad)
            pos = _cell(r, ci_pos)
            is_sup = "supervisor" in norm(cargo) or "supervisor" in norm(pos)
            cur = agg.get(nm)
            if cur is None:
                agg[nm] = {
                    "nome": nome, "cargo": cargo, "squad": squad, "posicao": pos,
                    "equipe": equipe, "is_supervisor": is_sup,
                }
            else:
                if not cur["cargo"] and cargo:
                    cur["cargo"] = cargo
                if not cur["squad"] and squad:
                    cur["squad"] = squad
                if not cur["posicao"] and pos:
                    cur["posicao"] = pos
                # equipe E is_supervisor: mantém os do PRIMEIRO setor — não deixa a
                # flag de supervisor de uma aba de outro time vazar (ex.: Fernanda,
                # supervisora em Acordos, aparece como linha "Acordo" em BB Réu).

    existing = {p.nome_norm: p for p in db.query(PerfPessoa).all()}
    for nm, d in agg.items():
        p = existing.get(nm)
        if p is None:
            p = PerfPessoa(nome_norm=nm)
            db.add(p)
        p.nome = d["nome"]
        p.cargo = _CARGO_OVERRIDE.get(nm) or _canon_cargo(d["cargo"])
        p.squad = d["squad"] or None
        p.posicao = d["posicao"] or None
        p.equipe = d["equipe"]
        p.is_supervisor = d["is_supervisor"]
        p.ativo = True
    db.commit()
    return {p.nome_norm: p.id for p in db.query(PerfPessoa).all()}


def seed_tarefas(db, name_to_id: dict, agenda_path: str = AGENDA_XLSX) -> int:
    """Replace total ATÔMICO: o delete e TODOS os inserts vão numa transação só,
    com commit ÚNICO no fim. Se cair no meio (container morto por redeploy, erro
    de parse), faz rollback e mantém os dados ANTIGOS completos — em vez de deixar
    a tabela parcial. (Já aconteceu: redeploy no meio do ingest = snapshot truncado
    com 78k de ~224k linhas e last_sync não gravado.)"""
    wb = openpyxl.load_workbook(agenda_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)
    n = 0
    try:
        db.query(PerfTarefa).delete()
        batch = []
        for r in rows:
            status = r[STATUS] if len(r) > STATUS else None
            if status not in ("Cumprido", "Pendente"):
                continue
            cumpr = norm(r[CUMPRIU]) if len(r) > CUMPRIU else ""
            env = norm(r[ENV]) if len(r) > ENV else ""
            # Cumprido -> executor (Cumprido por); Pendente -> responsável (Envolvido).
            pid = name_to_id.get(cumpr) if status == "Cumprido" else name_to_id.get(env)
            if not pid:
                continue  # fora do roster = ruído (escopo: só a planilha)
            batch.append(
                PerfTarefa(
                    pessoa_id=pid,
                    cumprido_por_nome=(str(r[CUMPRIU]).strip() if len(r) > CUMPRIU and r[CUMPRIU] else None),
                    envolvido_nome=(str(r[ENV]).strip() if len(r) > ENV and r[ENV] else None),
                    escritorio=(str(r[ESC]).strip() if len(r) > ESC and r[ESC] else None),
                    tipo=(str(r[TIPO]).strip() if len(r) > TIPO and r[TIPO] else None),
                    subtipo=(str(r[SUBTIPO]).strip() if len(r) > SUBTIPO and r[SUBTIPO] else None),
                    status=status,
                    concluido_em=_aware(r[CONCL]) if len(r) > CONCL else None,
                    cadastrado_em=_aware(r[CAD]) if len(r) > CAD else None,
                    prazo_previsto=_aware(r[PRAZO]) if len(r) > PRAZO else None,
                    l1_task_id=(int(r[ID]) if len(r) > ID and isinstance(r[ID], (int, float)) else None),
                    pasta=(str(r[PASTA]).strip() if len(r) > PASTA and r[PASTA] else None),
                    cnj=(str(r[CNJ]).strip() if len(r) > CNJ and r[CNJ] else None),
                    uf=(str(r[UF]).strip() if len(r) > UF and r[UF] else None),
                )
            )
            n += 1
            if len(batch) >= 2000:
                db.bulk_save_objects(batch)  # insere DENTRO da transação (sem commit)
                batch = []
        if batch:
            db.bulk_save_objects(batch)
        db.commit()  # único commit: delete + todos os inserts juntos (atômico)
    except Exception:
        db.rollback()
        raise
    return n


def classify_subtipos(db) -> None:
    rows = db.execute(
        text(
            """
            SELECT subtipo,
                   COUNT(*) FILTER (WHERE status='Cumprido') AS vol,
                   COUNT(DISTINCT (pessoa_id::text || ':' ||
                         (date(concluido_em AT TIME ZONE 'America/Sao_Paulo'))::text))
                     FILTER (WHERE status='Cumprido') AS pdias
            FROM perf_l1_tarefa
            WHERE subtipo IS NOT NULL
            GROUP BY subtipo
            """
        )
    ).fetchall()
    db.query(PerfSubtipoCategoria).delete()
    db.commit()
    for subtipo, vol, pdias in rows:
        vol = vol or 0
        pdias = pdias or 0
        dens = (vol / pdias) if pdias else 0.0
        if vol < 40:
            cat = CAT_RUIDO
        elif dens >= 6.0:
            cat = CAT_OPERACIONAL
        else:
            cat = CAT_PROFUNDO
        db.add(
            PerfSubtipoCategoria(subtipo=subtipo, categoria=cat, volume=vol, densidade=round(dens, 2))
        )
    db.commit()


def run() -> None:
    db = SessionLocal()
    try:
        ids = seed_pessoas(db)
        print("perf_pessoa:", len(ids))
        n = seed_tarefas(db, ids)
        print("perf_l1_tarefa:", n)
        classify_subtipos(db)
        cats = db.execute(
            text("SELECT categoria, count(*) FROM perf_subtipo_categoria GROUP BY categoria")
        ).fetchall()
        print("categorias:", {c: v for c, v in cats})
    finally:
        db.close()


if __name__ == "__main__":
    run()
