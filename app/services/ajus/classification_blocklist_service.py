"""
Servico do blocklist de classificacao pendente.

Operador sobe uma planilha XLSX gerada do Legal One com a lista de
processos com classificacao pendente (coluna "Numeros Processo" ou
similar — detectamos por header OU por conteudo via regex CNJ). Cada
upload SUBSTITUI o blocklist atomicamente:

  - CNJs novos no upload -> INSERT com first_seen_at = now.
  - CNJs que ja existiam -> UPDATE last_seen_at.
  - CNJs que sumiram do upload -> DELETE (classificacao concluida,
    libera disparo no dispatch).

Por que nao mudar status do item da fila?
- Status `pendente`/`erro`/`enviando`/`sucesso`/`cancelado` representa
  o ciclo de vida do envio AJUS. Bloqueio por classificacao pendente
  e' uma condicao ortogonal, externa ao ciclo de envio. Misturar os
  dois deixa a maquina de estados confusa e exigiria status extras
  ('bloqueado_classificacao_pendente') que sumiriam quando o operador
  sobe um upload novo sem o CNJ — perdendo o status anterior.
- Solucao: dispatch consulta a tabela na hora de mandar; item bloqueado
  e' simplesmente PULADO, mantem status atual.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timezone
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.ajus import AjusClassificationBlocklist

logger = logging.getLogger(__name__)


# Regex CNJ (mesmo padrao usado em queue_service.py): captura 20 digitos
# em qualquer formato (com ou sem mascara). Greedy o suficiente pra
# pegar a primeira ocorrencia em uma string com ruido.
_CNJ_REGEX = re.compile(
    r"(\d{7})[-.\s]?(\d{2})[-.\s]?(\d{4})[-.\s]?(\d{1})[-.\s]?(\d{2})[-.\s]?(\d{4})",
)


def _normalize_cnj(value) -> Optional[str]:
    """Aceita string com/sem mascara, numero, etc. Retorna 20 digitos
    ou None se nao bater."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Tenta extrair via regex (cobre o caso "linha tem texto extra")
    m = _CNJ_REGEX.search(s)
    if m:
        return "".join(m.groups())
    # Fallback: pega so digitos e valida tamanho
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 20:
        return digits
    return None


# Headers-padrao que indicam coluna de CNJ. Match case-insensitive +
# ignora acentos (a planilha vem com encoding torto as vezes).
_CNJ_HEADER_PATTERNS = (
    "numero processo",
    "numeros processo",
    "n processo",
    "no processo",
    "cnj",
    "processo",  # fallback amplo (ultima opcao)
)


def _norm_header(value) -> str:
    if value is None:
        return ""
    s = str(value).lower().strip()
    # Remove acentos comuns + caracteres soltos do CP1252 quebrado
    for a, b in (("á", "a"), ("ã", "a"), ("â", "a"), ("é", "e"),
                 ("ê", "e"), ("í", "i"), ("ó", "o"), ("ô", "o"),
                 ("ú", "u"), ("ç", "c"), ("ñ", "n"),
                 # CP1252 lixo quando file vem em encoding errado
                 ("�", ""), ("º", ""), ("°", "")):
        s = s.replace(a, b)
    # Colapsa whitespace
    return " ".join(s.split())


def _norm_cod_ajus(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in ("0", "0.0"):
        return None
    # Float vindo do openpyxl? Normaliza pra inteiro string
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s[:32]


def _norm_str_field(value, max_len: int) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s[:max_len]


class BlocklistParseError(RuntimeError):
    """Erro de parsing do XLSX (formato invalido, sem coluna CNJ, etc.)."""


def parse_xlsx(content: bytes) -> list[dict]:
    """
    Parseia o XLSX e devolve lista de dicts {cnj_number, cod_ajus, materia}.

    Detecta a coluna do CNJ:
      1. Pelo header (linha que tem 'Numero Processo'/'CNJ'/etc). Procura
         nas primeiras 5 linhas — header pode nao estar na linha 1 (caso
         da planilha real do L1, que tem linha 0 vazia).
      2. Se nao achar header, tenta detectar pela 1a linha que tem >= 1
         celula com CNJ valido — assume que essa coluna e' o CNJ.

    Tambem captura `cod_ajus` e `materia` se as colunas existirem (case-
    insensitive header match), pra debug/UI. Demais colunas ignoradas.

    Raises:
        BlocklistParseError: arquivo nao e' XLSX, planilha vazia, ou
            nao foi possivel achar coluna de CNJ.
    """
    if not content:
        raise BlocklistParseError("Arquivo vazio.")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BlocklistParseError(
            f"Nao foi possivel abrir o arquivo como XLSX: {exc}"
        ) from exc
    sheets = wb.sheetnames
    if not sheets:
        raise BlocklistParseError("XLSX sem abas.")
    ws = wb[sheets[0]]
    if ws.max_row is None or ws.max_row < 1:
        raise BlocklistParseError("Aba vazia.")

    # Le primeiras 10 linhas pra achar header
    head_rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        head_rows.append(row)
        if i >= 10:
            break

    cnj_col_idx: Optional[int] = None
    cod_col_idx: Optional[int] = None
    materia_col_idx: Optional[int] = None
    header_row_idx: Optional[int] = None

    # 1. Tenta achar header
    for ridx, row in enumerate(head_rows):
        if not row:
            continue
        norm_cells = [_norm_header(c) for c in row]
        for cidx, norm in enumerate(norm_cells):
            if not norm:
                continue
            for pattern in _CNJ_HEADER_PATTERNS:
                if pattern in norm:
                    cnj_col_idx = cidx
                    header_row_idx = ridx
                    break
            if cnj_col_idx is not None:
                break
        if cnj_col_idx is not None:
            # Acha cod_ajus e materia na MESMA linha de header
            for cidx, norm in enumerate(norm_cells):
                if not norm:
                    continue
                if cod_col_idx is None and ("cod ajus" in norm
                                            or "codigo ajus" in norm):
                    cod_col_idx = cidx
                if materia_col_idx is None and "materia" in norm:
                    materia_col_idx = cidx
            break

    # 2. Fallback: primeira linha com CNJ valido
    if cnj_col_idx is None:
        for ridx, row in enumerate(head_rows):
            if not row:
                continue
            for cidx, cell in enumerate(row):
                if _normalize_cnj(cell):
                    cnj_col_idx = cidx
                    header_row_idx = ridx - 1 if ridx > 0 else -1
                    break
            if cnj_col_idx is not None:
                break

    if cnj_col_idx is None:
        raise BlocklistParseError(
            "Nao foi possivel detectar a coluna de CNJ. Esperado header "
            "tipo 'Numeros Processo'/'CNJ'/'Processo' OU pelo menos uma "
            "celula com CNJ valido nas 10 primeiras linhas."
        )

    logger.info(
        "Blocklist XLSX: cnj_col=%d header_row=%s cod_col=%s materia_col=%s",
        cnj_col_idx, header_row_idx, cod_col_idx, materia_col_idx,
    )

    # Itera todas as linhas a partir da pos-header
    out: dict[str, dict] = {}
    skip_until = (header_row_idx + 1) if header_row_idx is not None else 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_until:
            continue
        if not row or len(row) <= cnj_col_idx:
            continue
        cnj = _normalize_cnj(row[cnj_col_idx])
        if not cnj:
            continue
        # Dedup intra-arquivo (mesmo CNJ aparece 2x na planilha)
        if cnj in out:
            continue
        cod = (_norm_cod_ajus(row[cod_col_idx])
               if cod_col_idx is not None and cod_col_idx < len(row) else None)
        materia = (_norm_str_field(row[materia_col_idx], 255)
                   if materia_col_idx is not None and materia_col_idx < len(row)
                   else None)
        out[cnj] = {
            "cnj_number": cnj,
            "cod_ajus": cod,
            "materia": materia,
        }

    if not out:
        raise BlocklistParseError(
            "Planilha lida mas nenhuma linha tinha CNJ valido. "
            "Confira a coluna de processo."
        )
    return list(out.values())


class AjusClassificationBlocklistService:
    """Operacoes do blocklist — replace (upload), query, stats."""

    def __init__(self, db: Session) -> None:
        self.db = db

    def replace_blocklist(self, parsed: list[dict]) -> dict:
        """
        Substitui o blocklist atomicamente. parsed = saida do parse_xlsx.
        Retorna summary {added, updated, removed, total_after}.

        Atomicidade: tudo numa transacao. Se algum INSERT/UPDATE der pau,
        rollback global — blocklist anterior fica intacto.
        """
        now = datetime.now(timezone.utc)
        new_cnjs: dict[str, dict] = {p["cnj_number"]: p for p in parsed}

        # Carrega o que ja' existe
        existing_rows = self.db.query(AjusClassificationBlocklist).all()
        existing_by_cnj = {r.cnj_number: r for r in existing_rows}

        added = 0
        updated = 0
        removed = 0

        # 1. Insert novos / Update existentes
        for cnj, payload in new_cnjs.items():
            row = existing_by_cnj.get(cnj)
            if row is None:
                row = AjusClassificationBlocklist(
                    cnj_number=cnj,
                    cod_ajus=payload.get("cod_ajus"),
                    materia=payload.get("materia"),
                    first_seen_at=now,
                    last_seen_at=now,
                )
                self.db.add(row)
                added += 1
            else:
                # Atualiza metadados (caso planilha nova tenha info nova)
                # e last_seen_at.
                row.cod_ajus = payload.get("cod_ajus")
                row.materia = payload.get("materia")
                row.last_seen_at = now
                updated += 1

        # 2. Delete os que sumiram (= classificacao concluida)
        for cnj, row in existing_by_cnj.items():
            if cnj not in new_cnjs:
                self.db.delete(row)
                removed += 1

        self.db.commit()

        total_after = (
            self.db.query(AjusClassificationBlocklist).count()
        )
        logger.info(
            "Blocklist replace: added=%d updated=%d removed=%d total_after=%d",
            added, updated, removed, total_after,
        )
        return {
            "added": added,
            "updated": updated,
            "removed": removed,
            "total_after": total_after,
        }

    def list_all(self, limit: int = 100, offset: int = 0,
                 cnj_filter: Optional[str] = None) -> tuple[int, list]:
        q = self.db.query(AjusClassificationBlocklist)
        if cnj_filter:
            digits = "".join(c for c in cnj_filter if c.isdigit())
            if digits:
                q = q.filter(
                    AjusClassificationBlocklist.cnj_number.like(
                        f"%{digits}%",
                    ),
                )
        total = q.count()
        rows = (
            q.order_by(AjusClassificationBlocklist.last_seen_at.desc())
            .offset(offset)
            .limit(limit)
            .all()
        )
        return total, rows

    def is_blocked(self, cnj_number: str) -> bool:
        """Test rapido — usado em dispatch_one."""
        if not cnj_number:
            return False
        digits = "".join(c for c in cnj_number if c.isdigit())
        if not digits:
            return False
        return self.db.query(
            AjusClassificationBlocklist.id,
        ).filter(
            AjusClassificationBlocklist.cnj_number == digits,
        ).first() is not None

    def get_blocked_set(self, cnjs: list[str]) -> set[str]:
        """
        Devolve subset dos CNJs informados que estao no blocklist.
        Recebe CNJs em qualquer formato (mascarado/digito); compara
        sempre por digito.
        """
        if not cnjs:
            return set()
        normalized = {
            "".join(c for c in (s or "") if c.isdigit()): (s or "")
            for s in cnjs
        }
        normalized.pop("", None)
        if not normalized:
            return set()
        rows = self.db.query(
            AjusClassificationBlocklist.cnj_number,
        ).filter(
            AjusClassificationBlocklist.cnj_number.in_(list(normalized.keys())),
        ).all()
        blocked_digits = {r[0] for r in rows}
        # Devolve nos formatos originais que entraram
        return {
            original
            for digits, original in normalized.items()
            if digits in blocked_digits
        }

    def stats(self) -> dict:
        """Resumo pra UI — total_no_blocklist, ultimo_upload_at, items_fila_bloqueados."""
        from app.models.ajus import (
            AJUS_QUEUE_PENDENTE,
            AJUS_QUEUE_ERRO,
            AjusAndamentoQueue,
        )

        total = self.db.query(AjusClassificationBlocklist).count()

        ultimo = (
            self.db.query(
                AjusClassificationBlocklist.last_seen_at,
            )
            .order_by(AjusClassificationBlocklist.last_seen_at.desc())
            .limit(1)
            .first()
        )
        ultimo_upload_at = ultimo[0] if ultimo else None

        # Conta items na fila (pendente/erro) cujo CNJ esta bloqueado
        items_fila_bloqueados = (
            self.db.query(AjusAndamentoQueue.id)
            .join(
                AjusClassificationBlocklist,
                AjusClassificationBlocklist.cnj_number ==
                AjusAndamentoQueue.cnj_number,
            )
            .filter(
                AjusAndamentoQueue.status.in_(
                    [AJUS_QUEUE_PENDENTE, AJUS_QUEUE_ERRO],
                ),
            )
            .count()
        )

        return {
            "total_no_blocklist": total,
            "ultimo_upload_at": ultimo_upload_at,
            "items_fila_bloqueados": items_fila_bloqueados,
        }

    def clear(self) -> int:
        """Apaga TODO o blocklist (escape hatch). Retorna quantos foram."""
        n = self.db.query(AjusClassificationBlocklist).delete()
        self.db.commit()
        logger.info("Blocklist clear: removidos=%d", n)
        return n
