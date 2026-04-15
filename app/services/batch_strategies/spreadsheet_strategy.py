import asyncio
import logging
import re
import unicodedata
from datetime import datetime, time, timezone
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy.orm import joinedload

from app.api.v1.schemas import BatchTaskCreationRequest
from app.models.batch_execution import (
    BATCH_STATUS_CANCELLED,
    BATCH_STATUS_PAUSED,
    BatchExecution,
    BatchExecutionItem,
)
from app.models.legal_one import LegalOneOffice, LegalOneTaskSubType, LegalOneUser
from app.services.batch_utils import build_task_fingerprint, load_successful_fingerprints

from .base_strategy import BaseStrategy

DEFAULT_TASK_STATUS_ID = 0


class SpreadsheetStrategy(BaseStrategy):
    REQUIRED_FIELDS = [
        "ESCRITORIO",
        "CNJ",
        "PUBLISH_DATE",
        "SUBTIPO",
        "EXECUTANTE",
        "DATA_TAREFA",
    ]
    COLUMN_KEYS = [
        "ESCRITORIO",
        "CNJ",
        "PUBLISH_DATE",
        "SUBTIPO",
        "EXECUTANTE",
        "PRAZO",
        "DATA_TAREFA",
        "HORARIO",
        "OBSERVACAO",
        "DESCRICAO",
    ]

    @staticmethod
    def _is_meaningful_row(row: tuple) -> bool:
        return any(value not in (None, "") and str(value).strip() for value in row)

    @staticmethod
    def _normalize_cnj_number(cnj_number: Any) -> str:
        if cnj_number is None:
            return ""
        return str(cnj_number).strip()

    @staticmethod
    def _normalize_lookup_value(value: Any) -> str:
        if value is None:
            return ""

        normalized = str(value).replace("\xa0", " ").strip().lower()
        normalized = unicodedata.normalize("NFKD", normalized)
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        normalized = re.sub(r"\s*/\s*", " / ", normalized)
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.strip()

    @staticmethod
    def _parse_task_time_value(time_value: Any) -> time:
        if isinstance(time_value, time):
            return time_value.replace(microsecond=0)

        if isinstance(time_value, datetime):
            return time_value.time().replace(microsecond=0)

        if isinstance(time_value, (int, float)):
            numeric_value = float(time_value)
        else:
            normalized_value = str(time_value).strip()
            if not normalized_value:
                raise ValueError("Valor de hora vazio.")

            try:
                return time.fromisoformat(normalized_value).replace(microsecond=0)
            except ValueError:
                pass

            try:
                numeric_value = float(normalized_value)
            except ValueError as exc:
                raise ValueError(f"Hora invalida: '{time_value}'") from exc

        if not 0 <= numeric_value < 1:
            raise ValueError(f"Hora invalida: '{time_value}'")

        total_seconds = min(round(numeric_value * 24 * 60 * 60), (24 * 60 * 60) - 1)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return time(hours, minutes, seconds)

    def _parse_and_format_date_to_utc(self, date_value: Any, time_value: Any = None) -> str:
        if not date_value:
            raise ValueError("Valor de data nao pode ser nulo.")

        try:
            if isinstance(date_value, datetime):
                local_date = date_value
            else:
                date_str = str(date_value).split(" ")[0]
                if "-" in date_str:
                    local_date = datetime.strptime(date_str, "%Y-%m-%d")
                else:
                    local_date = datetime.strptime(date_str, "%d/%m/%Y")

            task_time = time(23, 59, 59)
            if time_value:
                try:
                    task_time = self._parse_task_time_value(time_value)
                except (TypeError, ValueError):
                    logging.warning("Formato de hora invalido: '%s'. Usando horario padrao.", time_value)

            local_tz = ZoneInfo("America/Sao_Paulo")
            aware_datetime = datetime.combine(local_date.date(), task_time).replace(tzinfo=local_tz)
            utc_datetime = aware_datetime.astimezone(timezone.utc)
            return utc_datetime.isoformat().replace("+00:00", "Z")
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Data invalida: '{date_value}'") from exc

    def _format_date_for_description(self, date_value: Any) -> str:
        if not date_value:
            return ""

        try:
            if isinstance(date_value, datetime):
                return date_value.strftime("%d/%m/%Y")

            date_str = str(date_value).split(" ")[0]
            if "-" in date_str:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            else:
                date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            return date_obj.strftime("%d/%m/%Y")
        except ValueError:
            return str(date_value).split(" ")[0]

    async def _load_caches(self):
        return {
            "users": {
                self._normalize_lookup_value(user.name): user
                for user in self.db.query(LegalOneUser).filter(LegalOneUser.is_active == True).all()
            },
            "offices": {
                self._normalize_lookup_value(office.path): office
                for office in self.db.query(LegalOneOffice).filter(LegalOneOffice.is_active == True).all()
            },
            "subtypes": {
                self._normalize_lookup_value(subtype.name): subtype
                for subtype in self.db.query(LegalOneTaskSubType)
                .options(joinedload(LegalOneTaskSubType.parent_type))
                .filter(LegalOneTaskSubType.is_active == True)
                .all()
            },
        }

    def _load_sheet_rows(self, file_content: bytes):
        workbook = load_workbook(
            filename=BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        header_map = {
            str(cell).strip().upper(): index
            for index, cell in enumerate(header_row)
            if cell is not None
        }
        rows = [
            row
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if self._is_meaningful_row(row)
        ]
        return workbook, headers, header_map, rows

    def _build_indices(self, header_map: dict[str, int]) -> dict[str, int | None]:
        return {key: header_map.get(key) for key in self.COLUMN_KEYS}

    def _build_row_data(self, row: tuple, indices: dict[str, int | None]) -> dict[str, str | None]:
        row_data = {}
        for key, index in indices.items():
            value = row[index] if index is not None and index < len(row) else None
            row_data[key] = str(value).strip() if value is not None else None
        return row_data

    def extract_rows_for_queue(self, file_content: bytes) -> dict[str, Any]:
        workbook = None
        try:
            workbook, headers, header_map, rows = self._load_sheet_rows(file_content)
            indices = self._build_indices(header_map)
            missing_cols = [column for column in self.REQUIRED_FIELDS if indices[column] is None]
            if missing_cols:
                raise ValueError(f"Colunas obrigatorias ausentes na planilha: {', '.join(missing_cols)}")

            normalized_rows = [self._build_row_data(row, indices) for row in rows]
            return {"headers": headers, "rows": normalized_rows}
        finally:
            if workbook is not None:
                workbook.close()

    def preload_lawsuits_by_cnj(self, rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], set[str]]:
        prefetched_cnj_numbers = []
        seen_cnj_numbers = set()
        for row_data in rows:
            normalized_cnj = self._normalize_cnj_number(row_data.get("CNJ"))
            if not normalized_cnj or normalized_cnj in seen_cnj_numbers:
                continue
            prefetched_cnj_numbers.append(normalized_cnj)
            seen_cnj_numbers.add(normalized_cnj)

        if not prefetched_cnj_numbers or self.client is None:
            return {}, set(prefetched_cnj_numbers)

        lawsuit_lookup = self.client.search_lawsuits_by_cnj_numbers(prefetched_cnj_numbers)
        return lawsuit_lookup, set(prefetched_cnj_numbers)

    def _resolve_row_context(self, data: dict[str, Any], caches: dict[str, Any]) -> dict[str, Any]:
        missing = [field for field in self.REQUIRED_FIELDS if not data.get(field)]
        if missing:
            raise ValueError(f"Dados essenciais faltando: {', '.join(missing)}")

        office_name = data.get("ESCRITORIO")
        office = caches["offices"].get(self._normalize_lookup_value(office_name))
        if not office:
            raise ValueError(f"Escritorio '{office_name}' nao encontrado.")

        user_name = data.get("EXECUTANTE")
        user = caches["users"].get(self._normalize_lookup_value(user_name))
        if not user:
            raise ValueError(f"Usuario '{user_name}' nao encontrado.")

        subtype_name = data.get("SUBTIPO")
        sub_type = caches["subtypes"].get(self._normalize_lookup_value(subtype_name))
        if not sub_type or not sub_type.parent_type:
            raise ValueError(f"Subtipo '{subtype_name}' nao encontrado.")

        end_datetime_iso = self._parse_and_format_date_to_utc(
            data.get("DATA_TAREFA"),
            data.get("HORARIO"),
        )
        publish_date_iso = self._parse_and_format_date_to_utc(data.get("PUBLISH_DATE"))
        fingerprint = build_task_fingerprint(
            process_number=data.get("CNJ"),
            subtype_identifier=sub_type.external_id,
            responsible_identifier=user.external_id,
            due_datetime_iso=end_datetime_iso,
            origin_identifier=office.external_id,
        )

        return {
            "office": office,
            "user": user,
            "sub_type": sub_type,
            "end_datetime_iso": end_datetime_iso,
            "publish_date_iso": publish_date_iso,
            "fingerprint": fingerprint,
        }

    async def _wait_for_execution_signal(self, execution_id: int) -> str:
        while True:
            execution = (
                self.db.query(BatchExecution)
                .filter(BatchExecution.id == execution_id)
                .first()
            )
            if not execution:
                return BATCH_STATUS_CANCELLED
            if execution.status == BATCH_STATUS_PAUSED:
                await asyncio.sleep(1)
                continue
            return execution.status

    async def build_preview(self, file_content: bytes) -> dict[str, Any]:
        extracted = self.extract_rows_for_queue(file_content)
        headers = extracted["headers"]
        rows = extracted["rows"]

        caches = await self._load_caches()
        existing_fingerprints = load_successful_fingerprints(self.db)
        seen_fingerprints = set()

        preview_rows = []
        valid_rows = 0
        invalid_rows = 0
        duplicate_rows_in_file = 0
        duplicate_rows_in_history = 0

        for row_index, row_data in enumerate(rows, start=2):
            errors = []
            warnings = []

            try:
                context = self._resolve_row_context(row_data, caches)
                fingerprint = context["fingerprint"]
                if fingerprint in seen_fingerprints:
                    warnings.append("Duplicada na propria planilha.")
                    duplicate_rows_in_file += 1
                else:
                    seen_fingerprints.add(fingerprint)

                if fingerprint in existing_fingerprints:
                    warnings.append("Ja existe um agendamento igual processado com sucesso.")
                    duplicate_rows_in_history += 1
            except Exception as exc:
                fingerprint = None
                errors.append(str(exc))

            is_valid = not errors and not warnings
            if is_valid:
                valid_rows += 1
            else:
                invalid_rows += 1

            preview_rows.append(
                {
                    "row_id": row_index,
                    "process_number": row_data.get("CNJ"),
                    "is_valid": is_valid,
                    "errors": errors,
                    "warnings": warnings,
                    "data": row_data,
                    "fingerprint": fingerprint,
                }
            )

        return {
            "headers": headers,
            "rows": preview_rows,
            "summary": {
                "total_rows": len(rows),
                "valid_rows": valid_rows,
                "invalid_rows": invalid_rows,
                "duplicate_rows_in_file": duplicate_rows_in_file,
                "duplicate_rows_in_history": duplicate_rows_in_history,
            },
        }

    async def process_batch(self, request: BatchTaskCreationRequest, execution_log: BatchExecution) -> dict[str, Any]:
        success_count = 0
        failed_items = []

        extracted = self.extract_rows_for_queue(request.file_content)
        rows = extracted["rows"]
        execution_log.total_items = len(rows)
        self.db.commit()

        caches = await self._load_caches()
        known_fingerprints = load_successful_fingerprints(self.db)
        lawsuit_lookup, prefetched_cnj_numbers = self.preload_lawsuits_by_cnj(rows)

        for row_data in rows:
            signal = await self._wait_for_execution_signal(execution_log.id)
            if signal == BATCH_STATUS_CANCELLED:
                return {"sucesso": success_count, "falhas": len(failed_items), "cancelled": True}

            cnj = row_data.get("CNJ")
            log_item = BatchExecutionItem(
                process_number=cnj or "N/A",
                execution_id=execution_log.id,
                input_data=row_data,
                status="PENDENTE",
            )
            self.db.add(log_item)
            self.db.commit()

            success = await self.process_single_item(
                log_item,
                row_data,
                caches,
                known_fingerprints=known_fingerprints,
                lawsuit_lookup=lawsuit_lookup,
                prefetched_cnj_numbers=prefetched_cnj_numbers,
            )
            if success:
                success_count += 1
            else:
                failed_items.append({"cnj": cnj, "motivo": log_item.error_message})

            await asyncio.sleep(0.01)

        return {
            "sucesso": success_count,
            "falhas": len(failed_items),
            "detalhes_falhas": failed_items,
            "cancelled": False,
        }

    async def process_single_item(
        self,
        log_item: BatchExecutionItem,
        data: dict[str, Any],
        caches: dict[str, Any],
        *,
        known_fingerprints: set[str] | None = None,
        lawsuit_lookup: dict[str, dict[str, Any]] | None = None,
        prefetched_cnj_numbers: set[str] | None = None,
    ) -> bool:
        try:
            context = self._resolve_row_context(data, caches)
            fingerprint = context["fingerprint"]
            log_item.fingerprint = fingerprint

            if known_fingerprints is not None and fingerprint in known_fingerprints:
                raise ValueError("Tarefa duplicada detectada: ja existe um agendamento igual processado com sucesso.")

            cnj = self._normalize_cnj_number(data.get("CNJ"))
            lawsuit = (lawsuit_lookup or {}).get(cnj)
            if lawsuit is None and (prefetched_cnj_numbers is None or cnj not in prefetched_cnj_numbers):
                lawsuit = self.client.search_lawsuit_by_cnj(cnj)
            if not lawsuit or not lawsuit.get("id"):
                raise Exception("Processo nao encontrado no Legal One.")

            lawsuit_id = lawsuit["id"]
            responsible_office_id = lawsuit.get("responsibleOfficeId")
            if not responsible_office_id:
                raise Exception("Processo sem escritorio responsavel.")

            sub_type = context["sub_type"]
            formatted_deadline = self._format_date_for_description(data.get("PRAZO"))
            base_description = f"{sub_type.name} - {formatted_deadline}"
            extra_description = data.get("DESCRICAO")
            final_description = (
                f"{base_description} - {extra_description}"
                if extra_description
                else base_description
            )

            task_payload = {
                "description": final_description,
                "priority": "Normal",
                "startDateTime": context["end_datetime_iso"],
                "endDateTime": context["end_datetime_iso"],
                "publishDate": context["publish_date_iso"],
                "notes": data.get("OBSERVACAO"),
                "status": {"id": DEFAULT_TASK_STATUS_ID},
                "typeId": sub_type.parent_type.external_id,
                "subTypeId": sub_type.external_id,
                "responsibleOfficeId": responsible_office_id,
                "originOfficeId": context["office"].external_id,
                "participants": [
                    {
                        "contact": {"id": context["user"].external_id},
                        "isResponsible": True,
                        "isExecuter": True,
                        "isRequester": True,
                    }
                ],
            }

            created_task = self.client.create_task(task_payload)
            if not created_task or not created_task.get("id"):
                raise Exception("API retornou sucesso mas sem ID da tarefa.")

            task_id = created_task["id"]
            self.client.link_task_to_lawsuit(task_id, {"linkType": "Litigation", "linkId": lawsuit_id})

            log_item.status = "SUCESSO"
            log_item.created_task_id = task_id
            log_item.error_message = None
            if known_fingerprints is not None:
                known_fingerprints.add(fingerprint)
            self.db.commit()
            return True
        except Exception as exc:
            log_item.status = "FALHA"
            log_item.error_message = str(exc)
            self.db.commit()
            return False
