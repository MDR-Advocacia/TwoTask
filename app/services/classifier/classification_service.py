"""
Serviço orquestrador de classificação de publicações judiciais.
Lê a planilha, envia cada publicação para o agente IA, e persiste os resultados.
"""

import asyncio
import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.classification import (
    CLF_ITEM_FAILED,
    CLF_ITEM_PENDING,
    CLF_ITEM_SUCCESS,
    CLF_STATUS_CANCELLED,
    CLF_STATUS_COMPLETED,
    CLF_STATUS_COMPLETED_WITH_ERRORS,
    CLF_STATUS_PROCESSING,
    ClassificationBatch,
    ClassificationItem,
    FINAL_CLF_STATUSES,
)

from .ai_client import AnthropicClassifierClient
from .prompts import SYSTEM_PROMPT, build_user_message
from .taxonomy import validate_classification, repair_classification

logger = logging.getLogger(__name__)

# Colunas da planilha de entrada (0-indexed)
COL_PROCESS_NUMBER = 1   # Coluna B
COL_PUBLICATION_TEXT = 9  # Coluna J


class ClassificationService:

    def __init__(self, db: Session):
        self.db = db
        self.ai_client = AnthropicClassifierClient()

    # ──────────────────────────────────────────────
    # Leitura da planilha
    # ──────────────────────────────────────────────

    @staticmethod
    def extract_rows_from_spreadsheet(file_content: bytes) -> list[dict[str, Any]]:
        """Lê a planilha e extrai número do processo (col B) e texto (col J)."""
        wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        sheet = wb.active
        rows = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            process_number = row[COL_PROCESS_NUMBER] if len(row) > COL_PROCESS_NUMBER else None
            publication_text = row[COL_PUBLICATION_TEXT] if len(row) > COL_PUBLICATION_TEXT else None
            if not process_number:
                continue
            rows.append({
                "row_index": row_index,
                "process_number": str(process_number).strip(),
                "publication_text": str(publication_text).strip() if publication_text else "",
            })
        wb.close()
        return rows

    # ──────────────────────────────────────────────
    # Preview (antes de iniciar classificação)
    # ──────────────────────────────────────────────

    def build_preview(self, file_content: bytes) -> dict[str, Any]:
        rows = self.extract_rows_from_spreadsheet(file_content)
        preview_rows = []
        for row in rows:
            has_text = bool(row["publication_text"])
            preview_rows.append({
                "row_index": row["row_index"],
                "process_number": row["process_number"],
                "has_publication_text": has_text,
                "text_preview": row["publication_text"][:200] + "..." if len(row["publication_text"]) > 200 else row["publication_text"],
            })
        return {
            "total_rows": len(rows),
            "rows_with_text": sum(1 for r in rows if r["publication_text"]),
            "rows_without_text": sum(1 for r in rows if not r["publication_text"]),
            "preview": preview_rows[:50],
        }

    # ──────────────────────────────────────────────
    # Criação do batch
    # ──────────────────────────────────────────────

    def create_batch(
        self,
        file_content: bytes,
        filename: str | None = None,
        requested_by: str | None = None,
    ) -> ClassificationBatch:
        rows = self.extract_rows_from_spreadsheet(file_content)
        batch = ClassificationBatch(
            source_filename=filename,
            requested_by_email=requested_by,
            status=CLF_STATUS_PROCESSING,
            model_used=settings.classifier_model,
            total_items=len(rows),
        )
        self.db.add(batch)
        self.db.flush()

        for row in rows:
            item = ClassificationItem(
                batch_id=batch.id,
                row_index=row["row_index"],
                process_number=row["process_number"],
                publication_text=row["publication_text"],
                status=CLF_ITEM_PENDING,
            )
            self.db.add(item)

        self.db.commit()
        self.db.refresh(batch)
        return batch

    # ──────────────────────────────────────────────
    # Processamento assíncrono com concorrência
    # ──────────────────────────────────────────────

    async def process_batch(self, batch_id: int) -> dict[str, Any]:
        batch = self.db.query(ClassificationBatch).filter_by(id=batch_id).first()
        if not batch:
            raise ValueError(f"Batch {batch_id} não encontrado.")

        items = (
            self.db.query(ClassificationItem)
            .filter_by(batch_id=batch_id, status=CLF_ITEM_PENDING)
            .all()
        )

        semaphore = asyncio.Semaphore(settings.classifier_max_concurrent)
        success = 0
        failed = 0

        async def classify_item(item: ClassificationItem):
            nonlocal success, failed
            async with semaphore:
                # Checar cancelamento
                self.db.refresh(batch)
                if batch.status == CLF_STATUS_CANCELLED:
                    return

                if not item.publication_text:
                    item.status = CLF_ITEM_SUCCESS
                    item.category = "Para análise"
                    item.subcategory = "-"
                    item.confidence = "baixa"
                    item.justification = "Sem texto de publicação"
                    item.processed_at = datetime.now(timezone.utc)
                    self.db.commit()
                    success += 1
                    return

                try:
                    user_msg = build_user_message(item.process_number, item.publication_text)
                    result = await self.ai_client.classify(SYSTEM_PROMPT, user_msg)

                    raw_cat = result.get("categoria", "Para análise")
                    raw_sub = result.get("subcategoria", "-")
                    fixed_cat, fixed_sub = repair_classification(raw_cat, raw_sub)
                    if (fixed_cat, fixed_sub) != (raw_cat, raw_sub):
                        logger.info(
                            "Classificação auto-corrigida item %s: (%s/%s) → (%s/%s)",
                            item.id, raw_cat, raw_sub, fixed_cat, fixed_sub,
                        )
                    item.category = fixed_cat
                    item.subcategory = fixed_sub
                    item.confidence = result.get("confianca", "baixa")
                    item.justification = result.get("justificativa", "")
                    item.raw_response = result
                    item.processed_at = datetime.now(timezone.utc)

                    if not validate_classification(item.category, item.subcategory):
                        logger.warning(
                            "Classificação inválida para item %s: %s / %s",
                            item.id, item.category, item.subcategory,
                        )
                        item.category = "Para análise"
                        item.subcategory = "-"
                        item.confidence = "baixa"
                        item.justification = f"Classificação original inválida: {result}"

                    item.status = CLF_ITEM_SUCCESS
                    success += 1

                except Exception as exc:
                    logger.error("Erro classificando item %s: %s", item.id, exc)
                    item.status = CLF_ITEM_FAILED
                    item.error_message = str(exc)[:500]
                    item.processed_at = datetime.now(timezone.utc)
                    failed += 1

                self.db.commit()

        tasks = [classify_item(item) for item in items]
        await asyncio.gather(*tasks)

        # Atualizar status final do batch
        self.db.refresh(batch)
        if batch.status != CLF_STATUS_CANCELLED:
            batch.success_count = success
            batch.failure_count = failed
            batch.finished_at = datetime.now(timezone.utc)
            batch.status = (
                CLF_STATUS_COMPLETED if failed == 0
                else CLF_STATUS_COMPLETED_WITH_ERRORS
            )
            self.db.commit()

        return {
            "batch_id": batch_id,
            "status": batch.status,
            "success": success,
            "failed": failed,
            "total": len(items),
        }

    # ──────────────────────────────────────────────
    # Consulta de status
    # ──────────────────────────────────────────────

    def get_batch_status(self, batch_id: int) -> dict[str, Any]:
        batch = self.db.query(ClassificationBatch).filter_by(id=batch_id).first()
        if not batch:
            raise ValueError(f"Batch {batch_id} não encontrado.")

        items = self.db.query(ClassificationItem).filter_by(batch_id=batch_id).all()
        processed = sum(1 for i in items if i.status != CLF_ITEM_PENDING)

        return {
            "batch_id": batch.id,
            "status": batch.status,
            "total_items": batch.total_items,
            "processed": processed,
            "success_count": batch.success_count,
            "failure_count": batch.failure_count,
            "created_at": batch.created_at.isoformat() if batch.created_at else None,
            "finished_at": batch.finished_at.isoformat() if batch.finished_at else None,
        }

    # ──────────────────────────────────────────────
    # Resultados detalhados
    # ──────────────────────────────────────────────

    def get_batch_results(self, batch_id: int) -> list[dict[str, Any]]:
        items = (
            self.db.query(ClassificationItem)
            .filter_by(batch_id=batch_id)
            .order_by(ClassificationItem.row_index)
            .all()
        )
        return [
            {
                "row_index": item.row_index,
                "process_number": item.process_number,
                "publication_text_preview": (item.publication_text or "")[:150],
                "status": item.status,
                "category": item.category,
                "subcategory": item.subcategory,
                "confidence": item.confidence,
                "justification": item.justification,
                "error_message": item.error_message,
            }
            for item in items
        ]

    # ──────────────────────────────────────────────
    # Exportação para planilha
    # ──────────────────────────────────────────────

    def export_results_to_xlsx(self, batch_id: int) -> bytes:
        """Gera uma planilha XLSX com os resultados da classificação."""
        items = (
            self.db.query(ClassificationItem)
            .filter_by(batch_id=batch_id)
            .order_by(ClassificationItem.row_index)
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Classificações"

        headers = [
            "Nº do Processo",
            "Categoria",
            "Subcategoria",
            "Confiança",
            "Justificativa",
            "Status",
            "Erro",
        ]
        ws.append(headers)

        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2B5797")
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for item in items:
            ws.append([
                item.process_number,
                item.category or "",
                item.subcategory or "",
                item.confidence or "",
                item.justification or "",
                item.status,
                item.error_message or "",
            ])

        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 40

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    # ──────────────────────────────────────────────
    # Cancelamento
    # ──────────────────────────────────────────────

    def cancel_batch(self, batch_id: int) -> bool:
        batch = self.db.query(ClassificationBatch).filter_by(id=batch_id).first()
        if not batch or batch.status in FINAL_CLF_STATUSES:
            return False
        batch.status = CLF_STATUS_CANCELLED
        batch.finished_at = datetime.now(timezone.utc)
        self.db.commit()
        return True

    # ──────────────────────────────────────────────
    # Listagem de batches
    # ──────────────────────────────────────────────

    def list_batches(self, limit: int = 20) -> list[dict[str, Any]]:
        batches = (
            self.db.query(ClassificationBatch)
            .order_by(ClassificationBatch.created_at.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "id": b.id,
                "source_filename": b.source_filename,
                "status": b.status,
                "total_items": b.total_items,
                "success_count": b.success_count,
                "failure_count": b.failure_count,
                "model_used": b.model_used,
                "created_at": b.created_at.isoformat() if b.created_at else None,
                "finished_at": b.finished_at.isoformat() if b.finished_at else None,
            }
            for b in batches
        ]
