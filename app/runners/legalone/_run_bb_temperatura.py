"""Varredura do Banco do Brasil / Reu com classificacao de TEMPERATURA
(probabilidade de encerramento) por regra deterministica.

Pipeline:
  1. Le /tmp/base-bb.xlsx + filtra 17743 processos principais ativos
     (Polo=REU, exclui incidentais)
  2. Resolve CNJs -> lawsuit_ids via API L1
  3. Loop: lotes de 500 — cria run, varre (janela 30d), gera XLSX do
     lote com temperatura + capa. Copia pro Desktop.
  4. Repete ate zerar pendentes.

Output por lote: /tmp/bb-temperatura-{run_id}.xlsx
Logs: /app/output/playwright/legalone/varredura-andamentos/bb-temperatura.log

Uso:
    docker cp local.xlsx onetask-api-1:/tmp/base-bb.xlsx
    docker exec -d onetask-api-1 python //app/app/runners/legalone/_run_bb_temperatura.py
"""

from __future__ import annotations

import json
import logging
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

INPUT_PATH = Path("/tmp/base-bb.xlsx")
CAPA_MAP_PATH = Path("/tmp/bb-capa-map.json")
RESOLVE_CACHE = Path("/tmp/bb-cnj-resolve.json")
LOG_PATH = Path(
    "/app/output/playwright/legalone/varredura-andamentos/bb-temperatura.log"
)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("bb.temperatura")

WINDOW_DAYS = 30
BATCH_SIZE = 500

EXCLUIR_TIPO_ACAO = {
    "EMBARGOS A EXECUCAO",
    "EMBARGOS DE TERCEIRO",
    "CUMPRIMENTO DA SENTENCA",
    "ALVARA JUDICIAL",
    "CAUTELAR",
    "ADJUDICACAO",
}


# ── Etapa 1: leitura + filtro ─────────────────────────────────────────

def ler_e_filtrar_planilha() -> tuple[list[str], dict[str, dict]]:
    """Devolve (cnjs_digits, mapa cnj_digits -> capa)."""
    import openpyxl

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    cnjs: list[str] = []
    mapa: dict[str, dict] = {}
    seen: set[str] = set()
    total = 0
    excl_nao_reu = 0
    excl_tipo = 0
    excl_sinopse = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        polo = str(row[17] or "")
        tipo = str(row[13] or "").strip().upper()
        sinopse = str(row[36] or "").strip()
        cnj_raw = row[2]
        if polo != "REU":
            excl_nao_reu += 1
            continue
        if "incidental" in sinopse.lower():
            excl_sinopse += 1
            continue
        if tipo in EXCLUIR_TIPO_ACAO:
            excl_tipo += 1
            continue
        if not cnj_raw:
            continue
        digits = "".join(ch for ch in str(cnj_raw) if ch.isdigit())
        if len(digits) < 15 or digits in seen:
            continue
        seen.add(digits)
        capa = {
            headers[i]: row[i] for i in range(min(len(headers), len(row)))
        }
        capa["__cnj_original"] = str(cnj_raw).strip()
        mapa[digits] = capa
        cnjs.append(digits)
    logger.info(
        "Planilha: total=%s, excl_nao_reu=%s, excl_tipo=%s, excl_sinopse=%s, mantidos=%s",
        total, excl_nao_reu, excl_tipo, excl_sinopse, len(cnjs),
    )
    return cnjs, mapa


# ── Etapa 2: resolver CNJ -> lawsuit_id ───────────────────────────────

def resolver_cnjs(cnjs: list[str]) -> dict[str, int]:
    """Com cache no /tmp/bb-cnj-resolve.json."""
    from app.services.legal_one_client import LegalOneApiClient

    cache: dict[str, int] = {}
    if RESOLVE_CACHE.exists():
        try:
            cache = json.loads(RESOLVE_CACHE.read_text(encoding="utf-8"))
            cache = {k: int(v) for k, v in cache.items()}
            logger.info("Cache resolve: %s entries", len(cache))
        except Exception:
            cache = {}

    pendentes = [c for c in cnjs if c not in cache]
    if pendentes:
        client = LegalOneApiClient()
        logger.info("Resolvendo %s CNJs via L1...", len(pendentes))
        matches = client.search_lawsuits_by_cnj_numbers(pendentes)
        for cnj_norm in pendentes:
            payload = None
            for k, v in matches.items():
                if "".join(ch for ch in str(k) if ch.isdigit()) == cnj_norm:
                    payload = v
                    break
            if payload is not None and payload.get("id") is not None:
                try:
                    cache[cnj_norm] = int(payload["id"])
                except (TypeError, ValueError):
                    pass
        RESOLVE_CACHE.write_text(json.dumps(cache), encoding="utf-8")

    resolvidos = {c: cache[c] for c in cnjs if c in cache}
    logger.info("Resolvidos: %s / %s", len(resolvidos), len(cnjs))
    return resolvidos


# ── Etapa 3: regra de temperatura ─────────────────────────────────────

# Achados por lawsuit_id usados pra classificar
def classificar_temperatura(tipos_evento: set[str]) -> tuple[str, str]:
    """Retorna (temperatura, justificativa)."""
    if "transito_julgado" in tipos_evento or "arquivamento" in tipos_evento:
        sinais = sorted(tipos_evento & {"transito_julgado", "arquivamento"})
        return "ALTA", f"sinal de encerramento: {', '.join(sinais)}"
    if "sentenca" in tipos_evento:
        extras = tipos_evento - {"sentenca"}
        if extras:
            return "MEDIA", f"sentença + {', '.join(sorted(extras))}"
        return "MEDIA", "sentença proferida"
    if "audiencia_designada" in tipos_evento or "audiencia_cancelada" in tipos_evento or "revelia" in tipos_evento:
        return "BAIXA", f"em andamento: {', '.join(sorted(tipos_evento))}"
    return "INDETERMINADO", "sem andamentos relevantes em 30d"


# ── Etapa 4: gerar XLSX do lote ───────────────────────────────────────

def gerar_xlsx_lote(
    db, run_id: int, capa_map: dict[str, dict], xlsx_path: Path,
) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.models.varredura import VarreduraAchado, VarreduraProcessado, VarreduraRun

    run = db.query(VarreduraRun).filter(VarreduraRun.id == run_id).first()
    processados = (
        db.query(VarreduraProcessado)
        .filter(VarreduraProcessado.run_id == run_id)
        .all()
    )
    achados = (
        db.query(VarreduraAchado)
        .filter(VarreduraAchado.run_id == run_id)
        .order_by(VarreduraAchado.andamento_data.desc().nullslast(), VarreduraAchado.id.desc())
        .all()
    )

    # Agrega tipos por lawsuit
    tipos_por_lawsuit: dict[int, set[str]] = {}
    achados_por_lawsuit: dict[int, list] = {}
    for a in achados:
        tipos_por_lawsuit.setdefault(a.lawsuit_id, set()).add(a.tipo_evento)
        achados_por_lawsuit.setdefault(a.lawsuit_id, []).append(a)

    # Classifica cada processo
    rows = []
    contadores = {"ALTA": 0, "MEDIA": 0, "BAIXA": 0, "INDETERMINADO": 0}
    for p in processados:
        tipos = tipos_por_lawsuit.get(p.lawsuit_id, set())
        temp, just = classificar_temperatura(tipos)
        contadores[temp] += 1
        # Pega CNJ canônico do capa_map se disponível
        digits = "".join(ch for ch in (p.cnj_number or "") if ch.isdigit())
        capa = capa_map.get(digits) or {}
        rows.append({
            "lawsuit_id": p.lawsuit_id,
            "cnj": p.cnj_number or capa.get("__cnj_original") or "",
            "temperatura": temp,
            "justificativa": just,
            "tipos_evento": ", ".join(sorted(tipos)),
            "qtd_achados": len(achados_por_lawsuit.get(p.lawsuit_id, [])),
            "capa": capa,
            "achados": achados_por_lawsuit.get(p.lawsuit_id, []),
        })

    # Build workbook
    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFE5E7EB")
    color_alta = PatternFill("solid", fgColor="FFFEE2E2")
    color_media = PatternFill("solid", fgColor="FFFEF3C7")
    color_baixa = PatternFill("solid", fgColor="FFDCFCE7")
    color_indef = PatternFill("solid", fgColor="FFF3F4F6")

    # ── Aba Resumo ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = f"Lote run #{run_id} — BB/Réu — Temperatura de encerramento"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    info = [
        ("Run ID", run_id),
        ("Lote", run.triggered_by if run else ""),
        ("Iniciada", run.started_at.strftime("%d/%m/%Y %H:%M") if run and run.started_at else ""),
        ("Concluída", run.completed_at.strftime("%d/%m/%Y %H:%M") if run and run.completed_at else ""),
        ("Janela (dias)", WINDOW_DAYS),
        ("Total processos no lote", len(processados)),
        ("Total achados", len(achados)),
        ("", ""),
        ("🔴 ALTA (apto ao encerramento)", contadores["ALTA"]),
        ("🟡 MÉDIA (sentença sem trânsito)", contadores["MEDIA"]),
        ("🟢 BAIXA (em andamento)", contadores["BAIXA"]),
        ("⚪ INDETERMINADO (sem sinais)", contadores["INDETERMINADO"]),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = bold if isinstance(v, str) and v == "" else None
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 30

    # ── Helper pra escrever aba de temperatura ───────────────────────
    def _write_temp_sheet(sheet_name: str, subset: list, fill: PatternFill):
        ws_t = wb.create_sheet(sheet_name)
        cols = [
            "Temperatura", "CNJ", "Lawsuit ID", "NPJ", "Tipos de evento",
            "Justificativa", "Qtd achados",
            "Tipo Ação", "Situação", "UF", "Comarca", "Vara", "Valor da Causa",
            "Advogado", "Matéria",
            "Datas achados", "Trechos detectados",
        ]
        for c, h in enumerate(cols, start=1):
            cell = ws_t.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        for i, r in enumerate(subset, start=2):
            capa = r["capa"]
            ws_t.cell(row=i, column=1, value=r["temperatura"]).fill = fill
            ws_t.cell(row=i, column=2, value=r["cnj"])
            ws_t.cell(row=i, column=3, value=r["lawsuit_id"])
            ws_t.cell(row=i, column=4, value=str(capa.get("NPJ") or ""))
            ws_t.cell(row=i, column=5, value=r["tipos_evento"])
            ws_t.cell(row=i, column=6, value=r["justificativa"])
            ws_t.cell(row=i, column=7, value=r["qtd_achados"])
            ws_t.cell(row=i, column=8, value=str(capa.get("Tipo de Ação") or ""))
            ws_t.cell(row=i, column=9, value=str(capa.get("Situação do Processo") or ""))
            ws_t.cell(row=i, column=10, value=str(capa.get("UF") or ""))
            ws_t.cell(row=i, column=11, value=str(capa.get("Comarca") or ""))
            ws_t.cell(row=i, column=12, value=str(capa.get("Vara") or ""))
            ws_t.cell(row=i, column=13, value=str(capa.get("Valor da Causa") or ""))
            ws_t.cell(row=i, column=14, value=str(capa.get("Advogado") or ""))
            ws_t.cell(row=i, column=15, value=str(capa.get("Matéria") or ""))
            # Lista de datas/trechos
            datas = []
            trechos = []
            for a in r["achados"]:
                d = a.andamento_data.strftime("%d/%m/%Y") if a.andamento_data else "—"
                datas.append(f"{d}: {a.tipo_evento}")
                trecho = (a.andamento_texto or "")[:300]
                trechos.append(f"[{d}] {trecho}")
            ws_t.cell(row=i, column=16, value="\n".join(datas))
            tc = ws_t.cell(row=i, column=17, value="\n\n".join(trechos))
            tc.alignment = Alignment(wrap_text=True, vertical="top")
        widths = [16, 26, 12, 12, 40, 38, 10, 24, 14, 6, 22, 30, 14, 30, 22, 30, 90]
        for c, w in enumerate(widths, start=1):
            ws_t.column_dimensions[get_column_letter(c)].width = w
        if subset:
            ws_t.freeze_panes = "A2"
            ws_t.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(subset) + 1}"

    rows_alta = [r for r in rows if r["temperatura"] == "ALTA"]
    rows_media = [r for r in rows if r["temperatura"] == "MEDIA"]
    rows_baixa = [r for r in rows if r["temperatura"] == "BAIXA"]
    rows_indef = [r for r in rows if r["temperatura"] == "INDETERMINADO"]
    _write_temp_sheet("🔴 ALTA (apto encerramento)", rows_alta, color_alta)
    _write_temp_sheet("🟡 MEDIA", rows_media, color_media)
    _write_temp_sheet("🟢 BAIXA", rows_baixa, color_baixa)
    _write_temp_sheet("⚪ INDETERMINADO", rows_indef, color_indef)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return contadores


# ── Etapa 5: orquestracao em lotes ────────────────────────────────────

def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import (
        QUEUE_STATUS_PENDING,
        RUN_STATUS_RUNNING,
        VarreduraProcessado,
        VarreduraRun,
    )
    from app.services.varredura.varredura_service import (
        _run_subprocess_worker_impl,
    )

    logger.info("=" * 60)
    logger.info("BB/Reu — TEMPERATURA — pipeline iniciado")
    logger.info("=" * 60)

    cnjs, capa_map = ler_e_filtrar_planilha()
    cnj_to_id = resolver_cnjs(cnjs)

    # Salva capa map por lawsuit_id e por cnj_digits
    id_to_cnj = {v: k for k, v in cnj_to_id.items()}
    capa_for_save = {}
    for digits, capa in capa_map.items():
        capa_for_save[digits] = capa
    CAPA_MAP_PATH.write_text(
        json.dumps(capa_for_save, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    # ORDENA por prioridade de encerramento:
    # Cumprimento > Sentença > Recurso > Inicial > demais
    PRIORIDADE = {
        "Cumprimento": 0,
        "Sentença": 1,
        "Recurso": 2,
        "Inicial": 3,
    }

    def _prio(lid: int) -> tuple[int, int]:
        digits = id_to_cnj.get(lid)
        capa = capa_map.get(digits) if digits else None
        sit = str((capa or {}).get("Situação do Processo") or "").strip()
        return (PRIORIDADE.get(sit, 99), lid)

    lawsuits_total = sorted(set(cnj_to_id.values()), key=_prio)
    # Conta distribuicao por situacao
    from collections import Counter as _C
    sit_count = _C()
    for lid in lawsuits_total:
        digits = id_to_cnj.get(lid)
        capa = capa_map.get(digits) if digits else None
        sit = str((capa or {}).get("Situação do Processo") or "").strip()
        sit_count[sit] += 1
    logger.info("Total lawsuit_ids resolvidos: %s", len(lawsuits_total))
    logger.info("Ordem de prioridade aplicada: Cumprimento > Sentença > Recurso > Inicial")
    for k, v in sit_count.most_common():
        logger.info("  %s: %s", k or "(vazio)", v)

    # Lawsuits ja varridos em runs anteriores desta feature
    db = SessionLocal()
    try:
        runs_anteriores = (
            db.query(VarreduraRun.id)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-%"))
            .all()
        )
        ja_varridos: set[int] = set()
        for (rid,) in runs_anteriores:
            ids = (
                db.query(VarreduraProcessado.lawsuit_id)
                .filter(VarreduraProcessado.run_id == rid)
                .all()
            )
            ja_varridos.update(r[0] for r in ids)
    finally:
        db.close()
    logger.info("Ja varridos em runs anteriores BB: %s", len(ja_varridos))

    restante = [lid for lid in lawsuits_total if lid not in ja_varridos]
    total_lotes = (len(restante) + BATCH_SIZE - 1) // BATCH_SIZE
    logger.info(
        "Restante pra varrer: %s — em %s lotes de %s",
        len(restante), total_lotes, BATCH_SIZE,
    )

    for idx in range(total_lotes):
        chunk = restante[idx * BATCH_SIZE : (idx + 1) * BATCH_SIZE]
        if not chunk:
            break
        logger.info("─── LOTE %s/%s — %s processos ───", idx + 1, total_lotes, len(chunk))

        db = SessionLocal()
        try:
            run = VarreduraRun(
                status=RUN_STATUS_RUNNING,
                started_at=datetime.now(timezone.utc),
                responsible_office_ids=[23],
                window_days=WINDOW_DAYS,
                triggered_by=f"bb-temperatura-lote-{idx + 1}-de-{total_lotes}",
                total_processos=len(chunk),
            )
            db.add(run)
            db.flush()
            for lid in chunk:
                digits = id_to_cnj.get(lid)
                cnj_orig = None
                if digits:
                    cnj_orig = capa_map.get(digits, {}).get("__cnj_original")
                db.add(
                    VarreduraProcessado(
                        run_id=run.id,
                        lawsuit_id=lid,
                        cnj_number=cnj_orig or digits,
                        office_id=23,
                        queue_status=QUEUE_STATUS_PENDING,
                    )
                )
            db.commit()
            run_id = run.id
            logger.info("Run #%s criada. Disparando subprocess...", run_id)
            try:
                _run_subprocess_worker_impl(db, run_id)
            except Exception as exc:
                logger.exception("Erro no worker run #%s: %s", run_id, exc)

            # Gera XLSX do lote
            try:
                xlsx_path = Path(f"/tmp/bb-temperatura-{run_id}.xlsx")
                contadores = gerar_xlsx_lote(db, run_id, capa_map, xlsx_path)
                size = xlsx_path.stat().st_size
                logger.info(
                    "XLSX LOTE %s pronto: %s bytes — ALTA=%s, MEDIA=%s, BAIXA=%s, INDEF=%s",
                    run_id, size, contadores["ALTA"], contadores["MEDIA"],
                    contadores["BAIXA"], contadores["INDETERMINADO"],
                )
            except Exception:
                logger.exception("Falha XLSX lote run #%s", run_id)
        finally:
            db.close()

        # Respiro entre lotes (sessao OnePass)
        time.sleep(10)

    logger.info("=" * 60)
    logger.info("=== PIPELINE BB/Reu CONCLUIDO ===")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
