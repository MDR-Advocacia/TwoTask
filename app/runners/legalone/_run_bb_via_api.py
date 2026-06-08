"""Varredura BB/Reu via API L1 /Updates (sem RPA).

Substitui o pipeline RPA Playwright (~40h) por chamadas HTTP paralelas
ao endpoint /Updates da API L1 (~30min-1h pra 17708 processos com 4-8
threads). Salva TUDO em varredura_andamento_raw + sync DuckDB.

Diferencas vs V3 (RPA):
  - Sem janela artificial: API devolve TODOS os andamentos do processo
    (se quiser cap, passa --since YYYY-MM-DD)
  - Sem lotes de 500: 1 run unica com todos os processos
  - Threading paralelo (default 4) pra respeitar rate limit
  - Texto integra direto do `description` (vs scraping HTML)
  - Capa_json no processado (snapshot da planilha)
  - Regex aplicado inline -> varredura_achado
  - Sync DuckDB no final + delete do PG (libera espaco)

NAO dispara automaticamente. Voce roda manualmente quando quiser:

    docker cp /caminho/base-atualizada.xlsx onetask-api-1:/tmp/base-bb-api.xlsx

    # Smoke test com 10 processos primeiro
    docker exec onetask-api-1 python //app/app/runners/legalone/_run_bb_via_api.py --limit 10

    # Se OK, varredura completa em background
    docker exec -d onetask-api-1 python //app/app/runners/legalone/_run_bb_via_api.py

Opcoes:
    --limit N          Processa so' os primeiros N (smoke test)
    --since YYYY-MM-DD Filtra andamentos a partir dessa data
    --workers N        Threads paralelas (default 4; cuidado com 429)
    --triggered-by STR Prefixo no triggered_by da run (default 'bb-via-api')
    --no-duckdb-sync   Pula sync DuckDB no final (debug)
    --no-delete-raw    NAO apaga raw do PG apos sync (pra debug)

Log: /app/output/playwright/legalone/varredura-andamentos/bb-via-api.log
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

INPUT_PATHS = [
    Path("/tmp/base-bb-api.xlsx"),
    Path("/tmp/base-bb-v3.xlsx"),
    Path("/tmp/base-bb-v2.xlsx"),
]
CAPA_MAP_PATH = Path("/tmp/bb-api-capa-map.json")
RESOLVE_CACHE = Path("/tmp/bb-api-cnj-resolve.json")
LOG_PATH = Path(
    "/app/output/playwright/legalone/varredura-andamentos/bb-via-api.log"
)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("bb.via_api")

EXCLUIR_TIPO_ACAO = {
    "EMBARGOS A EXECUCAO",
    "EMBARGOS DE TERCEIRO",
    "CUMPRIMENTO DA SENTENCA",
    "ALVARA JUDICIAL",
    "CAUTELAR",
    "ADJUDICACAO",
}


# ── 1. Leitura + filtro ───────────────────────────────────────────────


def _input_path() -> Path:
    for p in INPUT_PATHS:
        if p.exists():
            logger.info("Base usada: %s", p)
            return p
    raise FileNotFoundError(
        f"Nenhuma base em {[str(p) for p in INPUT_PATHS]}. "
        "Copie com: docker cp local.xlsx onetask-api-1:/tmp/base-bb-api.xlsx"
    )


def ler_e_filtrar_planilha() -> tuple[list[str], dict[str, dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(_input_path(), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    cnjs: list[str] = []
    mapa: dict[str, dict] = {}
    seen: set[str] = set()
    total = excl_nao_reu = excl_tipo = excl_sinopse = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        total += 1
        polo = str(row[17] or "")
        tipo = str(row[13] or "").strip().upper()
        sinopse = str(row[36] or "").strip()
        cnj_raw = row[2]
        if polo != "REU":
            excl_nao_reu += 1
            continue
        if "incidental" in sinopse.lower():
            excl_sinopse += 1
            continue
        if tipo in EXCLUIR_TIPO_ACAO:
            excl_tipo += 1
            continue
        digits = "".join(ch for ch in str(cnj_raw) if ch.isdigit())
        if len(digits) < 15 or digits in seen:
            continue
        seen.add(digits)
        capa = {
            headers[i]: row[i] for i in range(min(len(headers), len(row)))
        }
        capa["__cnj_original"] = str(cnj_raw).strip()
        mapa[digits] = capa
        cnjs.append(digits)
    logger.info(
        "Planilha: total=%s, excl_nao_reu=%s, excl_tipo=%s, excl_sinopse=%s, mantidos=%s",
        total, excl_nao_reu, excl_tipo, excl_sinopse, len(cnjs),
    )
    return cnjs, mapa


# ── 2. Resolver CNJ -> lawsuit_id ─────────────────────────────────────


def resolver_cnjs(cnjs: list[str]) -> dict[str, int]:
    from app.services.legal_one_client import LegalOneApiClient

    cache: dict[str, int] = {}
    if RESOLVE_CACHE.exists():
        try:
            cache = json.loads(RESOLVE_CACHE.read_text(encoding="utf-8"))
            cache = {k: int(v) for k, v in cache.items()}
            logger.info("Cache resolve API: %s entries", len(cache))
        except Exception:
            cache = {}

    # Reaproveita caches anteriores (V2/V3)
    for old in (
        Path("/tmp/bb-v3-cnj-resolve.json"),
        Path("/tmp/bb-v2-cnj-resolve.json"),
    ):
        if old.exists():
            try:
                extra = json.loads(old.read_text(encoding="utf-8"))
                extra = {k: int(v) for k, v in extra.items()}
                count_before = len(cache)
                for k, v in extra.items():
                    cache.setdefault(k, v)
                logger.info(
                    "Cache %s importado: +%s entries (%s totais)",
                    old.name, len(cache) - count_before, len(cache),
                )
            except Exception:
                pass

    pendentes = [c for c in cnjs if c not in cache]
    if pendentes:
        client = LegalOneApiClient()
        logger.info("Resolvendo %s CNJs via L1...", len(pendentes))
        matches = client.search_lawsuits_by_cnj_numbers(pendentes)
        for cnj_norm in pendentes:
            payload = None
            for k, v in matches.items():
                if "".join(ch for ch in str(k) if ch.isdigit()) == cnj_norm:
                    payload = v
                    break
            if payload is not None and payload.get("id") is not None:
                try:
                    cache[cnj_norm] = int(payload["id"])
                except (TypeError, ValueError):
                    pass
        RESOLVE_CACHE.write_text(json.dumps(cache), encoding="utf-8")

    resolvidos = {c: cache[c] for c in cnjs if c in cache}
    logger.info("Resolvidos: %s / %s", len(resolvidos), len(cnjs))
    return resolvidos


# ── 3. Worker de fetch + persist ──────────────────────────────────────


def _parse_iso_date(s: Optional[str]):
    if not s:
        return None
    try:
        # Formato comum: "2026-02-23T00:00:00Z" ou "2025-01-13T17:04:31..."
        from datetime import datetime as _dt
        return _dt.fromisoformat(s.replace("Z", "+00:00")).date()
    except Exception:
        return None


def _parse_iso_time(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    try:
        from datetime import datetime as _dt
        t = _dt.fromisoformat(s.replace("Z", "+00:00"))
        return t.strftime("%H:%M:%S")
    except Exception:
        return None


def _persist_processo(
    processado_id: int,
    run_id: int,
    lawsuit_id: int,
    cnj_number: Optional[str],
    office_id: Optional[int],
    updates: list[dict],
) -> tuple[int, int]:
    """Persiste andamentos brutos + achados (regex). Retorna (qtd_raw, qtd_achados)."""
    from app.db.session import SessionLocal
    from app.models.varredura import (
        QUEUE_STATUS_COMPLETED,
        VarreduraAchado,
        VarreduraAndamentoRaw,
        VarreduraProcessado,
    )
    from app.services.varredura.regex_eventos import detect_eventos

    db = SessionLocal()
    try:
        qtd_raw = 0
        qtd_achados = 0
        for ordem, u in enumerate(updates):
            texto = (u.get("description") or "").strip()
            if not texto:
                continue
            d = _parse_iso_date(u.get("date") or u.get("creationDate"))
            hora = _parse_iso_time(u.get("date") or u.get("creationDate"))
            tipo_raw = u.get("originType") or ""
            type_id = u.get("typeId")
            tipo = f"{tipo_raw}#{type_id}" if type_id is not None else tipo_raw
            tipo = tipo[:64] or None

            db.add(
                VarreduraAndamentoRaw(
                    run_id=run_id,
                    processado_id=processado_id,
                    lawsuit_id=lawsuit_id,
                    cnj_number=cnj_number,
                    office_id=office_id,
                    andamento_data=d,
                    andamento_hora=(hora or "")[:8] or None,
                    andamento_tipo=tipo,
                    andamento_texto=texto,
                    andamento_movimentado_por=None,  # API L1 nao traz nome direto
                    ordem=ordem,
                )
            )
            qtd_raw += 1

            # Aplica regex
            for det in detect_eventos(texto):
                db.add(
                    VarreduraAchado(
                        run_id=run_id,
                        processado_id=processado_id,
                        lawsuit_id=lawsuit_id,
                        cnj_number=cnj_number,
                        andamento_data=d,
                        andamento_hora=(hora or "")[:8] or None,
                        andamento_tipo=tipo,
                        andamento_texto=texto,
                        andamento_movimentado_por=None,
                        tipo_evento=det.tipo,
                        regex_matched=det.matched_text[:500],
                    )
                )
                qtd_achados += 1

        # Atualiza processado
        p = (
            db.query(VarreduraProcessado)
            .filter(VarreduraProcessado.id == processado_id)
            .first()
        )
        if p is not None:
            p.queue_status = QUEUE_STATUS_COMPLETED
            p.total_andamentos_lidos = qtd_raw
            p.total_achados = qtd_achados
            p.completed_at = datetime.now(timezone.utc)
            p.last_attempt_at = datetime.now(timezone.utc)
            p.last_reason = "ok"

        db.commit()
        return qtd_raw, qtd_achados
    except Exception as exc:
        db.rollback()
        logger.warning("persist lid=%s falhou: %s", lawsuit_id, exc)
        return 0, 0
    finally:
        db.close()


def _mark_processado_falhou(processado_id: int, msg: str) -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import QUEUE_STATUS_FAILED, VarreduraProcessado

    db = SessionLocal()
    try:
        p = (
            db.query(VarreduraProcessado)
            .filter(VarreduraProcessado.id == processado_id)
            .first()
        )
        if p is not None:
            p.queue_status = QUEUE_STATUS_FAILED
            p.last_error = (msg or "")[:1000]
            p.last_reason = "api_error"
            p.last_attempt_at = datetime.now(timezone.utc)
            db.commit()
    finally:
        db.close()


def _worker_fetch(
    args: tuple[int, int, int, Optional[str], Optional[int], Optional[str]],
) -> tuple[int, int, int, Optional[str]]:
    """Thread worker. Args: (run_id, processado_id, lawsuit_id, cnj, office_id, since).
    Retorna (lawsuit_id, qtd_raw, qtd_achados, error_msg)."""
    from app.services.legal_one_client import LegalOneApiClient

    run_id, processado_id, lawsuit_id, cnj, office_id, since = args
    try:
        client = LegalOneApiClient()
        updates = client.fetch_updates_by_lawsuit(lawsuit_id, since=since)
    except Exception as exc:
        _mark_processado_falhou(processado_id, str(exc))
        return (lawsuit_id, 0, 0, str(exc)[:200])

    raw, ach = _persist_processo(
        processado_id, run_id, lawsuit_id, cnj, office_id, updates
    )
    return (lawsuit_id, raw, ach, None)


# ── 4. Main ───────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None,
                        help="Smoke test: processa só os primeiros N processos")
    parser.add_argument("--since", type=str, default=None,
                        help="Filtra andamentos a partir dessa data (YYYY-MM-DD)")
    parser.add_argument("--workers", type=int, default=4,
                        help="Threads paralelas (default 4)")
    parser.add_argument("--triggered-by", type=str, default="bb-via-api",
                        help="Prefixo no triggered_by da VarreduraRun")
    parser.add_argument("--no-duckdb-sync", action="store_true",
                        help="Pula sync DuckDB no final")
    parser.add_argument("--no-delete-raw", action="store_true",
                        help="NAO apaga raw do PG apos sync DuckDB")
    args = parser.parse_args()

    from app.db.session import SessionLocal
    from app.models.varredura import (
        QUEUE_STATUS_PENDING,
        RUN_STATUS_DONE,
        RUN_STATUS_FAILED,
        RUN_STATUS_RUNNING,
        VarreduraProcessado,
        VarreduraRun,
    )

    logger.info("=" * 60)
    logger.info(
        "BB/Reu via API L1 — triggered_by=%s, workers=%s, since=%s, limit=%s",
        args.triggered_by, args.workers, args.since, args.limit,
    )
    logger.info("=" * 60)

    cnjs, capa_map = ler_e_filtrar_planilha()
    cnj_to_id = resolver_cnjs(cnjs)
    CAPA_MAP_PATH.write_text(
        json.dumps(capa_map, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    id_to_cnj = {v: k for k, v in cnj_to_id.items()}

    # Prioriza Cumprimento > Sentença > Recurso > Inicial
    PRIORIDADE = {"Cumprimento": 0, "Sentença": 1, "Recurso": 2, "Inicial": 3}

    def _prio(lid: int) -> tuple[int, int]:
        digits = id_to_cnj.get(lid)
        capa = capa_map.get(digits) if digits else None
        sit = str((capa or {}).get("Situação do Processo") or "").strip()
        return (PRIORIDADE.get(sit, 99), lid)

    lawsuits_total = sorted(set(cnj_to_id.values()), key=_prio)
    if args.limit:
        lawsuits_total = lawsuits_total[: args.limit]
        logger.info("LIMIT aplicado: %s processos", len(lawsuits_total))

    # Cria run unica
    db = SessionLocal()
    try:
        run = VarreduraRun(
            status=RUN_STATUS_RUNNING,
            started_at=datetime.now(timezone.utc),
            responsible_office_ids=[23],
            window_days=0,  # API L1 nao usa janela artificial; 0 = sem janela
            triggered_by=args.triggered_by,
            total_processos=len(lawsuits_total),
        )
        db.add(run)
        db.flush()
        run_id = run.id
        # Cria processados com capa_json
        for lid in lawsuits_total:
            digits = id_to_cnj.get(lid)
            capa = capa_map.get(digits) if digits else None
            cnj_orig = (capa or {}).get("__cnj_original") if capa else None
            db.add(
                VarreduraProcessado(
                    run_id=run_id,
                    lawsuit_id=lid,
                    cnj_number=cnj_orig or digits,
                    office_id=23,
                    queue_status=QUEUE_STATUS_PENDING,
                    capa_json=capa,
                )
            )
        db.commit()
        logger.info(
            "Run #%s criada com %s processos. Disparando threads...",
            run_id, len(lawsuits_total),
        )

        # Mapa lawsuit_id -> processado_id
        proc_map = {
            p.lawsuit_id: p.id
            for p in db.query(VarreduraProcessado)
            .filter(VarreduraProcessado.run_id == run_id)
            .all()
        }
    finally:
        db.close()

    # Threading
    work_items = [
        (
            run_id,
            proc_map[lid],
            lid,
            id_to_cnj.get(lid),
            23,
            args.since,
        )
        for lid in lawsuits_total
        if lid in proc_map
    ]

    t0 = time.monotonic()
    counter = {"done": 0, "fail": 0, "raw": 0, "ach": 0}
    lock = threading.Lock()

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = [pool.submit(_worker_fetch, w) for w in work_items]
        for fut in as_completed(futures):
            lid, raw, ach, err = fut.result()
            with lock:
                if err:
                    counter["fail"] += 1
                else:
                    counter["done"] += 1
                    counter["raw"] += raw
                    counter["ach"] += ach
                if (counter["done"] + counter["fail"]) % 200 == 0:
                    dt = time.monotonic() - t0
                    n = counter["done"] + counter["fail"]
                    eta_min = (len(work_items) - n) * (dt / max(n, 1)) / 60
                    logger.info(
                        "Progresso: %s/%s (%s falhas, %s andamentos brutos, %s achados) — ETA %.1fmin",
                        n, len(work_items), counter["fail"],
                        counter["raw"], counter["ach"], eta_min,
                    )

    dt = time.monotonic() - t0
    logger.info(
        "Threading concluido em %.1fs (%.1fmin): %s OK, %s falhas, %s andamentos, %s achados",
        dt, dt / 60, counter["done"], counter["fail"],
        counter["raw"], counter["ach"],
    )

    # Atualiza run
    db = SessionLocal()
    try:
        run = db.query(VarreduraRun).filter(VarreduraRun.id == run_id).first()
        if run:
            run.status = RUN_STATUS_DONE if counter["fail"] == 0 else RUN_STATUS_DONE  # DONE mesmo com falhas parciais
            run.completed_at = datetime.now(timezone.utc)
            run.total_processados = counter["done"]
            run.total_falhas = counter["fail"]
            run.total_achados = counter["ach"]
            if counter["fail"] > 0:
                run.error_message = (
                    f"{counter['fail']} processos falharam na API L1 (ver varredura_processado)"
                )
            db.commit()
    finally:
        db.close()

    # Sync DuckDB
    if not args.no_duckdb_sync:
        try:
            from app.services.varredura.analytics_duckdb import (
                get_db_size_mb,
                sync_run,
            )
            res = sync_run(run_id, delete_after_sync=not args.no_delete_raw)
            size_mb = get_db_size_mb()
            logger.info(
                "Sync DuckDB run #%s: %s | duckdb=%.1fMB",
                run_id, res, size_mb,
            )
        except Exception:
            logger.exception("Falha sync DuckDB run #%s", run_id)
    else:
        logger.info("Sync DuckDB pulado (--no-duckdb-sync)")

    logger.info("=" * 60)
    logger.info("=== PIPELINE VIA API CONCLUIDO ===")
    logger.info("Run #%s — Tempo total: %.1fmin", run_id, dt / 60)
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
