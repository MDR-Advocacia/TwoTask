"""Consolida TODOS os lotes bb-temperatura-* (runs DONE) num XLSX unico.

Output: /tmp/bb-temperatura-CONSOLIDADO.xlsx
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("bb.consolidar")

CAPA_MAP_PATH = Path("/tmp/bb-capa-map.json")


def classificar_temperatura(tipos_evento: set[str]) -> tuple[str, str]:
    if "transito_julgado" in tipos_evento or "arquivamento" in tipos_evento:
        sinais = sorted(tipos_evento & {"transito_julgado", "arquivamento"})
        return "ALTA", f"sinal de encerramento: {', '.join(sinais)}"
    if "sentenca" in tipos_evento:
        extras = tipos_evento - {"sentenca"}
        if extras:
            return "MEDIA", f"sentença + {', '.join(sorted(extras))}"
        return "MEDIA", "sentença proferida"
    if {"audiencia_designada", "audiencia_cancelada", "revelia"} & tipos_evento:
        return "BAIXA", f"em andamento: {', '.join(sorted(tipos_evento))}"
    return "INDETERMINADO", "sem andamentos relevantes em 30d"


def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import (
        VarreduraAchado,
        VarreduraProcessado,
        VarreduraRun,
    )
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    capa_map = {}
    if CAPA_MAP_PATH.exists():
        capa_map = json.loads(CAPA_MAP_PATH.read_text(encoding="utf-8"))
    logger.info("Capa map: %s entries", len(capa_map))

    db = SessionLocal()
    try:
        runs = (
            db.query(VarreduraRun)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-%"))
            .filter(VarreduraRun.status == "DONE")
            .order_by(VarreduraRun.id)
            .all()
        )
        run_ids = [r.id for r in runs]
        logger.info("Runs DONE: %s (%s..%s)", len(runs), run_ids[0] if run_ids else "-", run_ids[-1] if run_ids else "-")

        processados = (
            db.query(VarreduraProcessado)
            .filter(VarreduraProcessado.run_id.in_(run_ids))
            .all()
        )
        achados = (
            db.query(VarreduraAchado)
            .filter(VarreduraAchado.run_id.in_(run_ids))
            .order_by(
                VarreduraAchado.andamento_data.desc().nullslast(),
                VarreduraAchado.id.desc(),
            )
            .all()
        )
        logger.info("Processados: %s · Achados: %s", len(processados), len(achados))

        # Dedup processados por lawsuit_id (caso algum apareça em 2 runs)
        proc_por_lid: dict[int, "VarreduraProcessado"] = {}
        for p in processados:
            proc_por_lid.setdefault(p.lawsuit_id, p)

        tipos_por_lid: dict[int, set[str]] = {}
        achados_por_lid: dict[int, list] = {}
        for a in achados:
            tipos_por_lid.setdefault(a.lawsuit_id, set()).add(a.tipo_evento)
            achados_por_lid.setdefault(a.lawsuit_id, []).append(a)

        rows = []
        contadores = {"ALTA": 0, "MEDIA": 0, "BAIXA": 0, "INDETERMINADO": 0}
        for lid, p in proc_por_lid.items():
            tipos = tipos_por_lid.get(lid, set())
            temp, just = classificar_temperatura(tipos)
            contadores[temp] += 1
            digits = "".join(ch for ch in (p.cnj_number or "") if ch.isdigit())
            capa = capa_map.get(digits) or {}
            rows.append({
                "lawsuit_id": lid,
                "cnj": p.cnj_number or capa.get("__cnj_original") or "",
                "temperatura": temp,
                "justificativa": just,
                "tipos_evento": ", ".join(sorted(tipos)),
                "qtd_achados": len(achados_por_lid.get(lid, [])),
                "run_id": p.run_id,
                "capa": capa,
                "achados": achados_por_lid.get(lid, []),
            })
        logger.info(
            "Distribuição: ALTA=%s, MEDIA=%s, BAIXA=%s, INDEF=%s",
            contadores["ALTA"], contadores["MEDIA"],
            contadores["BAIXA"], contadores["INDETERMINADO"],
        )

        wb = Workbook()
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="FFE5E7EB")
        color_alta = PatternFill("solid", fgColor="FFFEE2E2")
        color_media = PatternFill("solid", fgColor="FFFEF3C7")
        color_baixa = PatternFill("solid", fgColor="FFDCFCE7")
        color_indef = PatternFill("solid", fgColor="FFF3F4F6")

        # ── Aba Resumo ──
        ws = wb.active
        ws.title = "Resumo Geral"
        ws["A1"] = "Consolidação BB/Réu — Temperatura de Encerramento"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")
        info = [
            ("Total lotes consolidados", len(runs)),
            ("Range de runs", f"#{run_ids[0]} a #{run_ids[-1]}"),
            ("Janela varrida", "30 dias"),
            ("Total processos (únicos)", len(proc_por_lid)),
            ("Total achados", len(achados)),
            ("", ""),
            ("🔴 ALTA (apto ao encerramento)", contadores["ALTA"]),
            ("🟡 MÉDIA (sentença sem trânsito)", contadores["MEDIA"]),
            ("🟢 BAIXA (em andamento)", contadores["BAIXA"]),
            ("⚪ INDETERMINADO (sem sinais)", contadores["INDETERMINADO"]),
            ("", ""),
            ("% ALTA", f"{contadores['ALTA']/max(1,len(proc_por_lid))*100:.1f}%"),
            ("% MEDIA", f"{contadores['MEDIA']/max(1,len(proc_por_lid))*100:.1f}%"),
            ("% BAIXA", f"{contadores['BAIXA']/max(1,len(proc_por_lid))*100:.1f}%"),
            ("% INDETERMINADO", f"{contadores['INDETERMINADO']/max(1,len(proc_por_lid))*100:.1f}%"),
        ]
        for i, (k, v) in enumerate(info, start=3):
            ws.cell(row=i, column=1, value=k).font = bold if isinstance(v, str) and v == "" else None
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 30

        # Aba "Resumo por Lote" — pra rastreio
        ws_lt = wb.create_sheet("Resumo por Lote")
        headers_lt = ["Run #", "Lote", "Total proc", "Achados", "ALTA", "MEDIA", "BAIXA", "INDEF"]
        for c, h in enumerate(headers_lt, start=1):
            cell = ws_lt.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        # Conta por run
        for i, r in enumerate(runs, start=2):
            lids_run = [p.lawsuit_id for p in processados if p.run_id == r.id]
            cont_r = {"ALTA": 0, "MEDIA": 0, "BAIXA": 0, "INDETERMINADO": 0}
            for lid in lids_run:
                tipos = tipos_por_lid.get(lid, set())
                t, _ = classificar_temperatura(tipos)
                cont_r[t] += 1
            ws_lt.cell(row=i, column=1, value=r.id)
            ws_lt.cell(row=i, column=2, value=r.triggered_by)
            ws_lt.cell(row=i, column=3, value=r.total_processos)
            ws_lt.cell(row=i, column=4, value=r.total_achados)
            ws_lt.cell(row=i, column=5, value=cont_r["ALTA"])
            ws_lt.cell(row=i, column=6, value=cont_r["MEDIA"])
            ws_lt.cell(row=i, column=7, value=cont_r["BAIXA"])
            ws_lt.cell(row=i, column=8, value=cont_r["INDETERMINADO"])
        for c, w in enumerate([8, 32, 12, 10, 8, 8, 8, 10], start=1):
            ws_lt.column_dimensions[get_column_letter(c)].width = w
        ws_lt.freeze_panes = "A2"

        # ── Helper aba temperatura ──
        def _write_temp_sheet(sheet_name: str, subset: list, fill: PatternFill):
            ws_t = wb.create_sheet(sheet_name)
            cols = [
                "Temperatura", "CNJ", "Lawsuit ID", "NPJ", "Tipos de evento",
                "Justificativa", "Qtd achados",
                "Tipo Ação", "Situação", "UF", "Comarca", "Vara", "Valor da Causa",
                "Advogado", "Matéria", "Run #",
                "Datas achados", "Trechos detectados",
            ]
            for c, h in enumerate(cols, start=1):
                cell = ws_t.cell(row=1, column=c, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")
            for i, r in enumerate(subset, start=2):
                capa = r["capa"]
                ws_t.cell(row=i, column=1, value=r["temperatura"]).fill = fill
                ws_t.cell(row=i, column=2, value=r["cnj"])
                ws_t.cell(row=i, column=3, value=r["lawsuit_id"])
                ws_t.cell(row=i, column=4, value=str(capa.get("NPJ") or ""))
                ws_t.cell(row=i, column=5, value=r["tipos_evento"])
                ws_t.cell(row=i, column=6, value=r["justificativa"])
                ws_t.cell(row=i, column=7, value=r["qtd_achados"])
                ws_t.cell(row=i, column=8, value=str(capa.get("Tipo de Ação") or ""))
                ws_t.cell(row=i, column=9, value=str(capa.get("Situação do Processo") or ""))
                ws_t.cell(row=i, column=10, value=str(capa.get("UF") or ""))
                ws_t.cell(row=i, column=11, value=str(capa.get("Comarca") or ""))
                ws_t.cell(row=i, column=12, value=str(capa.get("Vara") or ""))
                ws_t.cell(row=i, column=13, value=str(capa.get("Valor da Causa") or ""))
                ws_t.cell(row=i, column=14, value=str(capa.get("Advogado") or ""))
                ws_t.cell(row=i, column=15, value=str(capa.get("Matéria") or ""))
                ws_t.cell(row=i, column=16, value=r["run_id"])
                datas = []
                trechos = []
                for a in r["achados"]:
                    d = a.andamento_data.strftime("%d/%m/%Y") if a.andamento_data else "—"
                    datas.append(f"{d}: {a.tipo_evento}")
                    trecho = (a.andamento_texto or "")[:300]
                    trechos.append(f"[{d}] {trecho}")
                ws_t.cell(row=i, column=17, value="\n".join(datas))
                tc = ws_t.cell(row=i, column=18, value="\n\n".join(trechos))
                tc.alignment = Alignment(wrap_text=True, vertical="top")
            widths = [16, 26, 12, 12, 40, 38, 10, 24, 14, 6, 22, 30, 14, 30, 22, 8, 30, 90]
            for c, w in enumerate(widths, start=1):
                ws_t.column_dimensions[get_column_letter(c)].width = w
            if subset:
                ws_t.freeze_panes = "A2"
                ws_t.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(subset) + 1}"

        rows_alta = sorted([r for r in rows if r["temperatura"] == "ALTA"], key=lambda x: -x["qtd_achados"])
        rows_media = sorted([r for r in rows if r["temperatura"] == "MEDIA"], key=lambda x: -x["qtd_achados"])
        rows_baixa = sorted([r for r in rows if r["temperatura"] == "BAIXA"], key=lambda x: -x["qtd_achados"])
        rows_indef = sorted([r for r in rows if r["temperatura"] == "INDETERMINADO"], key=lambda x: x["cnj"])
        _write_temp_sheet("🔴 ALTA (apto encerramento)", rows_alta, color_alta)
        _write_temp_sheet("🟡 MEDIA", rows_media, color_media)
        _write_temp_sheet("🟢 BAIXA", rows_baixa, color_baixa)
        _write_temp_sheet("⚪ INDETERMINADO", rows_indef, color_indef)

        out = Path("/tmp/bb-temperatura-CONSOLIDADO.xlsx")
        wb.save(out)
        logger.info("XLSX consolidado: %s (%s bytes)", out, out.stat().st_size)
    finally:
        db.close()


if __name__ == "__main__":
    main()
