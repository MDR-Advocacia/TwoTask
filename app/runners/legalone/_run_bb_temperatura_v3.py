"""V3 da varredura BB/Reu — janela 180d + sync automatico para DuckDB.

Diferencas vs V2:
  - Janela: 180d (vs 60d)
  - Captura TODOS os andamentos brutos no DB (var002 + capa_json)
  - Apos cada lote: sync varredura_andamento_raw -> DuckDB
    (output/analytics/varredura.duckdb) + apaga raw do PG (libera espaco)
  - XLSX por lote igual V2 (temperatura + delta)

Uso:
    docker cp local.xlsx onetask-api-1:/tmp/base-bb-v3.xlsx
    docker exec -d onetask-api-1 python //app/app/runners/legalone/_run_bb_temperatura_v3.py
"""

from __future__ import annotations

import json
import logging
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

INPUT_PATH_V3 = Path("/tmp/base-bb-v3.xlsx")
INPUT_PATH_V2 = Path("/tmp/base-bb-v2.xlsx")
CAPA_MAP_PATH = Path("/tmp/bb-v3-capa-map.json")
RESOLVE_CACHE = Path("/tmp/bb-v3-cnj-resolve.json")
LOG_PATH = Path(
    "/app/output/playwright/legalone/varredura-andamentos/bb-temperatura-v3.log"
)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("bb.temperatura.v3")

WINDOW_DAYS = 180
BATCH_SIZE = 500

EXCLUIR_TIPO_ACAO = {
    "EMBARGOS A EXECUCAO",
    "EMBARGOS DE TERCEIRO",
    "CUMPRIMENTO DA SENTENCA",
    "ALVARA JUDICIAL",
    "CAUTELAR",
    "ADJUDICACAO",
}


def _input_path() -> Path:
    if INPUT_PATH_V3.exists():
        return INPUT_PATH_V3
    if INPUT_PATH_V2.exists():
        logger.info("Reusando base v2 (base-bb-v2.xlsx) — sem v3 dedicada.")
        return INPUT_PATH_V2
    raise FileNotFoundError("Nenhuma base BB encontrada em /tmp/base-bb-*.xlsx")


def ler_e_filtrar_planilha() -> tuple[list[str], dict[str, dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(_input_path(), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    cnjs: list[str] = []
    mapa: dict[str, dict] = {}
    seen: set[str] = set()
    total = excl_nao_reu = excl_tipo = excl_sinopse = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        total += 1
        polo = str(row[17] or "")
        tipo = str(row[13] or "").strip().upper()
        sinopse = str(row[36] or "").strip()
        cnj_raw = row[2]
        if polo != "REU":
            excl_nao_reu += 1
            continue
        if "incidental" in sinopse.lower():
            excl_sinopse += 1
            continue
        if tipo in EXCLUIR_TIPO_ACAO:
            excl_tipo += 1
            continue
        digits = "".join(ch for ch in str(cnj_raw) if ch.isdigit())
        if len(digits) < 15 or digits in seen:
            continue
        seen.add(digits)
        capa = {
            headers[i]: row[i] for i in range(min(len(headers), len(row)))
        }
        capa["__cnj_original"] = str(cnj_raw).strip()
        mapa[digits] = capa
        cnjs.append(digits)
    logger.info(
        "Planilha v3: total=%s, excl_nao_reu=%s, excl_tipo=%s, excl_sinopse=%s, mantidos=%s",
        total, excl_nao_reu, excl_tipo, excl_sinopse, len(cnjs),
    )
    return cnjs, mapa


def resolver_cnjs(cnjs: list[str]) -> dict[str, int]:
    from app.services.legal_one_client import LegalOneApiClient

    cache: dict[str, int] = {}
    if RESOLVE_CACHE.exists():
        try:
            cache = json.loads(RESOLVE_CACHE.read_text(encoding="utf-8"))
            cache = {k: int(v) for k, v in cache.items()}
            logger.info("Cache resolve v3: %s entries", len(cache))
        except Exception:
            cache = {}

    # Tenta importar cache da V2 (CNJs ja resolvidos antes)
    v2_cache = Path("/tmp/bb-v2-cnj-resolve.json")
    if v2_cache.exists():
        try:
            extra = json.loads(v2_cache.read_text(encoding="utf-8"))
            extra = {k: int(v) for k, v in extra.items()}
            for k, v in extra.items():
                cache.setdefault(k, v)
            logger.info("Cache v2 importado: %s entries adicionais", len(extra))
        except Exception:
            pass

    pendentes = [c for c in cnjs if c not in cache]
    if pendentes:
        client = LegalOneApiClient()
        logger.info("Resolvendo %s CNJs via L1 (v3)...", len(pendentes))
        matches = client.search_lawsuits_by_cnj_numbers(pendentes)
        for cnj_norm in pendentes:
            payload = None
            for k, v in matches.items():
                if "".join(ch for ch in str(k) if ch.isdigit()) == cnj_norm:
                    payload = v
                    break
            if payload is not None and payload.get("id") is not None:
                try:
                    cache[cnj_norm] = int(payload["id"])
                except (TypeError, ValueError):
                    pass
        RESOLVE_CACHE.write_text(json.dumps(cache), encoding="utf-8")

    resolvidos = {c: cache[c] for c in cnjs if c in cache}
    logger.info("Resolvidos v3: %s / %s", len(resolvidos), len(cnjs))
    return resolvidos


def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import (
        QUEUE_STATUS_PENDING,
        RUN_STATUS_RUNNING,
        VarreduraProcessado,
        VarreduraRun,
    )
    from app.services.varredura.analytics_duckdb import (
        get_db_size_mb,
        sync_run,
    )
    from app.services.varredura.varredura_service import (
        _run_subprocess_worker_impl,
    )

    logger.info("=" * 60)
    logger.info("BB/Reu V3 — TEMPERATURA + DuckDB (janela %sd)", WINDOW_DAYS)
    logger.info("=" * 60)

    cnjs, capa_map = ler_e_filtrar_planilha()
    cnj_to_id = resolver_cnjs(cnjs)
    CAPA_MAP_PATH.write_text(json.dumps(capa_map, ensure_ascii=False, default=str), encoding="utf-8")

    id_to_cnj = {v: k for k, v in cnj_to_id.items()}

    PRIORIDADE = {"Cumprimento": 0, "Sentença": 1, "Recurso": 2, "Inicial": 3}

    def _prio(lid: int) -> tuple[int, int]:
        digits = id_to_cnj.get(lid)
        capa = capa_map.get(digits) if digits else None
        sit = str((capa or {}).get("Situação do Processo") or "").strip()
        return (PRIORIDADE.get(sit, 99), lid)

    lawsuits_total = sorted(set(cnj_to_id.values()), key=_prio)
    logger.info("Total lawsuit_ids resolvidos: %s", len(lawsuits_total))

    # Skip lawsuits ja varridos nesta V3 (runs anteriores)
    db = SessionLocal()
    try:
        runs_anteriores = (
            db.query(VarreduraRun.id)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-v3-%"))
            .all()
        )
        ja_varridos: set[int] = set()
        for (rid,) in runs_anteriores:
            ids = (
                db.query(VarreduraProcessado.lawsuit_id)
                .filter(VarreduraProcessado.run_id == rid)
                .all()
            )
            ja_varridos.update(r[0] for r in ids)
    finally:
        db.close()
    logger.info("Ja varridos em runs V3 anteriores: %s", len(ja_varridos))

    restante = [lid for lid in lawsuits_total if lid not in ja_varridos]
    total_lotes = (len(restante) + BATCH_SIZE - 1) // BATCH_SIZE
    logger.info(
        "Restante pra varrer: %s — em %s lotes de %s",
        len(restante), total_lotes, BATCH_SIZE,
    )

    for idx in range(total_lotes):
        chunk = restante[idx * BATCH_SIZE : (idx + 1) * BATCH_SIZE]
        if not chunk:
            break
        logger.info("─── LOTE V3 %s/%s — %s processos ───", idx + 1, total_lotes, len(chunk))

        db = SessionLocal()
        try:
            run = VarreduraRun(
                status=RUN_STATUS_RUNNING,
                started_at=datetime.now(timezone.utc),
                responsible_office_ids=[23],
                window_days=WINDOW_DAYS,
                triggered_by=f"bb-temperatura-v3-lote-{idx + 1}-de-{total_lotes}",
                total_processos=len(chunk),
            )
            db.add(run)
            db.flush()
            for lid in chunk:
                digits = id_to_cnj.get(lid)
                capa = capa_map.get(digits) if digits else None
                cnj_orig = (capa or {}).get("__cnj_original") if capa else None
                db.add(
                    VarreduraProcessado(
                        run_id=run.id,
                        lawsuit_id=lid,
                        cnj_number=cnj_orig or digits,
                        office_id=23,
                        queue_status=QUEUE_STATUS_PENDING,
                        capa_json=capa,  # ← Salva capa direto no processado
                    )
                )
            db.commit()
            run_id = run.id
            logger.info("Run #%s criada (com capa_json). Disparando subprocess...", run_id)

            try:
                _run_subprocess_worker_impl(db, run_id)
            except Exception as exc:
                logger.exception("Erro no worker run #%s: %s", run_id, exc)

            # Sync DuckDB + delete raw do PG
            try:
                res = sync_run(run_id, delete_after_sync=True)
                size_mb = get_db_size_mb()
                logger.info(
                    "Sync DuckDB run #%s: %s | duckdb=%.1fMB",
                    run_id, res, size_mb,
                )
            except Exception:
                logger.exception("Falha sync DuckDB lote run #%s", run_id)
        finally:
            db.close()

        time.sleep(10)

    logger.info("=" * 60)
    logger.info("=== PIPELINE BB/Reu V3 CONCLUIDO ===")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
