"""Consolida TODOS os achados de varredura num unico XLSX + classifica
revelias via Anthropic Sonnet (efetiva vs mencao passageira).

Output:
    /tmp/varredura-consolidado-master.xlsx
"""

from __future__ import annotations

import json
import logging
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("varredura.consolidar")


SYSTEM_PROMPT_REVELIA = """Voce e' analista juridico do escritorio MDR
Advocacia (Banco Master). Sua tarefa: classificar trechos de andamentos
processuais quanto a aplicacao DA REVELIA.

REGRA DE OURO: marque "EFETIVA" SOMENTE quando o juiz EFETIVAMENTE
DECRETOU/RECONHECEU/APLICOU a revelia do reu — i.e., a parte ficou
formalmente revel no processo, com os efeitos materiais e processuais
respectivos (presuncao de veracidade dos fatos, dispensa de intimacao
para atos subsequentes etc.).

Marque "MENCAO" quando:
- O termo "revelia" aparece de forma incidental/abstrata (citacao
  doutrinaria, peca da parte argumentando contra a revelia, recurso
  discutindo decisao anterior, mencao em ata de audiencia sem decreto,
  certidao informativa que apenas FALA sobre o instituto, etc.).
- Adverte sobre revelia futura ("sob pena de revelia") sem ela ter
  acontecido.
- A parte protestou contra a revelia ou pediu sua nao-aplicacao.
- Mencao em sentenca/decisao que NEGA a revelia, ou reconhece que
  nao houve revelia.

Marque "INCONCLUSIVO" se o texto for tao truncado/ambiguo que voce
nao consegue afirmar com seguranca.

Output: JSON ESTRITO no formato:
{
  "classificacoes": [
    {"id": 123, "classificacao": "EFETIVA"|"MENCAO"|"INCONCLUSIVO", "justificativa": "1 frase curta"},
    ...
  ]
}

NAO adicione texto fora do JSON. NAO use markdown. Resposta = APENAS o JSON.
"""


def call_anthropic_classify(items: list[dict]) -> list[dict]:
    """Manda batch de revelias pro Sonnet e retorna classificacoes."""
    import httpx
    from app.core.config import settings

    api_key = settings.anthropic_api_key
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY nao configurada")

    user_msg_lines = [
        "Classifique cada andamento abaixo. Output APENAS JSON.",
        "",
    ]
    for it in items:
        user_msg_lines.append(f"--- id={it['id']} ---")
        user_msg_lines.append(f"Data: {it.get('data') or '-'}")
        user_msg_lines.append(f"Tipo: {it.get('tipo') or '-'}")
        user_msg_lines.append(
            f"Movimentado por: {it.get('movimentado_por') or '-'}"
        )
        user_msg_lines.append(f"Texto: {it['texto']}")
        user_msg_lines.append("")

    user_msg = "\n".join(user_msg_lines)

    payload = {
        "model": settings.classifier_model or "claude-sonnet-4-5-20250929",
        "max_tokens": 8000,
        "temperature": 0,
        "system": SYSTEM_PROMPT_REVELIA,
        "messages": [{"role": "user", "content": user_msg}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }

    with httpx.Client(timeout=180.0) as client:
        resp = client.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json=payload,
        )
    if resp.status_code != 200:
        raise RuntimeError(
            f"Anthropic HTTP {resp.status_code}: {resp.text[:500]}"
        )
    data = resp.json()
    text = (data.get("content", [{}])[0].get("text") or "").strip()
    # Tenta parsear JSON. Se vier com markdown fence, remove.
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
        text = text.strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        logger.error("Falha parseando JSON do Sonnet: %s", exc)
        logger.error("Texto retornado: %s", text[:500])
        raise
    return parsed.get("classificacoes", [])


def classify_revelias(achados: list) -> dict[int, dict]:
    """Classifica revelias via Sonnet em batches de 30. Retorna mapa
    {achado_id: {'classificacao': 'EFETIVA'|'MENCAO'|'INCONCLUSIVO',
                 'justificativa': '...'}}."""
    cache_path = Path("/tmp/revelia-classificacoes.json")
    cache: dict = {}
    if cache_path.exists():
        try:
            cache = json.loads(cache_path.read_text(encoding="utf-8"))
            cache = {int(k): v for k, v in cache.items()}
            logger.info("Cache de classificacoes: %s entries", len(cache))
        except Exception:
            cache = {}

    pendentes = [a for a in achados if a.id not in cache]
    if not pendentes:
        logger.info("Todas %s revelias ja' estao no cache.", len(achados))
        return cache

    BATCH = 30
    for i in range(0, len(pendentes), BATCH):
        chunk = pendentes[i : i + BATCH]
        logger.info(
            "Classificando batch %s/%s (%s revelias)...",
            i // BATCH + 1,
            (len(pendentes) + BATCH - 1) // BATCH,
            len(chunk),
        )
        items_payload = [
            {
                "id": a.id,
                "data": (
                    a.andamento_data.strftime("%d/%m/%Y")
                    if a.andamento_data
                    else None
                ),
                "tipo": a.andamento_tipo,
                "movimentado_por": a.andamento_movimentado_por,
                "texto": (a.andamento_texto or "")[:2500],
            }
            for a in chunk
        ]
        try:
            results = call_anthropic_classify(items_payload)
        except Exception as exc:
            logger.exception("Falha no batch: %s — tentando individual", exc)
            results = []
            for it in items_payload:
                try:
                    r = call_anthropic_classify([it])
                    results.extend(r)
                except Exception:
                    cache[it["id"]] = {
                        "classificacao": "INCONCLUSIVO",
                        "justificativa": "Falha na chamada do classificador.",
                    }
        for r in results:
            aid = int(r.get("id") or 0)
            if aid:
                cache[aid] = {
                    "classificacao": (
                        r.get("classificacao") or "INCONCLUSIVO"
                    ).upper(),
                    "justificativa": (r.get("justificativa") or "").strip()[
                        :500
                    ],
                }
        # Salva cache parcial a cada batch
        cache_path.write_text(
            json.dumps(
                {str(k): v for k, v in cache.items()},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

    return cache


_EVENTO_LABEL = {
    "audiencia_designada": "Audiência designada",
    "audiencia_cancelada": "Audiência cancelada",
    "sentenca": "Sentença",
    "revelia": "Revelia",
    "transito_julgado": "Trânsito em julgado",
    "arquivamento": "Arquivamento",
}


def gerar_xlsx_consolidado(
    db,
    output_path: Path,
    revelia_class: dict[int, dict],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlalchemy import func as sa_func

    from app.models.varredura import VarreduraAchado, VarreduraRun

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid",
    )

    runs = (
        db.query(VarreduraRun)
        .filter(VarreduraRun.id <= 15)
        .order_by(VarreduraRun.id)
        .all()
    )
    achados = (
        db.query(VarreduraAchado)
        .filter(VarreduraAchado.run_id <= 15)
        .order_by(
            VarreduraAchado.andamento_data.desc().nullslast(),
            VarreduraAchado.id.desc(),
        )
        .all()
    )

    # ── Aba 1: Resumo geral ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo Geral"
    ws["A1"] = "Varredura de Andamentos — Consolidado Carteira Banco Master"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    total_proc = sum(r.total_processos for r in runs)
    total_ach = len(achados)
    revelias = [a for a in achados if a.tipo_evento == "revelia"]
    revelia_efetiva_qtd = sum(
        1
        for a in revelias
        if revelia_class.get(a.id, {}).get("classificacao") == "EFETIVA"
    )
    revelia_mencao_qtd = sum(
        1
        for a in revelias
        if revelia_class.get(a.id, {}).get("classificacao") == "MENCAO"
    )
    revelia_incl_qtd = len(revelias) - revelia_efetiva_qtd - revelia_mencao_qtd

    rows = [
        ("Total de processos varridos", total_proc),
        ("Total de varreduras (runs)", len(runs)),
        ("Total de achados", total_ach),
        ("", ""),
        ("Achados por tipo:", ""),
    ]
    by_type: dict[str, int] = {}
    for a in achados:
        by_type[a.tipo_evento] = by_type.get(a.tipo_evento, 0) + 1
    for k, v in sorted(by_type.items(), key=lambda x: -x[1]):
        rows.append((f"  {_EVENTO_LABEL.get(k, k)}", v))
    rows.append(("", ""))
    rows.append(("Revelias — revisao via IA:", ""))
    rows.append(("  EFETIVA (revelia aplicada)", revelia_efetiva_qtd))
    rows.append(("  MENCAO (citacao passageira)", revelia_mencao_qtd))
    rows.append(("  INCONCLUSIVO", revelia_incl_qtd))

    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = bold if v == "" else None
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14

    # ── Aba 2: Resumo por lote/run ───────────────────────────────────
    ws_l = wb.create_sheet("Resumo por Lote")
    headers = [
        "Run #",
        "Lote",
        "Iniciada",
        "Concluída",
        "Total processos",
        "Achados",
        "Falhas",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws_l.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
    for i, r in enumerate(runs, start=2):
        ws_l.cell(row=i, column=1, value=r.id)
        ws_l.cell(row=i, column=2, value=r.triggered_by or "")
        ws_l.cell(
            row=i,
            column=3,
            value=r.started_at.strftime("%d/%m/%Y %H:%M") if r.started_at else "",
        )
        ws_l.cell(
            row=i,
            column=4,
            value=(
                r.completed_at.strftime("%d/%m/%Y %H:%M") if r.completed_at else ""
            ),
        )
        ws_l.cell(row=i, column=5, value=r.total_processos)
        ws_l.cell(row=i, column=6, value=r.total_achados)
        ws_l.cell(row=i, column=7, value=r.total_falhas)
    widths = [8, 30, 18, 18, 16, 10, 10]
    for c, w in enumerate(widths, start=1):
        ws_l.column_dimensions[get_column_letter(c)].width = w
    ws_l.freeze_panes = "A2"

    # ── Aba 3: TODOS os achados ──────────────────────────────────────
    def _write_achados_sheet(
        ws_a,
        achados_subset: list,
        include_revelia_class: bool = False,
    ):
        cols = [
            "Run #",
            "CNJ",
            "Lawsuit ID",
            "Data",
            "Hora",
            "Tipo evento",
            "Trecho (regex)",
            "Movimentado por",
            "Texto completo",
        ]
        if include_revelia_class:
            cols.extend(["Classificação IA", "Justificativa"])
        for c, h in enumerate(cols, start=1):
            cell = ws_a.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for i, a in enumerate(achados_subset, start=2):
            ws_a.cell(row=i, column=1, value=a.run_id)
            ws_a.cell(row=i, column=2, value=a.cnj_number or "")
            ws_a.cell(row=i, column=3, value=a.lawsuit_id)
            ws_a.cell(
                row=i,
                column=4,
                value=(
                    a.andamento_data.strftime("%d/%m/%Y")
                    if a.andamento_data
                    else ""
                ),
            )
            ws_a.cell(row=i, column=5, value=a.andamento_hora or "")
            ws_a.cell(
                row=i,
                column=6,
                value=_EVENTO_LABEL.get(a.tipo_evento, a.tipo_evento),
            )
            ws_a.cell(row=i, column=7, value=a.regex_matched or "")
            ws_a.cell(row=i, column=8, value=a.andamento_movimentado_por or "")
            cell_txt = ws_a.cell(row=i, column=9, value=a.andamento_texto or "")
            cell_txt.alignment = Alignment(wrap_text=True, vertical="top")
            if include_revelia_class:
                cls = revelia_class.get(a.id, {})
                ws_a.cell(row=i, column=10, value=cls.get("classificacao") or "")
                ws_a.cell(row=i, column=11, value=cls.get("justificativa") or "")
        base_widths = [8, 28, 12, 12, 8, 22, 32, 26, 70]
        if include_revelia_class:
            base_widths.extend([18, 60])
        for c, w in enumerate(base_widths, start=1):
            ws_a.column_dimensions[get_column_letter(c)].width = w
        ws_a.freeze_panes = "A2"
        if achados_subset:
            ws_a.auto_filter.ref = (
                f"A1:{get_column_letter(len(cols))}{len(achados_subset) + 1}"
            )

    ws_all = wb.create_sheet("Todos os Achados")
    _write_achados_sheet(ws_all, achados)

    # Abas por tipo
    for tipo, label in _EVENTO_LABEL.items():
        subset = [a for a in achados if a.tipo_evento == tipo]
        if not subset:
            continue
        sheet_name = label[:31]
        ws_t = wb.create_sheet(sheet_name)
        is_rev = tipo == "revelia"
        _write_achados_sheet(ws_t, subset, include_revelia_class=is_rev)

    # ── Aba especial: Revelias EFETIVAS ──────────────────────────────
    rev_efetivas = [
        a
        for a in achados
        if a.tipo_evento == "revelia"
        and revelia_class.get(a.id, {}).get("classificacao") == "EFETIVA"
    ]
    ws_re = wb.create_sheet("Revelias EFETIVAS")
    _write_achados_sheet(ws_re, rev_efetivas, include_revelia_class=True)

    wb.save(output_path)


def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import VarreduraAchado

    db = SessionLocal()
    try:
        revelias = (
            db.query(VarreduraAchado)
            .filter(VarreduraAchado.tipo_evento == "revelia")
            .filter(VarreduraAchado.run_id <= 15)
            .all()
        )
        logger.info("Encontradas %s revelias. Classificando via Sonnet...", len(revelias))
        revelia_class = classify_revelias(revelias)
        logger.info(
            "Classificacoes prontas. EFETIVAS=%s, MENCAO=%s, INCONCLUSIVO=%s",
            sum(1 for v in revelia_class.values() if v["classificacao"] == "EFETIVA"),
            sum(1 for v in revelia_class.values() if v["classificacao"] == "MENCAO"),
            sum(
                1
                for v in revelia_class.values()
                if v["classificacao"] == "INCONCLUSIVO"
            ),
        )

        output = Path("/tmp/varredura-consolidado-master.xlsx")
        gerar_xlsx_consolidado(db, output, revelia_class)
        size = output.stat().st_size
        logger.info("XLSX consolidado gerado: %s (%s bytes)", output, size)
    finally:
        db.close()


if __name__ == "__main__":
    main()
