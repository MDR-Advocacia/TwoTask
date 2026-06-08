"""Consolida runs bb-temperatura-v2-* num XLSX unico + delta vs V1.

Output: /tmp/bb-temperatura-v2-CONSOLIDADO.xlsx
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", stream=sys.stdout)
logger = logging.getLogger("bb.consolidar.v2")

CAPA_MAP_PATH = Path("/tmp/bb-v2-capa-map.json")
LABEL_TEMP = {
    "ENCERRAMENTO_IMINENTE": "🔴 ENCERRAMENTO IMINENTE",
    "AGUARDANDO_CUMPRIMENTO": "🟠 AGUARDANDO CUMPRIMENTO",
    "MEDIA": "🟡 MEDIA",
    "BAIXA": "🟢 BAIXA",
    "INDETERMINADO": "⚪ INDETERMINADO",
}


import re as _re
_ILLEGAL_XLSX_RE = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(v):
    """Remove caracteres de controle nao permitidos em XLSX."""
    if v is None:
        return v
    if isinstance(v, str):
        return _ILLEGAL_XLSX_RE.sub("", v)
    return v


def classificar_temperatura(tipos: set[str]) -> tuple[str, str]:
    if "cumprimento_extinto" in tipos:
        return "ENCERRAMENTO_IMINENTE", "cumprimento extinto/satisfeito"
    if "cumprimento_iniciado" in tipos:
        return "AGUARDANDO_CUMPRIMENTO", "sinais de cumprimento ativos (sem extincao)"
    if "transito_julgado" in tipos or "arquivamento" in tipos:
        sinais = sorted(tipos & {"transito_julgado", "arquivamento"})
        return "ENCERRAMENTO_IMINENTE", f"sinal de encerramento: {', '.join(sinais)}"
    if "sentenca" in tipos:
        extras = tipos - {"sentenca"}
        if extras:
            return "MEDIA", f"sentenca + {', '.join(sorted(extras))}"
        return "MEDIA", "sentenca proferida"
    if {"audiencia_designada", "audiencia_cancelada", "revelia"} & tipos:
        return "BAIXA", f"em andamento: {', '.join(sorted(tipos))}"
    return "INDETERMINADO", "sem andamentos relevantes em 60d"


def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import VarreduraAchado, VarreduraProcessado, VarreduraRun
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    capa_map = json.loads(CAPA_MAP_PATH.read_text(encoding="utf-8")) if CAPA_MAP_PATH.exists() else {}
    logger.info("Capa map: %s entries", len(capa_map))

    db = SessionLocal()
    try:
        # Runs V2 DONE
        runs_v2 = (
            db.query(VarreduraRun)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-v2-%"))
            .filter(VarreduraRun.status == "DONE")
            .order_by(VarreduraRun.id)
            .all()
        )
        v2_ids = [r.id for r in runs_v2]
        logger.info("Runs V2 DONE: %s (#%s..#%s)", len(runs_v2), v2_ids[0], v2_ids[-1])

        proc_v2 = db.query(VarreduraProcessado).filter(VarreduraProcessado.run_id.in_(v2_ids)).all()
        ach_v2 = (
            db.query(VarreduraAchado)
            .filter(VarreduraAchado.run_id.in_(v2_ids))
            .order_by(VarreduraAchado.andamento_data.desc().nullslast(), VarreduraAchado.id.desc())
            .all()
        )
        logger.info("V2: %s proc, %s achados", len(proc_v2), len(ach_v2))

        # Dedup processados por lawsuit_id
        proc_por_lid_v2 = {}
        for p in proc_v2:
            proc_por_lid_v2.setdefault(p.lawsuit_id, p)
        tipos_v2 = {}
        ach_por_lid_v2 = {}
        for a in ach_v2:
            tipos_v2.setdefault(a.lawsuit_id, set()).add(a.tipo_evento)
            ach_por_lid_v2.setdefault(a.lawsuit_id, []).append(a)

        # Runs V1 (carteira anterior) pra delta
        runs_v1 = (
            db.query(VarreduraRun)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-%"))
            .filter(~VarreduraRun.triggered_by.like("bb-temperatura-v2-%"))
            .filter(VarreduraRun.status == "DONE")
            .all()
        )
        v1_ids = [r.id for r in runs_v1]
        tipos_v1 = {}
        if v1_ids:
            for a in db.query(VarreduraAchado).filter(VarreduraAchado.run_id.in_(v1_ids)):
                tipos_v1.setdefault(a.lawsuit_id, set()).add(a.tipo_evento)
            # garante que processados sem achados também estão no mapa V1
            for p in db.query(VarreduraProcessado).filter(VarreduraProcessado.run_id.in_(v1_ids)):
                tipos_v1.setdefault(p.lawsuit_id, set())
        logger.info("V1: %s lawsuits indexados", len(tipos_v1))

        # Classifica V1 com mesma regra V2 pra comparacao justa
        temp_v1_por_lid = {lid: classificar_temperatura(t)[0] for lid, t in tipos_v1.items()}

        rows = []
        contadores = {k: 0 for k in LABEL_TEMP}
        delta_novo = 0
        delta_mudou = 0
        delta_manteve = 0
        for lid, p in proc_por_lid_v2.items():
            tipos = tipos_v2.get(lid, set())
            temp, just = classificar_temperatura(tipos)
            contadores[temp] += 1
            digits = "".join(ch for ch in (p.cnj_number or "") if ch.isdigit())
            capa = capa_map.get(digits) or {}
            t_v1 = temp_v1_por_lid.get(lid)
            if t_v1 is None:
                delta = "NOVO"
                delta_novo += 1
            elif t_v1 != temp:
                delta = f"MUDOU ({LABEL_TEMP.get(t_v1, t_v1)} → {LABEL_TEMP.get(temp, temp)})"
                delta_mudou += 1
            else:
                delta = "MANTEVE"
                delta_manteve += 1
            rows.append({
                "lawsuit_id": lid,
                "cnj": p.cnj_number or capa.get("__cnj_original") or "",
                "temperatura": temp,
                "justificativa": just,
                "tipos_evento": ", ".join(sorted(tipos)),
                "qtd_achados": len(ach_por_lid_v2.get(lid, [])),
                "run_id": p.run_id,
                "delta": delta,
                "temp_v1": t_v1,
                "tipos_v1": tipos_v1.get(lid),
                "capa": capa,
                "achados": ach_por_lid_v2.get(lid, []),
            })
        logger.info("Distribuição V2: %s", {k: v for k, v in contadores.items() if v > 0})
        logger.info("Delta: NOVO=%s, MUDOU=%s, MANTEVE=%s", delta_novo, delta_mudou, delta_manteve)

        # Workbook
        wb = Workbook()
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="FFE5E7EB")
        cores = {
            "ENCERRAMENTO_IMINENTE": PatternFill("solid", fgColor="FFFEE2E2"),
            "AGUARDANDO_CUMPRIMENTO": PatternFill("solid", fgColor="FFFFEDD5"),
            "MEDIA": PatternFill("solid", fgColor="FFFEF3C7"),
            "BAIXA": PatternFill("solid", fgColor="FFDCFCE7"),
            "INDETERMINADO": PatternFill("solid", fgColor="FFF3F4F6"),
        }

        # ── Aba Resumo ──
        ws = wb.active
        ws.title = "Resumo Geral"
        ws["A1"] = "Consolidação BB/Réu V2 — Temperatura de Encerramento (janela 60d)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")
        info = [
            ("Total lotes V2", len(runs_v2)),
            ("Range de runs V2", f"#{v2_ids[0]} a #{v2_ids[-1]}"),
            ("Janela varrida", "60 dias"),
            ("Total processos (únicos)", len(proc_por_lid_v2)),
            ("Total achados V2", len(ach_v2)),
            ("Total achados V1 (referencia)", sum(len(t) for t in tipos_v1.values())),
            ("", ""),
            ("🔴 ENCERRAMENTO IMINENTE", contadores["ENCERRAMENTO_IMINENTE"]),
            ("🟠 AGUARDANDO CUMPRIMENTO", contadores["AGUARDANDO_CUMPRIMENTO"]),
            ("🟡 MEDIA (sentença sem trânsito)", contadores["MEDIA"]),
            ("🟢 BAIXA (em andamento)", contadores["BAIXA"]),
            ("⚪ INDETERMINADO", contadores["INDETERMINADO"]),
            ("", ""),
            ("DELTA vs V1", ""),
            ("  🆕 NOVOS (não estavam V1)", delta_novo),
            ("  🔄 MUDARAM temperatura", delta_mudou),
            ("  MANTIVERAM", delta_manteve),
        ]
        for i, (k, v) in enumerate(info, start=3):
            ws.cell(row=i, column=1, value=k).font = bold if isinstance(v, str) and v == "" else None
            ws.cell(row=i, column=2, value=v)
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 36

        # Helper aba temperatura
        def _write(sheet_name: str, subset: list, fill: PatternFill):
            ws_t = wb.create_sheet(sheet_name)
            cols = [
                "Temperatura", "CNJ", "Lawsuit ID", "NPJ", "Tipos de evento",
                "Justificativa", "Qtd achados", "Delta vs V1", "Temp V1",
                "Tipo Ação", "Situação", "UF", "Comarca", "Vara", "Valor da Causa",
                "Advogado", "Matéria", "Run #",
                "Datas achados", "Trechos detectados",
            ]
            for c, h in enumerate(cols, start=1):
                cell = ws_t.cell(row=1, column=c, value=h)
                cell.font = bold
                cell.fill = header_fill
            for i, r in enumerate(subset, start=2):
                capa = r["capa"]
                ws_t.cell(row=i, column=1, value=LABEL_TEMP.get(r["temperatura"], r["temperatura"])).fill = fill
                ws_t.cell(row=i, column=2, value=_clean(r["cnj"]))
                ws_t.cell(row=i, column=3, value=r["lawsuit_id"])
                ws_t.cell(row=i, column=4, value=_clean(str(capa.get("NPJ") or "")))
                ws_t.cell(row=i, column=5, value=_clean(r["tipos_evento"]))
                ws_t.cell(row=i, column=6, value=_clean(r["justificativa"]))
                ws_t.cell(row=i, column=7, value=r["qtd_achados"])
                ws_t.cell(row=i, column=8, value=_clean(r["delta"]))
                ws_t.cell(row=i, column=9, value=LABEL_TEMP.get(r["temp_v1"], r["temp_v1"]) if r["temp_v1"] else "—")
                ws_t.cell(row=i, column=10, value=_clean(str(capa.get("Tipo de Ação") or "")))
                ws_t.cell(row=i, column=11, value=_clean(str(capa.get("Situação do Processo") or "")))
                ws_t.cell(row=i, column=12, value=_clean(str(capa.get("UF") or "")))
                ws_t.cell(row=i, column=13, value=_clean(str(capa.get("Comarca") or "")))
                ws_t.cell(row=i, column=14, value=_clean(str(capa.get("Vara") or "")))
                ws_t.cell(row=i, column=15, value=_clean(str(capa.get("Valor da Causa") or "")))
                ws_t.cell(row=i, column=16, value=_clean(str(capa.get("Advogado") or "")))
                ws_t.cell(row=i, column=17, value=_clean(str(capa.get("Matéria") or "")))
                ws_t.cell(row=i, column=18, value=r["run_id"])
                datas, trechos = [], []
                for a in r["achados"]:
                    d = a.andamento_data.strftime("%d/%m/%Y") if a.andamento_data else "—"
                    datas.append(f"{d}: {a.tipo_evento}")
                    trechos.append(f"[{d}] {_clean((a.andamento_texto or ''))[:300]}")
                ws_t.cell(row=i, column=19, value=_clean("\n".join(datas)))
                tc = ws_t.cell(row=i, column=20, value=_clean("\n\n".join(trechos)))
                tc.alignment = Alignment(wrap_text=True, vertical="top")
            widths = [22, 26, 12, 12, 40, 38, 10, 40, 22, 24, 14, 6, 22, 30, 14, 30, 22, 8, 30, 90]
            for c, w in enumerate(widths, start=1):
                ws_t.column_dimensions[get_column_letter(c)].width = w
            if subset:
                ws_t.freeze_panes = "A2"
                ws_t.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(subset) + 1}"

        # Ordena por qtd_achados desc nas abas quentes
        def _sub(temp):
            return sorted([r for r in rows if r["temperatura"] == temp], key=lambda x: -x["qtd_achados"])

        _write("🔴 ENCERRAMENTO IMINENTE", _sub("ENCERRAMENTO_IMINENTE"), cores["ENCERRAMENTO_IMINENTE"])
        _write("🟠 AGUARDANDO CUMPRIMENTO", _sub("AGUARDANDO_CUMPRIMENTO"), cores["AGUARDANDO_CUMPRIMENTO"])
        _write("🟡 MEDIA", _sub("MEDIA"), cores["MEDIA"])
        _write("🟢 BAIXA", _sub("BAIXA"), cores["BAIXA"])
        _write("⚪ INDETERMINADO", _sub("INDETERMINADO"), cores["INDETERMINADO"])

        # Delta abas
        novos = sorted([r for r in rows if r["delta"] == "NOVO"], key=lambda x: -x["qtd_achados"])
        mudaram = sorted([r for r in rows if r["delta"].startswith("MUDOU")], key=lambda x: -x["qtd_achados"])
        if novos:
            _write("🆕 NOVOS (vs V1)", novos, header_fill)
        if mudaram:
            _write("🔄 MUDARAM (V1→V2)", mudaram, header_fill)

        out = Path("/tmp/bb-temperatura-v2-CONSOLIDADO.xlsx")
        wb.save(out)
        logger.info("XLSX V2 CONSOLIDADO: %s (%s bytes)", out, out.stat().st_size)
    finally:
        db.close()


if __name__ == "__main__":
    main()
