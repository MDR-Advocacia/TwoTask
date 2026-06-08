"""Varredura V2 do Banco do Brasil / Reu com classificacao de TEMPERATURA
expandida (inclui 🟠 AGUARDANDO CUMPRIMENTO) + comparacao com varredura
V1 anterior (runs bb-temperatura-* DONE).

Mudancas vs V1:
  - Janela: 60d (vs 30d)
  - 5 temperaturas (inclui 🟠 AGUARDANDO CUMPRIMENTO)
  - Regex: + cumprimento_iniciado, cumprimento_extinto
  - Delta vs V1: aba mostrando processos NOVOS ou que mudaram de temperatura

Pipeline:
  1. Le /tmp/base-bb-v2.xlsx + filtra
  2. Resolve CNJs -> lawsuit_ids
  3. Loop de lotes de 500 (janela 60d)
  4. Cada lote: XLSX com nova classificacao + delta vs V1

Output: /tmp/bb-temperatura-v2-{run_id}.xlsx (por lote)
"""

from __future__ import annotations

import json
import logging
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

INPUT_PATH = Path("/tmp/base-bb-v2.xlsx")
CAPA_MAP_PATH = Path("/tmp/bb-v2-capa-map.json")
RESOLVE_CACHE = Path("/tmp/bb-v2-cnj-resolve.json")
LOG_PATH = Path(
    "/app/output/playwright/legalone/varredura-andamentos/bb-temperatura-v2.log"
)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("bb.temperatura.v2")

WINDOW_DAYS = 60
BATCH_SIZE = 500

EXCLUIR_TIPO_ACAO = {
    "EMBARGOS A EXECUCAO",
    "EMBARGOS DE TERCEIRO",
    "CUMPRIMENTO DA SENTENCA",
    "ALVARA JUDICIAL",
    "CAUTELAR",
    "ADJUDICACAO",
}


# ── classificacao nova (5 temperaturas) ───────────────────────────────

def classificar_temperatura(tipos: set[str]) -> tuple[str, str]:
    """Retorna (temperatura, justificativa)."""
    if "cumprimento_extinto" in tipos:
        return "ENCERRAMENTO_IMINENTE", "cumprimento extinto/satisfeito"
    if "cumprimento_iniciado" in tipos:
        return "AGUARDANDO_CUMPRIMENTO", "sinais de cumprimento ativos (sem extincao)"
    if "transito_julgado" in tipos or "arquivamento" in tipos:
        sinais = sorted(tipos & {"transito_julgado", "arquivamento"})
        return "ENCERRAMENTO_IMINENTE", f"sinal de encerramento: {', '.join(sinais)}"
    if "sentenca" in tipos:
        extras = tipos - {"sentenca"}
        if extras:
            return "MEDIA", f"sentenca + {', '.join(sorted(extras))}"
        return "MEDIA", "sentenca proferida"
    if {"audiencia_designada", "audiencia_cancelada", "revelia"} & tipos:
        return "BAIXA", f"em andamento: {', '.join(sorted(tipos))}"
    return "INDETERMINADO", f"sem andamentos relevantes em {WINDOW_DAYS}d"


LABEL_TEMP = {
    "ENCERRAMENTO_IMINENTE": "🔴 ENCERRAMENTO IMINENTE",
    "AGUARDANDO_CUMPRIMENTO": "🟠 AGUARDANDO CUMPRIMENTO",
    "MEDIA": "🟡 MEDIA",
    "BAIXA": "🟢 BAIXA",
    "INDETERMINADO": "⚪ INDETERMINADO",
}


# ── Etapa 1: leitura + filtro ─────────────────────────────────────────

def ler_e_filtrar_planilha() -> tuple[list[str], dict[str, dict]]:
    import openpyxl

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    cnjs: list[str] = []
    mapa: dict[str, dict] = {}
    seen: set[str] = set()
    total = excl_nao_reu = excl_tipo = excl_sinopse = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        total += 1
        polo = str(row[17] or "")
        tipo = str(row[13] or "").strip().upper()
        sinopse = str(row[36] or "").strip()
        cnj_raw = row[2]
        if polo != "REU":
            excl_nao_reu += 1
            continue
        if "incidental" in sinopse.lower():
            excl_sinopse += 1
            continue
        if tipo in EXCLUIR_TIPO_ACAO:
            excl_tipo += 1
            continue
        digits = "".join(ch for ch in str(cnj_raw) if ch.isdigit())
        if len(digits) < 15 or digits in seen:
            continue
        seen.add(digits)
        capa = {
            headers[i]: row[i] for i in range(min(len(headers), len(row)))
        }
        capa["__cnj_original"] = str(cnj_raw).strip()
        mapa[digits] = capa
        cnjs.append(digits)
    logger.info(
        "Planilha v2: total=%s, excl_nao_reu=%s, excl_tipo=%s, excl_sinopse=%s, mantidos=%s",
        total, excl_nao_reu, excl_tipo, excl_sinopse, len(cnjs),
    )
    return cnjs, mapa


# ── Etapa 2: resolver CNJ -> lawsuit_id ───────────────────────────────

def resolver_cnjs(cnjs: list[str]) -> dict[str, int]:
    from app.services.legal_one_client import LegalOneApiClient

    cache: dict[str, int] = {}
    if RESOLVE_CACHE.exists():
        try:
            cache = json.loads(RESOLVE_CACHE.read_text(encoding="utf-8"))
            cache = {k: int(v) for k, v in cache.items()}
            logger.info("Cache resolve v2: %s entries", len(cache))
        except Exception:
            cache = {}

    pendentes = [c for c in cnjs if c not in cache]
    if pendentes:
        client = LegalOneApiClient()
        logger.info("Resolvendo %s CNJs via L1 (v2)...", len(pendentes))
        matches = client.search_lawsuits_by_cnj_numbers(pendentes)
        for cnj_norm in pendentes:
            payload = None
            for k, v in matches.items():
                if "".join(ch for ch in str(k) if ch.isdigit()) == cnj_norm:
                    payload = v
                    break
            if payload is not None and payload.get("id") is not None:
                try:
                    cache[cnj_norm] = int(payload["id"])
                except (TypeError, ValueError):
                    pass
        RESOLVE_CACHE.write_text(json.dumps(cache), encoding="utf-8")

    resolvidos = {c: cache[c] for c in cnjs if c in cache}
    logger.info("Resolvidos v2: %s / %s", len(resolvidos), len(cnjs))
    return resolvidos


# ── Etapa 3: pega mapa V1 (varredura anterior) pra delta ─────────────

def carregar_v1() -> dict[int, dict]:
    """Mapa lawsuit_id -> {temperatura_v1, achados_v1: set(tipos), run_v1}."""
    from app.db.session import SessionLocal
    from app.models.varredura import VarreduraAchado, VarreduraProcessado, VarreduraRun

    out: dict[int, dict] = {}
    db = SessionLocal()
    try:
        runs_v1 = (
            db.query(VarreduraRun)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-%"))
            .filter(~VarreduraRun.triggered_by.like("bb-temperatura-v2-%"))
            .filter(VarreduraRun.status == "DONE")
            .all()
        )
        v1_ids = [r.id for r in runs_v1]
        if not v1_ids:
            logger.warning("Nenhuma run V1 encontrada — sem delta.")
            return out
        logger.info("Runs V1 pra delta: %s", len(v1_ids))

        # Processados V1
        for p in db.query(VarreduraProcessado).filter(
            VarreduraProcessado.run_id.in_(v1_ids)
        ):
            out.setdefault(p.lawsuit_id, {"tipos_v1": set(), "run_v1": p.run_id})

        # Tipos por lawsuit V1
        for a in db.query(VarreduraAchado).filter(
            VarreduraAchado.run_id.in_(v1_ids)
        ):
            if a.lawsuit_id in out:
                out[a.lawsuit_id]["tipos_v1"].add(a.tipo_evento)

        # Classifica V1 (com a regra V2 pra comparacao justa)
        for lid, d in out.items():
            temp, just = classificar_temperatura(d["tipos_v1"])
            d["temp_v1"] = temp
            d["just_v1"] = just
    finally:
        db.close()
    logger.info("Mapa V1 carregado: %s lawsuits", len(out))
    return out


# ── Etapa 4: gerar XLSX do lote ───────────────────────────────────────

def gerar_xlsx_lote(
    db, run_id: int, capa_map: dict[str, dict],
    v1_map: dict[int, dict], xlsx_path: Path,
) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.models.varredura import VarreduraAchado, VarreduraProcessado, VarreduraRun

    run = db.query(VarreduraRun).filter(VarreduraRun.id == run_id).first()
    processados = (
        db.query(VarreduraProcessado)
        .filter(VarreduraProcessado.run_id == run_id)
        .all()
    )
    achados = (
        db.query(VarreduraAchado)
        .filter(VarreduraAchado.run_id == run_id)
        .order_by(VarreduraAchado.andamento_data.desc().nullslast(), VarreduraAchado.id.desc())
        .all()
    )

    tipos_por_lid: dict[int, set[str]] = {}
    achados_por_lid: dict[int, list] = {}
    for a in achados:
        tipos_por_lid.setdefault(a.lawsuit_id, set()).add(a.tipo_evento)
        achados_por_lid.setdefault(a.lawsuit_id, []).append(a)

    rows = []
    contadores = {
        "ENCERRAMENTO_IMINENTE": 0,
        "AGUARDANDO_CUMPRIMENTO": 0,
        "MEDIA": 0,
        "BAIXA": 0,
        "INDETERMINADO": 0,
    }
    # delta counts
    delta_novo = 0  # lawsuit nao estava na V1
    delta_mudou = 0  # mudou de temperatura V1 -> V2

    for p in processados:
        tipos = tipos_por_lid.get(p.lawsuit_id, set())
        temp, just = classificar_temperatura(tipos)
        contadores[temp] += 1
        digits = "".join(ch for ch in (p.cnj_number or "") if ch.isdigit())
        capa = capa_map.get(digits) or {}
        v1 = v1_map.get(p.lawsuit_id)
        if v1 is None:
            delta_status = "NOVO"
            delta_novo += 1
            temp_v1 = None
        else:
            temp_v1 = v1.get("temp_v1")
            if temp_v1 != temp:
                delta_status = f"MUDOU ({LABEL_TEMP.get(temp_v1, temp_v1)} → {LABEL_TEMP.get(temp, temp)})"
                delta_mudou += 1
            else:
                delta_status = "MANTEVE"
        rows.append({
            "lawsuit_id": p.lawsuit_id,
            "cnj": p.cnj_number or capa.get("__cnj_original") or "",
            "temperatura": temp,
            "justificativa": just,
            "tipos_evento": ", ".join(sorted(tipos)),
            "qtd_achados": len(achados_por_lid.get(p.lawsuit_id, [])),
            "delta": delta_status,
            "temp_v1": temp_v1,
            "tipos_v1": v1.get("tipos_v1") if v1 else None,
            "capa": capa,
            "achados": achados_por_lid.get(p.lawsuit_id, []),
        })

    # Workbook
    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFE5E7EB")
    cores = {
        "ENCERRAMENTO_IMINENTE": PatternFill("solid", fgColor="FFFEE2E2"),
        "AGUARDANDO_CUMPRIMENTO": PatternFill("solid", fgColor="FFFFEDD5"),
        "MEDIA": PatternFill("solid", fgColor="FFFEF3C7"),
        "BAIXA": PatternFill("solid", fgColor="FFDCFCE7"),
        "INDETERMINADO": PatternFill("solid", fgColor="FFF3F4F6"),
    }

    # ── Aba Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = f"Lote V2 run #{run_id} — BB/Réu — Temperatura (janela 60d)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    info = [
        ("Run ID", run_id),
        ("Lote", run.triggered_by if run else ""),
        ("Iniciada", run.started_at.strftime("%d/%m/%Y %H:%M") if run and run.started_at else ""),
        ("Concluída", run.completed_at.strftime("%d/%m/%Y %H:%M") if run and run.completed_at else ""),
        ("Janela (dias)", WINDOW_DAYS),
        ("Total processos no lote", len(processados)),
        ("Total achados", len(achados)),
        ("", ""),
        ("🔴 ENCERRAMENTO IMINENTE", contadores["ENCERRAMENTO_IMINENTE"]),
        ("🟠 AGUARDANDO CUMPRIMENTO", contadores["AGUARDANDO_CUMPRIMENTO"]),
        ("🟡 MEDIA (sentença sem trânsito)", contadores["MEDIA"]),
        ("🟢 BAIXA (em andamento)", contadores["BAIXA"]),
        ("⚪ INDETERMINADO", contadores["INDETERMINADO"]),
        ("", ""),
        ("DELTA vs V1", ""),
        ("  Lawsuits novos (não estavam na V1)", delta_novo),
        ("  Lawsuits que MUDARAM de temperatura", delta_mudou),
        ("  Lawsuits que MANTIVERAM", len(rows) - delta_novo - delta_mudou),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = bold if isinstance(v, str) and v == "" else None
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 36

    # ── Helper: aba por temperatura ──
    def _write_temp_sheet(sheet_name: str, subset: list, fill: PatternFill):
        ws_t = wb.create_sheet(sheet_name)
        cols = [
            "Temperatura", "CNJ", "Lawsuit ID", "NPJ", "Tipos de evento",
            "Justificativa", "Qtd achados", "Delta vs V1", "Temp V1",
            "Tipo Ação", "Situação", "UF", "Comarca", "Vara", "Valor da Causa",
            "Advogado", "Matéria",
            "Datas achados", "Trechos detectados",
        ]
        for c, h in enumerate(cols, start=1):
            cell = ws_t.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        for i, r in enumerate(subset, start=2):
            capa = r["capa"]
            ws_t.cell(row=i, column=1, value=LABEL_TEMP.get(r["temperatura"], r["temperatura"])).fill = fill
            ws_t.cell(row=i, column=2, value=r["cnj"])
            ws_t.cell(row=i, column=3, value=r["lawsuit_id"])
            ws_t.cell(row=i, column=4, value=str(capa.get("NPJ") or ""))
            ws_t.cell(row=i, column=5, value=r["tipos_evento"])
            ws_t.cell(row=i, column=6, value=r["justificativa"])
            ws_t.cell(row=i, column=7, value=r["qtd_achados"])
            ws_t.cell(row=i, column=8, value=r["delta"])
            ws_t.cell(row=i, column=9, value=LABEL_TEMP.get(r["temp_v1"], r["temp_v1"]) if r["temp_v1"] else "—")
            ws_t.cell(row=i, column=10, value=str(capa.get("Tipo de Ação") or ""))
            ws_t.cell(row=i, column=11, value=str(capa.get("Situação do Processo") or ""))
            ws_t.cell(row=i, column=12, value=str(capa.get("UF") or ""))
            ws_t.cell(row=i, column=13, value=str(capa.get("Comarca") or ""))
            ws_t.cell(row=i, column=14, value=str(capa.get("Vara") or ""))
            ws_t.cell(row=i, column=15, value=str(capa.get("Valor da Causa") or ""))
            ws_t.cell(row=i, column=16, value=str(capa.get("Advogado") or ""))
            ws_t.cell(row=i, column=17, value=str(capa.get("Matéria") or ""))
            datas = []
            trechos = []
            for a in r["achados"]:
                d = a.andamento_data.strftime("%d/%m/%Y") if a.andamento_data else "—"
                datas.append(f"{d}: {a.tipo_evento}")
                trecho = (a.andamento_texto or "")[:300]
                trechos.append(f"[{d}] {trecho}")
            ws_t.cell(row=i, column=18, value="\n".join(datas))
            tc = ws_t.cell(row=i, column=19, value="\n\n".join(trechos))
            tc.alignment = Alignment(wrap_text=True, vertical="top")
        widths = [22, 26, 12, 12, 40, 38, 10, 40, 22, 24, 14, 6, 22, 30, 14, 30, 22, 30, 90]
        for c, w in enumerate(widths, start=1):
            ws_t.column_dimensions[get_column_letter(c)].width = w
        if subset:
            ws_t.freeze_panes = "A2"
            ws_t.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(subset) + 1}"

    # Ordena cada temperatura por qtd_achados desc
    def _sub(temp):
        return sorted([r for r in rows if r["temperatura"] == temp], key=lambda x: -x["qtd_achados"])

    _write_temp_sheet("🔴 ENCERRAMENTO IMINENTE", _sub("ENCERRAMENTO_IMINENTE"), cores["ENCERRAMENTO_IMINENTE"])
    _write_temp_sheet("🟠 AGUARDANDO CUMPRIMENTO", _sub("AGUARDANDO_CUMPRIMENTO"), cores["AGUARDANDO_CUMPRIMENTO"])
    _write_temp_sheet("🟡 MEDIA", _sub("MEDIA"), cores["MEDIA"])
    _write_temp_sheet("🟢 BAIXA", _sub("BAIXA"), cores["BAIXA"])
    _write_temp_sheet("⚪ INDETERMINADO", _sub("INDETERMINADO"), cores["INDETERMINADO"])

    # Aba NOVOS achados vs V1
    novos = [r for r in rows if r["delta"] == "NOVO"]
    if novos:
        _write_temp_sheet("🆕 NOVOS (não estavam V1)", novos, header_fill)
    mudaram = [r for r in rows if r["delta"].startswith("MUDOU")]
    if mudaram:
        _write_temp_sheet("🔄 MUDARAM (V1→V2)", mudaram, header_fill)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return contadores


# ── Etapa 5: orquestracao em lotes ────────────────────────────────────

def main() -> None:
    from app.db.session import SessionLocal
    from app.models.varredura import (
        QUEUE_STATUS_PENDING,
        RUN_STATUS_RUNNING,
        VarreduraProcessado,
        VarreduraRun,
    )
    from app.services.varredura.varredura_service import (
        _run_subprocess_worker_impl,
    )

    logger.info("=" * 60)
    logger.info("BB/Reu V2 — TEMPERATURA — pipeline iniciado")
    logger.info("=" * 60)

    cnjs, capa_map = ler_e_filtrar_planilha()
    cnj_to_id = resolver_cnjs(cnjs)

    CAPA_MAP_PATH.write_text(
        json.dumps(capa_map, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    id_to_cnj = {v: k for k, v in cnj_to_id.items()}
    PRIORIDADE = {"Cumprimento": 0, "Sentença": 1, "Recurso": 2, "Inicial": 3}

    def _prio(lid: int) -> tuple[int, int]:
        digits = id_to_cnj.get(lid)
        capa = capa_map.get(digits) if digits else None
        sit = str((capa or {}).get("Situação do Processo") or "").strip()
        return (PRIORIDADE.get(sit, 99), lid)

    lawsuits_total = sorted(set(cnj_to_id.values()), key=_prio)
    logger.info("Total lawsuit_ids resolvidos: %s", len(lawsuits_total))

    # Mapa V1 pra delta
    v1_map = carregar_v1()

    # Skip lawsuits ja varridos nesta V2 (em runs anteriores)
    db = SessionLocal()
    try:
        runs_anteriores = (
            db.query(VarreduraRun.id)
            .filter(VarreduraRun.triggered_by.like("bb-temperatura-v2-%"))
            .all()
        )
        ja_varridos: set[int] = set()
        for (rid,) in runs_anteriores:
            ids = (
                db.query(VarreduraProcessado.lawsuit_id)
                .filter(VarreduraProcessado.run_id == rid)
                .all()
            )
            ja_varridos.update(r[0] for r in ids)
    finally:
        db.close()
    logger.info("Já varridos em runs V2 anteriores: %s", len(ja_varridos))

    restante = [lid for lid in lawsuits_total if lid not in ja_varridos]
    total_lotes = (len(restante) + BATCH_SIZE - 1) // BATCH_SIZE
    logger.info(
        "Restante pra varrer: %s — em %s lotes de %s",
        len(restante), total_lotes, BATCH_SIZE,
    )

    for idx in range(total_lotes):
        chunk = restante[idx * BATCH_SIZE : (idx + 1) * BATCH_SIZE]
        if not chunk:
            break
        logger.info("─── LOTE V2 %s/%s — %s processos ───", idx + 1, total_lotes, len(chunk))

        db = SessionLocal()
        try:
            run = VarreduraRun(
                status=RUN_STATUS_RUNNING,
                started_at=datetime.now(timezone.utc),
                responsible_office_ids=[23],
                window_days=WINDOW_DAYS,
                triggered_by=f"bb-temperatura-v2-lote-{idx + 1}-de-{total_lotes}",
                total_processos=len(chunk),
            )
            db.add(run)
            db.flush()
            for lid in chunk:
                digits = id_to_cnj.get(lid)
                cnj_orig = None
                if digits:
                    cnj_orig = capa_map.get(digits, {}).get("__cnj_original")
                db.add(
                    VarreduraProcessado(
                        run_id=run.id,
                        lawsuit_id=lid,
                        cnj_number=cnj_orig or digits,
                        office_id=23,
                        queue_status=QUEUE_STATUS_PENDING,
                    )
                )
            db.commit()
            run_id = run.id
            logger.info("Run #%s criada. Disparando subprocess...", run_id)
            try:
                _run_subprocess_worker_impl(db, run_id)
            except Exception as exc:
                logger.exception("Erro no worker run #%s: %s", run_id, exc)

            try:
                xlsx_path = Path(f"/tmp/bb-temperatura-v2-{run_id}.xlsx")
                contadores = gerar_xlsx_lote(db, run_id, capa_map, v1_map, xlsx_path)
                size = xlsx_path.stat().st_size
                logger.info(
                    "XLSX V2 LOTE %s pronto: %s bytes — ENC=%s, AGUARD=%s, MED=%s, BAI=%s, INDEF=%s",
                    run_id, size,
                    contadores["ENCERRAMENTO_IMINENTE"],
                    contadores["AGUARDANDO_CUMPRIMENTO"],
                    contadores["MEDIA"], contadores["BAIXA"],
                    contadores["INDETERMINADO"],
                )
            except Exception:
                logger.exception("Falha XLSX V2 lote run #%s", run_id)
        finally:
            db.close()

        time.sleep(10)

    logger.info("=" * 60)
    logger.info("=== PIPELINE BB/Reu V2 CONCLUIDO ===")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
