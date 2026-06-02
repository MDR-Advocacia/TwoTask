"""Le /tmp/listagem.xlsx (5965 CNJs + capa), resolve lawsuit_ids,
dispara varredura com janela 90 dias, depois processa cada andamento
com Sonnet pra gerar:
  - Lista de prazos em aberto categorizados (URGENTE_ATE_22_05 | PROX_SEMANA_25_05)
  - Relatorio processual (resumo executivo)

Output: /tmp/varredura-relatorios-master.xlsx

Uso:
    docker cp local.xlsx onetask-api-1:/tmp/listagem.xlsx
    docker exec -d onetask-api-1 python //app/app/runners/legalone/_run_planilha_relatorios.py
"""

from __future__ import annotations

import json
import logging
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from pathlib import Path

INPUT_PATH = Path("/tmp/listagem.xlsx")
CAPA_MAP_PATH = Path("/tmp/varredura-capa-map.json")
RELATORIOS_CACHE = Path("/tmp/varredura-relatorios-cache.json")
LOG_PATH = Path(
    "/app/output/playwright/legalone/varredura-andamentos/planilha-relatorios.log"
)
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, mode="a"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("varredura.planilha")

WINDOW_DAYS = 90
HOJE = date(2026, 5, 20)
URGENTE_LIMITE = date(2026, 5, 22)
PROX_SEMANA_INICIO = date(2026, 5, 25)


def ler_planilha() -> tuple[list[str], dict[str, dict]]:
    """Le a planilha e devolve (cnjs_digits, mapa { cnj_digits -> capa })."""
    import openpyxl

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb.active
    headers = [(c.value or "").strip() for c in ws[2]]

    cnjs: list[str] = []
    mapa: dict[str, dict] = {}
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        cnj_raw = row[6] if len(row) > 6 else None
        if not cnj_raw:
            continue
        cnj_str = str(cnj_raw).strip()
        digits = "".join(ch for ch in cnj_str if ch.isdigit())
        if len(digits) < 15:
            continue
        if digits in seen:
            continue
        seen.add(digits)
        capa = {
            headers[i]: row[i] for i in range(min(len(headers), len(row)))
        }
        capa["__cnj_original"] = cnj_str
        mapa[digits] = capa
        cnjs.append(digits)
    logger.info("Planilha: %s CNJs unicos com >=15 digitos.", len(cnjs))
    return cnjs, mapa


def resolver_cnjs(cnjs: list[str]) -> dict[str, int]:
    """Resolve CNJs em lawsuit_ids via API L1."""
    from app.services.legal_one_client import LegalOneApiClient

    client = LegalOneApiClient()
    logger.info("Resolvendo %s CNJs via API L1...", len(cnjs))
    matches = client.search_lawsuits_by_cnj_numbers(cnjs)
    cnj_to_id: dict[str, int] = {}
    for cnj_norm in cnjs:
        payload = None
        for k, v in matches.items():
            if "".join(ch for ch in str(k) if ch.isdigit()) == cnj_norm:
                payload = v
                break
        if payload is None:
            continue
        pid = payload.get("id")
        if pid is not None:
            try:
                cnj_to_id[cnj_norm] = int(pid)
            except (TypeError, ValueError):
                pass
    logger.info("Resolvidos: %s / %s", len(cnj_to_id), len(cnjs))
    return cnj_to_id


def disparar_varredura(
    db, capa_map: dict[str, dict], cnj_to_id: dict[str, int]
) -> int:
    """Cria run + items, dispara subprocess Node sincrono. Retorna run_id."""
    from app.models.varredura import (
        QUEUE_STATUS_PENDING,
        RUN_STATUS_RUNNING,
        VarreduraProcessado,
        VarreduraRun,
    )
    from app.services.varredura.varredura_service import (
        _run_subprocess_worker_impl,
    )

    lawsuit_ids = sorted(set(cnj_to_id.values()))
    id_to_cnj = {v: k for k, v in cnj_to_id.items()}

    run = VarreduraRun(
        status=RUN_STATUS_RUNNING,
        started_at=datetime.now(timezone.utc),
        responsible_office_ids=[],
        window_days=WINDOW_DAYS,
        triggered_by="planilha-relatorios-master",
        total_processos=len(lawsuit_ids),
    )
    db.add(run)
    db.flush()
    for lid in lawsuit_ids:
        db.add(
            VarreduraProcessado(
                run_id=run.id,
                lawsuit_id=lid,
                cnj_number=id_to_cnj.get(lid),
                queue_status=QUEUE_STATUS_PENDING,
            )
        )
    db.commit()
    logger.info("Run #%s criada com %s processos. Disparando subprocess...", run.id, len(lawsuit_ids))

    # Salva mapa lawsuit_id -> capa pra IA usar depois
    capa_by_lawsuit = {
        str(lid): capa_map.get(id_to_cnj.get(lid), {})
        for lid in lawsuit_ids
    }
    CAPA_MAP_PATH.write_text(
        json.dumps(capa_by_lawsuit, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    _run_subprocess_worker_impl(db, run.id)
    return run.id


# ── IA: prompt pra relatorio + prazos ────────────────────────────────


SYSTEM_PROMPT = """Voce e' analista juridico do escritorio MDR Advocacia.
O escritorio defende o REU em processos do Banco Master S/A (cliente) e
suas empresas vinculadas. Sua tarefa em DUAS partes:

1. EXTRAIR PRAZOS EM ABERTO **DO REU/BANCO MASTER**.
2. GERAR UM RESUMO PROCESSUAL.

PARTE 1 — PRAZOS EM ABERTO DO REU (REGRA ESTRITA)
=================================================
APENAS prazos processuais cujo CUMPRIMENTO recai sobre o **REU**
(Banco Master S/A e/ou seus advogados representando-o). Tudo o mais
deve ser IGNORADO.

INCLUA:
- Intimacoes pessoais ao REU/advogados do REU pra manifestar,
  contestar, recorrer, impugnar, oferecer embargos, juntar
  documentos, comparecer a audiencia, etc.
- Decisoes que abriram prazo para o REU se pronunciar.
- Intimacoes de sentenca/decisao com prazo recursal a contar
  contra o REU.

IGNORE OBRIGATORIAMENTE (NAO INCLUA):
- Qualquer prazo do AUTOR/POLO ATIVO (replica, impugnacao a
  contestacao, manifestacao sobre laudo solicitada ao autor, etc.).
- Prazos de TERCEIROS, peritos, partes nao representadas pelo MDR.
- Decorridos prazos (texto "DECORRIDO PRAZO DE ...") — sao apenas
  certidoes do que JA passou, nao geram novo prazo em aberto.
- Eventos meramente informativos: publicacoes que so' replicam o ato,
  certidoes de cartorio, expedicoes de mandado, distribuicao, conclusao
  ao juiz, juntada de aviso de recebimento, decurso de prazo certificado.
- Prazos cuja resposta JA foi juntada em andamento posterior
  (procurar "JUNTADA DE PETICAO DE CONTESTACAO/MANIFESTACAO/RECURSO"
  do polo passivo, "PROTOCOLADA PETICAO", etc., apos a data do prazo).
- Intimacoes generica/coletivas sem identificacao do REU como destinatario.

CALCULO DO VENCIMENTO:
- Conte dias UTEIS (padrao CPC), excluindo sabados, domingos e
  feriados nacionais. A contagem inicia no PRIMEIRO DIA UTIL apos
  a publicacao/intimacao (CPC art. 224 §3°).
- Hoje e' 20/05/2026 (quarta-feira).

CATEGORIZACAO:
- "URGENTE_ATE_22_05": vencimento entre 20/05 e 22/05/2026 (inclusive)
- "PROX_SEMANA_25_05": vencimento >= 25/05/2026 (segunda) e <= 31/12/2026
- "JA_VENCIDO": vencimento < 20/05/2026 — INCLUA APENAS se ainda
  estiver pendente cumprimento; se ja' houve manifestacao posterior,
  marque ja_cumprido=true e omita.

PARTE 2 — RESUMO PROCESSUAL (TOM PROFISSIONAL)
==============================================
Escreva um resumo do estado do processo em 3-5 paragrafos
(max 350 palavras). Cobertura:
- Capa: partes, polo, comarca, materia, valor (use dados da capa).
- Fase atual e principais movimentacoes recentes (90 dias).
- Decisoes ja' proferidas (sentenca, liminares, interlocutorias relevantes).
- Pontos de atencao (audiencias designadas, periodos suspeitos, recursos pendentes).
- Status objetivo do processo.

REGRAS DE REDACAO:
- Tom: tecnico-juridico, objetivo, conciso. Sem floreio, sem
  superlativos, sem juizo de valor.
- NAO mencione: "este relatorio", "este resumo", "analise", "IA",
  "inteligencia artificial", "modelo", "automatizado", "gerado por".
- Escreva como se fosse um advogado lendo os autos e descrevendo
  o estado processual ao colega. Comece direto pelo fato, nao
  pelo metalinguistico.
- Use 3a pessoa. Nao use "vamos", "analisamos", "verifica-se que
  conforme nossa avaliacao".

OUTPUT
======
JSON ESTRITO sem markdown fences. Resposta = APENAS o JSON:
{
  "relatorio": "texto direto do estado processual (max 350 palavras)",
  "fase_atual": "postulatorio|instrutorio|decisorio|recursal|cumprimento|arquivado|outro",
  "status_resumo": "1 frase objetiva do estado atual",
  "prazos_em_aberto": [
    {
      "data_andamento": "DD/MM/YYYY",
      "evento": "trecho do andamento que abriu o prazo do REU",
      "tipo_prazo": "contestacao|recurso|manifestacao|impugnacao|alegacoes_finais|embargos|outros",
      "prazo_dias": 15,
      "tipo_dias": "uteis|corridos",
      "data_vencimento": "DD/MM/YYYY",
      "categoria": "URGENTE_ATE_22_05|PROX_SEMANA_25_05|JA_VENCIDO",
      "ja_cumprido": false,
      "destinatario": "REU (Banco Master)",
      "justificativa": "1 frase explicando por que e' prazo do REU em aberto"
    }
  ]
}
"""


def chamar_sonnet(processo_payload: dict) -> dict:
    """1 chamada Sonnet sincrona pra 1 processo."""
    import httpx
    from app.core.config import settings

    api_key = settings.anthropic_api_key
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY ausente")

    capa = processo_payload["capa"]
    andamentos = processo_payload["andamentos"]
    cnj = processo_payload.get("cnj") or "-"
    lawsuit_id = processo_payload.get("lawsuit_id")

    capa_str = "\n".join(
        f"- {k}: {v}" for k, v in capa.items() if v not in (None, "", 0)
    )[:3000]
    andamentos_lines = []
    for a in andamentos:
        d = a.get("data") or "?"
        h = a.get("hora") or ""
        t = a.get("tipo") or ""
        texto = (a.get("texto") or "").strip()
        mov = a.get("movimentadoPor") or ""
        andamentos_lines.append(
            f"[{d} {h}] {t}: {texto[:800]}"
            + (f"\n  (Movimentado por: {mov})" if mov else "")
        )
    andamentos_str = "\n".join(andamentos_lines)[:18000]

    user_msg = f"""CNJ: {cnj}
Lawsuit ID interno: {lawsuit_id}
HOJE: 20/05/2026 (quarta)

=== CAPA DO PROCESSO ===
{capa_str}

=== ANDAMENTOS ({len(andamentos)} eventos, ordem cronologica decrescente) ===
{andamentos_str}

Gere o JSON conforme as instrucoes do system."""

    payload = {
        # Haiku 4.5 — mais rapido e barato pra tarefa estruturada (relatorio + extracao de prazos).
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 4000,
        "temperature": 0,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_msg}],
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    with httpx.Client(timeout=120.0) as client:
        resp = client.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json=payload,
        )
    if resp.status_code != 200:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:500]}")
    text = (resp.json().get("content", [{}])[0].get("text") or "").strip()
    if text.startswith("```"):
        text = text.strip("`").lstrip("json").strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Tenta extrair JSON do texto
        ini = text.find("{")
        fim = text.rfind("}")
        if ini >= 0 and fim > ini:
            return json.loads(text[ini : fim + 1])
        raise


def processar_processos_com_ia(
    run_id: int, capa_by_lawsuit: dict
) -> dict[int, dict]:
    """Pra cada processo concluido da run, le andamentos do status.json,
    chama Sonnet (8 threads), retorna mapa { lawsuit_id -> dict }."""
    status_path = Path(
        f"/app/output/playwright/legalone/varredura-andamentos/run-{run_id}/status.json"
    )
    if not status_path.exists():
        raise FileNotFoundError(f"status.json nao encontrado em {status_path}")
    status_data = json.loads(status_path.read_text(encoding="utf-8"))
    items = status_data.get("items") or []

    # Cache pra retomada
    cache: dict[int, dict] = {}
    if RELATORIOS_CACHE.exists():
        try:
            raw = json.loads(RELATORIOS_CACHE.read_text(encoding="utf-8"))
            cache = {int(k): v for k, v in raw.items()}
            logger.info("Cache de relatorios: %s entries", len(cache))
        except Exception:
            cache = {}

    todo = []
    for it in items:
        if (it.get("status") or "").lower() != "ok":
            continue
        lid = int(it.get("lawsuitId") or 0)
        if not lid or lid in cache:
            continue
        ands = it.get("andamentos") or []
        capa = capa_by_lawsuit.get(str(lid)) or {}
        todo.append(
            {
                "lawsuit_id": lid,
                "cnj": it.get("cnjNumber") or capa.get("__cnj_original"),
                "andamentos": ands,
                "capa": capa,
            }
        )
    logger.info("Processos pra IA: %s (cache: %s)", len(todo), len(cache))

    lock = threading.Lock()
    counter = {"done": 0, "fail": 0}

    def _work(p):
        lid = p["lawsuit_id"]
        try:
            r = chamar_sonnet(p)
            with lock:
                cache[lid] = r
                counter["done"] += 1
                if counter["done"] % 50 == 0:
                    RELATORIOS_CACHE.write_text(
                        json.dumps(
                            {str(k): v for k, v in cache.items()},
                            ensure_ascii=False,
                        ),
                        encoding="utf-8",
                    )
                    logger.info(
                        "IA: %s processados (%s falhas)",
                        counter["done"], counter["fail"],
                    )
        except Exception as exc:
            with lock:
                counter["fail"] += 1
                cache[lid] = {"error": str(exc)[:300]}

    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(_work, p) for p in todo]
        for _ in as_completed(futures):
            pass

    RELATORIOS_CACHE.write_text(
        json.dumps(
            {str(k): v for k, v in cache.items()}, ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    logger.info("IA concluida. %s sucesso, %s falhas", counter["done"], counter["fail"])
    return cache


def gerar_xlsx_final(
    run_id: int,
    capa_by_lawsuit: dict,
    relatorios: dict[int, dict],
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFE5E7EB", end_color="FFE5E7EB", fill_type="solid",
    )

    # Compila listas
    prazos_urgentes = []
    prazos_prox = []
    prazos_vencidos = []
    relatorio_rows = []
    sem_prazo = []

    for lid_str, capa in capa_by_lawsuit.items():
        lid = int(lid_str)
        r = relatorios.get(lid) or {}
        cnj = capa.get("__cnj_original") or ""
        if "error" in r:
            sem_prazo.append(
                {
                    "lawsuit_id": lid,
                    "cnj": cnj,
                    "capa": capa,
                    "motivo": r["error"],
                }
            )
            continue
        prazos = r.get("prazos_em_aberto") or []
        if not prazos:
            sem_prazo.append(
                {
                    "lawsuit_id": lid,
                    "cnj": cnj,
                    "capa": capa,
                    "motivo": "Sem prazo identificado pela IA",
                }
            )
        for p in prazos:
            row = {
                "lawsuit_id": lid,
                "cnj": cnj,
                "capa": capa,
                **p,
            }
            cat = (p.get("categoria") or "").upper()
            if cat == "URGENTE_ATE_22_05":
                prazos_urgentes.append(row)
            elif cat == "PROX_SEMANA_25_05":
                prazos_prox.append(row)
            elif cat == "JA_VENCIDO":
                prazos_vencidos.append(row)

        relatorio_rows.append(
            {
                "lawsuit_id": lid,
                "cnj": cnj,
                "capa": capa,
                "relatorio": r.get("relatorio") or "",
                "fase_atual": r.get("fase_atual") or "",
                "status_resumo": r.get("status_resumo") or "",
                "qtd_prazos": len(prazos),
            }
        )

    # ── Aba 1: Resumo Geral ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo Geral"
    ws["A1"] = "Varredura — Carteira Master com Prazos e Relatorios"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")
    rows = [
        ("Run ID", run_id),
        ("Hoje (referencia)", "20/05/2026 (quarta)"),
        ("Janela varrida", f"{WINDOW_DAYS} dias"),
        ("", ""),
        ("Total processos na planilha", len(capa_by_lawsuit)),
        ("Processos com relatorio IA", len([r for r in relatorios.values() if "error" not in r])),
        ("Processos com falha na IA", len([r for r in relatorios.values() if "error" in r])),
        ("", ""),
        ("PRAZOS ATE 22/05 (urgentes)", len(prazos_urgentes)),
        ("PRAZOS A PARTIR DE 25/05", len(prazos_prox)),
        ("Prazos JA VENCIDOS (apos verificacao)", len(prazos_vencidos)),
        ("Processos SEM prazo identificado", len(sem_prazo)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = bold if isinstance(v, str) and v == "" else None
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30

    # Helper pra escrever aba de prazos
    def _write_prazos(ws_p, lista: list[dict], titulo: str):
        cols = [
            "CNJ",
            "Lawsuit ID",
            "Codigo AJUS",
            "Cliente (Reu)",
            "Acao",
            "Materia",
            "Vara/Foro",
            "Comarca/UF",
            "Responsavel",
            "Valor Causa",
            "Data andamento",
            "Evento que gerou prazo",
            "Tipo prazo",
            "Prazo (dias)",
            "Tipo dias",
            "Data vencimento",
            "Ja cumprido?",
            "Justificativa IA",
        ]
        for c, h in enumerate(cols, start=1):
            cell = ws_p.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        for i, r in enumerate(lista, start=2):
            capa = r["capa"]
            ws_p.cell(row=i, column=1, value=r["cnj"])
            ws_p.cell(row=i, column=2, value=r["lawsuit_id"])
            ws_p.cell(row=i, column=3, value=str(capa.get("Cód AJUS") or ""))
            reus = str(capa.get("Réus - CNPJCPF") or "")[:100]
            ws_p.cell(row=i, column=4, value=reus)
            ws_p.cell(row=i, column=5, value=str(capa.get("Tipo de Ação") or ""))
            ws_p.cell(row=i, column=6, value=str(capa.get("Matéria") or ""))
            ws_p.cell(
                row=i,
                column=7,
                value=f"{capa.get('Nº Vara') or ''} - {capa.get('Foro') or ''}",
            )
            ws_p.cell(
                row=i,
                column=8,
                value=f"{capa.get('Comarca') or ''}/{capa.get('UF') or ''}",
            )
            ws_p.cell(row=i, column=9, value=str(capa.get("Usuário Responsável") or ""))
            ws_p.cell(row=i, column=10, value=str(capa.get("Valor Causa") or ""))
            ws_p.cell(row=i, column=11, value=r.get("data_andamento") or "")
            ws_p.cell(row=i, column=12, value=(r.get("evento") or "")[:200])
            ws_p.cell(row=i, column=13, value=r.get("tipo_prazo") or "")
            ws_p.cell(row=i, column=14, value=r.get("prazo_dias"))
            ws_p.cell(row=i, column=15, value=r.get("tipo_dias") or "")
            ws_p.cell(row=i, column=16, value=r.get("data_vencimento") or "")
            ws_p.cell(row=i, column=17, value="Sim" if r.get("ja_cumprido") else "Nao")
            jt = ws_p.cell(row=i, column=18, value=r.get("justificativa") or "")
            jt.alignment = Alignment(wrap_text=True, vertical="top")
        widths = [22, 10, 10, 35, 22, 18, 30, 24, 22, 14, 14, 50, 22, 8, 10, 14, 10, 60]
        for c, w in enumerate(widths, start=1):
            ws_p.column_dimensions[get_column_letter(c)].width = w
        if lista:
            ws_p.freeze_panes = "A2"
            ws_p.auto_filter.ref = (
                f"A1:{get_column_letter(len(cols))}{len(lista) + 1}"
            )

    ws_u = wb.create_sheet("Prazos ATE 22-05 (URGENTE)")
    _write_prazos(ws_u, prazos_urgentes, "")

    ws_n = wb.create_sheet("Prazos a partir 25-05")
    _write_prazos(ws_n, prazos_prox, "")

    if prazos_vencidos:
        ws_v = wb.create_sheet("Prazos JA VENCIDOS")
        _write_prazos(ws_v, prazos_vencidos, "")

    # ── Aba: Relatorio Processual ──────────────────────────────────
    ws_r = wb.create_sheet("Relatorio Processual")
    cols_r = [
        "CNJ",
        "Lawsuit ID",
        "Cod AJUS",
        "Cliente (Reu)",
        "Acao",
        "Materia",
        "Comarca/UF",
        "Polo",
        "Situacao",
        "Valor Causa",
        "Responsavel",
        "Fase atual",
        "Status (1 frase)",
        "Qtd prazos abertos",
        "Relatorio processual (IA)",
    ]
    for c, h in enumerate(cols_r, start=1):
        cell = ws_r.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
    for i, r in enumerate(relatorio_rows, start=2):
        capa = r["capa"]
        ws_r.cell(row=i, column=1, value=r["cnj"])
        ws_r.cell(row=i, column=2, value=r["lawsuit_id"])
        ws_r.cell(row=i, column=3, value=str(capa.get("Cód AJUS") or ""))
        ws_r.cell(row=i, column=4, value=str(capa.get("Réus - CNPJCPF") or "")[:100])
        ws_r.cell(row=i, column=5, value=str(capa.get("Tipo de Ação") or ""))
        ws_r.cell(row=i, column=6, value=str(capa.get("Matéria") or ""))
        ws_r.cell(
            row=i,
            column=7,
            value=f"{capa.get('Comarca') or ''}/{capa.get('UF') or ''}",
        )
        ws_r.cell(row=i, column=8, value=str(capa.get("Polo") or ""))
        ws_r.cell(row=i, column=9, value=str(capa.get("Situação Processo") or ""))
        ws_r.cell(row=i, column=10, value=str(capa.get("Valor Causa") or ""))
        ws_r.cell(row=i, column=11, value=str(capa.get("Usuário Responsável") or ""))
        ws_r.cell(row=i, column=12, value=r["fase_atual"])
        ws_r.cell(row=i, column=13, value=r["status_resumo"])
        ws_r.cell(row=i, column=14, value=r["qtd_prazos"])
        rt = ws_r.cell(row=i, column=15, value=r["relatorio"])
        rt.alignment = Alignment(wrap_text=True, vertical="top")
    widths_r = [22, 10, 10, 35, 22, 18, 24, 12, 14, 14, 22, 18, 50, 10, 100]
    for c, w in enumerate(widths_r, start=1):
        ws_r.column_dimensions[get_column_letter(c)].width = w
    if relatorio_rows:
        ws_r.freeze_panes = "A2"
        ws_r.auto_filter.ref = (
            f"A1:{get_column_letter(len(cols_r))}{len(relatorio_rows) + 1}"
        )

    # ── Aba: Sem prazo identificado ─────────────────────────────────
    ws_s = wb.create_sheet("Sem prazo identificado")
    cols_s = ["CNJ", "Lawsuit ID", "Cod AJUS", "Reu", "Acao", "Motivo"]
    for c, h in enumerate(cols_s, start=1):
        cell = ws_s.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
    for i, r in enumerate(sem_prazo, start=2):
        capa = r["capa"]
        ws_s.cell(row=i, column=1, value=r["cnj"])
        ws_s.cell(row=i, column=2, value=r["lawsuit_id"])
        ws_s.cell(row=i, column=3, value=str(capa.get("Cód AJUS") or ""))
        ws_s.cell(row=i, column=4, value=str(capa.get("Réus - CNPJCPF") or "")[:100])
        ws_s.cell(row=i, column=5, value=str(capa.get("Tipo de Ação") or ""))
        ws_s.cell(row=i, column=6, value=r["motivo"])
    widths_s = [22, 10, 10, 35, 22, 60]
    for c, w in enumerate(widths_s, start=1):
        ws_s.column_dimensions[get_column_letter(c)].width = w
    if sem_prazo:
        ws_s.freeze_panes = "A2"

    out_path = Path("/tmp/varredura-relatorios-master.xlsx")
    wb.save(out_path)
    return out_path


def main() -> None:
    from app.db.session import SessionLocal

    logger.info("============================================")
    logger.info("=== Varredura PLANILHA + RELATORIOS IA ===")
    logger.info("============================================")

    cnjs, capa_map = ler_planilha()
    cnj_to_id = resolver_cnjs(cnjs)
    unresolved = [c for c in cnjs if c not in cnj_to_id]
    logger.info("Nao resolvidos (placeholders ou nao encontrados): %s", len(unresolved))

    db = SessionLocal()
    try:
        run_id = disparar_varredura(db, capa_map, cnj_to_id)
        logger.info("=== VARREDURA CONCLUIDA. Run #%s ===", run_id)

        # IA: prazos + relatorios
        capa_by_lawsuit = json.loads(CAPA_MAP_PATH.read_text(encoding="utf-8"))
        relatorios = processar_processos_com_ia(run_id, capa_by_lawsuit)

        # Gera XLSX
        out = gerar_xlsx_final(run_id, capa_by_lawsuit, relatorios)
        logger.info("=== XLSX FINAL gerada: %s (%s bytes) ===", out, out.stat().st_size)
    finally:
        db.close()


if __name__ == "__main__":
    main()
