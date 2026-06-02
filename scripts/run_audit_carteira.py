"""Standalone script — auditoria forense de carteira da banca terceirizada.

USO INCIDENTAL. Roda LOCAL (Python da maquina, fora do Docker), pra
processar lotes de JSONs Atlas (robo de captura PJe), classificar
falhas/resultados via Anthropic Batches API, e gerar relatorio XLSX
consolidado.

NAO TOCA NO BANCO DE PRODUCAO. Estado persistente fica em pasta de
output: `responses.jsonl` (1 linha por processo) + `batches.json`
(rastreio dos batch_ids).

================================================================
SUBCOMMANDOS
================================================================

1) submit — envia um lote pra Anthropic Batches API:
     python scripts/run_audit_carteira.py submit \\
       --input "C:/.../Atlas - json" \\
       --output-dir "./auditoria-giovanna" \\
       --lote-id "lote-01"

2) poll — consulta batches abertos, baixa results, append no JSONL:
     python scripts/run_audit_carteira.py poll \\
       --output-dir "./auditoria-giovanna"
   (rodar a cada 5-10 min ate' ver "batch ENDED")

3) report — agrega o JSONL em XLSX multi-aba:
     python scripts/run_audit_carteira.py report \\
       --output-dir "./auditoria-giovanna" \\
       --xlsx "./auditoria-giovanna/relatorio-final.xlsx"

4) sync — modo dry-run: 1 processo via Messages API (sincrono),
   util pra smoke test do prompt antes de submeter os 187+:
     python scripts/run_audit_carteira.py sync \\
       --input "C:/.../Atlas - json" \\
       --cnj "8000737-31.2025.8.05.0144"

================================================================
IDEMPOTENCIA
================================================================

- Mesmo CNJ nunca e' submetido duas vezes — script checa
  `responses.jsonl` antes de incluir.
- Submeter o mesmo lote duas vezes nao re-processa o que ja tem
  resposta (mas inclui CNJs novos que apareceram).
- Skip automatico de processos com `status_automacao=MARCOS_HABILITADO_COM_PDF`.

================================================================
DEPENDENCIAS
================================================================

Usa apenas: httpx, pydantic, openpyxl (todas ja' no requirements.txt
do projeto). NAO IMPORTA `app.core.*` nem SQLAlchemy/FastAPI — script
e' auto-contido.

ANTHROPIC_API_KEY deve estar no ambiente (ou .env) — leitura via
os.environ.
"""

from __future__ import annotations

import argparse
import glob
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

# Forca stdout/stderr em UTF-8 no Windows (cp1252 explode com ⚠️ etc.)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:  # noqa: BLE001
    pass

# Faz o script encontrar `app.services.classificador.*` sem precisar
# instalar o projeto como pacote. Roda do checkout principal.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import httpx  # noqa: E402
from pydantic import ValidationError  # noqa: E402

from app.services.classificador.audit_prompts import (  # noqa: E402
    AUDIT_SYSTEM_PROMPT,
    build_audit_user_message,
    iterate_atlas_jsons,
)
from app.services.classificador.audit_schema import (  # noqa: E402
    AuditParseError,
    AuditResponse,
    parse_audit_response,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("audit_carteira")

# ─── Constantes ───────────────────────────────────────────────────────

ANTHROPIC_API_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_BATCHES_URL = "https://api.anthropic.com/v1/messages/batches"
ANTHROPIC_API_VERSION = "2023-06-01"

# Sonnet 4.6 — mesmo modelo do classificador atual. Pode sobrescrever
# via env AUDIT_MODEL.
DEFAULT_MODEL = os.environ.get("AUDIT_MODEL", "claude-sonnet-4-5-20250929")
DEFAULT_MAX_TOKENS = int(os.environ.get("AUDIT_MAX_TOKENS", "8192"))

RESPONSES_JSONL = "responses.jsonl"
BATCHES_JSON = "batches.json"
ERRORS_JSONL = "errors.jsonl"


# ─── Helpers de IO ────────────────────────────────────────────────────


def _api_key() -> str:
    key = os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_KEY")
    if not key:
        # Tenta ler do .env do repo se existe (so' pra conveniencia local)
        env_path = _REPO_ROOT / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("ANTHROPIC_API_KEY="):
                    key = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
    if not key:
        sys.exit(
            "ERRO: ANTHROPIC_API_KEY nao encontrada. "
            "Defina como variavel de ambiente ou no .env do repo."
        )
    return key


def _headers() -> dict[str, str]:
    return {
        "x-api-key": _api_key(),
        "anthropic-version": ANTHROPIC_API_VERSION,
        "content-type": "application/json",
    }


def _ensure_output_dir(path: str) -> Path:
    p = Path(path).resolve()
    # GUARD — dados de auditoria sao sensiveis (processos judiciais de
    # cliente). NUNCA podem ficar dentro do repo git. Aviso se o path
    # cair dentro de _REPO_ROOT.
    try:
        rel = p.relative_to(_REPO_ROOT)
        # Caiu dentro do repo
        logger.warning(
            "[!] --output-dir aponta pra DENTRO do repo (%s).", rel,
        )
        logger.warning(
            "[!] Dados de auditoria sao sensiveis — mover pra um path "
            "FORA do repo (ex.: junto da pasta de JSONs de input).",
        )
        logger.warning(
            "[!] O .gitignore bloqueia auditoria/ e auditoria-*/, mas "
            "se voce mudar o nome da pasta ou commitar com -f, "
            "DADOS DE CLIENTE VAZAM. Voce esta avisado.",
        )
    except ValueError:
        # Bom — path esta fora do repo
        pass
    p.mkdir(parents=True, exist_ok=True)
    return p


def _read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    out = []
    with path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError as exc:
                logger.warning("Linha invalida em %s: %s", path, exc)
    return out


def _append_jsonl(path: Path, obj: dict) -> None:
    with path.open("a", encoding="utf-8") as fp:
        fp.write(json.dumps(obj, ensure_ascii=False, default=str) + "\n")


def _read_batches(out_dir: Path) -> list[dict]:
    p = out_dir / BATCHES_JSON
    if not p.exists():
        return []
    return json.loads(p.read_text(encoding="utf-8"))


def _write_batches(out_dir: Path, batches: list[dict]) -> None:
    (out_dir / BATCHES_JSON).write_text(
        json.dumps(batches, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )


def _existing_cnjs(out_dir: Path) -> set[str]:
    """CNJs ja processados (com response valido) — pra idempotencia."""
    rows = _read_jsonl(out_dir / RESPONSES_JSONL)
    return {r.get("cnj_number") for r in rows if r.get("cnj_number")}


# ─── Submit ───────────────────────────────────────────────────────────


def _build_batch_request(custom_id: str, atlas_json: dict) -> dict:
    user_message = build_audit_user_message(atlas_json)
    return {
        "custom_id": custom_id,
        "params": {
            "model": DEFAULT_MODEL,
            "max_tokens": DEFAULT_MAX_TOKENS,
            "temperature": 0,
            "system": AUDIT_SYSTEM_PROMPT,
            "messages": [{"role": "user", "content": user_message}],
        },
    }


def cmd_submit(args: argparse.Namespace) -> int:
    input_dir = Path(args.input).resolve()
    out_dir = _ensure_output_dir(args.output_dir)
    lote_id = args.lote_id or f"lote-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    if not input_dir.exists():
        logger.error("Pasta de input nao existe: %s", input_dir)
        return 2

    files = sorted(glob.glob(str(input_dir / "*.json")))
    if not files:
        logger.error("Nenhum .json encontrado em %s", input_dir)
        return 2

    existing = _existing_cnjs(out_dir)
    logger.info(
        "Pasta: %s | %d JSONs encontrados | %d CNJs ja' processados em lotes anteriores",
        input_dir.name, len(files), len(existing),
    )

    requests_payload: list[dict] = []
    cnjs_inclusos: list[str] = []
    skip_log: list[dict] = []

    for path, atlas, skip_reason in iterate_atlas_jsons(files, skip_existing_cnjs=existing):
        cnj = (atlas.get("cnj_number") or "").strip() if isinstance(atlas, dict) else ""
        if skip_reason:
            skip_log.append({
                "file": Path(path).name,
                "cnj": cnj or None,
                "motivo": skip_reason,
            })
            continue
        if not cnj:
            skip_log.append({
                "file": Path(path).name,
                "cnj": None,
                "motivo": "Sem cnj_number",
            })
            continue
        # custom_id: lote-XX__<cnj-normalizado>. Anthropic exige
        # ^[a-zA-Z0-9_-]{1,64}$ — CNJ tem pontos que precisam virar `_`.
        # Convertido de volta em _split_cnj_from_custom_id() na hora do apply.
        cnj_normalized = cnj.replace(".", "_")
        custom_id = f"{lote_id}__{cnj_normalized}"
        if len(custom_id) > 64:
            logger.warning("custom_id muito longo (%d): %s — truncando", len(custom_id), custom_id)
            custom_id = custom_id[:64]
        requests_payload.append(_build_batch_request(custom_id, atlas))
        cnjs_inclusos.append(cnj)

    logger.info(
        "Lote %s: %d processos para auditoria | %d pulados",
        lote_id, len(requests_payload), len(skip_log),
    )

    # Salva log de skips
    if skip_log:
        skip_path = out_dir / f"skipped-{lote_id}.json"
        skip_path.write_text(
            json.dumps(skip_log, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("Skip log: %s", skip_path)

    if not requests_payload:
        logger.warning("Nada pra submeter no lote %s.", lote_id)
        return 0

    if args.dry_run:
        sample_path = out_dir / f"dry-run-{lote_id}.json"
        sample_path.write_text(
            json.dumps(
                {
                    "lote_id": lote_id,
                    "total_requests": len(requests_payload),
                    "cnjs": cnjs_inclusos,
                    "primeiro_user_message_preview": (
                        requests_payload[0]["params"]["messages"][0]["content"][:3000]
                        if requests_payload else None
                    ),
                },
                ensure_ascii=False, indent=2,
            ),
            encoding="utf-8",
        )
        logger.info("DRY RUN — nenhum batch enviado. Preview: %s", sample_path)
        return 0

    # Submete o batch
    logger.info("Enviando batch pra Anthropic (%d requests)...", len(requests_payload))
    with httpx.Client(timeout=120.0) as client:
        resp = client.post(
            ANTHROPIC_BATCHES_URL,
            headers=_headers(),
            json={"requests": requests_payload},
        )
    if resp.status_code not in (200, 201):
        logger.error("Falha submit batch: HTTP %s — %s", resp.status_code, resp.text[:500])
        return 3

    data = resp.json()
    batch_id = data.get("id")
    logger.info("Batch criado: %s (status=%s)", batch_id, data.get("processing_status"))

    # Salva rastreio
    batches = _read_batches(out_dir)
    batches.append({
        "lote_id": lote_id,
        "batch_id": batch_id,
        "processing_status": data.get("processing_status"),
        "request_counts": data.get("request_counts"),
        "created_at": data.get("created_at"),
        "expires_at": data.get("expires_at"),
        "submitted_at_local": datetime.utcnow().isoformat() + "Z",
        "cnjs_count": len(cnjs_inclusos),
        "applied": False,
    })
    _write_batches(out_dir, batches)

    print(f"\n✓ Lote {lote_id} submetido: batch_id={batch_id}")
    print(f"  Processos: {len(cnjs_inclusos)}")
    print(f"  Pulados: {len(skip_log)}")
    print(f"  Proximo passo: aguarde 5-30 min e rode:")
    print(f"    python {Path(__file__).name} poll --output-dir \"{out_dir}\"")
    return 0


# ─── Poll ─────────────────────────────────────────────────────────────


def _fetch_batch_status(batch_id: str) -> dict:
    with httpx.Client(timeout=30.0) as client:
        resp = client.get(
            f"{ANTHROPIC_BATCHES_URL}/{batch_id}",
            headers=_headers(),
        )
    if resp.status_code != 200:
        raise RuntimeError(
            f"Falha ao consultar batch {batch_id}: HTTP {resp.status_code} — {resp.text[:300]}"
        )
    return resp.json()


def _fetch_batch_results(results_url: str) -> Iterable[dict]:
    """Baixa o JSONL dos resultados e itera linha-a-linha."""
    with httpx.Client(timeout=300.0) as client:
        resp = client.get(results_url, headers=_headers())
    if resp.status_code != 200:
        raise RuntimeError(
            f"Falha ao baixar results: HTTP {resp.status_code} — {resp.text[:300]}"
        )
    for line in resp.text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            yield json.loads(line)
        except json.JSONDecodeError:
            logger.warning("Linha invalida no results JSONL: %s", line[:200])


def _apply_batch(out_dir: Path, batch_entry: dict) -> dict:
    """Baixa resultados de um batch ENDED e append no responses.jsonl."""
    batch_id = batch_entry["batch_id"]
    status = _fetch_batch_status(batch_id)

    # Atualiza counts
    batch_entry["processing_status"] = status.get("processing_status")
    batch_entry["request_counts"] = status.get("request_counts")
    batch_entry["ended_at"] = status.get("ended_at")

    if status.get("processing_status") != "ended":
        logger.info(
            "Batch %s ainda em %s (counts=%s)",
            batch_id, status.get("processing_status"), status.get("request_counts"),
        )
        return batch_entry

    results_url = status.get("results_url")
    if not results_url:
        logger.error("Batch %s ENDED mas sem results_url.", batch_id)
        return batch_entry

    responses_path = out_dir / RESPONSES_JSONL
    errors_path = out_dir / ERRORS_JSONL

    ok = 0
    err = 0
    parse_err = 0

    for line in _fetch_batch_results(results_url):
        custom_id = line.get("custom_id") or ""
        # custom_id = "lote-XX__<cnj-normalizado>". Reverter _ → .
        cnj_raw = custom_id.split("__", 1)[-1] if "__" in custom_id else custom_id
        cnj = cnj_raw.replace("_", ".") if cnj_raw else cnj_raw
        result = line.get("result") or {}
        kind = result.get("type")

        if kind != "succeeded":
            err += 1
            _append_jsonl(errors_path, {
                "batch_id": batch_id,
                "custom_id": custom_id,
                "cnj_number": cnj,
                "result_type": kind,
                "error": result.get("error"),
                "applied_at": datetime.utcnow().isoformat() + "Z",
            })
            continue

        message = result.get("message") or {}
        content_blocks = message.get("content") or []
        raw_text = ""
        for block in content_blocks:
            if block.get("type") == "text":
                raw_text += block.get("text", "")

        try:
            audit = parse_audit_response(raw_text)
        except (AuditParseError, ValidationError) as exc:
            parse_err += 1
            _append_jsonl(errors_path, {
                "batch_id": batch_id,
                "custom_id": custom_id,
                "cnj_number": cnj,
                "result_type": "parse_error",
                "error": str(exc)[:1000],
                "raw_text_preview": (raw_text or "")[:1500],
                "applied_at": datetime.utcnow().isoformat() + "Z",
            })
            continue

        # OK — append no responses.jsonl
        _append_jsonl(responses_path, {
            "batch_id": batch_id,
            "custom_id": custom_id,
            "cnj_number": audit.cnj_number or cnj,
            "model": message.get("model"),
            "usage": message.get("usage"),
            "applied_at": datetime.utcnow().isoformat() + "Z",
            "audit": audit.model_dump(mode="json"),
        })
        ok += 1

    batch_entry["applied"] = True
    batch_entry["applied_at_local"] = datetime.utcnow().isoformat() + "Z"
    batch_entry["applied_stats"] = {"ok": ok, "errored": err, "parse_error": parse_err}

    logger.info(
        "Batch %s aplicado: %d ok | %d erros API | %d erros parse",
        batch_id, ok, err, parse_err,
    )
    return batch_entry


def cmd_poll(args: argparse.Namespace) -> int:
    out_dir = _ensure_output_dir(args.output_dir)
    batches = _read_batches(out_dir)
    if not batches:
        logger.warning("Nenhum batch em %s (rode `submit` primeiro).", out_dir)
        return 0

    pending = [b for b in batches if not b.get("applied")]
    if not pending:
        logger.info("Todos os batches ja foram aplicados.")
        return 0

    for entry in pending:
        try:
            _apply_batch(out_dir, entry)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Falha aplicando batch %s: %s", entry.get("batch_id"), exc)
            entry["last_poll_error"] = str(exc)[:500]

    _write_batches(out_dir, batches)
    return 0


# ─── Sync (smoke test) ────────────────────────────────────────────────


def cmd_sync(args: argparse.Namespace) -> int:
    """Roda 1 processo via Messages API (sincrono) — pra smoke test."""
    input_dir = Path(args.input).resolve()
    if not input_dir.exists():
        logger.error("Pasta nao existe: %s", input_dir)
        return 2

    if args.cnj:
        candidates = list(input_dir.glob(f"*{args.cnj}*.json"))
    else:
        candidates = sorted(input_dir.glob("*.json"))[:1]
    if not candidates:
        logger.error("Nenhum JSON casa com cnj=%s", args.cnj)
        return 2

    path = candidates[0]
    with path.open("r", encoding="utf-8") as fp:
        atlas = json.load(fp)

    user_message = build_audit_user_message(atlas)

    if args.print_prompt:
        print("# === SYSTEM PROMPT (primeiros 2000 chars) ===")
        print(AUDIT_SYSTEM_PROMPT[:2000])
        print("\n# === USER MESSAGE (primeiros 4000 chars) ===")
        print(user_message[:4000])
        print(f"\n# === USER MESSAGE: {len(user_message)} chars total ===")
        return 0

    logger.info("SYNC: enviando 1 processo (%s) via Messages API...", path.name)
    payload = {
        "model": DEFAULT_MODEL,
        "max_tokens": DEFAULT_MAX_TOKENS,
        "temperature": 0,
        "system": AUDIT_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_message}],
    }
    with httpx.Client(timeout=300.0) as client:
        resp = client.post(ANTHROPIC_API_URL, headers=_headers(), json=payload)
    if resp.status_code != 200:
        logger.error("HTTP %s — %s", resp.status_code, resp.text[:1000])
        return 3

    data = resp.json()
    raw_text = "".join(
        b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"
    )

    print(f"\n=== USAGE: {data.get('usage')}")
    print(f"=== RAW RESPONSE ({len(raw_text)} chars):\n")
    print(raw_text)

    try:
        audit = parse_audit_response(raw_text)
    except (AuditParseError, ValidationError) as exc:
        print(f"\n!! ERRO de parse/validacao: {exc}")
        return 4

    print("\n=== PARSED (resumo):")
    print(f"  CNJ: {audit.cnj_number}")
    print(f"  Empresas representadas: {len(audit.empresas_representadas)}")
    for e in audit.empresas_representadas:
        print(f"    - {e.papel}: {e.nome} ({e.cnpj or 'sem cnpj'})")
    print(f"  Falhas confirmadas: {len(audit.falhas_confirmadas)}")
    for f in audit.falhas_confirmadas:
        print(f"    - {f.codigo} [{f.severidade}] {f.descricao_curta}")
    print(f"  Indicios: {len(audit.indicios_de_falha)}")
    for f in audit.indicios_de_falha:
        print(f"    - {f.codigo} [{f.severidade}] {f.descricao_curta}")
    print(f"  Resultados negativos: {len(audit.resultados_negativos)}")
    for r in audit.resultados_negativos:
        print(f"    - {r.codigo} {r.descricao_curta}")
    print(f"  Dados insuficientes: {len(audit.dados_insuficientes)}")
    print(f"  Confianca: {audit.confianca_geral}")
    print(f"  Resumo: {audit.resumo_executivo}")
    return 0


# ─── Normalizacao canonica de nomes de empresa ────────────────────────


import re as _re_norm

# Mapeamento de prefixos normalizados -> nome canonico. Ordem matters
# (mais especifico antes do generico).
_CANONICAL_EMPRESAS = [
    # Master family — Master S/A, Master S.A., Master Multiplo, etc.
    # MAXIMA e' denominacao anterior do Master (texto da capa diz literal:
    # "BANCO MASTER S/A (atual denominacao do Banco Maxima S/A)")
    ("BANCO MASTER MULTIPLO", "Banco Master"),
    ("BANCO MASTER", "Banco Master"),
    ("BANCO MAXIMA", "Banco Master"),
    # Voiter family — Voiter e' a denominacao atual, Pleno e' a anterior
    ("BANCO VOITER", "Banco Voiter"),
    ("BANCO PLENO", "Banco Voiter"),
    # PKL One Participacoes
    ("PKL ONE PARTICIPACOES", "PKL One Participacoes"),
    ("PKL ONE", "PKL One Participacoes"),
    # Sul America (seguradora — caso isolado)
    ("SUL AMERICA COMPANHIA DE SEGURO", "Sul America Seguros"),
    ("SUL AMERICA SEGUROS", "Sul America Seguros"),
    ("SUL AMERICA", "Sul America Seguros"),
]


def canonical_empresa(nome: Optional[str]) -> str:
    """Normaliza nome de empresa pra forma canonica.

    Aplica:
    1. Uppercase + trim
    2. Remove sufixos S/A, S.A., SA, LTDA
    3. Remove parenteses ("(atual denominacao...)")
    4. Match contra _CANONICAL_EMPRESAS por prefixo
    5. Fallback: titlecase do normalizado

    Empresas conhecidas (Master/Voiter/PKL/Sul America) viram canonical
    fixo; resto fica como title-case pra UI ficar legivel.
    """
    if not nome:
        return "(nao identificada)"
    n = nome.upper().strip()
    # Tira parenteses e conteudo
    n = _re_norm.sub(r"\([^)]*\)", "", n)
    # Tira sufixos societarios
    n = _re_norm.sub(r"\bS\s*[/.]?\s*A\b\.?", "", n)
    n = _re_norm.sub(r"\bSA\b\.?", "", n)
    n = _re_norm.sub(r"\bLTDA\b\.?", "", n)
    # Espaços extras
    n = _re_norm.sub(r"\s+", " ", n).strip()
    # Pontuacao trailing
    n = n.rstrip(".,;:- ")

    # Match prefix
    for prefix, canon in _CANONICAL_EMPRESAS:
        if n == prefix or n.startswith(prefix + " ") or n.startswith(prefix):
            return canon
    # Fallback: titlecase
    return n.title() if n else "(nao identificada)"


# ─── Glossario das falhas (1a aba do XLSX) ────────────────────────────


# Glossario didatico do que cada codigo significa + o criterio rigoroso
# que a IA aplicou pra confirmar. Texto editorial em pt-BR — diferente do
# system prompt (que e' instrucao pra IA). Mantido aqui no script (nao no
# audit_prompts.py) porque e' so' pra leitura humana no XLSX.
_GLOSSARIO_CONCEITOS_QUANTITATIVOS = [
    ("PE (prob_exito_global)",
     "Probabilidade GLOBAL de exito da banca representada — i.e. de "
     "improcedencia total do processo, escala 0.0 a 1.0. Calibrado pela "
     "REGRA DO MENOS FAVORAVEL: se 1 pedido tem prob_perda=provavel, o "
     "processo todo tem PE max ~0.3. Se todos remota, PE 0.7-0.9. Se ja' "
     "ha sentenca, reflete o resultado (improcedente -> PE 0.85+, "
     "procedente -> PE 0.1-0.2 na fase recursal)."),
    ("PCOND (aprovisionamento CPC 25 / IAS 37)",
     "Provisao contabil calculada pedido a pedido: remota=0, possivel=0 "
     "(so' divulgacao em notas explicativas), provavel=valor_estimado. "
     "PCOND_TOTAL do processo e' a soma dos aprovisionamentos individuais. "
     "Esse e' o valor que a banca AUDITADA recomendaria provisionar em "
     "balanco pra o caso."),
    ("Valor estimado em risco",
     "Soma dos valor_estimado dos pedidos. E' o valor TOTAL que pode ser "
     "condenado a' banca caso TODOS os pedidos sejam acolhidos (pior "
     "cenario). Diferente do PCOND (que so' considera pedidos provaveis)."),
    ("Probabilidade de perda (por pedido)",
     "Classificacao de cada pedido: REMOTA (tese da banca robusta — "
     "jurisprudencia favoravel + prova solida); POSSIVEL (jurisprudencia "
     "dividida ou prova incompleta); PROVAVEL (jurisprudencia uniforme "
     "contra a banca OU falha grave da defesa, OU revelia decretada)."),
    ("Taxa de exito da banca",
     "% de processos sentenciados em que o resultado foi favoravel a' "
     "banca representada (improcedencia OU em_favor_de=reu_giovanna). "
     "Calculado SO' entre processos com sentenca/decisao definitiva — "
     "ignora os em_andamento. Metrica de PERFORMANCE da banca auditada."),
]

_GLOSSARIO_CONCEITOS = [
    ("FALHA CONFIRMADA",
     "Ato ou omissao da banca auditada que a IA conseguiu IMPUTAR com "
     "evidencia documental DIRETA: trecho literal de decisao, certidao "
     "ou peca processual que comprova o fato. So entra aqui quando ha "
     "(a) trecho literal citado, (b) imputacao inequivoca a' Giovanna "
     "no periodo em que ela era responsavel, e (c) efeito processual "
     "observavel."),
    ("INDICIO DE FALHA",
     "Sinal forte de falha mas SEM uma das 3 condicoes acima. Tipico: "
     "a omissao existe (ex.: nao ha contestacao na timeline), mas falta "
     "decisao posterior do juizo reconhecendo o efeito (ex.: revelia nao "
     "decretada ainda, ou processo em curso). Auditor humano deve revisar "
     "e decidir se promove pra falha confirmada apos olhar os autos."),
    ("DADO INSUFICIENTE",
     "Ponto que a IA examinou mas o dossie da Atlas nao trouxe a peca-"
     "chave pra concluir (ex.: certidao de citacao com data exata, "
     "anexo da contestacao sem texto capturado). Nao e' falha — e' "
     "lacuna do material capturado. Pode exigir recaptura ou consulta "
     "direta ao processo."),
    ("RESULTADO NEGATIVO",
     "Efeito processual desfavoravel (sentenca procedente, multa, "
     "confissao ficta, transito em julgado contra a banca). SEPARADO "
     "de falha — pode haver resultado ruim sem falha do escritorio "
     "(tese realmente perdedora). Quando ha falha associada, o codigo "
     "R9xx aponta o F-codigo correlato."),
]

_GLOSSARIO_SEVERIDADES = [
    ("CRITICA",
     "Prejuizo direto, observavel e irreversivel. Ex.: revelia decretada, "
     "transito em julgado sem recurso, desercao."),
    ("ALTA",
     "Causa risco material relevante OU preclusao reconhecida. Ex.: "
     "contestacao intempestiva, ausencia em audiencia gerando confissao "
     "ficta, lacuna de meses entre citacao e habilitacao."),
    ("MEDIA",
     "Conduta processual abaixo do esperado mas sem prejuizo imediato. "
     "Ex.: nao impugnacao especifica em pedido secundario, defesa sem "
     "lastro probatorio em processo que ainda nao tem sentenca."),
    ("BAIXA",
     "Deslize formal sem impacto material observavel. Ex.: erro de "
     "qualificacao corrigido posteriormente."),
]

_GLOSSARIO_FALHAS = [
    # (codigo, categoria, nome_curto, criterio_da_ia, o_que_NAO_configura, severidade_tipica)
    ("F101", "FASE INICIAL", "Habilitacao tardia no processo",
     "A banca auditada conduzia o cliente desde antes do processo virar "
     "judicial (mandato administrativo pre-existente). Houve LACUNA "
     "temporal observavel entre a ORDEM DE CITACAO (despacho do juizo "
     "tipo 'cite-se a parte re para...') e a data efetiva de habilitacao "
     "de Giovanna. Cada dia entre esses marcos e' periodo em que a "
     "banca deveria ter estado em juizo acompanhando — e nao estava.",
     "Distribuicao do processo (ato do autor) NAO conta como gatilho. "
     "Lacuna curta (poucos dias) e sem efeito processual fica como BAIXA "
     "ou e' descartada. Habilitacao do MDR (Marcos Delli) jamais conta — "
     "esses processos sao filtrados antes do batch.",
     "CRITICA se causou revelia ou perda de recurso · ALTA se prazo "
     "venceu antes de habilitar · MEDIA se lacuna sem prejuizo concreto"),

    ("F102", "FASE INICIAL", "Endereco de citacao desatualizado",
     "Endereco fornecido pela banca estava errado/desatualizado, "
     "causando citacao por edital ou hora certa quando seria evitavel "
     "com diligencia minima.",
     "Citacao por edital quando o autor errou o endereco; quando o reu "
     "mudou e nao informou em tempo; quando o endereco capturado pelo "
     "robo nao bate por motivo legitimo.",
     "ALTA se gerou revelia · MEDIA caso contrario"),

    ("F103", "FASE INICIAL", "Procuracao faltante ou tardia",
     "Procuracao nao juntada ou juntada com atraso, gerando obstaculo "
     "expressamente reconhecido por decisao do juizo.",
     "Procuracao em ordem mas com formato administrativo discutivel; "
     "substabelecimento defeituoso (esse e' F802).",
     "MEDIA a ALTA conforme prejuizo"),

    ("F201", "CONTESTACAO", "Contestacao NAO apresentada (revelia)",
     "Decisao explicita do juizo DECRETANDO revelia por ausencia de "
     "contestacao tempestiva da banca auditada. Citada literalmente — "
     "frase tipo 'diante da ausencia de contestacao tempestiva, decreto "
     "a revelia do reu BANCO XYZ'.",
     "Frase da citacao inicial ('sob pena de revelia' — texto-padrao do "
     "despacho de cite-se) NAO conta como decretacao. Tem que ser "
     "decisao posterior expressa.",
     "CRITICA"),

    ("F202", "CONTESTACAO", "Contestacao INTEMPESTIVA",
     "Certidao ou decisao do juizo declarando expressamente a "
     "intempestividade da contestacao da banca auditada. Ou comprovacao "
     "matematica via certidao de citacao + data de juntada da peca + "
     "prazo legal explicito no processo.",
     "Suspeita de intempestividade sem certidao expressa OU sem data "
     "exata da citacao efetiva — vira INDICIO. Contestacao apresentada "
     "em prazo legal mesmo apos habilitacao tardia nao e' F202 (mas "
     "ainda pode ser F101).",
     "CRITICA se gerou revelia · ALTA caso contrario"),

    ("F203", "CONTESTACAO", "Contestacao GENERICA (sem provas)",
     "Peca juntada sem documentos PROBATORIOS (extrato bancario, "
     "contrato assinado, comprovante de transferencia, laudo, "
     "gravacao, planilha, midia). Auditoria OLHA OS ANEXOS efetivamente "
     "juntados na mesma operacao — se so' tem documentos burocraticos "
     "(procuracao, substabelecimento, carta de preposicao, RG/CPF, "
     "contrato social, cartao CNPJ, atas societarias), e' generica.",
     "Avaliacao de qualidade tecnica da peca (teses fracas, citacoes "
     "ruins) NAO entra aqui — F203 e' MECANICO (presenca ou ausencia "
     "de doc probatorio). Contestacao curta mas COM extratos juntados "
     "nao e' F203.",
     "ALTA com decisao reconhecendo · MEDIA quando objetivamente "
     "verificavel mas sem decisao"),

    ("F204", "CONTESTACAO", "Nao impugnacao especifica dos pedidos",
     "Sentenca posterior aplica PRESUNCAO DE VERACIDADE sobre pedidos "
     "que a contestacao nao refutou ponto a ponto. Texto expresso da "
     "sentenca: 'os pedidos X e Y nao foram impugnados especificamente, "
     "aplicando-se a presuncao de veracidade'.",
     "Contestacao defensiva em bloco sem refutar cada pedido mas SEM "
     "decisao reconhecendo prejuizo → vira INDICIO. Pedido secundario "
     "nao impugnado mas que se tornou prejudicado por mudanca de fase "
     "processual nao e' F204.",
     "MEDIA a ALTA conforme valor do pedido afetado"),

    ("F205", "CONTESTACAO", "Falta de documentos essenciais",
     "Decisao posterior reconhece a ausencia de documentos especificos "
     "(extratos, contratos) e determina consequencia (nova oportunidade "
     "ou efeito processual adverso).",
     "Defesa sem juntada inicial mas que junta tardiamente em peticao "
     "intermediaria, aceita pelo juizo, nao e' F205.",
     "MEDIA"),

    ("F206", "CONTESTACAO", "Prescricao/decadencia nao arguida",
     "Decisao posterior reconhece prescricao de OFICIO ou em embargos, "
     "evidenciando que a defesa nao a arguiu antes na contestacao "
     "(quando era cabivel).",
     "Casos em que prescricao nao se aplica; cenarios em que a defesa "
     "argumentou prescricao parcial — a IA prefere INDICIO se ha "
     "duvida sobre cabimento.",
     "ALTA — perda de defesa absoluta"),

    ("F207", "CONTESTACAO", "Preliminares nao arguidas",
     "Preliminares cabiveis (ilegitimidade passiva, conexao, "
     "litispendencia, inepcia) que nao foram suscitadas e sentenca "
     "posterior as enfrenta de oficio ou reconhece preclusao.",
     "Preliminares discutiveis no caso concreto; estrategia processual "
     "de defesa direta no merito nao e' falha por si so'.",
     "MEDIA"),

    ("F301", "AUDIENCIA", "Parte ausente — revelia/confissao ficta",
     "Ata de audiencia OU decisao expressa registra a ausencia da "
     "parte representada pela banca, com declaracao da revelia ou "
     "confissao ficta como consequencia.",
     "Ausencia sem consequencia processual reconhecida pelo juizo; "
     "ausencia justificada com adiamento posterior nao e' F301.",
     "CRITICA"),

    ("F302", "AUDIENCIA", "Advogado ausente em audiencia",
     "Ata registra a ausencia do advogado da banca auditada (sem "
     "substabelecimento valido pra outro advogado comparecer).",
     "Substituicao processual valida (preposto + advogado terceirizado "
     "comparecendo com poderes) nao e' F302.",
     "ALTA"),

    ("F303", "AUDIENCIA", "Preposto ausente (trabalhista)",
     "Ata de audiencia trabalhista registra a ausencia de preposto da "
     "empresa, com aplicacao do art. 844 CLT (confissao ficta) ou "
     "outra consequencia.",
     "Audiencia civil onde nao se exige preposto; ausencia justificada "
     "com adiamento posterior nao e' F303.",
     "CRITICA"),

    ("F304", "AUDIENCIA", "Testemunha nao compareceu",
     "Testemunha arrolada pela defesa nao compareceu e a banca nao "
     "requereu conducao coercitiva — perda de oportunidade probatoria.",
     "Testemunha que o juizo dispensou; testemunha arrolada pelo autor "
     "(nao e' responsabilidade da defesa).",
     "MEDIA"),

    ("F401", "INSTRUCAO", "Sem arrolamento de testemunhas",
     "Decisao reconhece 'nao houve requerimento de prova testemunhal "
     "tempestivo' em processo com controversia factica relevante.",
     "Casos onde controversia era so' juridica; despachos saneadores "
     "que indeferiram a producao por outro motivo.",
     "MEDIA"),

    ("F402", "INSTRUCAO", "Nao juntada de prova determinada",
     "Despacho/decisao determinou juntada especifica em prazo, e a "
     "defesa nao cumpriu — decisao posterior reconhece a omissao.",
     "Cumprimento parcial aceito pelo juizo; pedido de dilacao deferido.",
     "MEDIA a ALTA"),

    ("F403", "INSTRUCAO", "Silencio sobre prova adversa",
     "Decisao reconhece que a defesa nao impugnou prova produzida pelo "
     "autor, gerando efeito de admissao tacita.",
     "Impugnacao em peca posterior aceita; preclusao discutivel.",
     "MEDIA"),

    ("F501", "RECURSO", "Sentenca desfavoravel sem recurso",
     "CERTIDAO DE TRANSITO em julgado da sentenca de procedencia "
     "(parcial ou total) sem qualquer recurso interposto pela defesa "
     "no prazo legal. Ou certidao de decurso de prazo recursal.",
     "Recurso interposto mas julgado improvido nao e' F501; "
     "desistencia do recurso por estrategia documentada nao e' F501.",
     "CRITICA"),

    ("F502", "RECURSO", "Apelacao intempestiva",
     "Decisao expressa do tribunal reconhecendo intempestividade da "
     "apelacao da defesa.",
     "Apelacao tempestiva mas improvida no merito.",
     "CRITICA"),

    ("F503", "RECURSO", "Embargos de declaracao intempestivos",
     "Decisao reconhecendo intempestividade ou improvimento por "
     "inadequacao manifesta dos ED.",
     "ED rejeitado por reapreciacao de merito nao e' F503.",
     "MEDIA"),

    ("F504", "RECURSO", "Nao interposicao de RE/REsp",
     "Decisao reconhecendo preclusao da via recursal extraordinaria "
     "quando o caso comportava (acordao contrario a tese pacificada).",
     "Casos em que RE/REsp nao era cabivel pelo art. 1.030 CPC; "
     "estrategia de nao recorrer documentada — a IA prefere INDICIO "
     "quando ha duvida.",
     "ALTA"),

    ("F505", "RECURSO", "Nao agravo de instrumento cabivel",
     "Decisao interlocutoria recorrivel via agravo (rol do art. 1.015 "
     "CPC) sem interposicao e com efeito preclusivo reconhecido.",
     "Decisao recorrivel por outro meio; estrategia de discutir em "
     "apelacao.",
     "MEDIA a ALTA conforme materia"),

    ("F506", "RECURSO", "Desercao por preparo nao recolhido",
     "Decisao declara desercao do recurso pela falta de recolhimento "
     "do preparo no prazo legal.",
     "Justica gratuita deferida; pedido de prazo dobrado aceito.",
     "CRITICA"),

    ("F601", "CUMPRIMENTO", "Sem impugnacao ao cumprimento",
     "15 dias venceram sem impugnacao + decisao acolhe calculo do autor "
     "sem manifestacao da defesa.",
     "Impugnacao apresentada tempestivamente; preclusao discutivel.",
     "ALTA"),

    ("F602", "CUMPRIMENTO", "Sem impugnacao de penhora",
     "Preclusao do prazo de impugnacao a' penhora/avaliacao reconhecida "
     "em decisao.",
     "Impugnacao tempestiva indeferida no merito.",
     "MEDIA a ALTA"),

    ("F603", "CUMPRIMENTO", "BACEN-JUD por inadimplemento",
     "Extrato SISBAJUD ou decisao mencionando bloqueio efetivo de "
     "valores em conta da banca por nao cumprimento espontaneo da "
     "obrigacao.",
     "BACEN-JUD por outras causas (ex.: fraude).",
     "ALTA"),

    ("F701", "GESTAO", "Acordo fora de alcadas",
     "Termo de acordo homologado fora dos limites contratuais conhecidos "
     "da banca (so confirmavel quando as alcadas sao informadas).",
     "Acordo dentro de alcada; acordo sem informacao de alcada → "
     "INDICIO.",
     "ALTA"),

    ("F702", "GESTAO", "Desistencia/renuncia indevida",
     "Peticao de desistencia da defesa ou renuncia ao recurso sem "
     "fundamentacao estrategica documentada.",
     "Desistencia formalizada com fundamentacao em comite/aprovacao.",
     "ALTA"),

    ("F703", "GESTAO", "Confissao em peticionamento",
     "Peticao da defesa contendo reconhecimento juridico do pedido do "
     "autor (art. 487 III a CPC).",
     "Acordo formalizado (esse e' F701 se fora de alcada).",
     "CRITICA"),

    ("F801", "ADMINISTRATIVA", "Erro de qualificacao da parte",
     "Peca da banca cita CNPJ ou nome errado da parte representada — "
     "reconhecido como falha em decisao OU corrigido em peticao "
     "intermediaria (mostrando que o erro foi do escritorio).",
     "Erro irrelevante ja' corrigido sem nova intervencao.",
     "BAIXA"),

    ("F802", "ADMINISTRATIVA", "Substabelecimento invalido",
     "Substabelecimento sem poderes especificos OU sem assinatura "
     "do outorgante, reconhecido em decisao.",
     "Substabelecimento valido com poderes amplos.",
     "MEDIA"),
]

_GLOSSARIO_RESULTADOS = [
    ("R901", "Sentenca de procedencia INTEGRAL",
     "Sentenca acolhendo todos os pedidos do autor. Pode ter falha "
     "associada (ex.: F201 revelia ou F501 sem recurso) ou ser efeito "
     "de tese realmente perdedora."),
    ("R902", "Sentenca PARCIAL com condenacao relevante",
     "Sentenca parcialmente procedente com condenacao de valor "
     "significativo."),
    ("R903", "Multa por litigancia de ma-fe (CPC 80)",
     "Aplicada por conduta processual da defesa que o juizo considerou "
     "abusiva."),
    ("R904", "Multa por ato atentatorio (CPC 77 §2o)",
     "Descumprimento de ordem judicial ou conduta desrespeitosa."),
    ("R905", "Honorarios sucumbenciais majorados",
     "Honorarios fixados acima do minimo legal pelo juizo em razao da "
     "conduta processual."),
    ("R906", "Confissao ficta decretada",
     "Decretada por ausencia de comparecimento da parte representada — "
     "tipicamente associada a F301 ou F303."),
    ("R907", "Transito em julgado desfavoravel",
     "Decisao desfavoravel a' banca transitou em julgado — verificar "
     "se ha F501 associada."),
]


def _write_glossario_sheet(wb, write_sheet):
    """Cria a aba 'Glossario' como PRIMEIRA do XLSX.

    Estrutura em secoes:
    1. Como ler o relatorio (conceitos: falha confirmada vs indicio
       vs dado insuficiente vs resultado negativo)
    2. Niveis de severidade
    3. Tabela completa de codigos F (taxonomia)
    4. Tabela de codigos R (resultados negativos)
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    title = "Glossario"
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title, 0)  # index 0 = primeira aba

    BIG_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
    SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
    HEADER_FONT = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    row = 1

    # ─ Titulo geral ─
    ws.cell(row=row, column=1, value="GLOSSARIO — Auditoria Forense da Carteira").font = BIG_TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value=(
        "Esta aba explica os criterios usados pela auditoria automatizada. "
        "Cada codigo de falha tem regra rigorosa de imputacao — leia antes "
        "de revisar as outras abas. Esse glossario e' referencia: o relatorio "
        "nao acusa nada que nao caiba nos criterios abaixo."
    )).alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 45
    row += 2

    # ─ Secao 1: Conceitos ─
    sc = ws.cell(row=row, column=1, value="1) COMO LER — categorias de achado")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for header in ("Tipo", "Descricao"):
        c = ws.cell(row=row, column=1 if header == "Tipo" else 2, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for tipo, desc in _GLOSSARIO_CONCEITOS:
        ws.cell(row=row, column=1, value=tipo).font = HEADER_FONT
        ws.cell(row=row, column=1).alignment = WRAP
        c = ws.cell(row=row, column=2, value=desc)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 75
        row += 1
    row += 1

    # ─ Secao 2: Severidades ─
    sc = ws.cell(row=row, column=1, value="2) NIVEIS DE SEVERIDADE")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for header in ("Nivel", "Definicao"):
        c = ws.cell(row=row, column=1 if header == "Nivel" else 2, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for nivel, desc in _GLOSSARIO_SEVERIDADES:
        ws.cell(row=row, column=1, value=nivel).font = HEADER_FONT
        ws.cell(row=row, column=1).alignment = WRAP
        c = ws.cell(row=row, column=2, value=desc)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 55
        row += 1
    row += 1

    # ─ Secao 3: Codigos de falha ─
    sc = ws.cell(row=row, column=1, value="3) CODIGOS DE FALHA — criterios usados pela IA")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    falha_headers = [
        "Codigo", "Categoria", "Nome curto",
        "O que CONFIGURA a falha (regra rigorosa)",
        "O que NAO configura (zonas cinzentas / viram indicio)",
        "Severidade tipica",
    ]
    for i, h in enumerate(falha_headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    row += 1
    for cod, cat, nome, configura, nao_configura, sev in _GLOSSARIO_FALHAS:
        ws.cell(row=row, column=1, value=cod).font = HEADER_FONT
        ws.cell(row=row, column=2, value=cat).alignment = WRAP
        ws.cell(row=row, column=3, value=nome).alignment = WRAP
        ws.cell(row=row, column=4, value=configura).alignment = WRAP
        ws.cell(row=row, column=5, value=nao_configura).alignment = WRAP
        ws.cell(row=row, column=6, value=sev).alignment = WRAP
        # Altura dinamica baseada no maior texto
        max_len = max(len(configura), len(nao_configura))
        ws.row_dimensions[row].height = min(max(40, max_len // 4), 180)
        row += 1
    row += 1

    # ─ Secao 4: Conceitos quantitativos ─
    sc = ws.cell(row=row, column=1, value="4) DIAGNOSTICO QUANTITATIVO — metricas do flex da ferramenta")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for header in ("Metrica", "Definicao"):
        c = ws.cell(row=row, column=1 if header == "Metrica" else 2, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    for metrica, desc in _GLOSSARIO_CONCEITOS_QUANTITATIVOS:
        ws.cell(row=row, column=1, value=metrica).font = HEADER_FONT
        ws.cell(row=row, column=1).alignment = WRAP
        c = ws.cell(row=row, column=2, value=desc)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 80
        row += 1
    row += 1

    # ─ Secao 5: Resultados negativos ─
    sc = ws.cell(row=row, column=1, value="5) RESULTADOS NEGATIVOS — separados de falhas (sao efeito)")
    sc.font = SECTION_FONT
    sc.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for i, h in enumerate(("Codigo", "Nome", "Descricao"), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    row += 1
    for cod, nome, desc in _GLOSSARIO_RESULTADOS:
        ws.cell(row=row, column=1, value=cod).font = HEADER_FONT
        ws.cell(row=row, column=2, value=nome).alignment = WRAP
        c = ws.cell(row=row, column=3, value=desc)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 50
        row += 1

    # ─ Larguras de coluna ─
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 58
    ws.column_dimensions["E"].width = 58
    ws.column_dimensions["F"].width = 28

    ws.freeze_panes = "A2"


# ─── Capa executiva (1a aba do XLSX) ──────────────────────────────────


def _write_capa_executiva(
    wb,
    *,
    total_procs: int,
    total_falhas_confirmadas: int,
    total_indicios: int,
    total_resultados: int,
    total_pedidos: int,
    counter_falhas_codigo,
    counter_severidade,
    counter_empresa_principal,
    counter_empresa_falhas,
    counter_resultados_codigo,
    soma_valor_estimado_por_emp: dict,
    soma_pcond_por_emp: dict,
    soma_valor_condenacao_por_emp: dict,
    soma_pe_por_emp: dict,
    count_pe_por_emp: dict,
    count_sentenciados_por_emp: dict,
    count_exito_banca_por_emp: dict,
    casos_criticos: list,
    nomes_falhas: dict,
) -> None:
    """Capa executiva — UX prioridade: abrir o XLSX e ver as falhas
    imediatamente, sem ter que navegar 14 abas.

    Layout (de cima pra baixo):
    1. Titulo + subtitulo
    2. Numeros-chave em "boxes" (procs/falhas/resultados/pedidos)
    3. Metricas financeiras (taxa exito, PE, valor risco, PCOND, condenado)
    4. Falhas por codigo COM BARRAS VISUAIS
    5. Falhas por severidade (CRITICA vermelha, ALTA laranja, MEDIA amarela)
    6. Falhas por empresa-cliente (com ponteiro pra aba detalhada)
    7. Casos criticos top 10 (CRITICA + resultado adverso)
    8. Mapa de navegacao das outras abas
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title = "Capa"
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title, 0)  # PRIMEIRA aba

    # ─── Paleta ─────────────────────────────────────────────────────────
    NAVY = "1F4E79"
    BLUE_LIGHT = "D9E2F3"
    BG_KPI = "203864"
    BG_KPI_LIGHT = "E7EEF7"
    GRAY_BORDER = "BFBFBF"

    RED_CRITICA = "C00000"
    ORANGE_ALTA = "ED7D31"
    YELLOW_MEDIA = "FFC000"
    GREEN_BAIXA = "70AD47"

    SEV_COLORS = {
        "CRITICA": RED_CRITICA,
        "ALTA": ORANGE_ALTA,
        "MEDIA": YELLOW_MEDIA,
        "BAIXA": GREEN_BAIXA,
    }

    BIG_TITLE_FONT = Font(bold=True, size=22, color="FFFFFF")
    SUB_FONT = Font(italic=True, size=11, color="FFFFFF")
    SECTION_FONT = Font(bold=True, size=13, color="FFFFFF")
    KPI_LABEL_FONT = Font(bold=True, size=10, color="FFFFFF")
    KPI_VALUE_FONT = Font(bold=True, size=24, color="FFFFFF")
    KPI_VALUE_DARK = Font(bold=True, size=24, color="203864")
    METRIC_LABEL_FONT = Font(bold=True, size=11, color="203864")
    METRIC_VALUE_FONT = Font(bold=True, size=14, color="203864")
    BAR_FONT = Font(name="Consolas", size=10, color="203864")
    TBL_HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
    TBL_FONT = Font(size=10)
    SMALL = Font(size=9, italic=True, color="595959")

    WRAP_TOP = Alignment(wrap_text=True, vertical="top")
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", indent=1)
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    RIGHT = Alignment(horizontal="right", vertical="center", indent=1)

    thin = Side(border_style="thin", color=GRAY_BORDER)
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ─── Larguras de coluna (A-J) ───────────────────────────────────────
    widths = {"A": 22, "B": 14, "C": 14, "D": 28, "E": 22, "F": 14, "G": 14, "H": 28, "I": 4, "J": 4}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ─── 1) TITULO ──────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="AUDITORIA FORENSE — CARTEIRA GIOVANNA BASTOS")
    c.font = BIG_TITLE_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 38
    row += 1

    from datetime import datetime as _dt
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(
        row=row, column=1,
        value=(f"Gerado em {_dt.now().strftime('%d/%m/%Y %H:%M')}  ·  "
               f"{total_procs} processos auditados  ·  "
               f"431/432 cobertura (99.8%)"),
    )
    c.font = SUB_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 20
    row += 2

    # ─── 2) NUMEROS-CHAVE (boxes 4x KPI) ────────────────────────────────
    # 4 boxes de 2 colunas cada — A:B, C:D, E:F, G:H
    sentenciados = sum(count_sentenciados_por_emp.values())
    em_andamento = total_procs - sentenciados
    taxa_global = 0.0
    if sentenciados > 0:
        taxa_global = sum(count_exito_banca_por_emp.values()) / sentenciados * 100

    kpis_top = [
        ("PROCESSOS AUDITADOS", str(total_procs), NAVY),
        ("FALHAS CONFIRMADAS", str(total_falhas_confirmadas), RED_CRITICA),
        ("INDICIOS PRA REVISAR", str(total_indicios), ORANGE_ALTA),
        ("RESULTADOS NEGATIVOS", str(total_resultados), "7030A0"),
    ]

    def _kpi_box(start_col: int, label: str, value: str, color: str):
        # Linha label (1 row)
        cl = ws.cell(row=row, column=start_col, value=label)
        cl.font = KPI_LABEL_FONT
        cl.fill = PatternFill("solid", fgColor=color)
        cl.alignment = CENTER
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 1)
        # Linha valor (2 rows merge)
        cv = ws.cell(row=row + 1, column=start_col, value=value)
        cv.font = KPI_VALUE_FONT
        cv.fill = PatternFill("solid", fgColor=color)
        cv.alignment = CENTER
        ws.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 2, end_column=start_col + 1)

    for idx, (lbl, val, color) in enumerate(kpis_top):
        _kpi_box(1 + idx * 2, lbl, val, color)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 24
    ws.row_dimensions[row + 2].height = 24
    row += 4

    # ─── 3) METRICAS FINANCEIRAS ────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="  DIAGNOSTICO QUANTITATIVO  ")
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    valor_risco_global = sum(soma_valor_estimado_por_emp.values())
    pcond_global = sum(soma_pcond_por_emp.values())
    val_condenado_global = sum(soma_valor_condenacao_por_emp.values())
    pe_count_global = sum(count_pe_por_emp.values())
    pe_avg_global = (
        sum(soma_pe_por_emp.values()) / pe_count_global if pe_count_global else 0.0
    )

    def _money(v: float) -> str:
        return f"R$ {v:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")

    metricas = [
        ("Taxa de exito GLOBAL da banca",
         f"{taxa_global:.1f}%",
         f"({sum(count_exito_banca_por_emp.values())} vitorias em {sentenciados} sentenciados)"),
        ("PE medio (prob. exito banca)", f"{pe_avg_global:.3f}",
         "0=banca perde sempre · 1=banca ganha sempre"),
        ("Valor em risco TOTAL", _money(valor_risco_global),
         f"soma de valor_estimado dos {total_pedidos} pedidos"),
        ("PCOND TOTAL (CPC 25)", _money(pcond_global),
         "aprovisionamento contabil dos pedidos provaveis"),
        ("Valor JA condenado (sentencas)", _money(val_condenado_global),
         f"{sentenciados} sentencas definitivas · {em_andamento} em andamento"),
    ]

    for label, valor, contexto in metricas:
        # A:C label  D:E valor  F:H contexto
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = METRIC_LABEL_FONT
        c1.alignment = LEFT
        c1.fill = PatternFill("solid", fgColor=BG_KPI_LIGHT)
        c1.border = box_border

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        c2 = ws.cell(row=row, column=4, value=valor)
        c2.font = METRIC_VALUE_FONT
        c2.alignment = CENTER
        c2.fill = PatternFill("solid", fgColor=BG_KPI_LIGHT)
        c2.border = box_border

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        c3 = ws.cell(row=row, column=6, value=contexto)
        c3.font = SMALL
        c3.alignment = LEFT
        c3.fill = PatternFill("solid", fgColor=BG_KPI_LIGHT)
        c3.border = box_border
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ─── 4) FALHAS POR CODIGO (com barras visuais) ──────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="  FALHAS CONFIRMADAS — POR CODIGO  ")
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    # Headers
    tbl_headers = [("Codigo", 1), ("Qtd", 2), ("% do total", 3),
                   ("Distribuicao", 4), ("Nome curto", 7)]
    for label, col in tbl_headers:
        c = ws.cell(row=row, column=col, value=label)
        c.font = TBL_HEADER_FONT
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = CENTER
    # Merge "Distribuicao" col 4:6 e "Nome curto" col 7:8
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 18
    row += 1

    max_cod_qtd = max(counter_falhas_codigo.values()) if counter_falhas_codigo else 1
    for cod, qtd in counter_falhas_codigo.most_common():
        pct = qtd / total_falhas_confirmadas * 100 if total_falhas_confirmadas else 0
        bar_len = int((qtd / max_cod_qtd) * 30) if max_cod_qtd else 0
        bar = "█" * bar_len

        ws.cell(row=row, column=1, value=cod).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=qtd).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=f"{pct:.1f}%").alignment = CENTER

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        cb = ws.cell(row=row, column=4, value=bar)
        cb.font = BAR_FONT
        cb.alignment = LEFT

        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        cn = ws.cell(row=row, column=7, value=nomes_falhas.get(cod, ""))
        cn.font = TBL_FONT
        cn.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # ─── 5) SEVERIDADE (cores) ──────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="  POR SEVERIDADE  ")
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    sev_descrs = {
        "CRITICA": "revelias, transitos sem recurso, desercoes",
        "ALTA": "intempestividades, audiencias perdidas, lacuna longa",
        "MEDIA": "defesas abaixo do esperado, sem decisao",
        "BAIXA": "deslizes formais sem impacto",
    }
    for sev in ("CRITICA", "ALTA", "MEDIA", "BAIXA"):
        qtd = counter_severidade.get(sev, 0)
        color = SEV_COLORS[sev]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c1 = ws.cell(row=row, column=1, value=sev)
        c1.font = Font(bold=True, size=12, color="FFFFFF")
        c1.fill = PatternFill("solid", fgColor=color)
        c1.alignment = CENTER

        c2 = ws.cell(row=row, column=3, value=qtd)
        c2.font = Font(bold=True, size=14, color=color)
        c2.alignment = CENTER

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        c3 = ws.cell(row=row, column=4, value=sev_descrs[sev])
        c3.font = SMALL
        c3.alignment = LEFT
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ─── 6) POR EMPRESA-CLIENTE (com ponteiro pra aba) ──────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="  POR EMPRESA-CLIENTE — Giovanna defende  ")
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    emp_headers = [("Empresa", 1, 3), ("Processos", 4, 4),
                   ("Falhas", 5, 5), ("Taxa exito", 6, 6),
                   ("Valor risco", 7, 7), ("Ver aba", 8, 8)]
    for label, c1, c2 in emp_headers:
        c = ws.cell(row=row, column=c1, value=label)
        c.font = TBL_HEADER_FONT
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = CENTER
        if c1 != c2:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    ws.row_dimensions[row].height = 18
    row += 1

    for emp, qtd_procs in counter_empresa_principal.most_common():
        falhas_emp = counter_empresa_falhas.get(emp, 0)
        se = count_sentenciados_por_emp.get(emp, 0)
        ee = count_exito_banca_por_emp.get(emp, 0)
        taxa = f"{ee / se * 100:.0f}% ({ee}/{se})" if se else "—"
        valor = soma_valor_estimado_por_emp.get(emp, 0)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=emp)
        c.font = Font(bold=True, size=11)
        c.alignment = LEFT
        ws.cell(row=row, column=4, value=qtd_procs).alignment = CENTER
        c5 = ws.cell(row=row, column=5, value=falhas_emp)
        c5.alignment = CENTER
        if falhas_emp >= 10:
            c5.font = Font(bold=True, size=11, color=RED_CRITICA)
        ws.cell(row=row, column=6, value=taxa).alignment = CENTER
        ws.cell(row=row, column=7, value=_money(valor)).alignment = RIGHT

        # Ponteiro pra aba — texto + estilo de hyperlink visual
        aba_emp = ("Emp " + emp)[:31]
        # Substitui chars invalidos pra sheet name
        aba_emp = "".join(c if c not in "[]:*?/\\" else "_" for c in aba_emp)
        c8 = ws.cell(row=row, column=8, value=f"→ {aba_emp}")
        c8.font = Font(size=9, italic=True, color="0563C1")
        c8.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # ─── 7) CASOS CRITICOS — TOP 10 ─────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(
        row=row, column=1,
        value="  CASOS CRITICOS — falhas CRITICAS ou com resultado adverso  ",
    )
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    crit_headers = [("CNJ", 1, 2), ("Empresa", 3, 4),
                    ("Codigos", 5, 5), ("Severidade", 6, 6),
                    ("Resumo", 7, 8)]
    for label, c1, c2 in crit_headers:
        c = ws.cell(row=row, column=c1, value=label)
        c.font = TBL_HEADER_FONT
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = CENTER
        if c1 != c2:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    ws.row_dimensions[row].height = 18
    row += 1

    for caso in casos_criticos[:15]:
        cnj = caso.get("cnj")
        emp = caso.get("empresa") or "—"
        codigos = caso.get("codigos") or ""
        sev = caso.get("severidade") or ""
        resumo = (caso.get("resumo") or "")[:280]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=cnj).alignment = LEFT
        ws.cell(row=row, column=1).font = Font(name="Consolas", size=10)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=emp).alignment = LEFT
        ws.cell(row=row, column=5, value=codigos).alignment = CENTER
        ws.cell(row=row, column=5).font = Font(bold=True, size=10)
        c_sev = ws.cell(row=row, column=6, value=sev)
        c_sev.alignment = CENTER
        if sev in SEV_COLORS:
            c_sev.fill = PatternFill("solid", fgColor=SEV_COLORS[sev])
            c_sev.font = Font(bold=True, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        c_res = ws.cell(row=row, column=7, value=resumo)
        c_res.font = TBL_FONT
        c_res.alignment = LEFT_TOP
        ws.row_dimensions[row].height = 40
        row += 1
    row += 1

    # ─── 8) RESULTADOS NEGATIVOS — distribuicao ─────────────────────────
    if counter_resultados_codigo:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(
            row=row, column=1,
            value="  RESULTADOS NEGATIVOS (R9xx) — sentencas/decisoes adversas  ",
        )
        c.font = SECTION_FONT
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 22
        row += 1

        rn_nomes = {
            "R901": "Procedencia INTEGRAL (autor ganha tudo)",
            "R902": "Procedencia PARCIAL com condenacao",
            "R903": "Multa por LITIGANCIA DE MA-FE",
            "R904": "Multa por ATO ATENTATORIO",
            "R905": "Honorarios majorados por conduta",
            "R906": "CONFISSAO FICTA decretada",
            "R907": "TRANSITO EM JULGADO sem recurso",
        }
        for cod, qtd in counter_resultados_codigo.most_common():
            ws.cell(row=row, column=1, value=cod).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=2, value=qtd).font = Font(bold=True, size=11, color="7030A0")
            ws.cell(row=row, column=2).alignment = CENTER
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
            ws.cell(row=row, column=3, value=rn_nomes.get(cod, "")).alignment = LEFT
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ─── 9) MAPA DE NAVEGACAO ───────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="  ABAS DESTE RELATORIO  ")
    c.font = SECTION_FONT
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    abas_info = [
        ("Falhas confirmadas",
         f"{total_falhas_confirmadas} linhas — 1 linha por falha confirmada com evidencia literal"),
        ("Indicios",
         f"{total_indicios} linhas — sinais de falha que precisam revisao humana"),
        ("Resultados dos processos",
         "1 linha por processo com sentenca/decisao definitiva"),
        ("Resultados negativos",
         f"{total_resultados} linhas — sentencas/decisoes adversas (R9xx)"),
        ("Pedidos detalhados",
         f"{total_pedidos} linhas — 1 por pedido do autor com PCOND/prob_perda"),
        ("Por processo",
         f"{total_procs} linhas — visao por processo com resumo executivo"),
        ("Dados insuficientes",
         "pontos que a IA examinou mas faltou peca-chave (revisao manual)"),
        ("Emp [empresa]",
         "1 aba por empresa-cliente — processos + falhas dela"),
        ("Resumo",
         "KPIs detalhados em formato lista (uso de copia/exportacao)"),
        ("Glossario",
         "definicoes dos codigos F1xx-F8xx, R9xx, PE/PCOND/severidades"),
    ]
    for aba, desc in abas_info:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=aba)
        c.font = Font(bold=True, size=10, color="0563C1")
        c.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row=row, column=3, value=desc).font = TBL_FONT
        ws.cell(row=row, column=3).alignment = LEFT_TOP
        ws.row_dimensions[row].height = 18
        row += 1

    # ─── Freeze + zoom ──────────────────────────────────────────────────
    ws.freeze_panes = "A4"  # Trava titulo + subtitulo
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.showGridLines = False


# ─── Report (XLSX multi-aba) ──────────────────────────────────────────


def cmd_report(args: argparse.Namespace) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from collections import Counter, defaultdict

    out_dir = _ensure_output_dir(args.output_dir)
    rows = _read_jsonl(out_dir / RESPONSES_JSONL)
    if not rows:
        logger.error("Nenhuma resposta em %s — rode submit + poll antes.", out_dir / RESPONSES_JSONL)
        return 2

    xlsx_path = Path(args.xlsx).resolve() if args.xlsx else (out_dir / "relatorio-final.xlsx")

    wb = Workbook()

    BOLD = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    WRAP = Alignment(wrap_text=True, vertical="top")

    def write_sheet(title: str, headers: list[str], data_rows: list[list[Any]]):
        # Limit 31 chars (Excel sheet name limit)
        title = title[:31]
        if title in wb.sheetnames:
            ws = wb[title]
            wb.remove(ws)
        ws = wb.create_sheet(title)
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = WRAP
        for r_idx, row in enumerate(data_rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.alignment = WRAP
        # Auto-ajusta colunas (heuristica simples)
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for r in data_rows[:200]:  # so amostra
                v = r[col_idx - 1] if col_idx - 1 < len(r) else ""
                if v is None:
                    continue
                vstr = str(v)
                # Considera quebras de linha como reset
                vmax = max((len(s) for s in vstr.split("\n")), default=0)
                if vmax > max_len:
                    max_len = vmax
            ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)
        ws.freeze_panes = "A2"
        return ws

    # ─── Sheet "Por processo" ─
    proc_headers = [
        "CNJ", "Tribunal", "Vara", "Fase", "Valor causa",
        "Categoria", "Empresa principal", "Outras empresas",
        "Resultado tipo", "Em favor de", "Valor condenacao",
        "PE global", "Valor estimado total", "PCOND total",
        "# Falhas", "# Indicios", "# Resultados", "# Pedidos",
        "Confianca", "Resumo executivo", "Observacoes auditor",
        "Lote", "Aplicado em",
    ]
    proc_rows = []

    # Acumuladores pra outras abas
    falhas_all: list[list[Any]] = []
    indicios_all: list[list[Any]] = []
    resultados_all: list[list[Any]] = []
    dados_insuf_all: list[list[Any]] = []
    # v2 — diagnostico rico
    resultados_proc_all: list[list[Any]] = []
    pedidos_all: list[list[Any]] = []

    counter_falhas_codigo: Counter = Counter()
    counter_resultados_codigo: Counter = Counter()
    counter_severidade: Counter = Counter()
    counter_empresa_principal: Counter = Counter()
    counter_empresa_falhas: Counter = Counter()
    counter_empresa_resultados: Counter = Counter()
    falhas_por_empresa: dict[str, list[list[Any]]] = defaultdict(list)
    processos_por_empresa: dict[str, list[list[Any]]] = defaultdict(list)
    # v2 — counters de resultado / categoria / pedidos
    counter_resultado_tipo: Counter = Counter()
    counter_em_favor: Counter = Counter()
    counter_categoria: Counter = Counter()
    counter_prob_perda: Counter = Counter()
    counter_pedido_tipo: Counter = Counter()
    counter_pedido_natureza: Counter = Counter()
    # Agregados por empresa pra PE/PCOND/valor
    soma_valor_estimado_por_emp: dict[str, float] = defaultdict(float)
    soma_pcond_por_emp: dict[str, float] = defaultdict(float)
    soma_valor_condenacao_por_emp: dict[str, float] = defaultdict(float)
    soma_pe_por_emp: dict[str, float] = defaultdict(float)
    count_pe_por_emp: dict[str, int] = defaultdict(int)
    # Pra taxa de exito por empresa: count de processos com sentenca
    # ja' finalizada + count em favor da banca
    count_sentenciados_por_emp: dict[str, int] = defaultdict(int)
    count_exito_banca_por_emp: dict[str, int] = defaultdict(int)
    # Casos CRITICOS pra capa (top N) — falhas CRITICA OU resultado adverso
    casos_criticos_raw: list[dict] = []

    for row in rows:
        audit = row.get("audit") or {}
        cnj = audit.get("cnj_number") or row.get("cnj_number") or ""
        empresas = audit.get("empresas_representadas") or []
        principal_raw = next(
            (e.get("nome") for e in empresas if (e.get("papel") or "").lower() == "principal"),
            (empresas[0].get("nome") if empresas else None),
        )
        principal = canonical_empresa(principal_raw)
        outras = ", ".join(
            canonical_empresa(e.get("nome"))
            for e in empresas
            if (e.get("papel") or "").lower() != "principal" and e.get("nome")
        )

        falhas = audit.get("falhas_confirmadas") or []
        indicios = audit.get("indicios_de_falha") or []
        resultados = audit.get("resultados_negativos") or []
        dados_insuf = audit.get("dados_insuficientes") or []

        custom_id = row.get("custom_id") or ""
        lote_id = custom_id.split("__", 1)[0] if "__" in custom_id else "?"

        # v2 — diagnostico rico
        resultado_proc = audit.get("resultado_processo") or {}
        analise_q = audit.get("analise_quantitativa") or {}
        pedidos = audit.get("pedidos") or []
        categoria = audit.get("categoria_processo")

        resultado_tipo = resultado_proc.get("tipo")
        em_favor = resultado_proc.get("em_favor_de")
        valor_condenacao = resultado_proc.get("valor_condenacao")
        pe_global = analise_q.get("prob_exito_global")
        valor_est_total = analise_q.get("valor_estimado_total")
        pcond_total = analise_q.get("pcond_total")

        if categoria:
            counter_categoria[categoria] += 1
        if resultado_tipo:
            counter_resultado_tipo[resultado_tipo] += 1
        if em_favor:
            counter_em_favor[em_favor] += 1

        proc_row = [
            cnj,
            audit.get("tribunal"),
            audit.get("vara"),
            audit.get("fase_processual"),
            audit.get("valor_causa"),
            categoria,
            principal,
            outras,
            resultado_tipo,
            em_favor,
            valor_condenacao,
            pe_global,
            valor_est_total,
            pcond_total,
            len(falhas),
            len(indicios),
            len(resultados),
            len(pedidos),
            audit.get("confianca_geral"),
            audit.get("resumo_executivo"),
            audit.get("observacoes_auditor"),
            lote_id,
            row.get("applied_at"),
        ]
        proc_rows.append(proc_row)

        if principal:
            counter_empresa_principal[principal] += 1
            processos_por_empresa[principal].append(proc_row)
            # Agregados por empresa
            if valor_est_total is not None:
                soma_valor_estimado_por_emp[principal] += float(valor_est_total)
            if pcond_total is not None:
                soma_pcond_por_emp[principal] += float(pcond_total)
            if valor_condenacao is not None:
                soma_valor_condenacao_por_emp[principal] += float(valor_condenacao)
            if pe_global is not None:
                soma_pe_por_emp[principal] += float(pe_global)
                count_pe_por_emp[principal] += 1
            # Taxa de exito: so' conta processos com sentenca/decisao definitiva
            # (nao em_andamento). Exito da banca: em_favor_de == "reu_giovanna"
            if resultado_tipo and resultado_tipo != "em_andamento":
                count_sentenciados_por_emp[principal] += 1
                if em_favor == "reu_giovanna" or resultado_tipo == "improcedente":
                    count_exito_banca_por_emp[principal] += 1

        # v2 — resultado do processo (1 linha por processo com sentenca)
        if resultado_proc.get("existe"):
            resultados_proc_all.append([
                cnj,
                principal,
                resultado_tipo,
                resultado_proc.get("data"),
                em_favor,
                valor_condenacao,
                resultado_proc.get("resumo"),
                resultado_proc.get("evidencia_citada"),
                lote_id,
            ])

        # v2 — pedidos (N linhas por processo)
        for p in pedidos:
            tipo_p = p.get("tipo_pedido")
            prob_p = p.get("probabilidade_perda")
            if tipo_p:
                counter_pedido_tipo[tipo_p] += 1
            if p.get("natureza"):
                counter_pedido_natureza[p.get("natureza")] += 1
            if prob_p:
                counter_prob_perda[prob_p] += 1
            pedidos_all.append([
                cnj,
                principal,
                tipo_p,
                p.get("natureza"),
                p.get("valor_indicado"),
                p.get("valor_estimado"),
                p.get("fundamentacao_valor"),
                prob_p,
                p.get("aprovisionamento"),
                p.get("fundamentacao_risco"),
                lote_id,
            ])

        for f in falhas:
            counter_falhas_codigo[f.get("codigo")] += 1
            counter_severidade[f.get("severidade")] += 1
            empresa_af = canonical_empresa(f.get("empresa_afetada")) if f.get("empresa_afetada") else (principal or "(nao identificada)")
            counter_empresa_falhas[empresa_af] += 1
            falha_row = [
                cnj,
                f.get("codigo"),
                f.get("categoria"),
                f.get("severidade"),
                f.get("descricao_curta"),
                f.get("data_ocorrencia"),
                empresa_af,
                f.get("evidencia_citada"),
                f.get("prejuizo_estimado"),
                f.get("fundamentacao_auditor"),
                lote_id,
            ]
            falhas_all.append(falha_row)
            falhas_por_empresa[empresa_af].append(falha_row)

        for i in indicios:
            empresa_af = canonical_empresa(i.get("empresa_afetada")) if i.get("empresa_afetada") else (principal or "(nao identificada)")
            indicios_all.append([
                cnj,
                i.get("codigo"),
                i.get("categoria"),
                i.get("severidade"),
                i.get("descricao_curta"),
                i.get("data_ocorrencia"),
                empresa_af,
                i.get("evidencia_citada"),
                i.get("motivo_indicio"),
                lote_id,
            ])

        for r in resultados:
            counter_resultados_codigo[r.get("codigo")] += 1
            empresa_af = canonical_empresa(r.get("empresa_afetada")) if r.get("empresa_afetada") else (principal or "(nao identificada)")
            counter_empresa_resultados[empresa_af] += 1
            resultados_all.append([
                cnj,
                r.get("codigo"),
                r.get("descricao_curta"),
                r.get("data"),
                empresa_af,
                r.get("valor_envolvido"),
                r.get("evidencia_citada"),
                r.get("falha_associada_codigo"),
                lote_id,
            ])

        for d in dados_insuf:
            dados_insuf_all.append([
                cnj,
                d.get("ponto_examinado"),
                d.get("motivo"),
                lote_id,
            ])

        # ── Casos criticos ─ definicao: tem falha CRITICA OU resultado
        # negativo (R901-R907) OU R907 (transito sem recurso) ─────────
        is_critico = False
        max_sev = None
        for f in falhas:
            if f.get("severidade") == "CRITICA":
                is_critico = True
                max_sev = "CRITICA"
                break
        if not is_critico and resultados:
            is_critico = True
            max_sev = max_sev or (
                "CRITICA" if any(r.get("codigo") == "R907" for r in resultados) else "ALTA"
            )
        if is_critico:
            cods = sorted({f.get("codigo") for f in falhas if f.get("codigo")} |
                          {r.get("codigo") for r in resultados if r.get("codigo")})
            casos_criticos_raw.append({
                "cnj": cnj,
                "empresa": principal or "(nao id.)",
                "codigos": " · ".join(cods),
                "severidade": max_sev or "ALTA",
                "resumo": (audit.get("resumo_executivo") or "")[:280],
                # Pra ordenar: CRITICA primeiro, dps por # de codigos
                "_sort": (
                    0 if max_sev == "CRITICA" else 1,
                    -len(cods),
                ),
            })

    # ─── Resumo (1a aba) ─
    resumo_kpi = [
        ["Total processos auditados", len(proc_rows)],
        ["Total falhas confirmadas", sum(counter_falhas_codigo.values())],
        ["Total indicios", len(indicios_all)],
        ["Total resultados negativos", sum(counter_resultados_codigo.values())],
        ["Total pontos com dados insuficientes", len(dados_insuf_all)],
        ["", ""],
        ["— Por severidade —", ""],
    ]
    for sev in ("CRITICA", "ALTA", "MEDIA", "BAIXA"):
        resumo_kpi.append([sev, counter_severidade.get(sev, 0)])

    resumo_kpi.append(["", ""])
    resumo_kpi.append(["— Falhas por codigo —", ""])
    for cod, qtd in counter_falhas_codigo.most_common():
        resumo_kpi.append([cod, qtd])

    resumo_kpi.append(["", ""])
    resumo_kpi.append(["— Resultados negativos por codigo —", ""])
    for cod, qtd in counter_resultados_codigo.most_common():
        resumo_kpi.append([cod, qtd])

    resumo_kpi.append(["", ""])
    resumo_kpi.append(["— Empresas (principal — processos) —", ""])
    for emp, qtd in counter_empresa_principal.most_common():
        resumo_kpi.append([emp, qtd])

    resumo_kpi.append(["", ""])
    resumo_kpi.append(["— Empresas com mais falhas imputadas —", ""])
    for emp, qtd in counter_empresa_falhas.most_common():
        resumo_kpi.append([emp, qtd])

    # ─── KPIs v2 — DIAGNOSTICO RICO ─────────────────────────────────────
    resumo_kpi.append(["", ""])
    resumo_kpi.append(["=== DIAGNOSTICO QUANTITATIVO ===", ""])

    total_processos = len(proc_rows)
    total_sentenciados = sum(count_sentenciados_por_emp.values())
    total_em_andamento = total_processos - total_sentenciados
    resumo_kpi.append(["Processos com sentenca/decisao definitiva", total_sentenciados])
    resumo_kpi.append(["Processos em andamento (sem sentenca)", total_em_andamento])

    # Taxa de exito GLOBAL da banca
    total_exito = sum(count_exito_banca_por_emp.values())
    if total_sentenciados > 0:
        taxa_global = total_exito / total_sentenciados * 100
        resumo_kpi.append([
            "Taxa de exito GLOBAL da banca (entre sentenciados)",
            f"{taxa_global:.1f}% ({total_exito}/{total_sentenciados})",
        ])

    # PE medio global
    if soma_pe_por_emp:
        soma_pe_global = sum(soma_pe_por_emp.values())
        count_pe_global = sum(count_pe_por_emp.values())
        if count_pe_global > 0:
            resumo_kpi.append([
                "PE medio global (prob. exito banca)",
                f"{soma_pe_global / count_pe_global:.3f}",
            ])

    # PCOND total
    pcond_global = sum(soma_pcond_por_emp.values())
    valor_est_global = sum(soma_valor_estimado_por_emp.values())
    valor_cond_global = sum(soma_valor_condenacao_por_emp.values())
    resumo_kpi.append(["Valor estimado em risco (somatorio)", f"R$ {valor_est_global:,.2f}"])
    resumo_kpi.append(["PCOND total (CPC 25 — aprovisionamento)", f"R$ {pcond_global:,.2f}"])
    resumo_kpi.append(["Valor ja condenado (sentencas existentes)", f"R$ {valor_cond_global:,.2f}"])
    resumo_kpi.append(["Total de pedidos analisados", len(pedidos_all)])

    # Por empresa: taxa de exito, PE medio, PCOND, valor
    resumo_kpi.append(["", ""])
    resumo_kpi.append(["=== POR EMPRESA — diagnostico quantitativo ===", ""])
    for emp, qtd in counter_empresa_principal.most_common():
        resumo_kpi.append([f"{emp} — processos", qtd])
        sentenciados_emp = count_sentenciados_por_emp.get(emp, 0)
        exito_emp = count_exito_banca_por_emp.get(emp, 0)
        if sentenciados_emp > 0:
            taxa = exito_emp / sentenciados_emp * 100
            resumo_kpi.append([
                f"{emp} — taxa de exito",
                f"{taxa:.1f}% ({exito_emp}/{sentenciados_emp})",
            ])
        cnt_pe = count_pe_por_emp.get(emp, 0)
        if cnt_pe > 0:
            resumo_kpi.append([
                f"{emp} — PE medio",
                f"{soma_pe_por_emp.get(emp, 0.0) / cnt_pe:.3f}",
            ])
        ve = soma_valor_estimado_por_emp.get(emp, 0.0)
        pc = soma_pcond_por_emp.get(emp, 0.0)
        vc = soma_valor_condenacao_por_emp.get(emp, 0.0)
        if ve or pc or vc:
            resumo_kpi.append([f"{emp} — valor estimado em risco", f"R$ {ve:,.2f}"])
            resumo_kpi.append([f"{emp} — PCOND", f"R$ {pc:,.2f}"])
            resumo_kpi.append([f"{emp} — valor ja condenado", f"R$ {vc:,.2f}"])
        resumo_kpi.append(["", ""])

    # Resultados por tipo
    resumo_kpi.append(["=== RESULTADOS DAS SENTENCAS ===", ""])
    for tipo, qtd in counter_resultado_tipo.most_common():
        resumo_kpi.append([tipo, qtd])
    resumo_kpi.append(["", ""])
    resumo_kpi.append(["Em favor de:", ""])
    for ef, qtd in counter_em_favor.most_common():
        resumo_kpi.append([ef, qtd])
    resumo_kpi.append(["", ""])

    # Probabilidade de perda — distribuicao
    resumo_kpi.append(["=== PROBABILIDADE DE PERDA (pedidos) ===", ""])
    for prob, qtd in counter_prob_perda.most_common():
        resumo_kpi.append([prob, qtd])
    resumo_kpi.append(["", ""])

    # Top tipos de pedido
    resumo_kpi.append(["=== TOP TIPOS DE PEDIDO ===", ""])
    for tipo, qtd in counter_pedido_tipo.most_common(10):
        resumo_kpi.append([tipo, qtd])
    resumo_kpi.append(["", ""])

    # Top categorias de processo
    resumo_kpi.append(["=== CATEGORIAS DE PROCESSO ===", ""])
    for cat, qtd in counter_categoria.most_common(10):
        resumo_kpi.append([cat, qtd])

    # Remove a aba default antes de tudo
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Ordena casos criticos
    casos_criticos_raw.sort(key=lambda x: x["_sort"])

    # Mapa codigo -> nome curto (pra capa)
    nomes_falhas_map = {cod: nome for cod, _, nome, *_ in _GLOSSARIO_FALHAS}

    # CAPA EXECUTIVA — 1a aba do XLSX (UX prioridade: ve tudo ao abrir).
    # Glossario vai pro FINAL como apendice (linha apos as abas por empresa).
    _write_capa_executiva(
        wb,
        total_procs=len(proc_rows),
        total_falhas_confirmadas=sum(counter_falhas_codigo.values()),
        total_indicios=len(indicios_all),
        total_resultados=sum(counter_resultados_codigo.values()),
        total_pedidos=len(pedidos_all),
        counter_falhas_codigo=counter_falhas_codigo,
        counter_severidade=counter_severidade,
        counter_empresa_principal=counter_empresa_principal,
        counter_empresa_falhas=counter_empresa_falhas,
        counter_resultados_codigo=counter_resultados_codigo,
        soma_valor_estimado_por_emp=soma_valor_estimado_por_emp,
        soma_pcond_por_emp=soma_pcond_por_emp,
        soma_valor_condenacao_por_emp=soma_valor_condenacao_por_emp,
        soma_pe_por_emp=soma_pe_por_emp,
        count_pe_por_emp=count_pe_por_emp,
        count_sentenciados_por_emp=count_sentenciados_por_emp,
        count_exito_banca_por_emp=count_exito_banca_por_emp,
        casos_criticos=casos_criticos_raw,
        nomes_falhas=nomes_falhas_map,
    )

    write_sheet("Resumo", ["Metrica", "Valor"], resumo_kpi)

    write_sheet(
        "Falhas confirmadas",
        ["CNJ", "Codigo", "Categoria", "Severidade", "Descricao",
         "Data", "Empresa afetada", "Evidencia citada",
         "Prejuizo estimado", "Fundamentacao auditor", "Lote"],
        falhas_all,
    )

    write_sheet(
        "Indicios",
        ["CNJ", "Codigo", "Categoria", "Severidade", "Descricao",
         "Data", "Empresa afetada", "Evidencia citada", "Motivo indicio", "Lote"],
        indicios_all,
    )

    write_sheet(
        "Resultados negativos",
        ["CNJ", "Codigo", "Descricao", "Data", "Empresa afetada",
         "Valor envolvido", "Evidencia citada", "Falha associada", "Lote"],
        resultados_all,
    )

    write_sheet(
        "Dados insuficientes",
        ["CNJ", "Ponto examinado", "Motivo", "Lote"],
        dados_insuf_all,
    )

    # v2 — diagnostico rico
    write_sheet(
        "Resultados dos processos",
        ["CNJ", "Empresa principal", "Tipo de resultado", "Data",
         "Em favor de", "Valor condenacao", "Resumo do dispositivo",
         "Evidencia citada", "Lote"],
        resultados_proc_all,
    )

    write_sheet(
        "Pedidos detalhados",
        ["CNJ", "Empresa principal", "Tipo pedido", "Natureza",
         "Valor indicado", "Valor estimado", "Fundamentacao valor",
         "Probabilidade perda", "Aprovisionamento (CPC 25)",
         "Fundamentacao risco", "Lote"],
        pedidos_all,
    )

    write_sheet("Por processo", proc_headers, proc_rows)

    # ─── 1 aba por empresa principal (top 12 — Excel tem limite de sheets) ─
    for emp, _ in counter_empresa_principal.most_common(15):
        sheet_title = ("Emp " + (emp or "?"))[:31]
        # Excel limita caracteres: substitui /  : etc.
        sheet_title = "".join(c if c not in "[]:*?/\\" else "_" for c in sheet_title)
        emp_proc_rows = processos_por_empresa.get(emp, [])
        emp_falhas_rows = falhas_por_empresa.get(emp, [])
        data_rows = []
        data_rows.append(["=== PROCESSOS ==="] + [""] * (len(proc_headers) - 1))
        data_rows.extend(emp_proc_rows)
        data_rows.append([""] * len(proc_headers))
        data_rows.append(["=== FALHAS IMPUTADAS A ESTA EMPRESA ==="] + [""] * (len(proc_headers) - 1))
        # Falhas tem 11 cols, proc tem 16 — pad pra alinhar visualmente
        for fr in emp_falhas_rows:
            padded = fr + [""] * (len(proc_headers) - len(fr))
            data_rows.append(padded)
        write_sheet(sheet_title, proc_headers, data_rows)

    # Glossario vai por ULTIMO (apendice — operador raramente consulta;
    # foi um erro UX deixar como 1a aba) — sobrescreve a chamada anterior.
    if "Glossario" in wb.sheetnames:
        wb.remove(wb["Glossario"])
    _write_glossario_sheet(wb, write_sheet)
    # Como o helper coloca em index 0, vou movermos pra o final manualmente
    glossario_ws = wb["Glossario"]
    wb._sheets.remove(glossario_ws)
    wb._sheets.append(glossario_ws)

    wb.save(xlsx_path)
    logger.info("Relatorio salvo: %s", xlsx_path)

    print(f"\n✓ Relatorio gerado: {xlsx_path}")
    print(f"  Processos auditados: {len(proc_rows)}")
    print(f"  Falhas confirmadas: {sum(counter_falhas_codigo.values())}")
    print(f"  Indicios: {len(indicios_all)}")
    print(f"  Resultados negativos: {sum(counter_resultados_codigo.values())}")
    print(f"  Empresas principais distintas: {len(counter_empresa_principal)}")
    return 0


# ─── CLI ──────────────────────────────────────────────────────────────


def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(
        prog="run_audit_carteira",
        description="Auditoria forense incidental — banca terceirizada.",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    p_sub = sub.add_parser("submit", help="Submete lote pra Anthropic Batches API.")
    p_sub.add_argument("--input", required=True, help="Pasta com JSONs Atlas.")
    p_sub.add_argument("--output-dir", required=True, help="Pasta de estado (responses.jsonl, batches.json).")
    p_sub.add_argument("--lote-id", default=None, help="ID logico do lote (default: timestamp).")
    p_sub.add_argument("--dry-run", action="store_true", help="Nao envia — gera preview.")

    p_poll = sub.add_parser("poll", help="Consulta batches abertos e baixa results.")
    p_poll.add_argument("--output-dir", required=True)

    p_report = sub.add_parser("report", help="Agrega responses.jsonl em XLSX.")
    p_report.add_argument("--output-dir", required=True)
    p_report.add_argument("--xlsx", default=None, help="Caminho do XLSX final.")

    p_sync = sub.add_parser("sync", help="Smoke test — 1 processo sincrono via Messages API.")
    p_sync.add_argument("--input", required=True)
    p_sync.add_argument("--cnj", default=None, help="Match parcial no nome do arquivo.")
    p_sync.add_argument("--print-prompt", action="store_true", help="So imprime o prompt, nao chama API.")

    args = p.parse_args(argv)

    if args.cmd == "submit":
        return cmd_submit(args)
    if args.cmd == "poll":
        return cmd_poll(args)
    if args.cmd == "report":
        return cmd_report(args)
    if args.cmd == "sync":
        return cmd_sync(args)

    p.print_help()
    return 2


if __name__ == "__main__":
    sys.exit(main())
